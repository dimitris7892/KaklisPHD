import numpy as np
import pandas as pd
from scipy.stats import ks_2samp,chisquare,chi2_contingency
import math
#import seaborn as sns
#from pyproj import Proj, transform
import pyearth as sp
import json
from datetime import date
import matplotlib
from sklearn.cluster import KMeans
from sklearn.model_selection import TimeSeriesSplit
from scipy.spatial import distance as dis
from sklearn.model_selection import train_test_split
import pyearth as SplineRegression
from sklearn.metrics import mean_absolute_error
from matplotlib.lines import Line2D
from decimal import Decimal
from scipy import stats
import random
from numpy import inf
#from coordinates.converter import CoordinateConverter, WGS84, L_Est97
import pyodbc
import csv
import locale
#locale.setlocale(locale.LC_NUMERIC, "en_DK.UTF-8")
import datetime
import cx_Oracle
from dateutil.rrule import rrule, DAILY, MINUTELY
from sympy.solvers import solve
from sympy import Symbol
from pathlib import Path
import itertools
from sympy import cos, sin , tan , exp , sqrt , E
from openpyxl import load_workbook
import glob, os
from pathlib import Path
from openpyxl.styles.colors import YELLOW
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import shutil
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
#import seaborn as sns
from openpyxl.drawing.image import Image
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression
from tensorflow import keras
from global_land_mask import globe
import Danaos_ML_Project.dataModeling as dModel

dm = dModel.BasePartitionModeler()
currModeler = keras.models.load_model('./DeployedModels/estimatorCl_Gen.h5')

class BaseProfileGenerator:

    def ConvertMSToBeaufort(self, ws):
        wsB = 0
        if ws >= 0 and ws < 0.2:
            wsB = 0
        elif ws >= 0.3 and ws < 1.5:
            wsB = 1
        elif ws >= 1.6 and ws < 3.3:
            wsB = 2
        elif ws >= 3.4 and ws < 5.4:
            wsB = 3
        elif ws >= 5.5 and ws < 7.9:
            wsB = 4

        elif ws >= 8 and ws < 10.7:
            wsB = 5
        elif ws >= 10.8 and ws < 13.8:
            wsB = 6
        elif ws >= 13.9 and ws < 17.1:
            wsB = 7
        elif ws >= 17.2 and ws < 20.7:
            wsB = 8
        elif ws >= 20.8 and ws < 24.4:
            wsB = 9
        elif ws >= 24.5 and ws < 28.4:
            wsB = 10
        elif ws >= 28.5 and ws < 32.6:
            wsB = 11
        elif ws >= 32.7:
            wsB = 12
        return wsB

    def calculateExcelStatistics(self, workbook, dtNew, velocities, draft, trim, velocitiesTlg, rawData, company,
                                 vessel, tlgDataset, period):

        if period == 'all':
            sheet = 6
        elif period == 'bdd':
            sheet = 5
        else:
            sheet = 6
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        ####STATISTICS TAB
        ##SPEED CELLS
        velocitiesTlg = np.array([k for k in tlgDataset if float(k[12]) < 27])[:, 12]
        velocitiesTlg = np.nan_to_num(velocitiesTlg.astype(float))
        velocitiesTlg = velocitiesTlg[velocitiesTlg > 0]
        velocitiesTlgAmount = velocitiesTlg.__len__()
        rowsAmount = 56
        maxSpeed = np.ceil(np.max(velocitiesTlg))
        minSpeed = np.floor(np.min(velocitiesTlg))
        i = minSpeed
        speedsApp = []
        k = 0
        workbook._sheets[sheet].row_dimensions[1].height = 35
        idh = 1

        workbook._sheets[sheet]['C' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh + 2)] = np.round(np.min(velocitiesTlg), 2)
        workbook._sheets[sheet]['C' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')

        workbook._sheets[sheet]['E' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['E' + str(idh + 2)] = np.round(np.max(velocitiesTlg), 2)
        workbook._sheets[sheet]['E' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')

        while i < maxSpeed:
            workbook._sheets[sheet]['A' + str(k + 3)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + 3)].font = Font(bold=True, name='Calibri', size='10.5')

            speedsApp.append(str(np.round(
                (np.array([k for k in velocitiesTlg if k >= i and k <= i + 0.5]).__len__()) / velocitiesTlgAmount * 100,
                2)) + '%')
            workbook._sheets[sheet]['A' + str(k + 3)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + 3)].border = thin_border

            i += 0.5
            k += 1

        speedDeletedRows = 0
        if k < rowsAmount:
            # for i in range(k-1,19+1):
            workbook._sheets[sheet].delete_rows(idx=3 + k, amount=rowsAmount - k + 1)
            speedDeletedRows = rowsAmount - k + 1

        elif k > rowsAmount:
            workbook._sheets[sheet].insert_rows(idx=3 + k + 6, amount=abs(rowsAmount - k + 1))
            speedDeletedRows = -(abs(rowsAmount - k + 1))

        for i in range(0, len(speedsApp)):
            workbook._sheets[sheet]['B' + str(i + 3)] = speedsApp[i]
            workbook._sheets[sheet]['B' + str(i + 3)].alignment = Alignment(horizontal='left')

        for i in range(3, 3 + len(speedsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split("%")[0]) > 2:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                c = 0

        ##### END OF SPEED CELLS

        ##DRAFT CELLS
        draft = np.array([k for k in tlgDataset if float(k[8]) > 0 and float(k[8]) < 20])[:, 8].astype(float)
        draftAmount = draft.__len__()
        minDraft = int(np.floor(np.min(draft)))
        maxDraft = int(np.ceil(np.max(draft)))
        draftsApp = []
        k = 0
        i = minDraft
        rowsAmount = 31
        id = 70 - speedDeletedRows
        workbook._sheets[sheet].row_dimensions[id - 2].height = 35

        idh = id - 2

        workbook._sheets[sheet].merge_cells('C' + str(idh) + ':' + 'F' + str(idh))
        workbook._sheets[sheet]['C' + str(idh)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh)] = 'Min and Max draft discovered'
        workbook._sheets[sheet]['C' + str(idh)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh)].border = thin_border

        workbook._sheets[sheet].merge_cells('C' + str(idh + 1) + ':' + 'D' + str(idh + 1))
        workbook._sheets[sheet]['C' + str(idh + 1)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh + 1)] = 'Min'
        workbook._sheets[sheet]['C' + str(idh + 1)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh + 1)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
        workbook._sheets[sheet]['C' + str(idh + 1)].border = thin_border

        workbook._sheets[sheet].merge_cells('E' + str(idh + 1) + ':' + 'F' + str(idh + 1))
        workbook._sheets[sheet]['E' + str(idh + 1)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['E' + str(idh + 1)] = 'Max'
        workbook._sheets[sheet]['E' + str(idh + 1)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
        workbook._sheets[sheet]['E' + str(idh + 1)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['E' + str(idh + 1)].border = thin_border

        workbook._sheets[sheet].merge_cells('C' + str(idh + 2) + ':' + 'D' + str(idh + 2))
        workbook._sheets[sheet]['C' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh + 2)] = np.round(np.min(draft), 2)
        workbook._sheets[sheet]['C' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh + 2)].border = thin_border

        workbook._sheets[sheet].merge_cells('E' + str(idh + 2) + ':' + 'F' + str(idh + 2))
        workbook._sheets[sheet]['E' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['E' + str(idh + 2)] = np.round(np.max(draft), 2)
        workbook._sheets[sheet]['E' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['E' + str(idh + 2)].border = thin_border

        while i < maxDraft:
            # workbook._sheets[4].insert_rows(k+41)
            workbook._sheets[sheet]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            draftsApp.append(str(np.round((np.array(
                [k for k in tlgDataset if float(k[8]) >= i and float(k[8]) <= i + 0.5]).__len__()) / draftAmount * 100,
                                          2)) + '%')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            i += 0.5
            k += 1

        draftDeletedRows = 0
        if k < rowsAmount:
            # for i in range(k-1,19+1):
            workbook._sheets[sheet].delete_rows(idx=id + k, amount=rowsAmount - k + 1)
            draftDeletedRows = speedDeletedRows + rowsAmount - k + 1
        elif k > rowsAmount:
            workbook._sheets[sheet].insert_rows(idx=id + k, amount=rowsAmount - k + 1)
            draftDeletedRows = -(abs(speedDeletedRows + rowsAmount - k + 1))

        for i in range(0, len(draftsApp)):
            workbook._sheets[sheet]['B' + str(i + id)] = draftsApp[i]
            workbook._sheets[sheet]['B' + str(i + id)].alignment = Alignment(horizontal='left')
            height = workbook._sheets[sheet].row_dimensions[i + id].height
            if height != None:
                if float(height) > 13.8:
                    workbook._sheets[sheet].row_dimensions[i + id].height = 13.8

        for i in range(id, id + len(draftsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split('%')[0]) > 2:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                c = 0
        ##### END OF DRAFT CELSS

        ##TRIM CELLS ######################################################################
        trim = np.array([k for k in tlgDataset if float(k[16]) > -4 and float(k[16]) < 9])[:, 16].astype(float)
        trimAmount = trim.__len__()
        minTrim = int(np.floor(np.min(trim)))
        maxTrim = int(np.ceil(np.max(trim)))
        trimsApp = []
        rowsAmount = 40
        k = 0
        i = minTrim
        id = 115 - draftDeletedRows
        workbook._sheets[sheet].row_dimensions[id - 2].height = 35
        idh = id - 2

        workbook._sheets[sheet].merge_cells('C' + str(idh) + ':' + 'F' + str(idh))
        workbook._sheets[sheet]['C' + str(idh)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh)] = 'Min and Max trim discovered'
        workbook._sheets[sheet]['C' + str(idh)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh)].border = thin_border

        workbook._sheets[sheet].merge_cells('C' + str(idh + 1) + ':' + 'D' + str(idh + 1))
        workbook._sheets[sheet]['C' + str(idh + 1)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh + 1)] = 'Min'
        workbook._sheets[sheet]['C' + str(idh + 1)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
        workbook._sheets[sheet]['C' + str(idh + 1)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh + 1)].border = thin_border

        workbook._sheets[sheet].merge_cells('E' + str(idh + 1) + ':' + 'F' + str(idh + 1))
        workbook._sheets[sheet]['E' + str(idh + 1)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['E' + str(idh + 1)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
        workbook._sheets[sheet]['E' + str(idh + 1)] = 'Max'
        workbook._sheets[sheet]['E' + str(idh + 1)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['E' + str(idh + 1)].border = thin_border

        workbook._sheets[sheet].merge_cells('C' + str(idh + 2) + ':' + 'D' + str(idh + 2))
        workbook._sheets[sheet]['C' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['C' + str(idh + 2)] = np.round(np.min(trim), 2)
        workbook._sheets[sheet]['C' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['C' + str(idh + 2)].border = thin_border

        workbook._sheets[sheet].merge_cells('E' + str(idh + 2) + ':' + 'F' + str(idh + 2))
        workbook._sheets[sheet]['E' + str(idh + 2)].alignment = Alignment(horizontal='center')
        workbook._sheets[sheet]['E' + str(idh + 2)] = np.round(np.max(trim), 2)
        workbook._sheets[sheet]['E' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')
        workbook._sheets[sheet]['E' + str(idh + 2)].border = thin_border

        while i < maxTrim:
            # workbook._sheets[sheet].insert_rows(k+27)
            if i < 0:
                workbook._sheets[sheet]['A' + str(k + id)] = ' (' + str(i) + ' - ' + str(i + 0.5) + ')'
                workbook._sheets[sheet]['A' + str(k + id)].font = Font(color='ce181e', bold=True, name='Calibri',
                                                                       size='10.5')
            else:
                workbook._sheets[sheet]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
                workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            trimsApp.append(str(np.round((np.array(
                [k for k in tlgDataset if float(k[16]) >= i and float(k[16]) <= i + 0.5]).__len__() / trimAmount) * 100,
                                         2)) + '%')
            i += 0.5
            k += 1

        trimDeletedRows = 0
        if k < rowsAmount:
            # for i in range(k-1,19+1):
            workbook._sheets[sheet].delete_rows(idx=id + k, amount=rowsAmount - k + 1)
            trimDeletedRows = draftDeletedRows + rowsAmount - k + 1
        if k > rowsAmount:
            workbook._sheets[sheet].insert_rows(idx=id + k, amount=rowsAmount - k + 1)
            trimDeletedRows = -(abs(draftDeletedRows + rowsAmount - k + 1))

            # speedDeletedRows+

        for i in range(0, len(trimsApp)):
            workbook._sheets[sheet]['B' + str(i + id)] = trimsApp[i]
            workbook._sheets[sheet]['B' + str(i + id)].alignment = Alignment(horizontal='left')
            height = workbook._sheets[sheet].row_dimensions[i + id].height
            if height != None:
                if float(height) > 13.8:
                    workbook._sheets[sheet].row_dimensions[i + id].height = 13.8

        for i in range(id, id + len(trimsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split('%')[0]) > 2:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                c = 0
        ##### END OF TRIM CELLS

        ##FOC TLG CELLS ######################################################################
        # foc = np.array([k for k in dtNew if float(k[16])>0])[:,16]
        foc = np.array([k for k in tlgDataset if float(k[15]) > 0 and float(k[15]) < 300 and float(k[12] )> 7 and float(k[12]) < 27])[:,15].astype(float)

        rowsAmount = 92
        focAmount = foc.__len__()
        minFOC = int(np.floor(np.min(foc)))
        maxFOC = int(np.ceil(np.max(foc)))
        focsApp = []
        meanSpeeds = []
        stdSpeeds = []
        ranges = []
        k = 0
        i = minFOC if minFOC > 0 else 1
        id = 171 - trimDeletedRows
        workbook._sheets[sheet].row_dimensions[id - 2].height = 35
        focsPLot = []
        speedsPlot = []
        workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
        foc = np.array([k for k in tlgDataset if float(k[15]) > 0 and float(k[15]) < 300 and float(k[12] )> 7 and float(k[12]) < 27])
        while i < maxFOC:
            # workbook._sheets[sheet].insert_rows(k+27)
            workbook._sheets[sheet]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 1) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border

            focArray = np.array([k for k in foc if float(k[15]) >= i and float(k[15]) <= i + 1])
            focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
            meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[sheet]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border

            if focArray.__len__() > 0:
                focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                ranges.append(i)
            i += 1
            k += 1

        xi = np.array(speedsPlot)
        yi = np.array(ranges)
        zi = np.array(focsPLot)

        plt.clf()
        # Change color with c and alpha
        p2 = np.poly1d(np.polyfit(xi, yi, 2))
        xp = np.linspace(min(xi), max(xi), 100)
        plt.plot([], [], '.', xp, p2(xp))

        plt.scatter(xi, yi, s=zi, c="red", alpha=0.4, linewidth=4)
        plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (MT / day)")
        plt.title("Density plot", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        zi = zi.reshape(-1, 1)
        dataModel.fit(zi)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0]),
                        label=str(int(np.floor(z[0]))) + ' obs.')
        plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title='# of obs')

        fig.savefig('./Figures/' + company + '_' + vessel + str(sheet) + '_1.png', dpi=96)
        # plt.clf()

        img = Image('./Figures/' + company + '_' + vessel + str(sheet) + '_1.png')

        workbook._sheets[sheet].add_image(img, 'F' + str(id - 2))
        # workbook._sheets[4].insert_image('G'+str(id+30), './Danaos_ML_Project/Figures/'+company+'_'+vessel+'.png', {'x_scale': 0.5, 'y_scale': 0.5})

        if k < rowsAmount:
            # for i in range(k-1,19+1):
            workbook._sheets[sheet].delete_rows(idx=id + k, amount=rowsAmount - k + 1)
            focDeletedRows = rowsAmount - k + 1

        for i in range(0, len(focsApp)):
            workbook._sheets[sheet]['B' + str(i + id)] = focsApp[i]

            workbook._sheets[sheet]['B' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['C' + str(i + id)] = meanSpeeds[i]
            workbook._sheets[sheet]['D' + str(i + id)] = stdSpeeds[i]

            workbook._sheets[sheet]['C' + str(i + id)].alignment = Alignment(horizontal='center')
            workbook._sheets[sheet]['D' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['B' + str(i + id)].border = thin_border
            workbook._sheets[sheet]['B' + str(i + id)].font = Font(name='Calibri', size='10.5')

            height = workbook._sheets[sheet].row_dimensions[i + id].height
            if height != None:
                if float(height) > 13.8:
                    workbook._sheets[sheet].row_dimensions[i + id].height = 13.8

        for i in range(id, id + len(focsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split('%')[0]) > 1.5:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['C' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['D' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                c = 0
        ##### END OF FOC TLG CELLS
        ########################################################################
        ##FOC PER MILE / SPEED RANGES / BEAUFORTS GRAPHS
        #focPerMile = np.array([k for k in tlgDataset if float(k[17]) > 0])[:, 17].astype(float)
        #steamTimeSum = np.sum(np.array([k for k in tlgDataset if float(k[17]) > 0])[:, 13].astype(float) + np.array(
            #[k for k in tlgDataset if float(k[17]) > 0])[:, 14].astype(float) / 60)
        # minsSlc = np.sum()
        # steamTimeSum = hoursSlc + minsSlc/60
        rowsAmount = 92
        '''focAmount = focPerMile.__len__()
        minFOC = int(np.floor(np.min(focPerMile)))
        maxFOC = int(np.ceil(np.max(focPerMile)))
        focsApp = []
        meanSpeeds = []
        steamTimeRange = []
        minSlc = []
        hoursSlc = []
        stdSpeeds = []
        ranges = []
        k = 0
        i = minFOC
        id = 415
        workbook._sheets[sheet].row_dimensions[id - 2].height = 35
        focsPLot = []
        speedsPlot = []
        workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
        while i < maxFOC:
            # workbook._sheets[sheet].insert_rows(k+27)
            i = np.round(i, 2)
            workbook._sheets[4]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.02) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            focArray = np.array([k for k in tlgDataset if float(k[17]) >= i and float(k[17]) <= i + 0.02])
            if focArray.__len__() > 0:
                hoursSlc.append(focArray[:, 13])
                minSlc.append(focArray[:, 14])
                steamTimeRange = np.sum(np.array(focArray[:, 13].astype(float) + focArray[:, 14].astype(float) / 60))

                focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
                meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                           2) if focArray.__len__() > 0 else 'N/A')
                stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                          2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[4]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border

            if focArray.__len__() > 0:
                focsPLot.append(np.round(steamTimeRange / steamTimeSum, 4))
                # focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                ranges.append(i)
            i += 0.02
            k += 1

        plt.clf()
        xi = np.array(speedsPlot)
        yi = np.array(ranges)
        zi = np.array(focsPLot)
        # Change color with c and alpha
        p2 = np.poly1d(np.polyfit(xi, yi, 2))
        xp = np.linspace(min(xi), max(xi), 100)
        plt.plot([], [], '.', xp, p2(xp))

        plt.scatter(xi, yi, s=zi * 2000, c="red", alpha=0.4, linewidth=4)
        plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (per / Mile)")
        plt.title("Density plot", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        zi = zi.reshape(-1, 1)
        dataModel.fit(zi)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='r', alpha=0.5, s=z[0] * 700,
                        label=str(np.round(z[0], 3)) + ' Hours')
        plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title=' Hours')

        fig.savefig('./Figures/' + company + '_' + vessel + '_3.png', dpi=96)

        img = Image('./Figures/' + company + '_' + vessel + '_3.png')

        workbook._sheets[4].add_image(img, 'F' + str(415))
        x = 0

        for i in range(0, len(focsApp)):
            workbook._sheets[sheet]['B' + str(i + id)] = focsApp[i]

            workbook._sheets[sheet]['B' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['C' + str(i + id)] = meanSpeeds[i]
            workbook._sheets[sheet]['D' + str(i + id)] = stdSpeeds[i]

            workbook._sheets[sheet]['C' + str(i + id)].alignment = Alignment(horizontal='center')
            workbook._sheets[sheet]['D' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['B' + str(i + id)].border = thin_border
            workbook._sheets[sheet]['B' + str(i + id)].font = Font(name='Calibri', size='10.5')

            height = workbook._sheets[sheet].row_dimensions[i + id].height
            if height != None:
                if float(height) > 13.8:
                    workbook._sheets[sheet].row_dimensions[i + id].height = 13.8

        for i in range(id, id + len(focsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split('%')[0]) > 1.5:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['C' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['D' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                o = 0
        ##########################FOC PER WIND RANGES / WIND ANGLE
        foc = np.array([k for k in tlgDataset if float(k[15]) > 0 and float(k[15]) < 40])[:, 15].astype(float)

        focAmount = foc.__len__()
        minFOC = int(np.floor(np.min(foc)))
        maxFOC = int(np.ceil(np.max(foc)))

        focsAppGen = []
        meanSpeedsGen = []
        stdSpeedsGen = []
        rangesGen = []
        focsPLotGen = []
        speedsPlotGen = []

        focsApp = []
        meanSpeeds = []
        stdSpeeds = []
        ranges = []
        focsPLot = []
        speedsPlot = []

        focsApp1 = []
        meanSpeeds1 = []
        stdSpeeds1 = []
        ranges1 = []
        focsPLot1 = []
        speedsPlot1 = []

        focsApp2 = []
        meanSpeeds2 = []
        stdSpeeds2 = []
        ranges2 = []
        focsPLot2 = []
        speedsPlot2 = []

        focsApp3 = []
        meanSpeeds3 = []
        stdSpeeds3 = []
        ranges3 = []
        focsPLot3 = []
        speedsPlot3 = []

        focsApp4 = []
        meanSpeeds4 = []
        stdSpeeds4 = []
        ranges4 = []
        focsPLot4 = []
        speedsPlot4 = []

        k = 0
        i = minFOC if minFOC > 0 else 1
        id = 171 - trimDeletedRows
        # workbook._sheets[sheet].row_dimensions[id - 2].height = 35

        # workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        # workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
        while i < maxFOC:
            # workbook._sheets[sheet].insert_rows(k+27)
            workbook._sheets[4]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            focArrayGen = np.array([k for k in tlgDataset if
                                    float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                        k[11]) <= 3])

            focArray = np.array([k for k in focArrayGen if
                                 float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                     k[11]) <= 3 and
                                 float(k[10]) >= 0 and float(k[10]) <= 22.5])

            focArray1 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                      k[11]) <= 3 and
                                  float(k[10]) >= 22.5 and float(k[10]) <= 67.5])

            focArray2 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                      k[11]) <= 3 and
                                  float(k[10]) >= 67.5 and float(k[10]) <= 112.5])

            focArray3 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                      k[11]) <= 3 and float(k[10]) >= 112.5 and float(k[10]) <= 157.5])

            focArray4 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 1 and float(
                                      k[11]) <= 3 and float(k[10]) >= 157.5 and float(k[10]) <= 180])

            focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
            meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[4]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border
            if focArrayGen.__len__() > 0:
                focsPLotGen.append(focArrayGen.__len__())
                speedsPlotGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 12].astype(float)))), 2))
                rangesGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 15].astype(float)))), 2))

            if focArray.__len__() > 0:
                focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                ranges.append(np.round((np.mean(np.nan_to_num(focArray[:, 15].astype(float)))), 2))

            if focArray1.__len__() > 0:
                focsPLot1.append(focArray1.__len__())
                speedsPlot1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 12].astype(float)))), 2))
                ranges1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 15].astype(float)))), 2))

            if focArray2.__len__() > 0:
                focsPLot2.append(focArray2.__len__())
                speedsPlot2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 12].astype(float)))), 2))
                ranges2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 15].astype(float)))), 2))

            if focArray3.__len__() > 0:
                focsPLot3.append(focArray3.__len__())
                speedsPlot3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 12].astype(float)))), 2))
                ranges3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 15].astype(float)))), 2))

            if focArray4.__len__() > 0:
                focsPLot4.append(focArray4.__len__())
                speedsPlot4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 12].astype(float)))), 2))
                ranges4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 15].astype(float)))), 2))

            i += 0.5
            k += 1

        plt.clf()
        xiGen = np.array(speedsPlotGen)
        yiGen = np.array(rangesGen)
        ziGen = np.array(focsPLotGen)
        # Change color with c and alpha
        p2Gen = np.poly1d(np.polyfit(xiGen, yiGen, 2))
        xpGen = np.linspace(min(xiGen), max(xiGen), 100)
        # plt.plot([], [], '.', xpGen, p2Gen(xpGen), color='black')

        # plt.scatter(xiGen, yiGen, s=ziGen , c="red", alpha=0.4, linewidth=4)
        #########################################################
        xi = np.array(speedsPlot)
        yi = np.array(ranges)
        zi = np.array(focsPLot)
        # Change color with c and alpha
        # p2 = np.poly1d(np.polyfit(xi, yi, 2))
        # xp = np.linspace(min(xi), max(xi), 100)
        # line1=plt.plot([], [], '.', xp, p2(xp),color='red')

        plt.scatter(xi, yi, s=zi * 10, c="red", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        #####################################################################
        xi1 = np.array(speedsPlot1)
        yi1 = np.array(ranges1)
        zi1 = np.array(focsPLot1)
        # Change color with c and alpha
        # p21 = np.poly1d(np.polyfit(xi1, yi1, 2))
        # xp1 = np.linspace(min(xi1), max(xi1), 100)
        # line2=plt.plot([], [], '.', xp1, p21(xp1),color='green')

        plt.scatter(xi1, yi1, s=zi1 * 10, c="green", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi1)) - 1, np.ceil(max(xi1)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi1)), np.ceil(max(yi1)) + 1, 5))
        ###################################################################################################################
        #####################################################################
        xi2 = np.array(speedsPlot2)
        yi2 = np.array(ranges2)
        zi2 = np.array(focsPLot2)
        # Change color with c and alpha
        # p22 = np.poly1d(np.polyfit(xi2, yi2, 2))
        # xp2 = np.linspace(min(xi2), max(xi2), 100)
        # plt.plot([], [], '.', xp2, p22(xp2),color='blue')
        plt.scatter(xi2, yi2, s=zi2 * 10, c="blue", alpha=0.4, linewidth=4)
        ###################################################################################################################
        xi3 = np.array(speedsPlot3)
        yi3 = np.array(ranges3)
        zi3 = np.array(focsPLot3)
        # Change color with c and alpha
        # p23 = np.poly1d(np.polyfit(xi3, yi3, 2))
        # xp3 = np.linspace(min(xi3), max(xi3), 100)
        # plt.plot([], [], '.', xp3, p23(xp3), color='orange')
        plt.scatter(xi3, yi3, s=zi3 * 10, c="orange", alpha=0.4, linewidth=4)
        ###################################################################################################################
        if speedsPlot4.__len__() > 0:
            xi4 = np.array(speedsPlot4)
            yi4 = np.array(ranges4)
            zi4 = np.array(focsPLot4)
            # Change color with c and alpha
            p24 = np.poly1d(np.polyfit(xi4, yi4, 2))
            xp4 = np.linspace(min(xi4), max(xi4), 100)
            # plt.plot([], [], '.', xp4, p24(xp4), color='purple')
            plt.scatter(xi4, yi4, s=zi4 * 10, c="purple", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi2)) - 1, np.ceil(max(xi2)) + 1, 1))
            plt.yticks(np.arange(minFOC, maxFOC + 1, 5))
        ###################################################################################################################

        ############################################
        ############################################
        ############################################
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (MT / day) ")
        plt.title("Density plot for Wind Speed (1 - 3) Beauforts", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        ziGen = ziGen.reshape(-1, 1)
        dataModel.fit(ziGen)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='grey', alpha=0.5, s=np.floor(z[0] * 10),
                        label=str(int(np.floor(z[0]))) + ' obs.')
        legend1 = plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title='# of obs', loc='upper right')

        custom_lines = [Line2D([0], [0], color='red', lw=4),
                        Line2D([0], [0], color='green', lw=4),
                        Line2D([0], [0], color='blue', lw=4),
                        Line2D([0], [0], color='orange', lw=4),
                        Line2D([0], [0], color='purple', lw=4)]

        plt.legend(custom_lines, ['(0 - 22.5)', '(22.5 - 67.5)', '(67.5 - 112.5)', '(112.5 - 157.5)', '(157.5 - 180)'],
                   title='Wind Dir', loc='upper left')

        plt.gca().add_artist(legend1)

        fig.savefig('./Figures/' + company + '_' + vessel + str(sheet) + '_4.png', dpi=96)

        # plt.clf()

        img = Image('./Figures/' + company + '_' + vessel + str(sheet) + '_4.png')
        workbook._sheets[sheet].add_image(img, 'A' + str(193))
        x = 0
        ###BEAFUORTS (3-5) #############################################################
        focsAppGen = []
        meanSpeedsGen = []
        stdSpeedsGen = []
        rangesGen = []
        focsPLotGen = []
        speedsPlotGen = []

        focsApp = []
        meanSpeeds = []
        stdSpeeds = []
        ranges = []
        focsPLot = []
        speedsPlot = []

        focsApp1 = []
        meanSpeeds1 = []
        stdSpeeds1 = []
        ranges1 = []
        focsPLot1 = []
        speedsPlot1 = []

        focsApp2 = []
        meanSpeeds2 = []
        stdSpeeds2 = []
        ranges2 = []
        focsPLot2 = []
        speedsPlot2 = []

        focsApp3 = []
        meanSpeeds3 = []
        stdSpeeds3 = []
        ranges3 = []
        focsPLot3 = []
        speedsPlot3 = []

        focsApp4 = []
        meanSpeeds4 = []
        stdSpeeds4 = []
        ranges4 = []
        focsPLot4 = []
        speedsPlot4 = []

        k = 0
        i = minFOC if minFOC > 0 else 1
        id = 171 - trimDeletedRows
        # workbook._sheets[sheet].row_dimensions[id - 2].height = 35'''

        # workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        # workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
        '''while i < maxFOC:
            # workbook._sheets[sheet].insert_rows(k+27)
            workbook._sheets[4]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            focArrayGen = np.array([k for k in tlgDataset if
                                    float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                        k[11]) <= 5])

            focArray = np.array([k for k in focArrayGen if
                                 float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                     k[11]) <= 5 and
                                 float(k[10]) >= 0 and float(k[10]) <= 22.5])

            focArray1 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                      k[11]) <= 5 and
                                  float(k[10]) >= 22.5 and float(k[10]) <= 67.5])

            focArray2 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                      k[11]) <= 5 and
                                  float(k[10]) >= 67.5 and float(k[10]) <= 112.5])

            focArray3 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                      k[11]) <= 5 and float(k[10]) >= 112.5 and float(k[10]) <= 157.5])

            focArray4 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 3 and float(
                                      k[11]) <= 5 and float(k[10]) >= 157.5 and float(k[10]) <= 180])

            focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
            meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[4]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border
            if focArrayGen.__len__() > 0:
                focsPLotGen.append(focArrayGen.__len__())
                speedsPlotGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 12].astype(float)))), 2))
                rangesGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 15].astype(float)))), 2))

            if focArray.__len__() > 0:
                focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                ranges.append(np.round((np.mean(np.nan_to_num(focArray[:, 15].astype(float)))), 2))

            if focArray1.__len__() > 0:
                focsPLot1.append(focArray1.__len__())
                speedsPlot1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 12].astype(float)))), 2))
                ranges1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 15].astype(float)))), 2))

            if focArray2.__len__() > 0:
                focsPLot2.append(focArray2.__len__())
                speedsPlot2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 12].astype(float)))), 2))
                ranges2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 15].astype(float)))), 2))

            if focArray3.__len__() > 0:
                focsPLot3.append(focArray3.__len__())
                speedsPlot3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 12].astype(float)))), 2))
                ranges3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 15].astype(float)))), 2))

            if focArray4.__len__() > 0:
                focsPLot4.append(focArray4.__len__())
                speedsPlot4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 12].astype(float)))), 2))
                ranges4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 15].astype(float)))), 2))

            i += 0.5
            k += 1

        plt.clf()
        xiGen = np.array(speedsPlotGen)
        yiGen = np.array(rangesGen)
        ziGen = np.array(focsPLotGen)
        # Change color with c and alpha
        p2Gen = np.poly1d(np.polyfit(xiGen, yiGen, 2))
        xpGen = np.linspace(min(xiGen), max(xiGen), 100)
        # plt.plot([], [], '.', xpGen, p2Gen(xpGen), color='black')

        # plt.scatter(xiGen, yiGen, s=ziGen , c="red", alpha=0.4, linewidth=4)
        #########################################################
        xi = np.array(speedsPlot)
        yi = np.array(ranges)
        zi = np.array(focsPLot)
        # Change color with c and alpha
        p2 = np.poly1d(np.polyfit(xi, yi, 2))
        xp = np.linspace(min(xi), max(xi), 100)
        # line1=plt.plot([], [], '.', xp, p2(xp),color='red')

        plt.scatter(xi, yi, s=zi * 10, c="red", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        #####################################################################
        xi1 = np.array(speedsPlot1)
        yi1 = np.array(ranges1)
        zi1 = np.array(focsPLot1)
        # Change color with c and alpha
        p21 = np.poly1d(np.polyfit(xi1, yi1, 2))
        xp1 = np.linspace(min(xi1), max(xi1), 100)
        # line2=plt.plot([], [], '.', xp1, p21(xp1),color='green')

        plt.scatter(xi1, yi1, s=zi1 * 10, c="green", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi1)) - 1, np.ceil(max(xi1)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi1)), np.ceil(max(yi1)) + 1, 5))
        ###################################################################################################################
        #####################################################################
        xi2 = np.array(speedsPlot2)
        yi2 = np.array(ranges2)
        zi2 = np.array(focsPLot2)
        # Change color with c and alpha
        p22 = np.poly1d(np.polyfit(xi2, yi2, 2))
        xp2 = np.linspace(min(xi2), max(xi2), 100)
        # plt.plot([], [], '.', xp2, p22(xp2),color='blue')
        plt.scatter(xi2, yi2, s=zi2 * 10, c="blue", alpha=0.4, linewidth=4)
        ###################################################################################################################
        xi3 = np.array(speedsPlot3)
        yi3 = np.array(ranges3)
        zi3 = np.array(focsPLot3)
        # Change color with c and alpha
        p23 = np.poly1d(np.polyfit(xi3, yi3, 2))
        xp3 = np.linspace(min(xi3), max(xi3), 100)
        # plt.plot([], [], '.', xp3, p23(xp3), color='orange')
        plt.scatter(xi3, yi3, s=zi3 * 10, c="orange", alpha=0.4, linewidth=4)
        ###################################################################################################################
        if speedsPlot4.__len__() > 0:
            xi4 = np.array(speedsPlot4)
            yi4 = np.array(ranges4)
            zi4 = np.array(focsPLot4)
            # Change color with c and alpha
            p24 = np.poly1d(np.polyfit(xi4, yi4, 2))
            xp4 = np.linspace(min(xi4), max(xi4), 100)
            # plt.plot([], [], '.', xp4, p24(xp4), color='purple')
            plt.scatter(xi4, yi4, s=zi4 * 10, c="purple", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi2)) - 1, np.ceil(max(xi2)) + 1, 1))
            plt.yticks(np.arange(minFOC, maxFOC + 1, 5))
        ###################################################################################################################

        ############################################
        ############################################
        ############################################
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (MT / day) ")
        plt.title("Density plot for Wind Speed (3 - 5) Beauforts", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        ziGen = ziGen.reshape(-1, 1)
        dataModel.fit(ziGen)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='grey', alpha=0.5, s=np.floor(z[0] * 10),
                        label=str(int(np.floor(z[0]))) + ' obs.')
        legend1 = plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title='# of obs', loc='upper right')

        custom_lines = [Line2D([0], [0], color='red', lw=4),
                        Line2D([0], [0], color='green', lw=4),
                        Line2D([0], [0], color='blue', lw=4),
                        Line2D([0], [0], color='orange', lw=4),
                        Line2D([0], [0], color='purple', lw=4)]

        plt.legend(custom_lines, ['(0 - 22.5)', '(22.5 - 67.5)', '(67.5 - 112.5)', '(112.5 - 157.5)', '(157.5 - 180)'],
                   title='Wind Dir', loc='upper left')

        plt.gca().add_artist(legend1)

        fig.savefig('./Figures/' + company + '_' + vessel + str(sheet) + '_5.png', dpi=96)

        # plt.clf()

        img = Image('./Figures/' + company + '_' + vessel + str(sheet) + '_5.png')
        workbook._sheets[sheet].add_image(img, 'T' + str(193))'''
        ##################### BEAUFORTS (5-8)
        ###BEAFUORTS (3-5) #############################################################
        '''focsAppGen = []
        meanSpeedsGen = []
        stdSpeedsGen = []
        rangesGen = []
        focsPLotGen = []
        speedsPlotGen = []

        focsApp = []
        meanSpeeds = []
        stdSpeeds = []
        ranges = []
        focsPLot = []
        speedsPlot = []

        focsApp1 = []
        meanSpeeds1 = []
        stdSpeeds1 = []
        ranges1 = []
        focsPLot1 = []
        speedsPlot1 = []

        focsApp2 = []
        meanSpeeds2 = []
        stdSpeeds2 = []
        ranges2 = []
        focsPLot2 = []
        speedsPlot2 = []

        focsApp3 = []
        meanSpeeds3 = []
        stdSpeeds3 = []
        ranges3 = []
        focsPLot3 = []
        speedsPlot3 = []

        focsApp4 = []
        meanSpeeds4 = []
        stdSpeeds4 = []
        ranges4 = []
        focsPLot4 = []
        speedsPlot4 = []

        k = 0
        i = minFOC if minFOC > 0 else 1
        id = 171 - trimDeletedRows
        # workbook._sheets[sheet].row_dimensions[id - 2].height = 35

        # workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        # workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
        while i < maxFOC:
            # workbook._sheets[sheet].insert_rows(k+27)
            workbook._sheets[4]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border
            focArrayGen = np.array([k for k in tlgDataset if
                                    float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                        k[11]) <= 9])

            focArray = np.array([k for k in focArrayGen if
                                 float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                     k[11]) <= 9 and
                                 float(k[10]) >= 0 and float(k[10]) <= 22.5])

            focArray1 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                      k[11]) <= 9 and
                                  float(k[10]) >= 22.5 and float(k[10]) <= 67.5])

            focArray2 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                      k[11]) <= 9 and
                                  float(k[10]) >= 67.5 and float(k[10]) <= 112.5])

            focArray3 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                      k[11]) <= 9 and float(k[10]) >= 112.5 and float(k[10]) <= 157.5])

            focArray4 = np.array([k for k in focArrayGen if
                                  float(k[15]) >= i and float(k[15]) <= i + 0.5 and float(k[11]) >= 5 and float(
                                      k[11]) <= 9 and float(k[10]) >= 157.5 and float(k[10]) <= 180])

            focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
            meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[4]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border
            if focArrayGen.__len__() > 0:
                focsPLotGen.append(focArrayGen.__len__())
                speedsPlotGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 12].astype(float)))), 2))
                rangesGen.append(np.round((np.mean(np.nan_to_num(focArrayGen[:, 15].astype(float)))), 2))

            if focArray.__len__() > 0:
                focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                ranges.append(np.round((np.mean(np.nan_to_num(focArray[:, 15].astype(float)))), 2))

            if focArray1.__len__() > 0:
                focsPLot1.append(focArray1.__len__())
                speedsPlot1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 12].astype(float)))), 2))
                ranges1.append(np.round((np.mean(np.nan_to_num(focArray1[:, 15].astype(float)))), 2))

            if focArray2.__len__() > 0:
                focsPLot2.append(focArray2.__len__())
                speedsPlot2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 12].astype(float)))), 2))
                ranges2.append(np.round((np.mean(np.nan_to_num(focArray2[:, 15].astype(float)))), 2))

            if focArray3.__len__() > 0:
                focsPLot3.append(focArray3.__len__())
                speedsPlot3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 12].astype(float)))), 2))
                ranges3.append(np.round((np.mean(np.nan_to_num(focArray3[:, 15].astype(float)))), 2))

            if focArray4.__len__() > 0:
                focsPLot4.append(focArray4.__len__())
                speedsPlot4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 12].astype(float)))), 2))
                ranges4.append(np.round((np.mean(np.nan_to_num(focArray4[:, 15].astype(float)))), 2))

            i += 0.5
            k += 1

        plt.clf()
        xiGen = np.array(speedsPlotGen)
        yiGen = np.array(rangesGen)
        ziGen = np.array(focsPLotGen)
        # Change color with c and alpha
        p2Gen = np.poly1d(np.polyfit(xiGen, yiGen, 2))
        xpGen = np.linspace(min(xiGen), max(xiGen), 100)
        # plt.plot([], [], '.', xpGen, p2Gen(xpGen), color='black')

        # plt.scatter(xiGen, yiGen, s=ziGen , c="red", alpha=0.4, linewidth=4)
        #########################################################
        xi = np.array(speedsPlot)
        yi = np.array(ranges)
        zi = np.array(focsPLot)
        # Change color with c and alpha
        p2 = np.poly1d(np.polyfit(xi, yi, 2))
        xp = np.linspace(min(xi), max(xi), 100)
        # line1=plt.plot([], [], '.', xp, p2(xp),color='red')

        plt.scatter(xi, yi, s=zi * 10, c="red", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        #####################################################################
        xi1 = np.array(speedsPlot1)
        yi1 = np.array(ranges1)
        zi1 = np.array(focsPLot1)
        # Change color with c and alpha
        p21 = np.poly1d(np.polyfit(xi1, yi1, 2))
        xp1 = np.linspace(min(xi1), max(xi1), 100)
        # line2=plt.plot([], [], '.', xp1, p21(xp1),color='green')

        plt.scatter(xi1, yi1, s=zi1 * 10, c="green", alpha=0.4, linewidth=4)
        # plt.xticks(np.arange(np.floor(min(xi1)) - 1, np.ceil(max(xi1)) + 1, 1))
        # plt.yticks(np.arange(np.floor(min(yi1)), np.ceil(max(yi1)) + 1, 5))
        ###################################################################################################################
        #####################################################################
        xi2 = np.array(speedsPlot2)
        yi2 = np.array(ranges2)
        zi2 = np.array(focsPLot2)
        # Change color with c and alpha
        p22 = np.poly1d(np.polyfit(xi2, yi2, 2))
        xp2 = np.linspace(min(xi2), max(xi2), 100)
        # plt.plot([], [], '.', xp2, p22(xp2),color='blue')
        plt.scatter(xi2, yi2, s=zi2 * 10, c="blue", alpha=0.4, linewidth=4)
        ###################################################################################################################
        if speedsPlot3.__len__() > 0:
            xi3 = np.array(speedsPlot3)
            yi3 = np.array(ranges3)
            zi3 = np.array(focsPLot3)
            # Change color with c and alpha
            p23 = np.poly1d(np.polyfit(xi3, yi3, 2))
            xp3 = np.linspace(min(xi3), max(xi3), 100)
            # plt.plot([], [], '.', xp3, p23(xp3), color='orange')
            plt.scatter(xi3, yi3, s=zi3 * 10, c="orange", alpha=0.4, linewidth=4)
        ###################################################################################################################
        if speedsPlot4.__len__() > 0:
            xi4 = np.array(speedsPlot4)
            yi4 = np.array(ranges4)
            zi4 = np.array(focsPLot4)
            # Change color with c and alpha
            p24 = np.poly1d(np.polyfit(xi4, yi4, 2))
            xp4 = np.linspace(min(xi4), max(xi4), 100)
            # plt.plot([], [], '.', xp4, p24(xp4), color='purple')
            plt.scatter(xi4, yi4, s=zi4 * 10, c="purple", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi2)) - 1, np.ceil(max(xi2)) + 1, 1))
            plt.yticks(np.arange(minFOC, maxFOC + 1, 5))
        ###################################################################################################################

        ############################################
        ############################################
        ############################################
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (MT / day) ")
        plt.title("Density plot for Wind Speed (5 - 8) Beauforts", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        ziGen = ziGen.reshape(-1, 1)
        dataModel.fit(ziGen)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='grey', alpha=0.5, s=np.floor(z[0] * 10),
                        label=str(int(np.floor(z[0]))) + ' obs.')
        legend1 = plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title='# of obs', loc='upper right')

        custom_lines = [Line2D([0], [0], color='red', lw=4),
                        Line2D([0], [0], color='green', lw=4),
                        Line2D([0], [0], color='blue', lw=4),
                        Line2D([0], [0], color='orange', lw=4),
                        Line2D([0], [0], color='purple', lw=4)]

        plt.legend(custom_lines, ['(0 - 22.5)', '(22.5 - 67.5)', '(67.5 - 112.5)', '(112.5 - 157.5)', '(157.5 - 180)'],
                   title='Wind Dir', loc='upper left')

        plt.gca().add_artist(legend1)

        fig.savefig('./Figures/' + company + '_' + vessel + str(sheet) + '_6.png', dpi=96)

        plt.clf()

        img = Image('./Figures/' + company + '_' + vessel + str(sheet) + '_6.png')
        workbook._sheets[sheet].add_image(img, 'H' + str(248))'''
        if rawData:
            ##############START OF RAW TAB ####################################################################
            ##############START OF RAW TAB ####################################################################
            ###START OF STW RAW #####################
            ##SPEED  RAW CELLS
            rowsAmount = 56
            velocities = np.array([k for k in dtNew])[:, 12]
            velocities = np.nan_to_num(velocities.astype(float))
            velocities = velocities[velocities >= 0]
            velocitiesAmount = velocities.__len__()
            maxSpeed = np.ceil(np.max(velocities))
            minSpeed = np.floor(np.min(velocities))
            speedsApp = []
            id = 3
            k = 0
            i = minSpeed
            idh = 1

            workbook._sheets[5]['C' + str(idh + 2)].alignment = Alignment(horizontal='center')
            workbook._sheets[5]['C' + str(idh + 2)] = np.round(np.min(velocities), 2)
            workbook._sheets[5]['C' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')

            workbook._sheets[5]['E' + str(idh + 2)].alignment = Alignment(horizontal='center')
            workbook._sheets[5]['E' + str(idh + 2)] = np.round(np.max(velocities), 2)
            workbook._sheets[5]['E' + str(idh + 2)].font = Font(bold=True, name='Calibri', size='10.5')

            workbook._sheets[5].row_dimensions[1].height = 35
            focPerMile = []
            while i < maxSpeed:
                workbook._sheets[5]['A' + str(k + 3)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
                workbook._sheets[5]['A' + str(k + 3)].font = Font(bold=True, name='Calibri', size='10.5')

                meanFoc = np.array([k for k in dtNew if float(k[12]) >= i and float(k[12]) <= i + 0.5])
                if meanFoc.__len__() > 0:
                    meanSpeed = i + 0.5
                    focPerMile.append(np.round(np.mean(meanFoc[:, 15]) / (meanSpeed), 2))
                speedsApp.append(str(np.round((np.array([k for k in dtNew if float(k[12]) >= i and float(
                    k[12]) <= i + 0.5]).__len__()) / velocitiesAmount * 100, 2)) + '%')
                i += 0.5
                k += 1

            speedDeletedRows = 0
            if k < rowsAmount:
                # for i in range(k-1,19+1):
                workbook._sheets[5].delete_rows(idx=3 + k, amount=rowsAmount - k + 1)
                speedDeletedRows = rowsAmount - k + 1
            for i in range(0, len(focPerMile)):
                workbook._sheets[5]['G' + str(i + 4)] = focPerMile[i]

            for i in range(0, len(speedsApp)):
                workbook._sheets[5]['B' + str(i + 3)] = speedsApp[i]
                workbook._sheets[5]['B' + str(k + 3)].alignment = Alignment(horizontal='left')
                height = workbook._sheets[5].row_dimensions[i + 3].height
                if height != None:
                    if float(height) > 13.8:
                        workbook._sheets[5].row_dimensions[i + 3].height = 13.8

            for i in range(3, 3 + len(speedsApp)):
                try:
                    if float(workbook._sheets[5]['B' + str(i)].value) > 500:
                        workbook._sheets[5]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                        workbook._sheets[5]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                except:
                    c = 0

            for i in range(id, id + len(speedsApp)):
                if float(str(workbook._sheets[5]['B' + str(i)].value).split('%')[0]) > 1.5:
                    workbook._sheets[5]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[5]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")

            ##### END OF SPEED RAW CELLS

            ##FOC RAW CELLS ######################################################################
            foc = np.array([k for k in dtNew if float(k[15]) > 0 and float(k[15] < 45)])[:, 15]
            rowsAmount = 92
            focAmount = foc.__len__()
            minFOC = int(np.floor(np.min(foc)))
            maxFOC = int(np.ceil(np.max(foc)))
            focsApp = []
            k = 0
            i = minFOC
            id = 171 - speedDeletedRows
            focsPLot = []
            speedsPlot = []
            focPerMile = []
            ranges = []
            meanSpeeds = []
            stdSpeeds = []
            workbook._sheets[5].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
            workbook._sheets[5]['C' + str(id - 2)].alignment = Alignment(horizontal='center')
            workbook._sheets[5].row_dimensions[id - 2].height = 35

            while i < maxFOC:
                workbook._sheets[5]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
                workbook._sheets[5]['A' + str(k + id)].alignment = Alignment(horizontal='center')

                workbook._sheets[5]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
                workbook._sheets[5]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
                workbook._sheets[5]['A' + str(k + id)].border = thin_border
                focArray = np.array([k for k in dtNew if float(k[15]) >= i and float(k[15]) <= i + 0.5])
                focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
                meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))),
                                           2) if focArray.__len__() > 0 else 'N/A')
                stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 12].astype(float)))),
                                          2) if focArray.__len__() > 0 else 'N/A')
                meanFoc = np.mean(focArray[:, 15])
                meanSpeed = np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2)
                focPerMile.append(np.round(meanFoc / (meanSpeed * 24), 2))
                workbook._sheets[5]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
                workbook._sheets[5]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
                workbook._sheets[5]['C' + str(k + id)].border = thin_border

                workbook._sheets[5]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
                workbook._sheets[5]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
                workbook._sheets[5]['D' + str(k + id)].border = thin_border

                if focArray.__len__() > 0:
                    focsPLot.append(focArray.__len__())
                    speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 12].astype(float)))), 2))
                    ranges.append(i)
                i += 0.5
                k += 1

            if k < rowsAmount:
                # for i in range(k-1,19+1):
                workbook._sheets[5].delete_rows(idx=id + k, amount=rowsAmount - k + 1)
                focDeletedRows = rowsAmount - k + 1

            for i in range(0, len(focsApp)):

                workbook._sheets[5]['E' + str(i + id)] = focPerMile[i]

                workbook._sheets[5]['B' + str(i + id)] = focsApp[i]

                workbook._sheets[5]['B' + str(i + id)].alignment = Alignment(horizontal='center')

                workbook._sheets[5]['C' + str(i + id)] = meanSpeeds[i]
                workbook._sheets[5]['D' + str(i + id)] = stdSpeeds[i]

                workbook._sheets[5]['C' + str(i + id)].alignment = Alignment(horizontal='center')
                workbook._sheets[5]['D' + str(i + id)].alignment = Alignment(horizontal='center')

                workbook._sheets[5]['B' + str(i + id)].border = thin_border
                workbook._sheets[5]['B' + str(i + id)].font = Font(name='Calibri', size='10.5')
                height = workbook._sheets[5].row_dimensions[i + id].height
                if height != None:
                    if float(height) > 13.8:
                        workbook._sheets[5].row_dimensions[i + id].height = 13.8

            for i in range(id, id + len(focsApp)):
                try:
                    if float(str(workbook._sheets[5]['B' + str(i)].value).split('%')[0]) > 1.5:
                        workbook._sheets[5]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                        workbook._sheets[5]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                        workbook._sheets[5]['C' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                        workbook._sheets[5]['D' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                except:
                    c = 0

            plt.clf()

            fig = matplotlib.pyplot.gcf()
            fig.set_size_inches(17.5, 9.5)

            xi = np.array(speedsPlot)
            yi = np.array(ranges)
            zi = np.array(focsPLot)
            # Change color with c and alpha
            p2 = np.poly1d(np.polyfit(xi, yi, 3))
            xp = np.linspace(min(xi), max(xi), 100)
            plt.plot([], [], '.', xp, p2(xp))

            plt.scatter(xi, yi, s=zi, c="red", alpha=0.4, linewidth=4)
            plt.xticks(np.arange(np.floor(min(xi)), np.ceil(max(xi)) + 1, 1))
            plt.yticks(np.arange(min(yi), max(yi) + 1, 5))
            plt.xlabel("Speed (knots)")
            plt.ylabel("FOC (MT / day)")
            plt.title("Density plot", loc="center")

            dataModel = KMeans(n_clusters=3)
            zi = zi.reshape(-1, 1)
            dataModel.fit(zi)
            # Extract centroid values
            centroids = dataModel.cluster_centers_
            ziSorted = np.sort(centroids, axis=0)

            for z in ziSorted:
                plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0]),
                            label='       ' + str(int(np.floor(z[0]))) + ' obs.')
            plt.legend(borderpad=4, scatterpoints=1, frameon=True, labelspacing=6, title='# of obs')

            fig.savefig('./Figures/' + company + '_' + vessel + '_2.png', dpi=96)
            # plt.clf()
            img = Image('./Figures/' + company + '_' + vessel + '_2.png')

            workbook._sheets[5].add_image(img, 'F' + str(id - 2))
            ##### END OF FOC RAW CELLS

        ########################################################## END OF STATISTICS TAB #########################################
        return workbook

    def fillPORTSCANALSsheet(self, workbook, dtNewPC ,json_decoded, company,vessel):

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        sheet = 4
        workbook._sheets[sheet]['B2'] = np.round(np.mean(dtNewPC[:,1]),2)
        ##FOC CANALS PORTS CELLS ######################################################################
        # foc = np.array([k for k in dtNew if float(k[16])>0])[:,16]
        foc = dtNewPC[:,8]
        speed = dtNewPC[:,5]
        id = 11
        rowsAmount = 92
        focAmount = foc.__len__()
        minFOC = int(np.floor(np.min(foc)))
        maxFOC = int(np.ceil(np.max(foc)))
        focsApp = []
        meanSpeeds = []
        stdSpeeds = []
        meanFocs = []
        stdFocs = []
        ranges = []
        k = 0

        minSpeed = int(np.min(speed))
        maxSpeed = int(np.max(speed))

        workbook._sheets[sheet].row_dimensions[id - 2].height = 35
        focsPLot = []
        speedsPlot = []
        avgFocsPlot = [ ]
        workbook._sheets[sheet].merge_cells('C' + str(id - 2) + ':' + 'D' + str(id - 2))
        workbook._sheets[sheet]['C' + str(id - 2)].alignment = Alignment(horizontal='center')

        foc = dtNewPC

        i = minSpeed

        while i < maxSpeed:
            outerItem = {"draft": np.round(np.mean(dtNewPC[:, 1]), 2), "speed": i, "cells": []}
            # workbook._sheets[sheet].insert_rows(k+27)
            workbook._sheets[sheet]['A' + str(k + id)] = ' (' + str(i) + '-' + str(i + 0.5) + ')'
            workbook._sheets[sheet]['A' + str(k + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['A' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['A' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['A' + str(k + id)].border = thin_border

            focArray = np.array([k for k in foc if float(k[5]) >= i-0.25 and float(k[5]) <= i + 0.25])
            focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
            meanSpeeds.append(np.round((np.mean(np.nan_to_num(focArray[:, 5].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdSpeeds.append(np.round((np.std(np.nan_to_num(focArray[:, 5].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            meanFocs.append(np.round((np.mean(np.nan_to_num(focArray[:, 8].astype(float)))),
                                       2) if focArray.__len__() > 0 else 'N/A')
            stdFocs.append(np.round((np.std(np.nan_to_num(focArray[:, 8].astype(float)))),
                                      2) if focArray.__len__() > 0 else 'N/A')

            workbook._sheets[sheet]['C' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['C' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['C' + str(k + id)].border = thin_border

            workbook._sheets[sheet]['D' + str(k + id)].font = Font(bold=True, name='Calibri', size='10.5')
            workbook._sheets[sheet]['D' + str(k + id)].fill = PatternFill(fgColor='dae3f3', fill_type="solid")
            workbook._sheets[sheet]['D' + str(k + id)].border = thin_border

            if focArray.__len__() > 0:
                focsPLot.append(focArray.__len__())
                speedsPlot.append(np.round((np.mean(np.nan_to_num(focArray[:, 5].astype(float)))), 2))
                avgFocsPlot.append(np.mean(focArray[:,8]))

                windSpeed = np.mean(focArray[:,4])
                windDir = np.mean(focArray[:, 3])
                swh = np.mean(focArray[:, 13])

                item = {"windBFT": windSpeed, "windDir": windDir, "swell": swh, "cons": np.mean(focArray[:,8])}
                outerItem['cells'].append(item)

                ranges.append(i)
            json_decoded['ConsumptionProfile']['consProfilePORTS'].append(outerItem)
            i += 0.5
            k += 1





        xi = np.array(speedsPlot)
        yi = np.array(avgFocsPlot)
        zi = np.array(focsPLot)

        plt.clf()
        # Change color with c and alpha
        p2 = np.poly1d(np.polyfit(xi, yi, 2))
        xp = np.linspace(min(xi), max(xi), 100)
        plt.plot([], [], '.', xp, p2(xp))

        plt.scatter(xi, yi, s=zi, c="red", alpha=0.4, linewidth=4)
        plt.xticks(np.arange(np.floor(min(xi)) - 1, np.ceil(max(xi)) + 1, 1))
        plt.yticks(np.arange(np.floor(min(yi)), np.ceil(max(yi)) + 1, 5))
        plt.xlabel("Speed (knots)")
        plt.ylabel("FOC (MT / day)")
        plt.title("Density plot", loc="center")
        fig = matplotlib.pyplot.gcf()
        fig.set_size_inches(17.5, 9.5)

        dataModel = KMeans(n_clusters=3)
        zi = zi.reshape(-1, 1)
        dataModel.fit(zi)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        ziSorted = np.sort(centroids, axis=0)

        for z in ziSorted:
            plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0]),
                        label=str(int(np.floor(z[0]))) + ' obs.')
        plt.legend(scatterpoints=1, frameon=True, labelspacing=2, title='# of obs')

        #fig.savefig('./Figures/' + company + '_' + vessel + str(sheet) + '_1.png', dpi=96)
        # plt.clf()

        img = Image('./Figures/' + company + '_' + vessel + str(sheet) + '_1.png')

        workbook._sheets[sheet].add_image(img, 'F' + str(id - 2))
        # workbook._sheets[4].insert_image('G'+str(id+30), './Danaos_ML_Project/Figures/'+company+'_'+vessel+'.png', {'x_scale': 0.5, 'y_scale': 0.5})

        if k < rowsAmount:
            # for i in range(k-1,19+1):
            workbook._sheets[sheet].delete_rows(idx=id + k, amount=rowsAmount - k + 1)
            focDeletedRows = rowsAmount - k + 1

        for i in range(0, len(focsApp)):
            workbook._sheets[sheet]['B' + str(i + id)] = focsApp[i]

            workbook._sheets[sheet]['B' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['C' + str(i + id)] = meanFocs[i]
            workbook._sheets[sheet]['D' + str(i + id)] = stdFocs[i]

            workbook._sheets[sheet]['C' + str(i + id)].alignment = Alignment(horizontal='center')
            workbook._sheets[sheet]['D' + str(i + id)].alignment = Alignment(horizontal='center')

            workbook._sheets[sheet]['B' + str(i + id)].border = thin_border
            workbook._sheets[sheet]['B' + str(i + id)].font = Font(name='Calibri', size='10.5')

            height = workbook._sheets[sheet].row_dimensions[i + id].height
            if height != None:
                if float(height) > 13.8:
                    workbook._sheets[sheet].row_dimensions[i + id].height = 13.8

        for i in range(id, id + len(focsApp)):
            try:
                if float(str(workbook._sheets[4]['B' + str(i)].value).split('%')[0]) > 1.5:
                    workbook._sheets[sheet]['B' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['A' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['C' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
                    workbook._sheets[sheet]['D' + str(i)].fill = PatternFill(fgColor=YELLOW, fill_type="solid")
            except:
                c = 0
        ##### END OF FOC TLG CELLS

        return json_decoded
    def fillExcelProfCons(self, company, vessel, pathToexcel, dataSet, rawData, tlgDataset, dataSetBDD, dataSetADD):
        ##FEATURE SET EXACT POSITION OF COLUMNS NEEDED IN ORDER TO PRODUCE EXCEL
        # 2nd place BALLAST FLAG
        # 8th place DRAFT
        # 10th place WD
        # 11th place WF
        # 12th place SPEED
        # 15th place ME FOC 24H
        # 16th place ME FOC 24H TLGS
        # 17th place TRIM
        # 19th place SteamHours
        # 18th place STW_TLG

        # if float(k[5])>6.5
        wd = np.array([k for k in dataSet])[:, 10]
        for i in range(0, len(wd)):
            if wd[i] > 180:
                wd[i] = wd[i] - 180  # and  float(k[8])<20
        dataSet[:, 10] = wd
        lenConditionTlg = 5
        dtNew = np.array([k for k in dataSet if float(k[15]) > 0 and float(k[12]) > 0])  # and  float(k[8])<20
        dtNewBDD = np.array([k for k in dataSetBDD if float(k[15]) > 0 and float(k[12]) > 0])
        dtNewADD = np.array([k for k in dataSetADD if float(k[15]) > 0 and float(k[12]) > 0])

        ballastDt = np.array([k for k in dtNew if k[2] == 'B' if float(k[8]) < 16])[:, 7:].astype(float)
        ladenDt = np.array([k for k in dtNew if k[2] == 'L' if float(k[8]) < 16])[:, 7:].astype(float)

        meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)
        meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        minDraftLadden = round(float(np.min(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        maxDraftBallast = round(float(np.max(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)

        for i in range(0, len(dtNew)):
            # tNew[i, 10] = self.getRelativeDirectionWeatherVessel(float(dtNew[i, 7]), float(dtNew[i, 10]))
            if str(dtNew[i, 2]) == 'nan' and float(dtNew[i, 8]) > 0:
                if float(dtNew[i, 8]) >= meanDraftLadden:
                    dtNew[i, 2] = 'L'
                else:
                    # if float(dtNew[i, 8]) <=meanDraftBallast+1:
                    dtNew[i, 2] = 'B'
        ########################################################################
        ########################################################################
        ########################################################################

        ballastDt = np.array([k for k in dtNew if k[2] == 'B' if float(k[8]) > 0])[:, 7:].astype(float)
        ladenDt = np.array([k for k in dtNew if k[2] == 'L' if float(k[8]) > 0])[:, 7:].astype(float)

        meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)
        meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        minDraftLadden = round(float(np.min(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        maxDraftBallast = round(float(np.max(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)

        draft = (np.array((np.array([k for k in dtNew if float(k[8]) > 0 and float(k[8]) < 20])[:, 8])).astype(float))
        trim = (np.array((np.array([k for k in dtNew if float(k[17]) < 20])[:, 17])).astype(float))
        velocities = (
            np.array((np.array([k for k in dtNew if float(k[12]) > 0])[:, 12])).astype(float))  # and float(k[12]) < 18
        velocitiesTlg = (
            np.array((np.array([k for k in dtNew if float(k[18]) > 0 and float(k[18]) < 35])[:, 12])).astype(float))
        '''if tlgDataset==[]:
              tlgDataset = dtNew
              tlgDatasetBDD = dtNewBDD
              tlgDatasetADD = dtNewADD

              #velocitiesTlgBDD = (
                  #np.array((np.array([k for k in dtNewBDD if float(k[18]) > 0 ])[:, 12])).astype(float))
              #velocitiesTlgADD = (
                  #np.array((np.array([k for k in dtNewADD if float(k[18]) > 0])[:, 12])).astype(float))
          else:
              velocitiesTlg = (np.array((np.array([k for k in dtNew if float(k[18]) > 0 and float(k[18])<35 ])[:, 18])).astype(float)) #and float(k[12]) < 18'''

        dataModel = KMeans(n_clusters=4)
        velocities = velocities.reshape(-1, 1)
        dataModel.fit(velocities)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        velocitiesSorted = np.sort(centroids, axis=0)
        ################################################################################################
        ballastDt = np.array([k for k in dtNew if k[2] == 'B'])[:, 7:].astype(float)
        ladenDt = np.array([k for k in dtNew if k[2] == 'L'])[:, 7:].astype(float)

        velocitiesB = np.array([k for k in ballastDt if k[5] > 6 and k[5] < 16])[:, 5]

        dataModel = KMeans(n_clusters=4)
        velocitiesB = velocitiesB.reshape(-1, 1)
        dataModel.fit(velocitiesB)
        labels = dataModel.predict(velocitiesB)
        # Extract centroid values

        centroidsB = dataModel.cluster_centers_
        centroidsB = np.sort(centroidsB, axis=0)
        ##LOAD EXCEL
        workbook = load_workbook(filename=pathToexcel)

        wind = [0, 22.5, 67.5, 112.5, 157.5, 180]

        ###SEA STATE
        ballastMaxSeaSate = []
        for i in range(0, len(wind) - 1):
            arrayWF = np.array(
                [k for k in ballastDt if k[3] >= wind[i] and k[3] <= wind[i + 1] and k[4] <= 9 and k[8] > 0])
            maxSS = np.max(arrayWF[:, 4]) if arrayWF.__len__() > 0 else 0
            ballastMaxSeaSate.append(round(maxSS, 2))

        for i in range(3, 8):
            workbook._sheets[3]['B' + str(i)] = ballastMaxSeaSate[i - 3]
        ####################################################
        laddenMaxSeaSate = []
        for i in range(0, len(wind) - 1):
            arrayWF = np.array(
                [k for k in ladenDt if k[3] >= wind[i] and k[3] <= wind[i + 1] and k[4] <= 9 and k[8] > 0])
            maxSS = np.max(arrayWF[:, 4]) if arrayWF.__len__() > 0 else 0
            laddenMaxSeaSate.append(round(maxSS, 2))

        for i in range(12, 17):
            workbook._sheets[3]['B' + str(i)] = laddenMaxSeaSate[i - 12]
        ###############################
        ##end of sea state
        ###
        # meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)
        workbook._sheets[2]['B2'] = np.floor(meanDraftBallast) + 0.5
        # meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        workbook._sheets[1]['B2'] = np.ceil(meanDraftLadden) + 0.5

        ##WRITE DRAFTS TO FILE
        drafts = []
        drafts.append(np.floor(meanDraftBallast) + 0.5)
        drafts.append(np.ceil(meanDraftLadden) + 0.5)
        with open('./data/' + company + '/' + vessel + '/ListOfDrafts.csv', mode='w') as data:
            data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            data_writer.writerow(
                ['Draft'])
            for i in range(0, 2):
                data_writer.writerow(
                    [drafts[i]])

        #############################################################################
        ##BALLAST
        ##delete ballast outliers
        # np.delete(ladenDt, [i for (i, v) in enumerate(ladenDt[:, 8]) if v < (
        # np.mean(ladenDt[:, 8]) - np.std(ladenDt[:, 8])) or v > np.mean(
        # ladenDt[:, 8]) + np.std(
        # ladenDt[:, 8])], 0)
        partitionsX = []
        partitionLabels = []
        # For each label
        for curLbl in np.unique(labels):
            # Create a partition for X using records with corresponding label equal to the current
            partitionsX.append(np.asarray(velocitiesB[labels == curLbl]))
            # Create a partition for Y using records with corresponding label equal to the current

            # Keep partition label to ascertain same order of results
            partitionLabels.append(curLbl)

        sorted = []
        initialX = len(partitionsX)
        initialXpart = partitionsX
        while sorted.__len__() < initialX:
            min = 100000000000
            for i in range(0, len(partitionsX)):
                mean = np.mean(partitionsX[i])
                if mean < min:
                    min = mean
                    minIndx = i
            sorted.append(partitionsX[minIndx])
            # partitionsX.remove(partitionsX[minIndx])
            partitionsX.pop(minIndx)

        ################################################################################################################

        workbook = self.calculateExcelStatistics(workbook, dtNew, velocities, draft, trim, velocitiesTlg, rawData,
                                                 company, vessel, tlgDataset, 'all')
        # workbook = self.calculateExcelStatistics(workbook, dtNewBDD, velocities, draft, trim, velocitiesTlgBDD, rawData,
        # company, vessel, tlgDatasetBDD,'bdd')
        # workbook = self.calculateExcelStatistics(workbook, dtNewADD, velocities, draft, trim, velocitiesTlgBDD, rawData,
        # company, vessel, tlgDatasetADD, 'Add')
        ##delete ladden outliers
        # np.delete(ladenDt, [i for (i, v) in enumerate(ladenDt[:, 8]) if v < (
        # np.mean(ladenDt[:, 8]) - np.std(ladenDt[:, 8])) or v > np.mean(
        # ladenDt[:, 8]) + np.std(
        # ladenDt[:, 8])], 0)

        ###SPEED 10 WIND <1.5
        vel0Min = np.floor(np.min(sorted[0])) + 0.5
        vel1Min = np.floor(np.min(sorted[1])) + 0.5
        vel2Min = np.floor(np.min(sorted[2])) + 0.5
        vel3Min = np.floor(np.min(sorted[3])) + 0.5

        vel0Max = np.floor(np.max(sorted[0])) + 0.5
        vel1Max = np.floor(np.max(sorted[1])) + 0.5
        vel2Max = np.floor(np.max(sorted[2])) + 0.5
        vel3Max = np.floor(np.max(sorted[3])) + 0.5

        vel0Mean = np.floor(np.mean(sorted[0])) + 0.5
        vel1Mean = np.floor(np.mean(sorted[1])) + 0.5
        vel2Mean = np.floor(np.mean(sorted[2])) + 0.5
        vel3Mean = np.floor(np.mean(sorted[3])) + 0.5

        ####VESSEL BASIC INFO
        blVelocities = []
        blVelocities.append(vel0Mean)
        blVelocities.append(vel1Mean)
        blVelocities.append(vel2Mean)
        blVelocities.append(vel3Mean)

        '''vel0Min = 3
          vel1Min = 12.5
          vel2Min =12.5
          vel3Min = 13.5

          vel0Max = 12
          vel1Max = 13
          vel2Max = 13.5
          vel3Max = 18'''
        # vel0Min = 6
        # vel0Max = 9

        '''workbook._sheets[2]['B6'] = str(vel0Mean) + '  ('+str(vel0Min)+' - '+str(vel0Max)+')'
          workbook._sheets[2]['B16'] = str(vel1Mean) + '  ('+str(vel1Min)+' - '+str(vel1Max)+')'
          workbook._sheets[2]['B26'] = str(vel2Mean) + '  ('+str(vel2Min)+' - '+str(vel2Max)+')'
          workbook._sheets[2]['B36'] = str(vel3Mean) + '  ('+str(vel3Min)+' - '+str(vel3Max)+')'''

        workbook._sheets[0]['B2'] = vessel
        workbook._sheets[0]['B5'] = round(velocitiesSorted[3][0], 1)
        workbook._sheets[0]['B7'] = np.min(draft)
        workbook._sheets[0]['B8'] = np.max(draft)
        ##END OF VESSEL BASIC INFO

        ballastDt10_0 = []
        numberOfApp10_0 = []

        minAccThres = 0

        FocCentral = np.array([k for k in ballastDt if
                               k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 8]

        steamTimeGen = np.array([k for k in ballastDt if
                                 k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 12]

        weighted_avgFocCentral = np.average(FocCentral, weights=steamTimeGen)

        centralMean = np.mean(np.array([k for k in ballastDt if
                                        k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 8])
        centralMean = weighted_avgFocCentral
        centralArray = np.array([k for k in ballastDt if
                                 k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 8]
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k >= 3])

            # tlgarrayFoc=np.mean(np.array([k for k in ballastDt if k[5] >= round(centroidsB[0][0], 2) and k[5] <= 10 and k[4] >= 0 and k[4] <= 1 and k[9] > 10])[:, 9])
            tlgarrayFoc = np.array([k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])

            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array([k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:,
                              9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8], weights=steamTime[:, 12
                                                                          ]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp10_0.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt10_0.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[2]['B' + str(i)] = ballastDt10_0[i - 9]

        ###SPEED 10  2 < WIND <3
        ballastDt10_3 = []
        numberOfApp10_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[
                                     8] > 5])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 5])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:

                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp10_3.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt10_3.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[2]['C' + str(i)] = ballastDt10_3[i - 9]

        ###SPEED 10  4 < WIND <5
        ballastDt10_5 = []
        numberOfApp10_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[
                                     8] > 5])
            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 5])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]

                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp10_5.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt10_5.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[2]['D' + str(i)] = ballastDt10_5[i - 9]

        ###SPEED 10  7 < WIND <8
        ballastDt10_8 = []
        numberOfApp10_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 6 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[3] <= wind[i + 1] and k[
                                     8] > 5])
            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 6 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 5])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp10_8.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt10_8.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[2]['E' + str(i)] = ballastDt10_8[i - 9]

        ##################################################################################################################
        ##################################################################################################################

        ###SPEED 11.5   WIND <1.5

        centralMean = np.mean(np.array([k for k in ballastDt if
                                        k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 8])
        centralArray = np.array([k for k in ballastDt if
                                 k[5] >= vel1Min and k[5] <= vel1Max and k[8] > 4])[:, 8]
        ballastDt11_0 = []
        numberOfApp11_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 5])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] >= vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 5])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp11_0.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt11_0.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[2]['B' + str(i)] = ballastDt11_0[i - 19]

        ###SPEED 11.5  2 < WIND <3
        ballastDt11_3 = []
        numberOfApp11_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] >= vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp11_3.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt11_3.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[2]['C' + str(i)] = ballastDt11_3[i - 19]

        ###SPEED 11.5  4 < WIND <5
        ballastDt11_5 = []
        numberOfApp11_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] >= vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp11_5.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt11_5.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[2]['D' + str(i)] = ballastDt11_5[i - 19]

            ###SPEED 11.5  7 < WIND <8
        ballastDt11_8 = []
        numberOfApp11_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 6 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 6 and k[5] >= vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(

                    [k for k in ballastDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp11_8.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt11_8.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[2]['E' + str(i)] = ballastDt11_8[i - 19]

        #################################

        ###SPEED 12.5 WIND <1.5
        centralMean = np.mean(np.array([k for k in ballastDt if
                                        k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 8])
        centralArray = np.array([k for k in ballastDt if
                                 k[5] >= vel2Min and k[5] <= vel2Max and k[8] > 4])[:, 8]
        ballastDt12_0 = []
        numberOfApp12_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_0.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt12_0.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[2]['B' + str(i)] = ballastDt12_0[i - 29]

        ###SPEED 11.5  2 < WIND <3
        ballastDt12_3 = []
        numberOfApp12_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_3.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt12_3.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[2]['C' + str(i)] = ballastDt12_3[i - 29]

        ###SPEED 11.5  4 < WIND <5
        ballastDt12_5 = []
        numberOfApp12_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            ballastDt12_5.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[2]['D' + str(i)] = ballastDt12_5[i - 29]

        ###SPEED 11.5  7 < WIND <8
        ballastDt12_8 = []
        numberOfApp12_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 6 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and k[3] <= wind[
                                     i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 6 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_8.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt12_8.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[2]['E' + str(i)] = ballastDt12_8[i - 29]

        #################################

        ###SPEED 13.5 WIND <1.5
        centralMean = np.array([k for k in ballastDt if
                                k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
        if centralMean.__len__() > 0:
            centralMean = np.mean(np.array([k for k in ballastDt if
                                            k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 8])

            centralArray = np.array([k for k in ballastDt if
                                     k[5] >= vel3Min and k[5] <= vel3Max and k[8] > 4])
        else:
            centralMean = np.mean(np.array([k for k in ballastDt if
                                            k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 8]) + 1
            centralArray = np.array([k for k in ballastDt if
                                     k[5] >= vel2Min and k[5] <= vel2Max and k[8] > 4])[:, 8] + 1

        ballastDt13_0 = []
        numberOfApp13_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >=
                                 wind[i] and k[3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_0.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt13_0.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[2]['B' + str(i)] = ballastDt13_0[i - 39]

        ###SPEED 13.5  2 < WIND <3
        ballastDt13_3 = []
        numberOfApp13_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_3.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt13_3.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[2]['C' + str(i)] = ballastDt13_3[i - 39]

        ###SPEED 13.5  4 < WIND <5
        ballastDt13_5 = []
        numberOfApp13_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_5.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt13_5.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[2]['D' + str(i)] = ballastDt13_5[i - 39]

        ###SPEED 13.5  7 < WIND <8
        ballastDt13_8 = []
        numberOfApp13_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ballastDt if
                                 k[4] >= 6 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ballastDt if
                                  k[4] >= 6 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array(
                [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            # tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ballastDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_8.append(arrayFoc.__len__() + centralArray.__len__())
            ballastDt13_8.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[2]['E' + str(i)] = ballastDt13_8[i - 39]

        ####END OF BALLAST #############################################################
        ####END OF BALLAST #############################################################
        ####END OF BALLAST #############################################################

        ###WRITE BALLAST CONS TO FILE

        ##TREAT outliers / missing values for ballast values

        # workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
        # return
        #####################################################################################################################
        ##REPLACE NULL ZERO VALUES
        #####################################################################################################################
        values = [k for k in ballastDt10_0 if k != 0]
        length = values.__len__()
        numberOfApp10_0 = [numberOfApp10_0[i] for i in range(0, len(numberOfApp10_0)) if ballastDt10_0[i] > 0]
        for i in range(0, len(ballastDt10_0)):
            if length > 0:
                if length > 0:
                    if ballastDt10_0[i] == 0:
                        ##find items !=0
                        ballastDt10_0[i] = np.array(np.sum(
                            [numberOfApp10_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_0)
                    elif np.isnan(ballastDt10_0[i]):
                        ballastDt10_0[i] = np.array(np.sum(
                            [numberOfApp10_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_0)

        values = [k for k in ballastDt11_0 if k != 0]
        length = values.__len__()
        numberOfApp11_0 = [numberOfApp11_0[i] for i in range(0, len(numberOfApp11_0)) if ballastDt11_0[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt11_0)):
                if length > 0:

                    if ballastDt11_0[i] == 0:
                        ##find items !=0
                        ballastDt11_0[i] = np.array(np.sum(
                            [numberOfApp11_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_0)
                    elif np.isnan(ballastDt10_0[i]):
                        ballastDt11_0[i] = np.array(np.sum(
                            [numberOfApp11_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_0)

        values = [k for k in ballastDt12_0 if k != 0]
        length = values.__len__()
        numberOfApp12_0 = [numberOfApp12_0[i] for i in range(0, len(numberOfApp12_0)) if ballastDt12_0[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt12_0)):
                if length > 0:
                    if ballastDt12_0[i] == 0:
                        ##find items !=0
                        ballastDt12_0[i] = np.array(np.sum(
                            [numberOfApp12_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_0)
                    elif np.isnan(ballastDt10_0[i]):
                        ballastDt12_0[i] = np.array(np.sum(
                            [numberOfApp12_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_0)

        values = [k for k in ballastDt13_0 if k != 0]
        length = values.__len__()
        numberOfApp13_0 = [numberOfApp13_0[i] for i in range(0, len(numberOfApp13_0)) if ballastDt13_0[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt13_0)):
                if length > 0:
                    if ballastDt13_0[i] == 0:
                        ##find items !=0
                        ballastDt13_0[i] = np.array(np.sum(
                            [numberOfApp13_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_0)
                    elif np.isnan(ballastDt10_0[i]):
                        ballastDt13_0[i] = np.array(np.sum(
                            [numberOfApp13_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_0)

        values = [k for k in ballastDt10_3 if k != 0]
        length = values.__len__()
        numberOfApp10_3 = [numberOfApp10_3[i] for i in range(0, len(numberOfApp10_3)) if ballastDt10_3[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt10_3)):
                if length > 0:
                    if ballastDt10_3[i] == 0:
                        ##find items !=0
                        ballastDt10_3[i] = np.array(np.sum(
                            [numberOfApp10_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_3)
                    elif np.isnan(ballastDt10_3[i]):
                        ballastDt10_0[i] = np.array(np.sum(
                            [numberOfApp10_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_3)

        values = [k for k in ballastDt11_3 if k != 0]
        length = values.__len__()
        if values.__len__() > 0:
            numberOfApp11_3 = [numberOfApp11_3[i] for i in range(0, len(numberOfApp11_3)) if ballastDt11_3[i] > 0]
            for i in range(0, len(ballastDt11_3)):
                if length > 0:
                    if ballastDt11_3[i] == 0:
                        ##find items !=0
                        ballastDt11_3[i] = np.array(np.sum(
                            [numberOfApp11_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_3)
                    elif np.isnan(ballastDt11_3[i]):
                        ballastDt11_3[i] = np.array(np.sum(
                            [numberOfApp11_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_3)

        values = [k for k in ballastDt12_3 if k != 0]
        length = values.__len__()
        numberOfApp12_3 = [numberOfApp12_3[i] for i in range(0, len(numberOfApp12_3)) if ballastDt12_3[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt12_3)):
                if length > 0:
                    if ballastDt12_3[i] == 0:
                        ##find items !=0
                        ballastDt12_3[i] = np.array(np.sum(
                            [numberOfApp12_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_3)
                    elif np.isnan(ballastDt12_3[i]):
                        ballastDt12_3[i] = np.array(np.sum(
                            [numberOfApp12_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_3)

        values = [k for k in ballastDt13_3 if k != 0]
        length = values.__len__()
        numberOfApp13_3 = [numberOfApp13_3[i] for i in range(0, len(numberOfApp13_3)) if ballastDt13_3[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt13_3)):
                if length > 0:
                    if ballastDt13_3[i] == 0:
                        ##find items !=0
                        ballastDt13_3[i] = np.array(np.sum(
                            [numberOfApp13_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_3)
                    elif np.isnan(ballastDt13_3[i]):
                        ballastDt12_3[i] = np.array(np.sum(
                            [numberOfApp13_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_3)

        values = [k for k in ballastDt10_5 if k != 0]
        length = values.__len__()
        numberOfApp10_5 = [numberOfApp10_5[i] for i in range(0, len(numberOfApp10_5)) if ballastDt10_5[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt10_5)):
                if length > 0:
                    if ballastDt10_5[i] == 0:
                        ##find items !=0
                        ballastDt10_5[i] = np.array(np.sum(
                            [numberOfApp10_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_5)
                    elif np.isnan(ballastDt12_3[i]):
                        ballastDt12_3[i] = np.array(np.sum(
                            [numberOfApp10_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_5)

        values = [k for k in ballastDt11_5 if k != 0]
        length = values.__len__()
        numberOfApp11_5 = [numberOfApp11_5[i] for i in range(0, len(numberOfApp11_5)) if ballastDt11_5[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt11_5)):
                if length > 0:
                    if ballastDt11_5[i] == 0:
                        ##find items !=0
                        ballastDt11_5[i] = np.array(np.sum(
                            [numberOfApp11_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_5)
                    elif np.isnan(ballastDt11_5[i]):
                        ballastDt11_5[i] = np.array(np.sum(
                            [numberOfApp10_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_5)

        values = [k for k in ballastDt12_5 if k != 0]
        length = values.__len__()
        numberOfApp12_5 = [numberOfApp12_5[i] for i in range(0, len(numberOfApp12_5)) if ballastDt12_5[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt12_5)):
                if length > 0:
                    if ballastDt12_5[i] == 0:
                        ##find items !=0
                        ballastDt12_5[i] = np.array(np.sum(
                            [numberOfApp12_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_5)
                    elif np.isnan(ballastDt12_5[i]):
                        ballastDt12_5[i] = np.array(np.sum(
                            [numberOfApp12_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_5)

        values = [k for k in ballastDt13_5 if k != 0]
        length = values.__len__()
        numberOfApp13_5 = [numberOfApp13_5[i] for i in range(0, len(numberOfApp13_5)) if ballastDt13_5[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt13_5)):
                if length > 0:
                    if ballastDt13_5[i] == 0:
                        ##find items !=0
                        ballastDt13_5[i] = np.array(np.sum(
                            [numberOfApp13_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_5)
                    elif np.isnan(ballastDt13_5[i]):
                        ballastDt13_5[i] = np.array(np.sum(
                            [numberOfApp13_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_5)

        values = [k for k in ballastDt10_8 if k != 0]
        length = values.__len__()
        numberOfApp10_8 = [numberOfApp10_8[i] for i in range(0, len(numberOfApp10_8)) if ballastDt10_8[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt10_8)):
                if length > 0:
                    if ballastDt10_8[i] == 0:
                        ##find items !=0
                        ballastDt10_8[i] = np.array(np.sum(
                            [numberOfApp10_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_8)
                    elif np.isnan(ballastDt10_8[i]):
                        ballastDt10_8[i] = np.array(np.sum(
                            [numberOfApp10_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp10_8)

        values = [k for k in ballastDt11_8 if k != 0]
        length = values.__len__()
        numberOfApp11_8 = [numberOfApp11_8[i] for i in range(0, len(numberOfApp11_8)) if ballastDt11_8[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt11_8)):
                if length > 0:
                    if ballastDt11_8[i] == 0:
                        ##find items !=0
                        ballastDt11_8[i] = np.array(np.sum(
                            [numberOfApp11_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_8)
                    elif np.isnan(ballastDt11_8[i]):
                        ballastDt11_8[i] = np.array(np.sum(
                            [numberOfApp11_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp11_8)

        values = [k for k in ballastDt12_8 if k != 0]
        length = values.__len__()
        numberOfApp12_8 = [numberOfApp12_8[i] for i in range(0, len(numberOfApp12_8)) if ballastDt12_8[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt12_8)):
                if length > 0:
                    if ballastDt12_8[i] == 0:
                        ##find items !=0
                        ballastDt12_8[i] = np.array(np.sum(
                            [numberOfApp12_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_8)
                    elif np.isnan(ballastDt12_8[i]):
                        ballastDt12_8[i] = np.array(np.sum(
                            [numberOfApp12_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp12_8)

        values = [k for k in ballastDt13_8 if k != 0]
        length = values.__len__()
        numberOfApp13_8 = [numberOfApp13_8[i] for i in range(0, len(numberOfApp13_8)) if ballastDt13_8[i] > 0]
        if values.__len__() > 0:
            for i in range(0, len(ballastDt13_8)):
                if length > 0:
                    if ballastDt13_8[i] == 0:
                        ##find items !=0
                        ballastDt13_8[i] = np.array(np.sum(
                            [numberOfApp13_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_8)
                    elif np.isnan(ballastDt13_8[i]):
                        ballastDt13_8[i] = np.array(np.sum(
                            [numberOfApp13_8[i] * values[i] for i in range(0, len(values))])) / np.sum(
                            numberOfApp13_8)

        ################################################################################################################
        if (np.array(ballastDt10_0) == 0).all() and (np.array(ballastDt10_3) == 0).all() and (
                np.array(ballastDt10_5) == 0).all() and (np.array(ballastDt10_8) == 0).all():
            ballastDt10_0 = np.array(ballastDt11_0) - 2
            ballastDt10_3 = np.array(ballastDt11_3) - 2
            ballastDt10_5 = np.array(ballastDt11_5) - 2
            ballastDt10_8 = np.array(ballastDt11_8) - 2

        values = [k for k in ballastDt10_0 if k != 0]
        if values.__len__() == 0 or (np.array(ballastDt10_0) <= 0).all():
            ballastDt10_0 = np.array(ballastDt10_3) - 1

        values = [k for k in ballastDt11_0 if k != 0]
        if values.__len__() == 0 or (np.array(ballastDt11_0) <= 0).all():
            ballastDt11_0 = np.array(ballastDt11_3) - 1
        values = [k for k in ballastDt12_0 if k != 0]

        if values.__len__() == 0 or (np.array(ballastDt12_0) <= 0).all():
            ballastDt12_0 = np.array(ballastDt12_3) - 1

        values = [k for k in ballastDt13_0 if k != 0]
        if values.__len__() == 0 or (np.array(ballastDt13_0) <= 0).all():
            ballastDt13_0 = np.array(ballastDt13_3) - 1
        #####################################################################################################################
        #####################################################################################################################

        #####################################################################################################################
        for i in range(0, len(ballastDt10_0)):

            if (ballastDt10_0[i] >= ballastDt11_0[i]) and ballastDt11_0[i] > 0:
                while (ballastDt10_0[i] >= ballastDt11_0[i]):
                    ballastDt11_0[i] = ballastDt11_0[i] + 0.1 * ballastDt11_0[i]

        for i in range(0, len(ballastDt11_0)):

            if (ballastDt11_0[i] >= ballastDt12_0[i]) and ballastDt12_0[i] > 0:
                while (ballastDt11_0[i] >= ballastDt12_0[i]):
                    ballastDt12_0[i] = ballastDt12_0[i] + 0.1 * ballastDt12_0[i]

        for i in range(0, len(ballastDt12_0)):

            if (ballastDt12_0[i] >= ballastDt13_0[i]) and ballastDt13_0[i] > 0:
                while (ballastDt12_0[i] >= ballastDt13_0[i]):
                    ballastDt13_0[i] = ballastDt13_0[i] + 0.1 * ballastDt13_0[i]

        #####################################################################################################################
        for i in range(0, len(ballastDt10_3)):

            if (ballastDt10_0[i] >= ballastDt10_3[i]):
                while (ballastDt10_0[i] >= ballastDt10_3[i]):
                    if ballastDt10_3[i] == 0:
                        ballastDt10_3[i] = ballastDt10_0[i] + 0.1 * ballastDt10_0[i]
                    else:
                        ballastDt10_3[i] = ballastDt10_3[i] + 0.1 * ballastDt10_3[i]

        for i in range(0, len(ballastDt10_5)):

            if (ballastDt10_3[i] >= ballastDt10_5[i]):
                while (ballastDt10_3[i] >= ballastDt10_5[i]):
                    if ballastDt10_5[i] == 0:
                        ballastDt10_5[i] = ballastDt10_3[i] + 0.1 * ballastDt10_3[i]
                    else:
                        ballastDt10_5[i] = ballastDt10_5[i] + 0.1 * ballastDt10_5[i]

        for i in range(0, len(ballastDt10_8)):

            if (ballastDt10_5[i] >= ballastDt10_8[i]):
                while (ballastDt10_5[i] >= ballastDt10_8[i]):
                    if ballastDt10_8[i] == 0:
                        ballastDt10_8[i] = ballastDt10_5[i] + 0.1 * ballastDt10_5[i]
                    else:
                        ballastDt10_8[i] = ballastDt10_8[i] + 0.1 * ballastDt10_8[i]
        #####################################################################################################################

        for i in range(0, len(ballastDt11_3)):

            if (ballastDt11_0[i] >= ballastDt11_3[i]):
                while (ballastDt11_0[i] >= ballastDt11_3[i]):
                    if ballastDt11_3[i] == 0:
                        ballastDt11_3[i] = ballastDt11_0[i] + 0.1 * ballastDt11_0[i]
                    else:
                        ballastDt11_3[i] = ballastDt11_3[i] + 0.1 * ballastDt11_3[i]

        for i in range(0, len(ballastDt11_5)):

            if (ballastDt11_3[i] >= ballastDt11_5[i]):
                while (ballastDt11_3[i] >= ballastDt11_5[i]):
                    if ballastDt11_5[i] == 0:
                        ballastDt11_5[i] = ballastDt11_3[i] + 0.1 * ballastDt11_3[i]
                    else:
                        ballastDt11_5[i] = ballastDt11_5[i] + 0.1 * ballastDt11_5[i]

        for i in range(0, len(ballastDt11_8)):

            if (ballastDt11_5[i] >= ballastDt11_8[i]):
                while (ballastDt11_5[i] > ballastDt11_8[i]):
                    if ballastDt11_8[i] == 0:
                        ballastDt11_8[i] = ballastDt11_5[i] + 0.1 * ballastDt11_5[i]
                    else:
                        ballastDt11_8[i] = ballastDt11_8[i] + 0.1 * ballastDt11_8[i]

        #####################################################################################################################

        for i in range(0, len(ballastDt10_3)):

            if (ballastDt10_3[i] > ballastDt11_3[i]) and ballastDt11_3[i] > 0:
                while (ballastDt10_3[i] > ballastDt11_3[i]):
                    ballastDt11_3[i] = ballastDt11_3[i] + 0.1 * ballastDt11_3[i]

        for i in range(0, len(ballastDt10_5)):

            if (ballastDt10_5[i] > ballastDt11_5[i]) and ballastDt11_5[i] > 0:
                while (ballastDt10_5[i] > ballastDt11_5[i]):
                    ballastDt11_5[i] = ballastDt11_5[i] + 0.1 * ballastDt11_5[i]

        for i in range(0, len(ballastDt10_5)):

            if (ballastDt10_8[i] > ballastDt11_8[i]) and ballastDt11_8[i] > 0:
                while (ballastDt10_8[i] > ballastDt11_8[i]):
                    ballastDt11_8[i] = ballastDt11_8[i] + 0.1 * ballastDt11_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        #####################################################################################################################

        for i in range(0, len(ballastDt12_3)):

            if (ballastDt12_0[i] > ballastDt12_3[i]):
                while (ballastDt12_0[i] > ballastDt12_3[i]):
                    if ballastDt12_3[i] == 0:
                        ballastDt12_3[i] = ballastDt12_0[i] + 0.1 * ballastDt12_0[i]
                    else:
                        ballastDt12_3[i] = ballastDt12_3[i] + 0.1 * ballastDt12_3[i]

        for i in range(0, len(ballastDt12_5)):

            if (ballastDt12_3[i] > ballastDt12_5[i]):
                while (ballastDt12_3[i] > ballastDt12_5[i]):
                    if ballastDt12_5[i] == 0:
                        ballastDt12_5[i] = ballastDt12_3[i] + 0.1 * ballastDt12_3[i]
                    else:
                        ballastDt12_5[i] = ballastDt12_5[i] + 0.1 * ballastDt12_5[i]

        for i in range(0, len(ballastDt12_8)):

            if (ballastDt12_5[i] > ballastDt12_8[i]):
                while (ballastDt12_5[i] > ballastDt12_8[i]):
                    if ballastDt12_8[i] == 0:
                        ballastDt12_8[i] = ballastDt12_5[i] + 0.1 * ballastDt12_5[i]
                    else:
                        ballastDt12_8[i] = ballastDt12_8[i] + 0.1 * ballastDt12_8[i]

        #####################################################################################################################
        for i in range(0, len(ballastDt12_3)):

            if (ballastDt11_3[i] > ballastDt12_3[i]) and ballastDt12_3[i] > 0:
                while (ballastDt11_3[i] > ballastDt12_3[i]):
                    ballastDt12_3[i] = ballastDt12_3[i] + 0.1 * ballastDt12_3[i]

        for i in range(0, len(ballastDt12_5)):

            if (ballastDt11_5[i] > ballastDt12_5[i]) and ballastDt12_5[i] > 0:
                while (ballastDt11_5[i] > ballastDt12_5[i]):
                    ballastDt12_5[i] = ballastDt12_5[i] + 0.1 * ballastDt12_5[i]

        for i in range(0, len(ballastDt12_8)):

            if (ballastDt11_8[i] > ballastDt12_8[i]) and ballastDt12_8[i] > 0:
                while (ballastDt11_8[i] > ballastDt12_8[i]):
                    ballastDt12_8[i] = ballastDt12_8[i] + 0.1 * ballastDt12_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        #####################################################################################################################

        for i in range(0, len(ballastDt13_3)):

            if (ballastDt13_0[i] > ballastDt13_3[i]):
                while (ballastDt13_0[i] > ballastDt13_3[i]):
                    if ballastDt13_3[i] == 0:
                        ballastDt13_3[i] = ballastDt12_0[i] + 0.1 * ballastDt12_0[i]
                    else:
                        ballastDt13_3[i] = ballastDt13_3[i] + 0.1 * ballastDt13_3[i]

        for i in range(0, len(ballastDt13_5)):

            if (ballastDt13_3[i] > ballastDt13_5[i]):
                while (ballastDt13_3[i] > ballastDt13_5[i]):
                    if ballastDt13_5[i] == 0:
                        ballastDt13_5[i] = ballastDt12_3[i] + 0.1 * ballastDt12_3[i]
                    else:
                        ballastDt13_5[i] = ballastDt13_5[i] + 0.1 * ballastDt13_5[i]

        for i in range(0, len(ballastDt13_8)):

            if (ballastDt13_5[i] > ballastDt13_8[i]):
                while (ballastDt13_5[i] > ballastDt13_8[i]):
                    if ballastDt13_8[i] == 0:
                        ballastDt13_8[i] = ballastDt12_5[i] + 0.1 * ballastDt12_5[i]
                    else:
                        ballastDt13_8[i] = ballastDt13_8[i] + 0.1 * ballastDt13_8[i]

        #####################################################################################################################

        for i in range(0, len(ballastDt13_3)):

            if (ballastDt12_3[i] > ballastDt13_3[i]) and ballastDt13_3[i] > 0:
                while (ballastDt12_3[i] > ballastDt13_3[i]):
                    ballastDt13_3[i] = ballastDt13_3[i] + 0.1 * ballastDt13_3[i]

        for i in range(0, len(ballastDt13_5)):

            if (ballastDt12_5[i] > ballastDt13_5[i]) and ballastDt13_5[i] > 0:
                while (ballastDt12_5[i] > ballastDt13_5[i]):
                    ballastDt13_5[i] = ballastDt13_5[i] + 0.1 * ballastDt13_5[i]

        for i in range(0, len(ballastDt13_8)):
            if (ballastDt12_8[i] > ballastDt13_8[i]) and ballastDt13_8[i] > 0:
                while (ballastDt12_8[i] > ballastDt13_8[i]):
                    ballastDt13_8[i] = ballastDt13_8[i] + 0.1 * ballastDt13_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        if (np.array(ballastDt10_3) == 0).all():
            ballastDt10_3 = np.array(ballastDt10_0) + 1.5

        if (np.array(ballastDt10_5) == 0).all():
            ballastDt10_5 = np.array(ballastDt10_3) + 1.5

        if (np.array(ballastDt10_8) == 0).all():
            ballastDt10_8 = np.array(ballastDt10_5) + 1.5

        if (np.array(ballastDt11_0) == 0).all():
            ballastDt11_0 = np.array(ballastDt10_0) + 1

        if (np.array(ballastDt11_3) == 0).all():
            ballastDt11_3 = np.array(ballastDt11_0) + 1.5

        if (np.array(ballastDt11_5) == 0).all():
            ballastDt11_5 = np.array(ballastDt11_3) + 1.5

        if (np.array(ballastDt11_8) == 0).all():
            ballastDt11_8 = np.array(ballastDt11_5) + 1.5

        if (np.array(ballastDt12_0) == 0).all() or (np.array(ballastDt12_0) < 0).any():
            ballastDt12_0 = np.array(ballastDt11_0) + 1

        if (np.array(ballastDt12_3) == 0).all():
            ballastDt12_3 = np.array(ballastDt12_0) + 1.5
        for i in range(0, len(ballastDt12_3)):
            if ballastDt12_3[i] == inf:
                ballastDt12_3[i] = 0
        if (np.array(ballastDt12_5) == 0).all():
            ballastDt12_5 = np.array(ballastDt12_3) + 1.5

        if (np.array(ballastDt12_8) == 0).all():
            ballastDt12_8 = np.array(ballastDt12_5) + 1.5

        if (np.array(ballastDt13_0) == 0).all() or (np.array(ballastDt13_0) < 0).any():
            ballastDt13_0 = np.array(ballastDt12_0) + 1

        if (np.array(ballastDt13_3) == 0).all():
            ballastDt13_3 = np.array(ballastDt13_0) + 1.5

        if (np.array(ballastDt13_5) == 0).all():
            ballastDt13_5 = np.array(ballastDt13_3) + 1.5

        if (np.array(ballastDt13_8) == 0).all():
            ballastDt13_8 = np.array(ballastDt13_5) + 1.5
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        if (np.array(ballastDt10_0) == 0).all():
            ballastDt10_0 = np.array(ballastDt10_3) - 1 if (np.array(ballastDt10_3) != 0).any() else np.array(
                ballastDt10_5) - 2
        # if (np.array(ballastDt10_0)!=0).all()
        for i in range(9, 14):
            ballastDt10_0[0] = ballastDt10_0[4] + 1 if ballastDt10_0[0] <= ballastDt10_0[4] else ballastDt10_0[0]
            ballastDt10_0[2] = ballastDt10_0[3] + 1 if ballastDt10_0[2] <= ballastDt10_0[3] else ballastDt10_0[2]
            workbook._sheets[2]['B' + str(i)] = round(ballastDt10_0[i - 9], 2)

        ##TREAT outliers / missing values for ballastt values
        if (np.array(ballastDt10_3) == 0).all():
            ballastDt10_3 = np.array(ballastDt10_5) - 1 if (np.array(ballastDt10_5) != 0).any() else np.array(
                ballastDt10_8) - 3
        for i in range(9, 14):
            ballastDt10_3[0] = ballastDt10_3[4] + 1 if ballastDt10_3[0] <= ballastDt10_3[4] else ballastDt10_3[0]
            ballastDt10_3[2] = ballastDt10_3[3] + 1 if ballastDt10_3[2] <= ballastDt10_3[3] else ballastDt10_3[2]
            workbook._sheets[2]['C' + str(i)] = round(ballastDt10_3[i - 9], 2)

            ##TREAT outliers / missing values for ballastt values
        if (np.array(ballastDt10_5) == 0).all():
            ballastDt10_5 = np.array(ballastDt10_8) - 1.5 if (np.array(ballastDt10_8) != 0).any() else np.array(
                ballastDt10_3) + 1
        for i in range(9, 14):
            ballastDt10_5[0] = ballastDt10_5[4] + 1 if ballastDt10_5[0] <= ballastDt10_5[4] else ballastDt10_5[0]
            ballastDt10_5[2] = ballastDt10_5[3] + 1 if ballastDt10_5[2] <= ballastDt10_5[3] else ballastDt10_5[2]
            workbook._sheets[2]['D' + str(i)] = round(ballastDt10_5[i - 9], 2)

        ##############################################################################################################################
        ##FIX SIDE  / AGAINST
        if (ballastDt10_0[0] <= ballastDt10_0[2]):
            while (ballastDt10_0[0] <= ballastDt10_0[2]):
                ballastDt10_0[0] = ballastDt10_0[0] + 0.1 * ballastDt10_0[0]

        if (ballastDt10_0[1] <= ballastDt10_0[3]):
            while (ballastDt10_0[1] <= ballastDt10_0[3]):
                ballastDt10_0[1] = ballastDt10_0[1] + 0.1 * ballastDt10_0[1]

        if (ballastDt10_3[0] <= ballastDt10_3[2]):
            while (ballastDt10_3[0] <= ballastDt10_3[2]):
                ballastDt10_3[0] = ballastDt10_3[0] + 0.1 * ballastDt10_3[0]

        if (ballastDt10_3[1] <= ballastDt10_3[3]):
            while (ballastDt10_3[1] <= ballastDt10_3[3]):
                ballastDt10_3[1] = ballastDt10_3[1] + 0.1 * ballastDt10_3[1]

        if (ballastDt10_5[0] <= ballastDt10_5[2]):
            while (ballastDt10_5[0] <= ballastDt10_5[2]):
                ballastDt10_5[0] = ballastDt10_5[0] + 0.1 * ballastDt10_5[0]

        if (ballastDt10_5[1] <= ballastDt10_5[3]):
            while (ballastDt10_5[1] <= ballastDt10_5[3]):
                ballastDt10_5[1] = ballastDt10_5[1] + 0.1 * ballastDt10_5[1]

        if (ballastDt10_8[0] <= ballastDt10_8[2]):
            while (ballastDt10_8[0] <= ballastDt10_8[2]):
                ballastDt10_8[0] = ballastDt10_8[0] + 0.1 * ballastDt10_8[0]

        if (ballastDt10_8[1] <= ballastDt10_8[3]):
            while (ballastDt10_8[1] <= ballastDt10_8[3]):
                ballastDt10_8[1] = ballastDt10_8[1] + 0.1 * ballastDt10_8[1]

        if (ballastDt11_0[0] <= ballastDt11_0[2]):
            while (ballastDt11_0[0] <= ballastDt11_0[2]):
                ballastDt11_0[0] = ballastDt11_0[0] + 0.1 * ballastDt11_0[0]

        if (ballastDt11_0[1] <= ballastDt11_0[3]):
            while (ballastDt11_0[1] <= ballastDt11_0[3]):
                ballastDt11_0[1] = ballastDt11_0[1] + 0.1 * ballastDt11_0[1]

        if (ballastDt11_3[0] <= ballastDt11_3[2]):
            while (ballastDt11_3[0] <= ballastDt11_3[2]):
                ballastDt11_3[0] = ballastDt11_3[0] + 0.1 * ballastDt11_3[0]

        if (ballastDt11_3[1] <= ballastDt11_3[3]):
            while (ballastDt11_3[1] <= ballastDt11_3[3]):
                ballastDt11_3[1] = ballastDt11_3[1] + 0.1 * ballastDt11_3[1]

        if (ballastDt11_5[0] <= ballastDt11_5[2]):
            while (ballastDt11_5[0] <= ballastDt11_5[2]):
                ballastDt11_5[0] = ballastDt11_5[0] + 0.1 * ballastDt11_5[0]

        if (ballastDt11_5[1] <= ballastDt11_5[3]):
            while (ballastDt11_5[1] <= ballastDt11_5[3]):
                ballastDt11_5[1] = ballastDt11_5[1] + 0.1 * ballastDt11_5[1]

        if (ballastDt11_8[0] <= ballastDt11_8[2]):
            while (ballastDt11_8[0] <= ballastDt11_8[2]):
                ballastDt11_8[0] = ballastDt11_8[0] + 0.1 * ballastDt11_8[0]

        if (ballastDt11_8[1] <= ballastDt11_8[3]):
            while (ballastDt11_8[1] <= ballastDt11_8[3]):
                ballastDt11_8[1] = ballastDt11_8[1] + 0.1 * ballastDt11_8[1]

        if (ballastDt12_0[0] <= ballastDt12_0[2]):
            while (ballastDt12_0[0] <= ballastDt12_0[2]):
                ballastDt12_0[0] = ballastDt12_0[0] + 0.1 * ballastDt12_0[0]

        if (ballastDt12_0[1] <= ballastDt12_0[3]):
            while (ballastDt12_0[1] <= ballastDt12_0[3]):
                ballastDt12_0[1] = ballastDt12_0[1] + 0.1 * ballastDt12_0[1]

        if (ballastDt12_3[0] <= ballastDt12_3[2]):
            while (ballastDt12_3[0] <= ballastDt12_3[2]):
                ballastDt12_3[0] = ballastDt12_3[0] + 0.1 * ballastDt12_3[0]

        if (ballastDt12_3[1] <= ballastDt12_3[3]):
            while (ballastDt12_3[1] <= ballastDt12_3[3]):
                ballastDt12_3[1] = ballastDt12_3[1] + 0.1 * ballastDt12_3[1]

        if (ballastDt12_5[0] <= ballastDt12_5[2]):
            while (ballastDt12_5[0] <= ballastDt12_5[2]):
                ballastDt12_5[0] = ballastDt12_5[0] + 0.1 * ballastDt12_5[0]

        if (ballastDt12_5[1] <= ballastDt12_5[3]):
            while (ballastDt12_5[1] <= ballastDt12_5[3]):
                ballastDt12_5[1] = ballastDt12_5[1] + 0.1 * ballastDt12_5[1]

        if (ballastDt12_8[0] <= ballastDt12_8[2]):
            while (ballastDt12_8[0] <= ballastDt12_8[2]):
                ballastDt12_8[0] = ballastDt12_8[0] + 0.1 * ballastDt12_8[0]

        if (ballastDt12_8[1] <= ballastDt12_8[3]):
            while (ballastDt12_8[1] <= ballastDt12_8[3]):
                ballastDt12_8[1] = ballastDt12_8[1] + 0.1 * ballastDt12_8[1]

        if (ballastDt13_0[0] <= ballastDt13_0[2]):
            while (ballastDt13_0[0] <= ballastDt13_0[2]):
                ballastDt13_0[0] = ballastDt13_0[0] + 0.1 * ballastDt13_0[0]

        if (ballastDt13_0[1] <= ballastDt13_0[3]):
            while (ballastDt13_0[1] <= ballastDt13_0[3]):
                ballastDt13_0[1] = ballastDt13_0[1] + 0.1 * ballastDt13_0[1]

        if (ballastDt13_3[0] <= ballastDt13_3[2]):
            while (ballastDt13_3[0] <= ballastDt13_3[2]):
                ballastDt13_3[0] = ballastDt13_3[0] + 0.1 * ballastDt13_3[0]

        if (ballastDt13_3[1] <= ballastDt13_3[3]):
            while (ballastDt13_3[1] <= ballastDt13_3[3]):
                ballastDt13_3[1] = ballastDt13_3[1] + 0.1 * ballastDt13_3[1]

        if (ballastDt13_5[0] <= ballastDt13_5[2]):
            while (ballastDt13_5[0] <= ballastDt13_5[2]):
                ballastDt13_5[0] = ballastDt13_5[0] + 0.1 * ballastDt13_5[0]

        if (ballastDt13_5[1] <= ballastDt13_5[3]):
            while (ballastDt13_5[1] <= ballastDt13_5[3]):
                ballastDt13_5[1] = ballastDt13_5[1] + 0.1 * ballastDt13_5[1]

        if (ballastDt13_8[0] <= ballastDt13_8[2]):
            while (ballastDt13_8[0] <= ballastDt13_8[2]):
                ballastDt13_8[0] = ballastDt13_8[0] + 0.1 * ballastDt13_8[0]

        if (ballastDt13_8[1] <= ballastDt13_8[3]):
            while (ballastDt13_8[1] <= ballastDt13_8[3]):
                ballastDt13_8[1] = ballastDt13_8[1] + 0.1 * ballastDt13_8[1]
        ##FIX SIDE  / AGAINST##FIX SIDE  / AGAINST##FIX SIDE  / AGAINST##FIX SIDE  / AGAINST##FIX SIDE  / AGAINST##FIX SIDE  / AGAINST

        ##TREAT outliers / missing values for ballastt values
        if (np.array(ballastDt10_8) == 0).all():
            ballastDt10_8 = np.array(ballastDt10_5) + 2 if (np.array(ballastDt10_5) != 0).any() else np.array(
                ballastDt10_3) + 3
        for i in range(9, 14):
            ballastDt10_8[0] = ballastDt10_8[4] + 1 if ballastDt10_8[0] <= ballastDt10_8[4] else ballastDt10_8[0]
            ballastDt10_8[2] = ballastDt10_8[2] + 1 if ballastDt10_8[2] <= ballastDt10_8[3] else ballastDt10_8[2]
            workbook._sheets[2]['E' + str(i)] = round(ballastDt10_8[i - 9], 2)

            ####################################################################################################

            # values = [k for k in ballastDt11_0 if k != 0]
            # length = values.__len__()

        for i in range(19, 24):
            ballastDt11_0[0] = ballastDt11_0[4] + 1 if ballastDt11_0[0] <= ballastDt11_0[4] else ballastDt11_0[0]
            ballastDt11_0[2] = ballastDt11_0[3] + 1 if ballastDt11_0[2] <= ballastDt11_0[3] else ballastDt11_0[2]
            workbook._sheets[2]['B' + str(i)] = round(ballastDt11_0[i - 19], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(19, 24):
            ballastDt11_3[0] = ballastDt11_3[4] + 1 if ballastDt11_3[0] <= ballastDt11_3[4] else ballastDt11_3[0]
            ballastDt11_3[2] = ballastDt11_3[3] + 1 if ballastDt11_3[2] <= ballastDt11_3[3] else ballastDt11_3[2]
            workbook._sheets[2]['C' + str(i)] = round(ballastDt11_3[i - 19], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(19, 24):
            ballastDt11_5[0] = ballastDt11_5[4] + 1 if ballastDt11_5[0] <= ballastDt11_5[4] else ballastDt11_5[0]
            ballastDt11_5[2] = ballastDt11_5[3] + 1 if ballastDt11_5[2] <= ballastDt11_5[3] else ballastDt11_5[2]
            workbook._sheets[2]['D' + str(i)] = round(ballastDt11_5[i - 19], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(19, 24):
            ballastDt11_8[0] = ballastDt11_8[4] + 1 if ballastDt11_8[0] <= ballastDt11_8[4] else ballastDt11_8[0]
            ballastDt11_8[2] = ballastDt11_8[3] + 1 if ballastDt11_8[2] <= ballastDt11_8[3] else ballastDt11_8[2]
            workbook._sheets[2]['E' + str(i)] = round(ballastDt11_8[i - 19], 2)

            ####################################################################################################

        for i in range(29, 34):
            ballastDt12_0[0] = ballastDt12_0[4] + 1 if ballastDt12_0[0] <= ballastDt12_0[4] else ballastDt12_0[0]
            ballastDt12_0[2] = ballastDt12_0[3] + 1 if ballastDt12_0[2] <= ballastDt12_0[3] else ballastDt12_0[2]
            workbook._sheets[2]['B' + str(i)] = round(ballastDt12_0[i - 29], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(29, 34):
            ballastDt12_3[0] = ballastDt12_3[4] + 1 if ballastDt12_3[0] <= ballastDt12_3[4] else ballastDt12_3[0]
            ballastDt12_3[2] = ballastDt12_3[3] + 1 if ballastDt12_3[2] <= ballastDt12_3[3] else ballastDt12_3[2]
            workbook._sheets[2]['C' + str(i)] = round(ballastDt12_3[i - 29], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(29, 34):
            ballastDt12_5[0] = ballastDt12_5[4] + 1 if ballastDt12_5[0] <= ballastDt12_5[4] else ballastDt12_5[0]
            ballastDt12_5[2] = ballastDt12_5[3] + 1 if ballastDt12_5[2] <= ballastDt12_5[3] else ballastDt12_5[2]
            workbook._sheets[2]['D' + str(i)] = round(ballastDt12_5[i - 29], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(29, 34):
            ballastDt12_8[0] = ballastDt12_8[4] + 1 if ballastDt12_8[0] <= ballastDt12_8[4] else ballastDt12_8[0]
            ballastDt12_8[2] = ballastDt12_8[3] + 1 if ballastDt12_8[2] <= ballastDt12_8[3] else ballastDt12_8[2]
            workbook._sheets[2]['E' + str(i)] = round(ballastDt12_8[i - 29], 2)

            ################################################################################################################
        for i in range(39, 44):
            ballastDt13_0[0] = ballastDt13_0[4] + 1 if ballastDt13_0[0] <= ballastDt13_0[4] else ballastDt13_0[0]
            ballastDt13_0[2] = ballastDt13_0[3] + 1 if ballastDt13_0[2] <= ballastDt13_0[3] else ballastDt13_0[2]
            workbook._sheets[2]['B' + str(i)] = round(ballastDt13_0[i - 39], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(39, 44):
            ballastDt13_3[0] = ballastDt13_3[4] + 1 if ballastDt13_3[0] <= ballastDt13_3[4] else ballastDt13_3[0]
            ballastDt13_3[2] = ballastDt13_3[3] + 1 if ballastDt13_3[2] <= ballastDt13_3[3] else ballastDt13_3[2]
            workbook._sheets[2]['C' + str(i)] = round(ballastDt13_3[i - 39], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(39, 44):
            ballastDt13_5[0] = ballastDt13_5[4] + 1 if ballastDt13_5[0] <= ballastDt13_5[4] else ballastDt13_5[0]
            ballastDt13_5[2] = ballastDt13_5[3] + 1 if ballastDt13_5[2] <= ballastDt13_5[3] else ballastDt13_5[2]
            workbook._sheets[2]['D' + str(i)] = round(ballastDt13_5[i - 39], 2)

            ##TREAT outliers / missing values for ballastt values

        for i in range(39, 44):
            ballastDt13_8[0] = ballastDt13_8[4] + 1 if ballastDt13_8[0] <= ballastDt13_8[4] else ballastDt13_8[0]
            ballastDt13_8[2] = ballastDt13_8[3] + 1 if ballastDt13_8[2] <= ballastDt13_8[3] else ballastDt13_8[2]
            workbook._sheets[2]['E' + str(i)] = round(ballastDt13_8[i - 39], 2)

        ballastCons = list(itertools.chain(ballastDt10_0, ballastDt10_3, ballastDt10_5, ballastDt10_8,
                                           ballastDt11_0, ballastDt11_3, ballastDt11_5, ballastDt10_8,
                                           ballastDt12_0, ballastDt12_3, ballastDt12_5, ballastDt12_8,
                                           ballastDt13_0, ballastDt13_3, ballastDt13_5, ballastDt13_8))

        ###START OF LADDEN##################################################################################
        ####################################################################################################
        ####################################################################################################
        ####################################################################################################
        ####################################################################################################
        ####################################################################################################
        ####################################################################################################
        ####################################################################################################

        velocitiesL = np.array([k for k in ladenDt if k[5] > 6 and k[5] < 16])[:, 5]
        dataModel = KMeans(n_clusters=4)
        velocitiesL = velocitiesL.reshape(-1, 1)
        dataModel.fit(velocitiesL)
        labels = dataModel.predict(velocitiesL)
        # Extract centroid values

        centroidsL = dataModel.cluster_centers_
        centroidsL = np.sort(centroidsL, axis=0)

        partitionsX = []
        partitionLabels = []
        # For each label
        for curLbl in np.unique(labels):
            # Create a partition for X using records with corresponding label equal to the current
            partitionsX.append(np.asarray(velocitiesL[labels == curLbl]))
            # Create a partition for Y using records with corresponding label equal to the current

            # Keep partition label to ascertain same order of results
            partitionLabels.append(curLbl)

        sorted = []
        initialX = len(partitionsX)
        while sorted.__len__() < initialX:
            min = 100000000000
            for i in range(0, len(partitionsX)):
                mean = np.mean(partitionsX[i])
                if mean < min:
                    min = mean
                    minIndx = i
            sorted.append(partitionsX[minIndx])
            # partitionsX.remove(partitionsX[minIndx])
            partitionsX.pop(minIndx)

        ##delete ladden outliers
        # np.delete(ladenDt, [i for (i, v) in enumerate(ladenDt[:, 8]) if v < (
        # np.mean(ladenDt[:, 8]) - np.std(ladenDt[:, 8])) or v > np.mean(
        # ladenDt[:, 8]) + np.std(
        # ladenDt[:, 8])], 0)

        ###SPEED 10 WIND <1.5
        vel0Min = np.floor(np.min(sorted[0])) + 0.5
        vel1Min = np.floor(np.min(sorted[1])) + 0.5
        vel2Min = np.floor(np.min(sorted[2])) + 0.5
        vel3Min = np.floor(np.min(sorted[3])) + 0.5

        vel0Max = np.floor(np.max(sorted[0])) + 0.5
        vel1Max = np.floor(np.max(sorted[1])) + 0.5
        vel2Max = np.floor(np.max(sorted[2])) + 0.5
        vel3Max = np.floor(np.max(sorted[3])) + 0.5

        vel0Mean = np.floor(np.mean(sorted[0])) + 0.5
        vel1Mean = np.floor(np.mean(sorted[1])) + 0.5
        vel2Mean = np.floor(np.mean(sorted[2])) + 0.5
        vel3Mean = np.floor(np.mean(sorted[3])) + 0.5

        ####VESSEL BASIC INFO
        ldVelocities = []
        ldVelocities.append(vel0Mean)
        ldVelocities.append(vel1Mean)
        ldVelocities.append(vel2Mean)
        ldVelocities.append(vel3Mean)

        # vel0Min = 7
        # vel1Min = 9
        # vel2Min = 12
        # vel3Min = 13.5

        # vel0Max = 12
        # vel1Max = 12.5
        # vel2Max = 13.5
        # vel3Max = 15

        '''workbook._sheets[1]['B6'] = vel0Mean
          workbook._sheets[1]['B16'] = vel1Mean
          workbook._sheets[1]['B26'] = vel2Mean
          workbook._sheets[1]['B36'] = vel3Mean'''

        workbook._sheets[1]['B6'] = str(vel0Mean) + '  (' + str(vel0Min) + ' - ' + str(vel0Max) + ')'
        workbook._sheets[1]['B16'] = str(vel1Mean) + '  (' + str(vel1Min) + ' - ' + str(vel1Max) + ')'
        workbook._sheets[1]['B26'] = str(vel2Mean) + '  (' + str(vel2Min) + ' - ' + str(vel2Max) + ')'
        workbook._sheets[1]['B36'] = str(vel3Mean) + '  (' + str(vel3Min) + ' - ' + str(vel3Max) + ')'

        ##END OF VESSEL BASIC INFO
        speeds = list(itertools.chain(blVelocities, ldVelocities))
        with open('./data/' + company + '/' + vessel + '/ListOfSpeeds.csv', mode='w') as data:
            data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            data_writer.writerow(
                ['Speeds'])
            for i in range(0, len(speeds)):
                data_writer.writerow(
                    [speeds[i]])

        # workbook._sheets[1]['B6'] = round(centroidsL[0][0], 2)
        # workbook._sheets[1]['B16'] = round(centroidsL[1][0], 2)
        # workbook._sheets[1]['B26'] = round(centroidsL[2][0], 2)
        # workbook._sheets[1]['B36'] = round(centroidsL[3][0], 2)

        ladenDt10_0 = []
        # vel0Min = vel0Min + 2
        centralMean = np.mean(np.array([k for k in ladenDt if
                                        k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 8])
        centralArray = np.array([k for k in ladenDt if
                                 k[5] >= vel0Min and k[5] <= vel0Max and k[8] > 10])[:, 8]
        numberOfApp10_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k >= 3])

            # tlgarrayFoc=np.mean(np.array([k for k in ladenDt if k[5] >= round(centroidsB[0][0], 2) and k[5] <= 10 and k[4] >= 0 and k[4] <= 1 and k[9] > 10])[:, 9])
            tlgarrayFoc = np.array([k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])

            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array([k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                numberOfApp10_0.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt10_0.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[1]['B' + str(i)] = ladenDt10_0[i - 9]

        ###SPEED 10  2 < WIND <3
        ladenDt10_3 = []
        numberOfApp10_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp10_3.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt10_3.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[1]['C' + str(i)] = ladenDt10_3[i - 9]

        ###SPEED 10  4 < WIND <5
        ladenDt10_5 = []
        numberOfApp10_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp10_5.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt10_5.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[1]['D' + str(i)] = ladenDt10_5[i - 9]

        ###SPEED 10  7 < WIND <8
        ladenDt10_8 = []
        numberOfApp10_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 6 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[i] and
                                 k[3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 6 and k[5] >= vel0Min and k[5] <= vel0Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])
            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 13])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] >= vel0Min and k[5] <= vel0Max and k[9] >= 3])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp10_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp10_8.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt10_8.append(round(meanFoc, 2))

        for i in range(9, 14):
            workbook._sheets[1]['E' + str(i)] = ladenDt10_8[i - 9]

        ##################################################################################################################
        ##################################################################################################################

        ###SPEED 11.5   WIND <1.5
        centralMean = np.mean(np.array([k for k in ladenDt if
                                        k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 8])
        centralArray = np.array([k for k in ladenDt if
                                 k[5] >= vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 8]
        ladenDt11_0 = []
        numberOfApp11_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp11_0.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt11_0.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[1]['B' + str(i)] = ladenDt11_0[i - 19]

        ###SPEED 11.5  2 < WIND <3
        ladenDt11_3 = []
        numberOfApp11_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp11_3.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt11_3.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[1]['C' + str(i)] = ladenDt11_3[i - 19]

        ###SPEED 11.5  4 < WIND <5
        ladenDt11_5 = []
        numberOfApp11_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp11_5.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt11_5.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[1]['D' + str(i)] = ladenDt11_5[i - 19]

            ###SPEED 11.5  7 < WIND <8
        ladenDt11_8 = []
        numberOfApp11_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 6 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 6 and k[5] > vel1Min and k[5] <= vel1Max and k[3] >= wind[i] and k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel1Min and k[5] <= vel1Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp11_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp11_8.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt11_8.append(round(meanFoc, 2))

        for i in range(19, 24):
            workbook._sheets[1]['E' + str(i)] = ladenDt11_8[i - 19]

        #################################

        ###SPEED 12.5 WIND <1.5
        centralMean = np.mean(np.array([k for k in ladenDt if
                                        k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 8])

        centralArray = np.array([k for k in ladenDt if
                                 k[5] >= vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 8]
        ladenDt12_0 = []
        numberOfApp12_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_0.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt12_0.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[1]['B' + str(i)] = ladenDt12_0[i - 29]

        ###SPEED 11.5  2 < WIND <3
        ladenDt12_3 = []
        numberOfApp12_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_3.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt12_3.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[1]['C' + str(i)] = ladenDt12_3[i - 29]

        ###SPEED 11.5  4 < WIND <5
        ladenDt12_5 = []
        numberOfApp12_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_5.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt12_5.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[1]['D' + str(i)] = ladenDt12_5[i - 29]

        ###SPEED 11.5  7 < WIND <8
        ladenDt12_8 = []
        numberOfApp12_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 6 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and k[3] <= wind[
                                     i + 1] and k[8] > 10])
            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 6 and k[5] > vel2Min and k[5] <= vel2Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp12_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp12_8.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt12_8.append(round(meanFoc, 2))

        for i in range(29, 34):
            workbook._sheets[1]['E' + str(i)] = ladenDt12_8[i - 29]

        #################################

        ###SPEED 13.5 WIND <1.5
        centralMean = np.array([k for k in ladenDt if
                                k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
        if centralMean.__len__() > 0:
            centralMean = np.mean(np.array([k for k in ladenDt if
                                            k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 8])

            centralArray = np.array([k for k in ladenDt if
                                     k[5] >= vel3Min and k[5] <= vel3Max and k[8] > 4])
        else:
            centralMean = np.mean(np.array([k for k in ladenDt if
                                            k[5] > vel2Min and k[5] <= vel2Max and k[8] > 10])[:, 8]) + 1
            centralArray = np.array([k for k in ladenDt if
                                     k[5] >= vel2Min and k[5] <= vel2Max and k[8] > 4])[:, 8] + 1
        ladenDt13_0 = []
        numberOfApp13_0 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 0 and k[4] <= 1 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >=
                                 wind[
                                     i] and
                                 k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 0 and k[4] <= 1 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_0.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_0.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt13_0.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[1]['B' + str(i)] = ladenDt13_0[i - 39]

        ###SPEED 13.5  2 < WIND <3
        ladenDt13_3 = []
        numberOfApp13_3 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 2 and k[4] <= 3 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 2 and k[4] <= 3 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_3.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_3.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt13_3.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[1]['C' + str(i)] = ladenDt13_3[i - 39]

        ###SPEED 13.5  4 < WIND <5
        ladenDt13_5 = []
        numberOfApp13_5 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 4 and k[4] <= 5 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])

            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 4 and k[4] <= 5 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_5.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_5.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt13_5.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[1]['D' + str(i)] = ladenDt13_5[i - 39]

        ###SPEED 13.5  7 < WIND <8
        ladenDt13_8 = []
        numberOfApp13_8 = []
        for i in range(0, len(wind) - 1):
            arrayFoc = np.array([k for k in ladenDt if
                                 k[4] >= 6 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[
                                     i] and k[
                                     3] <= wind[i + 1] and k[8] > 10])
            steamTime = np.array([k for k in ladenDt if
                                  k[4] >= 6 and k[5] > vel3Min and k[5] <= vel3Max and k[3] >= wind[i] and
                                  k[3] >= wind[
                                      i] and
                                  k[3] <= wind[i + 1] and k[8] > 10])

            tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
            tlgarrayFoc = np.array(
                [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])
            # tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
            if tlgarrayFoc.__len__() > lenConditionTlg:
                tlgarrayFoc = np.array(
                    [k for k in ladenDt if k[5] > vel3Min and k[5] <= vel3Max and k[8] > 10])[:, 9]
                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (centralMean + np.mean(
                    tlgarrayFoc)) / 2
                numberOfApp13_8.append(arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
            else:
                weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                                  weights=steamTime[:, 12]) if arrayFoc.__len__() > minAccThres else 0
                meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else 0
                numberOfApp13_8.append(arrayFoc.__len__() + centralArray.__len__())
            ladenDt13_8.append(round(meanFoc, 2))

        for i in range(39, 44):
            workbook._sheets[1]['E' + str(i)] = ladenDt13_8[i - 39]

        ####END OF LADDEN #############################################################
        ####END OF LADDEN #############################################################
        ####END OF LADDEN #############################################################
        ##TREAT outliers / missing values for LADDEN values

        # workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
        # return
        #####################################################################################################################
        ##REPLACE NULL ZERO VALUES
        #####################################################################################################################
        ##WEIGHTED MEANS
        values = [k for k in ladenDt10_0 if k != 0]
        length = values.__len__()
        numberOfApp10_0 = [numberOfApp10_0[i] for i in range(0, len(numberOfApp10_0)) if ladenDt10_0[i] > 0]
        for i in range(0, len(ladenDt10_0)):
            if length > 0:
                if ladenDt10_0[i] == 0:
                    ##find items !=0
                    ladenDt10_0[i] = np.array(np.sum(
                        [numberOfApp10_0[i] * values[i] for i in range(0, len(values)) if values[i] > 0])) / np.sum(
                        numberOfApp10_0)
                elif np.isnan(ladenDt10_0[i]):
                    ladenDt10_0[i] = np.array(np.sum(
                        [numberOfApp10_0[i] * values[i] for i in range(0, len(values)) if values[i] > 0])) / np.sum(
                        numberOfApp10_0)

        values = [k for k in ladenDt11_0 if k != 0]
        length = values.__len__()
        numberOfApp11_0 = [numberOfApp11_0[i] for i in range(0, len(numberOfApp11_0)) if ladenDt11_0[i] > 0]
        for i in range(0, len(ladenDt11_0)):
            if length > 0:
                if ladenDt11_0[i] == 0:
                    ##find items !=0
                    ladenDt11_0[i] = np.array(
                        np.sum(
                            [numberOfApp11_0[i] * values[i] for i in range(0, len(values)) if values[i] > 0])) / np.sum(
                        numberOfApp11_0)
                elif np.isnan(ladenDt11_0[i]):
                    ladenDt11_0[i] = np.array(
                        np.sum(
                            [numberOfApp11_0[i] * values[i] for i in range(0, len(values)) if values[i] > 0])) / np.sum(
                        numberOfApp11_0)

        values = [k for k in ladenDt12_0 if k != 0]
        length = values.__len__()
        numberOfApp12_0 = [numberOfApp12_0[i] for i in range(0, len(numberOfApp12_0)) if ladenDt12_0[i] > 0]
        for i in range(0, len(ladenDt12_0)):
            if length > 0:
                if ladenDt12_0[i] == 0:
                    ##find items !=0
                    ladenDt12_0[i] = np.array(
                        np.sum([numberOfApp12_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp12_0)
                elif np.isnan(ladenDt11_0[i]):
                    ladenDt12_0[i] = np.array(
                        np.sum([numberOfApp12_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp12_0)

        values = [k for k in ladenDt13_0 if k != 0]
        length = values.__len__()
        numberOfApp13_0 = [numberOfApp13_0[i] for i in range(0, len(numberOfApp13_0)) if ladenDt13_0[i] > 0]
        for i in range(0, len(ladenDt13_0)):
            if length > 0:
                if ladenDt13_0[i] == 0:
                    ladenDt13_0[i] = np.array(
                        np.sum([numberOfApp13_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp13_0)
                elif np.isnan(ladenDt13_0[i]):
                    ladenDt13_0[i] = np.array(
                        np.sum([numberOfApp13_0[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp13_0)

        values = [k for k in ladenDt10_3 if k != 0]
        length = values.__len__()
        numberOfApp10_3 = [numberOfApp10_3[i] for i in range(0, len(numberOfApp10_3)) if ladenDt10_3[i] > 0]
        for i in range(0, len(ladenDt10_3)):
            if length > 0:
                if ladenDt10_3[i] == 0:
                    ladenDt10_3[i] = np.array(
                        np.sum([numberOfApp10_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp10_3)
                elif np.isnan(ladenDt10_3[i]):
                    ladenDt10_3[i] = np.array(
                        np.sum([numberOfApp10_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp10_3)

        values = [k for k in ladenDt11_3 if k != 0]
        length = values.__len__()
        numberOfApp11_3 = [numberOfApp11_3[i] for i in range(0, len(numberOfApp11_3)) if ladenDt11_3[i] > 0]
        for i in range(0, len(ladenDt11_3)):
            if length > 0:
                if ladenDt11_3[i] == 0:
                    ladenDt11_3[i] = np.array(
                        np.sum([numberOfApp11_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp11_3)
                elif np.isnan(ladenDt11_3[i]):
                    ladenDt11_3[i] = np.array(
                        np.sum([numberOfApp11_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp11_3)

        values = [k for k in ladenDt12_3 if k != 0]
        length = values.__len__()
        numberOfApp12_3 = [numberOfApp12_3[i] for i in range(0, len(numberOfApp12_3)) if ladenDt12_3[i] > 0]
        for i in range(0, len(ladenDt12_3)):
            if length > 0:
                if ladenDt12_3[i] == 0:
                    ladenDt12_3[i] = np.array(
                        np.sum([numberOfApp12_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp12_3)
                elif np.isnan(ladenDt12_3[i]):
                    ladenDt11_3[i] = np.array(
                        np.sum([numberOfApp12_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp12_3)

        values = [k for k in ladenDt13_3 if k != 0]
        length = values.__len__()
        numberOfApp13_3 = [numberOfApp13_3[i] for i in range(0, len(numberOfApp13_3)) if ladenDt13_3[i] > 0]
        for i in range(0, len(ladenDt13_3)):
            if length > 0:
                if ladenDt13_3[i] == 0:
                    ladenDt13_3[i] = np.array(
                        np.sum([numberOfApp13_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp13_3)
                elif np.isnan(ladenDt13_3[i]):
                    ladenDt13_3[i] = np.array(
                        np.sum([numberOfApp13_3[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp13_3)

        values = [k for k in ladenDt10_5 if k != 0]
        length = values.__len__()
        numberOfApp10_5 = [numberOfApp10_5[i] for i in range(0, len(numberOfApp10_5)) if ladenDt10_5[i] > 0]
        for i in range(0, len(ladenDt10_5)):
            if length > 0:
                if ladenDt10_5[i] == 0:
                    ladenDt10_5[i] = np.array(
                        np.sum([numberOfApp10_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt10_5[i] > 0])) / np.sum(
                        numberOfApp10_5)
                elif np.isnan(ladenDt10_5[i]):
                    ladenDt10_5[i] = np.array(
                        np.sum([numberOfApp10_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt10_5[i] > 0])) / np.sum(
                        numberOfApp10_5)

        values = [k for k in ladenDt11_5 if k != 0]
        length = values.__len__()
        numberOfApp11_5 = [numberOfApp11_5[i] for i in range(0, len(numberOfApp11_5)) if ladenDt11_5[i] > 0]
        for i in range(0, len(ladenDt11_5)):
            if length > 0:
                if ladenDt11_5[i] == 0:
                    ladenDt11_5[i] = np.array(
                        np.sum([numberOfApp11_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp11_5)
                elif np.isnan(ladenDt11_5[i]):
                    ladenDt11_5[i] = np.array(
                        np.sum([numberOfApp11_5[i] * values[i] for i in range(0, len(values))])) / np.sum(
                        numberOfApp11_5)

        values = [k for k in ladenDt12_5 if k != 0]
        length = values.__len__()
        numberOfApp12_5 = [numberOfApp12_5[i] for i in range(0, len(numberOfApp12_5)) if ladenDt12_5[i] > 0]
        for i in range(0, len(ladenDt12_5)):
            if length > 0:
                if ladenDt12_5[i] == 0:
                    ladenDt12_5[i] = np.array(
                        np.sum([numberOfApp12_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt12_5[i] > 0])) / np.sum(
                        numberOfApp12_5)
                elif np.isnan(ladenDt12_5[i]):
                    ladenDt12_5[i] = np.array(
                        np.sum([numberOfApp12_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt12_5[i] > 0])) / np.sum(
                        numberOfApp12_5)

        values = [k for k in ladenDt13_5 if k != 0]
        length = values.__len__()
        numberOfApp13_5 = [numberOfApp13_5[i] for i in range(0, len(numberOfApp13_5)) if ladenDt12_5[i] > 0]
        for i in range(0, len(ladenDt13_5)):
            if length > 0:
                if ladenDt13_5[i] == 0:
                    ladenDt13_5[i] = np.array(
                        np.sum([numberOfApp13_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt13_5[i] > 0])) / np.sum(
                        numberOfApp13_5)
                elif np.isnan(ladenDt13_5[i]):
                    ladenDt13_5[i] = np.array(
                        np.sum([numberOfApp13_5[i] * values[i] for i in range(0, len(values)) if
                                ladenDt13_5[i] > 0])) / np.sum(
                        numberOfApp13_5)

        values = [k for k in ladenDt10_8 if k != 0]
        length = values.__len__()
        numberOfApp10_8 = [numberOfApp10_8[i] for i in range(0, len(numberOfApp10_8)) if ladenDt10_8[i] > 0]
        for i in range(0, len(ladenDt10_8)):
            if length > 0:
                if ladenDt10_8[i] == 0:
                    ladenDt10_8[i] = np.array(
                        np.sum([numberOfApp10_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt10_8[i] > 0])) / np.sum(
                        numberOfApp10_8)
                elif np.isnan(ladenDt10_8[i]):
                    ladenDt10_8[i] = np.array(
                        np.sum([numberOfApp10_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt10_8[i] > 0])) / np.sum(
                        numberOfApp10_8)

        values = [k for k in ladenDt11_8 if k != 0]
        length = values.__len__()
        numberOfApp11_8 = [numberOfApp11_8[i] for i in range(0, len(numberOfApp11_8)) if ladenDt11_8[i] > 0]
        for i in range(0, len(ladenDt11_8)):
            if length > 0:
                if length > 0:
                    if ladenDt11_8[i] == 0:
                        ladenDt11_8[i] = np.array(
                            np.sum([numberOfApp11_8[i] * values[i] for i in range(0, len(values)) if
                                    ladenDt11_8[i] > 0])) / np.sum(
                            numberOfApp11_8)
                    elif np.isnan(ladenDt11_8[i]):
                        ladenDt11_8[i] = np.array(
                            np.sum([numberOfApp11_8[i] * values[i] for i in range(0, len(values)) if
                                    ladenDt11_8[i] > 0])) / np.sum(
                            numberOfApp11_8)

        values = [k for k in ladenDt12_8 if k != 0]
        length = values.__len__()
        numberOfApp12_8 = [numberOfApp12_8[i] for i in range(0, len(numberOfApp12_8)) if ladenDt12_8[i] > 0]
        for i in range(0, len(ladenDt12_8)):
            if length > 0:
                if ladenDt12_8[i] == 0:
                    ladenDt12_8[i] = np.array(
                        np.sum([numberOfApp12_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt12_8[i] > 0])) / np.sum(
                        numberOfApp12_8)
                elif np.isnan(ladenDt12_8[i]):
                    ladenDt12_8[i] = np.array(
                        np.sum([numberOfApp12_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt12_8[i] > 0])) / np.sum(
                        numberOfApp12_8)

        values = [k for k in ladenDt13_8 if k != 0]
        length = values.__len__()
        numberOfApp13_8 = [numberOfApp13_8[i] for i in range(0, len(numberOfApp13_8)) if ladenDt13_8[i] > 0]
        for i in range(0, len(ladenDt13_8)):
            if length > 0:
                if ladenDt13_8[i] == 0:
                    ladenDt13_8[i] = np.array(
                        np.sum([numberOfApp13_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt13_8[i] > 0])) / np.sum(
                        numberOfApp13_8)
                elif np.isnan(ladenDt13_8[i]):
                    ladenDt13_8[i] = np.array(
                        np.sum([numberOfApp12_8[i] * values[i] for i in range(0, len(values)) if
                                ladenDt13_8[i] > 0])) / np.sum(
                        numberOfApp13_8)

        #####################################################################################################################
        if (np.array(ladenDt10_0) == 0).all() and (np.array(ladenDt10_3) == 0).all() and (
                np.array(ladenDt10_5) == 0).all() and (np.array(ladenDt10_8) == 0).all():
            ladenDt10_0 = np.array(ladenDt11_0) - 2
            ladenDt10_3 = np.array(ladenDt11_3) - 2
            ladenDt10_5 = np.array(ladenDt11_5) - 2
            ladenDt10_8 = np.array(ladenDt11_8) - 2

        if (np.array(ladenDt10_3) == 0).all():
            ladenDt10_3 = np.array(ladenDt10_0) + 1.5

        if (np.array(ladenDt10_5) == 0).all():
            ladenDt10_5 = np.array(ladenDt10_3) + 1.5

        if (np.array(ladenDt10_8) == 0).all():
            ladenDt10_8 = np.array(ladenDt10_5) + 1.5

        if (np.array(ladenDt11_0) == 0).all():
            ladenDt11_0 = np.array(ladenDt10_0) + 1

        if (np.array(ladenDt11_3) == 0).all():
            ladenDt11_3 = np.array(ladenDt11_0) + 1.5

        if (np.array(ladenDt11_5) == 0).all():
            ladenDt11_5 = np.array(ladenDt11_3) + 1.5

        if (np.array(ladenDt11_8) == 0).all():
            ladenDt11_8 = np.array(ladenDt11_5) + 1.5

        if (np.array(ladenDt12_0) == 0).all():
            ladenDt12_0 = np.array(ladenDt11_0) + 1

        if (np.array(ladenDt12_3) == 0).all():
            ladenDt12_3 = np.array(ladenDt12_0) + 1.5
        for i in range(0, len(ladenDt12_3)):
            if ladenDt12_3[i] == inf:
                ladenDt12_3[i] = 0
        if (np.array(ladenDt12_5) == 0).all():
            ladenDt12_5 = np.array(ladenDt12_3) + 1.5

        if (np.array(ladenDt12_8) == 0).all():
            ladenDt12_8 = np.array(ladenDt12_5) + 1.5

        if (np.array(ladenDt13_0) == 0).all():
            ladenDt13_0 = np.array(ladenDt12_0) + 1

        if (np.array(ladenDt13_3) == 0).all():
            ladenDt13_3 = np.array(ladenDt13_0) + 1.5

        if (np.array(ladenDt13_5) == 0).all():
            ladenDt13_5 = np.array(ladenDt13_3) + 1.5

        if (np.array(ladenDt13_8) == 0).all():
            ladenDt13_8 = np.array(ladenDt13_5) + 1.5
        ######################
        ####################

        values = [k for k in ladenDt10_0 if k != 0]
        if values.__len__() == 0:
            if (np.array(ladenDt10_3) != 0).all():
                ladenDt10_0 = np.array(ladenDt10_3) - 1
            elif (np.array(ladenDt10_5) != 0).all():
                ladenDt10_0 = np.array(ladenDt10_5) - 2
            else:
                ladenDt10_0 = np.array(ladenDt10_8) - 3

        values = [k for k in ladenDt11_0 if k != 0]
        if values.__len__() == 0:
            ladenDt11_0 = np.array(ladenDt11_3) - 1
        values = [k for k in ladenDt12_0 if k != 0]

        if values.__len__() == 0:
            ladenDt12_0 = np.array(ladenDt12_3) - 1

        values = [k for k in ladenDt13_0 if k != 0]
        if values.__len__() == 0:
            ladenDt13_0 = np.array(ladenDt13_3) - 1

        for i in range(0, len(ladenDt10_0)):

            if (ladenDt10_0[i] >= ladenDt11_0[i]) and ladenDt11_0[i] > 0:
                while (ladenDt10_0[i] >= ladenDt11_0[i]):
                    ladenDt11_0[i] = ladenDt11_0[i] + 0.1 * ladenDt11_0[i]

        for i in range(0, len(ladenDt11_0)):

            if (ladenDt11_0[i] >= ladenDt12_0[i]) and ladenDt12_0[i] > 0:
                while (ladenDt11_0[i] >= ladenDt12_0[i]):
                    ladenDt12_0[i] = ladenDt12_0[i] + 0.1 * ladenDt12_0[i]

        for i in range(0, len(ladenDt12_0)):

            if (ladenDt12_0[i] >= ladenDt13_0[i]) and ladenDt13_0[i] > 0:
                while (ladenDt12_0[i] >= ladenDt13_0[i]):
                    ladenDt13_0[i] = ladenDt13_0[i] + 0.1 * ladenDt13_0[i]

        #####################################################################################################################

        for i in range(0, len(ladenDt10_3)):

            if (ladenDt10_0[i] >= ladenDt10_3[i]):
                while (ladenDt10_0[i] >= ladenDt10_3[i]):
                    if ladenDt10_3[i] == 0:
                        ladenDt10_3[i] = ladenDt10_0[i] + 0.1 * ladenDt10_0[i]
                    else:
                        ladenDt10_3[i] = ladenDt10_3[i] + 0.1 * ladenDt10_3[i]

        for i in range(0, len(ladenDt10_5)):

            if (ladenDt10_3[i] >= ladenDt10_5[i]):
                while (ladenDt10_3[i] >= ladenDt10_5[i]):
                    if ladenDt10_5[i] == 0:
                        ladenDt10_5[i] = ladenDt10_3[i] + 0.1 * ladenDt10_3[i]
                    else:
                        ladenDt10_5[i] = ladenDt10_5[i] + 0.1 * ladenDt10_5[i]

        for i in range(0, len(ladenDt10_8)):

            if (ladenDt10_5[i] >= ladenDt10_8[i]):
                while (ladenDt10_5[i] >= ladenDt10_8[i]):
                    if ladenDt10_8[i] == 0:
                        ladenDt10_8[i] = ladenDt10_5[i] + 0.1 * ladenDt10_5[i]
                    else:
                        ladenDt10_8[i] = ladenDt10_8[i] + 0.1 * ladenDt10_8[i]
        #####################################################################################################################

        for i in range(0, len(ladenDt11_3)):

            if (ladenDt11_0[i] >= ladenDt11_3[i]):
                while (ladenDt11_0[i] >= ladenDt11_3[i]):
                    if ladenDt11_3[i] == 0:
                        ladenDt11_3[i] = ladenDt11_0[i] + 0.1 * ladenDt11_0[i]
                    else:
                        ladenDt11_3[i] = ladenDt11_3[i] + 0.1 * ladenDt11_3[i]

        for i in range(0, len(ladenDt11_5)):

            if (ladenDt11_3[i] >= ladenDt11_5[i]):
                while (ladenDt11_3[i] >= ladenDt11_5[i]):
                    if ladenDt11_5[i] == 0:
                        ladenDt11_5[i] = ladenDt11_3[i] + 0.1 * ladenDt11_3[i]
                    else:
                        ladenDt11_5[i] = ladenDt11_5[i] + 0.1 * ladenDt11_5[i]

        for i in range(0, len(ladenDt11_8)):

            if (ladenDt11_5[i] >= ladenDt11_8[i]):
                while (ladenDt11_5[i] >= ladenDt11_8[i]):
                    if ladenDt11_8[i] == 0:
                        ladenDt11_8[i] = ladenDt11_5[i] + 0.1 * ladenDt11_5[i]
                    else:
                        ladenDt11_8[i] = ladenDt11_8[i] + 0.1 * ladenDt11_8[i]

        #####################################################################################################################

        for i in range(0, len(ladenDt10_3)):

            if (ladenDt10_3[i] > ladenDt11_3[i]) and ladenDt11_3[i] > 0:
                while (ladenDt10_3[i] > ladenDt11_3[i]):
                    ladenDt11_3[i] = ladenDt11_3[i] + 0.1 * ladenDt11_3[i]

        for i in range(0, len(ladenDt10_5)):

            if (ladenDt10_5[i] > ladenDt11_5[i]) and ladenDt11_5[i] > 0:
                while (ladenDt10_5[i] > ladenDt11_5[i]):
                    ladenDt11_5[i] = ladenDt11_5[i] + 0.1 * ladenDt11_5[i]

        for i in range(0, len(ladenDt10_5)):

            if (ladenDt10_8[i] > ladenDt11_8[i]) and ladenDt11_8[i] > 0:
                while (ladenDt10_8[i] > ladenDt11_8[i]):
                    ladenDt11_8[i] = ladenDt11_8[i] + 0.1 * ladenDt11_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

        #####################################################################################################################

        for i in range(0, len(ladenDt12_3)):

            if (ladenDt12_0[i] > ladenDt12_3[i]):
                while (ladenDt12_0[i] > ladenDt12_3[i]):
                    if ladenDt12_3[i] == 0:
                        ladenDt12_3[i] = ladenDt12_0[i] + 0.1 * ladenDt12_0[i]
                    else:
                        ladenDt12_3[i] = ladenDt12_3[i] + 0.1 * ladenDt12_3[i]

        for i in range(0, len(ladenDt12_5)):

            if (ladenDt12_3[i] > ladenDt12_5[i]):
                while (ladenDt12_3[i] > ladenDt12_5[i]):
                    if ladenDt12_5[i] == 0:
                        ladenDt12_5[i] = ladenDt12_3[i] + 0.1 * ladenDt12_3[i]
                    else:
                        ladenDt12_5[i] = ladenDt12_5[i] + 0.1 * ladenDt12_5[i]

        for i in range(0, len(ladenDt12_8)):

            if (ladenDt12_5[i] > ladenDt12_8[i]):
                while (ladenDt12_5[i] > ladenDt12_8[i]):
                    if ladenDt12_8[i] == 0:
                        ladenDt12_8[i] = ladenDt12_5[i] + 0.1 * ladenDt12_5[i]
                    else:
                        ladenDt12_8[i] = ladenDt12_8[i] + 0.1 * ladenDt12_8[i]

        #####################################################################################################################
        for i in range(0, len(ladenDt12_3)):

            if (ladenDt11_3[i] > ladenDt12_3[i]) and ladenDt12_3[i] > 0:
                while (ladenDt11_3[i] > ladenDt12_3[i]):
                    ladenDt12_3[i] = ladenDt12_3[i] + 0.1 * ladenDt12_3[i]

        for i in range(0, len(ladenDt12_5)):

            if (ladenDt11_5[i] > ladenDt12_5[i]) and ladenDt12_5[i] > 0:
                while (ladenDt11_5[i] > ladenDt12_5[i]):
                    ladenDt12_5[i] = ladenDt12_5[i] + 0.1 * ladenDt12_5[i]

        for i in range(0, len(ladenDt12_8)):

            if (ladenDt11_8[i] > ladenDt12_8[i]) and ladenDt12_8[i] > 0:
                while (ladenDt11_8[i] > ladenDt12_8[i]):
                    ladenDt12_8[i] = ladenDt12_8[i] + 0.1 * ladenDt12_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        # STANDARIZE WEATHER
        #####################################################################################################################

        for i in range(0, len(ladenDt13_3)):

            if (ladenDt13_0[i] > ladenDt13_3[i]):
                while (ladenDt13_0[i] > ladenDt13_3[i]):
                    if ladenDt13_3[i] == 0:
                        ladenDt13_3[i] = ladenDt12_0[i] + 0.1 * ladenDt12_0[i]
                    else:
                        ladenDt13_3[i] = ladenDt13_3[i] + 0.1 * ladenDt13_3[i]

        for i in range(0, len(ladenDt13_5)):

            if (ladenDt13_3[i] > ladenDt13_5[i]):
                while (ladenDt13_3[i] > ladenDt13_5[i]):
                    if ladenDt13_5[i] == 0:
                        ladenDt13_5[i] = ladenDt12_3[i] + 0.1 * ladenDt12_3[i]
                    else:
                        ladenDt13_5[i] = ladenDt13_5[i] + 0.1 * ladenDt13_5[i]

        for i in range(0, len(ladenDt13_8)):

            if (ladenDt13_5[i] > ladenDt13_8[i]):
                while (ladenDt13_5[i] > ladenDt13_8[i]):
                    if ladenDt13_8[i] == 0:
                        ladenDt13_8[i] = ladenDt12_5[i] + 0.1 * ladenDt12_5[i]
                    else:
                        ladenDt13_8[i] = ladenDt13_8[i] + 0.1 * ladenDt13_8[i]

        #####################################################################################################################

        for i in range(0, len(ladenDt13_3)):

            if (ladenDt12_3[i] > ladenDt13_3[i]) and ladenDt13_3[i] > 0:
                while (ladenDt12_3[i] > ladenDt13_3[i]):
                    ladenDt13_3[i] = ladenDt13_3[i] + 0.1 * ladenDt13_3[i]

        for i in range(0, len(ladenDt13_5)):

            if (ladenDt12_5[i] > ladenDt13_5[i]) and ladenDt13_5[i] > 0:
                while (ladenDt12_5[i] > ladenDt13_5[i]):
                    ladenDt13_5[i] = ladenDt13_5[i] + 0.1 * ladenDt13_5[i]

        for i in range(0, len(ladenDt13_8)):

            if (ladenDt12_8[i] > ladenDt13_8[i]) and ladenDt13_8[i] > 0:
                while (ladenDt12_8[i] > ladenDt13_8[i]):
                    ladenDt13_8[i] = ladenDt13_8[i] + 0.1 * ladenDt13_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        ##LAST CHECKPOINT
        for i in range(0, len(ladenDt13_8)):

            if (ladenDt13_5[i] > ladenDt13_8[i]):
                while (ladenDt13_5[i] > ladenDt13_8[i]):
                    if ladenDt13_8[i] == 0:
                        ladenDt13_8[i] = ladenDt12_5[i] + 0.1 * ladenDt12_5[i]
                    else:
                        ladenDt13_8[i] = ladenDt13_8[i] + 0.1 * ladenDt13_8[i]
        # 1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#1!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
        ### fix side /against LADEN
        if (ladenDt10_0[0] <= ladenDt10_0[2]):
            while (ladenDt10_0[0] <= ladenDt10_0[2]):
                ladenDt10_0[0] = ladenDt10_0[0] + 0.1 * ladenDt10_0[0]

        if (ladenDt10_0[1] <= ladenDt10_0[3]):
            while (ladenDt10_0[1] <= ladenDt10_0[3]):
                ladenDt10_0[1] = ladenDt10_0[1] + 0.1 * ladenDt10_0[1]

        if (ladenDt10_3[0] <= ladenDt10_3[2]):
            while (ladenDt10_3[0] <= ladenDt10_3[2]):
                ladenDt10_3[0] = ladenDt10_3[0] + 0.1 * ladenDt10_3[0]

        if (ladenDt10_3[1] <= ladenDt10_3[3]):
            while (ladenDt10_3[1] <= ladenDt10_3[3]):
                ladenDt10_3[1] = ladenDt10_3[1] + 0.1 * ladenDt10_3[1]

        if (ladenDt10_5[0] <= ladenDt10_5[2]):
            while (ladenDt10_5[0] <= ladenDt10_5[2]):
                ladenDt10_5[0] = ladenDt10_5[0] + 0.1 * ladenDt10_5[0]

        if (ladenDt10_5[1] <= ladenDt10_5[3]):
            while (ladenDt10_5[1] <= ladenDt10_5[3]):
                ladenDt10_5[1] = ladenDt10_5[1] + 0.1 * ladenDt10_5[1]

        if (ladenDt10_8[0] <= ladenDt10_8[2]):
            while (ladenDt10_8[0] <= ladenDt10_8[2]):
                ladenDt10_8[0] = ladenDt10_8[0] + 0.1 * ladenDt10_8[0]

        if (ladenDt10_8[1] <= ladenDt10_8[3]):
            while (ladenDt10_8[1] <= ladenDt10_8[3]):
                ladenDt10_8[1] = ladenDt10_8[1] + 0.1 * ladenDt10_8[1]

        if (ladenDt11_0[0] <= ladenDt11_0[2]):
            while (ladenDt11_0[0] <= ladenDt11_0[2]):
                ladenDt11_0[0] = ladenDt11_0[0] + 0.1 * ladenDt11_0[0]

        if (ladenDt11_0[1] <= ladenDt11_0[3]):
            while (ladenDt11_0[1] <= ladenDt11_0[3]):
                ladenDt11_0[1] = ladenDt11_0[1] + 0.1 * ladenDt11_0[1]

        if (ladenDt11_3[0] <= ladenDt11_3[2]):
            while (ladenDt11_3[0] <= ladenDt11_3[2]):
                ladenDt11_3[0] = ladenDt11_3[0] + 0.1 * ladenDt11_3[0]

        if (ladenDt11_3[1] <= ladenDt11_3[3]):
            while (ladenDt11_3[1] <= ladenDt11_3[3]):
                ladenDt11_3[1] = ladenDt11_3[1] + 0.1 * ladenDt11_3[1]

        if (ladenDt11_5[0] <= ladenDt11_5[2]):
            while (ladenDt11_5[0] <= ladenDt11_5[2]):
                ladenDt11_5[0] = ladenDt11_5[0] + 0.1 * ladenDt11_5[0]

        if (ladenDt11_5[1] <= ladenDt11_5[3]):
            while (ladenDt11_5[1] <= ladenDt11_5[3]):
                ladenDt11_5[1] = ladenDt11_5[1] + 0.1 * ladenDt11_5[1]

        if (ladenDt11_8[0] <= ladenDt11_8[2]):
            while (ladenDt11_8[0] <= ladenDt11_8[2]):
                ladenDt11_8[0] = ladenDt11_8[0] + 0.1 * ladenDt11_8[0]

        if (ladenDt11_8[1] <= ladenDt11_8[3]):
            while (ladenDt11_8[1] <= ladenDt11_8[3]):
                ladenDt11_8[1] = ladenDt11_8[1] + 0.1 * ladenDt11_8[1]

        if (ladenDt12_0[0] <= ladenDt12_0[2]):
            while (ladenDt12_0[0] <= ladenDt12_0[2]):
                ladenDt12_0[0] = ladenDt12_0[0] + 0.1 * ladenDt12_0[0]

        if (ladenDt12_0[1] <= ladenDt12_0[3]):
            while (ladenDt12_0[1] <= ladenDt12_0[3]):
                ladenDt12_0[1] = ladenDt12_0[1] + 0.1 * ladenDt12_0[1]

        if (ladenDt12_3[0] <= ladenDt12_3[2]):
            while (ladenDt12_3[0] <= ladenDt12_3[2]):
                ladenDt12_3[0] = ladenDt12_3[0] + 0.1 * ladenDt12_3[0]

        if (ladenDt12_3[1] <= ladenDt12_3[3]):
            while (ladenDt12_3[1] <= ladenDt12_3[3]):
                ladenDt12_3[1] = ladenDt12_3[1] + 0.1 * ladenDt12_3[1]

        if (ladenDt12_5[0] <= ladenDt12_5[2]):
            while (ladenDt12_5[0] <= ladenDt12_5[2]):
                ladenDt12_5[0] = ladenDt12_5[0] + 0.1 * ladenDt12_5[0]

        if (ladenDt12_5[1] <= ladenDt12_5[3]):
            while (ladenDt12_5[1] <= ladenDt12_5[3]):
                ladenDt12_5[1] = ladenDt12_5[1] + 0.1 * ladenDt12_5[1]

        if (ladenDt12_8[0] <= ladenDt12_8[2]):
            while (ladenDt12_8[0] <= ladenDt12_8[2]):
                ladenDt12_8[0] = ladenDt12_8[0] + 0.1 * ladenDt12_8[0]

        if (ladenDt12_8[1] <= ladenDt12_8[3]):
            while (ladenDt12_8[1] <= ladenDt12_8[3]):
                ladenDt12_8[1] = ladenDt12_8[1] + 0.1 * ladenDt12_8[1]

        if (ladenDt13_0[0] <= ladenDt13_0[2]):
            while (ladenDt13_0[0] <= ladenDt13_0[2]):
                ladenDt13_0[0] = ladenDt13_0[0] + 0.1 * ladenDt13_0[0]

        if (ladenDt13_0[1] <= ladenDt13_0[3]):
            while (ladenDt13_0[1] <= ladenDt13_0[3]):
                ladenDt13_0[1] = ladenDt13_0[1] + 0.1 * ladenDt13_0[1]

        if (ladenDt13_3[0] <= ladenDt13_3[2]):
            while (ladenDt13_3[0] <= ladenDt13_3[2]):
                ladenDt13_3[0] = ladenDt13_3[0] + 0.1 * ladenDt13_3[0]

        if (ladenDt13_3[1] <= ladenDt13_3[3]):
            while (ladenDt13_3[1] <= ladenDt13_3[3]):
                ladenDt13_3[1] = ladenDt13_3[1] + 0.1 * ladenDt13_3[1]

        if (ladenDt13_5[0] <= ladenDt13_5[2]):
            while (ladenDt13_5[0] <= ladenDt13_5[2]):
                ladenDt13_5[0] = ladenDt13_5[0] + 0.1 * ladenDt13_5[0]

        if (ladenDt13_5[1] <= ladenDt13_5[3]):
            while (ladenDt13_5[1] <= ladenDt13_5[3]):
                ladenDt13_5[1] = ladenDt13_5[1] + 0.1 * ladenDt13_5[1]

        if (ladenDt13_8[0] <= ladenDt13_8[2]):
            while (ladenDt13_8[0] <= ladenDt13_8[2]):
                ladenDt13_8[0] = ladenDt13_8[0] + 0.1 * ladenDt13_8[0]

        if (ladenDt13_8[1] <= ladenDt13_8[3]):
            while (ladenDt13_8[1] <= ladenDt13_8[3]):
                ladenDt13_8[1] = ladenDt13_8[1] + 0.1 * ladenDt13_8[1]
        ##################################################END AGAINST SIDE#################################
        if (np.array(ladenDt10_0) == 0).all():
            ladenDt10_0 = np.array(ladenDt10_3) - 1 if (np.array(ladenDt10_0) != 3).all() else np.array(ladenDt10_5) - 2
        for i in range(9, 14):
            ladenDt10_0[0] = ladenDt10_0[4] + 1 if ladenDt10_0[0] <= ladenDt10_0[4] else ladenDt10_0[0]
            ladenDt10_0[2] = ladenDt10_0[3] + 1 if ladenDt10_0[2] <= ladenDt10_0[3] else ladenDt10_0[2]
            workbook._sheets[1]['B' + str(i)] = round(ladenDt10_0[i - 9], 2)

            ##TREAT outliers / missing values for ladent values
        if (np.array(ladenDt10_3) == 0).all():
            ladenDt10_3 = np.array(ladenDt10_5) - 1 if (np.array(ladenDt10_5) != 0).any() else np.array(ladenDt10_8) - 2
        for i in range(9, 14):
            ladenDt10_3[0] = ladenDt10_3[4] + 1 if ladenDt10_3[0] <= ladenDt10_3[4] else ladenDt10_3[0]
            ladenDt10_3[2] = ladenDt10_3[3] + 1 if ladenDt10_3[2] <= ladenDt10_3[3] else ladenDt10_3[2]
            workbook._sheets[1]['C' + str(i)] = round(ladenDt10_3[i - 9], 2)

            ##TREAT outliers / missing values for ladent values
        if (np.array(ladenDt10_5) == 0).all():
            ladenDt10_5 = np.array(ladenDt10_8) - 1 if (np.array(ladenDt10_8) != 0).any() else np.array(ladenDt10_3) + 1
        for i in range(9, 14):
            ladenDt10_5[0] = ladenDt10_5[4] + 1 if ladenDt10_5[0] <= ladenDt10_5[4] else ladenDt10_5[0]
            ladenDt10_5[2] = ladenDt10_5[3] + 1 if ladenDt10_5[2] <= ladenDt10_5[3] else ladenDt10_5[2]
            workbook._sheets[1]['D' + str(i)] = round(ladenDt10_5[i - 9], 2)

            ##TREAT outliers / missing values for ladent values
        if (np.array(ladenDt10_8) == 0).all():
            ladenDt10_8 = np.array(ladenDt10_5) + 1 if (np.array(ladenDt10_5) != 0).any() else np.array(ladenDt10_3) + 3
        for i in range(9, 14):
            ladenDt10_8[0] = ladenDt10_8[4] + 1 if ladenDt10_8[0] <= ladenDt10_8[4] else ladenDt10_8[0]
            ladenDt10_8[2] = ladenDt10_8[3] + 1 if ladenDt10_8[2] <= ladenDt10_8[3] else ladenDt10_8[2]
            workbook._sheets[1]['E' + str(i)] = round(ladenDt10_8[i - 9], 2)

            ####################################################################################################

            # values = [k for k in ladenDt11_0 if k != 0]
            # length = values.__len__()

        for i in range(19, 24):
            ladenDt11_0[0] = ladenDt11_0[4] + 1 if ladenDt11_0[0] <= ladenDt11_0[4] else ladenDt11_0[0]
            ladenDt11_0[2] = ladenDt11_0[3] + 1 if ladenDt11_0[2] <= ladenDt11_0[3] else ladenDt11_0[2]
            workbook._sheets[1]['B' + str(i)] = round(ladenDt11_0[i - 19], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(19, 24):
            ladenDt11_8[0] = ladenDt11_8[4] + 1 if ladenDt11_8[0] <= ladenDt11_8[4] else ladenDt11_8[0]
            ladenDt11_8[2] = ladenDt11_8[3] + 1 if ladenDt11_8[2] <= ladenDt11_8[3] else ladenDt11_8[2]
            workbook._sheets[1]['C' + str(i)] = round(ladenDt11_3[i - 19], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(19, 24):
            ladenDt11_5[0] = ladenDt11_5[4] + 1 if ladenDt11_5[0] <= ladenDt11_5[4] else ladenDt11_5[0]
            ladenDt11_5[2] = ladenDt11_5[3] + 1 if ladenDt11_5[2] <= ladenDt11_5[3] else ladenDt11_5[2]
            workbook._sheets[1]['D' + str(i)] = round(ladenDt11_5[i - 19], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(19, 24):
            ladenDt11_8[0] = ladenDt11_8[4] + 1 if ladenDt11_8[0] <= ladenDt11_8[4] else ladenDt11_8[0]
            ladenDt11_8[2] = ladenDt11_8[3] + 1 if ladenDt11_8[2] <= ladenDt11_8[3] else ladenDt11_8[2]
            workbook._sheets[1]['E' + str(i)] = round(ladenDt11_8[i - 19], 2)

            ####################################################################################################

        for i in range(29, 34):
            ladenDt12_0[0] = ladenDt12_0[4] + 1 if ladenDt12_0[0] <= ladenDt12_0[4] else ladenDt12_0[0]
            ladenDt12_0[2] = ladenDt12_0[3] + 1 if ladenDt12_0[2] <= ladenDt12_0[3] else ladenDt12_0[2]
            workbook._sheets[1]['B' + str(i)] = round(ladenDt12_0[i - 29], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(29, 34):
            ladenDt12_3[0] = ladenDt12_3[4] + 1 if ladenDt12_3[0] <= ladenDt12_3[4] else ladenDt12_3[0]
            ladenDt12_3[2] = ladenDt12_3[3] + 1 if ladenDt12_3[2] <= ladenDt12_3[3] else ladenDt12_3[2]
            workbook._sheets[1]['C' + str(i)] = round(ladenDt12_3[i - 29], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(29, 34):
            ladenDt12_5[0] = ladenDt12_5[4] + 1 if ladenDt12_5[0] <= ladenDt12_5[4] else ladenDt12_5[0]
            ladenDt12_5[2] = ladenDt12_5[3] + 1 if ladenDt12_5[2] <= ladenDt12_5[3] else ladenDt12_5[2]
            workbook._sheets[1]['D' + str(i)] = round(ladenDt12_5[i - 29], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(29, 34):
            ladenDt12_8[0] = ladenDt12_8[4] + 1 if ladenDt12_8[0] <= ladenDt12_8[4] else ladenDt12_8[0]
            ladenDt12_8[2] = ladenDt12_8[3] + 1 if ladenDt12_8[2] <= ladenDt12_8[3] else ladenDt12_8[2]
            workbook._sheets[1]['E' + str(i)] = round(ladenDt12_8[i - 29], 2)

            ################################################################################################################
        for i in range(39, 44):
            ladenDt13_0[0] = ladenDt13_0[4] + 1 if ladenDt13_0[0] <= ladenDt13_0[4] else ladenDt13_0[0]
            ladenDt13_0[2] = ladenDt13_0[3] + 1 if ladenDt13_0[2] <= ladenDt13_0[3] else ladenDt13_0[2]
            workbook._sheets[1]['B' + str(i)] = round(ladenDt13_0[i - 39], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(39, 44):
            ladenDt13_3[0] = ladenDt13_3[4] + 1 if ladenDt13_3[0] <= ladenDt13_3[4] else ladenDt13_3[0]
            ladenDt13_3[2] = ladenDt13_3[3] + 1 if ladenDt13_3[2] <= ladenDt13_3[3] else ladenDt13_3[2]
            workbook._sheets[1]['C' + str(i)] = round(ladenDt13_3[i - 39], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(39, 44):
            ladenDt13_5[0] = ladenDt13_5[4] + 1 if ladenDt13_5[0] <= ladenDt13_5[4] else ladenDt13_5[0]
            ladenDt13_5[2] = ladenDt13_5[3] + 1 if ladenDt13_5[2] <= ladenDt13_5[3] else ladenDt13_5[2]
            workbook._sheets[1]['D' + str(i)] = round(ladenDt13_5[i - 39], 2)

            ##TREAT outliers / missing values for ladent values

        for i in range(39, 44):
            ladenDt13_8[0] = ladenDt13_8[4] + 1 if ladenDt13_8[0] <= ladenDt13_8[4] else ladenDt13_8[0]
            ladenDt13_8[2] = ladenDt13_8[3] + 1 if ladenDt13_8[2] <= ladenDt13_8[3] else ladenDt13_8[2]
            workbook._sheets[1]['E' + str(i)] = round(ladenDt13_8[i - 39], 2)

        ladenCons = list(itertools.chain(ladenDt10_0, ladenDt10_3, ladenDt10_5, ladenDt10_8,
                                         ladenDt11_0, ladenDt11_3, ladenDt11_5, ladenDt10_8,
                                         ladenDt12_0, ladenDt12_3, ladenDt12_5, ladenDt12_8,
                                         ladenDt13_0, ladenDt13_3, ladenDt13_5, ladenDt13_8))

        profCons = list(itertools.chain(ballastCons, ladenCons))

        '''with open('./data/' + company + '/' + vessel + '/ListOfCons.csv', mode='w') as data:
            data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            data_writer.writerow(
                ['FOC'])
            for i in range(0, len(profCons)):
                data_writer.writerow(
                    [profCons[i]])'''

        workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
        return


    def findCloseToLandDataPoints(self,trData):


        notIncOcean = []
        inOcean = []
        for i in range(0, len(trData)):
            lat = trData[i, 6]
            lon = trData[i, 7]
            is_in_ocean = globe.is_ocean(lat, lon)
            if is_in_ocean == False and trData[i,3]<9:
                notIncOcean.append(trData[i])
            elif is_in_ocean==True:
                inOcean.append(trData[i])
        trData1 = notIncOcean

        trData1 = np.array(trData1)
        trData2 = np.array(inOcean)
        return trData1 , trData2

    def fillDetailedExcelProfCons(self, company, vessel, pathToexcel, dataSet, rawData, tlgDataset, dataSetBDD,
                                  dataSetADD, imo):
        ##FEATURE SET EXACT POSITION OF COLUMNS NEEDED IN ORDER TO PRODUCE EXCEL
        # 2nd place BALLAST FLAG
        # 8th place DRAFT
        # 10th place WD
        # 11th place WF
        # 12th place SPEED
        # 15th place ME FOC 24H
        # 16th place ME FOC 24H TLGS
        # 17th place TRIM
        # 19th place SteamHours
        # 18th place STW_TLG
        # 20st place swellSWH
        #dataSet = dataSet[:10000, :]
        #dataSet = dataSet[:,:dataSet.shape[1]-1]
        #rawData = pd.read_csv('./data/DANAOS/'+vessel+'/'+'LEO_Cdata.csv').values
        #rawDraft = rawData[:,7]
        #dataSet[:,8]= rawDraft[:len(dataSet)]
        '''foc = np.array([.k for k in dataSet if float(k[15]) > 0])[:, 15]
          df = pd.DataFrame({
              'foc': foc,
          })
          sns.displot(df, x="foc")
          plt.show()'''
        d=0
        n_steps = 6
        #COVNERT kg/min to MT/day
        #dataSet[:,15]= ( (dataSet[:,15]  ) / 1000 )* 1440
        #dataSet[:, 15] = ((dataSet[:, 15]) ) * 24
        # COVNERT kg/min to MT/day

        #COVNERT kg/h to MT/day
        #dataSet[:,15]= ( (dataSet[:,15]  ) / 1000 )* 24
        #COVNERT kg/h to MT/day

        #dataSet = dataSet[0:30000]
        # if float(k[5])>6.5
        # dataSet = np.array(dataSet).astype(float)
        '''bl = np.array([k for k in dataSet])[:, 2]
        for i in range(0, len(bl)):
            bl[i] = 'L'
        dataSet[:, 2] = bl'''

        '''wd = np.array([k for k in dataSet])[:, 10]
        for i in range(0, len(wd)):
            if float(wd[i]) > 180:
                wd[i] = float(wd[i]) - 180  # and  float(k[8])<20
        dataSet[:, 10] = wd'''

        #wf = np.array([k for k in dataSet])[:, 11]/ (1.944)
        #for i in range(0, len(wf)):
              #wf[i] = self.ConvertMSToBeaufort(float(float(wf[i])))
        #dataSet[:, 11] = wf

        lenConditionTlg = 5000000
        dtNew = np.nan_to_num(np.array([k for k in dataSet if float(k[15]) > 0 and float(k[12]) > 0 ])[:,7:].astype(float)) # and  float(k[8])<20

        #for i in range(0, len(dtNew)):
              #ballastDt[i,12] = 1 if ballastDt[i,12]==0 else 1
              #dtNew[i] = np.mean(dtNew[i:i + 15], axis=0)
        # dtNewBDD = np.array([k for k in dataSetBDD if float(k[15]) > 0 and float(k[12]) > 0])
        # dtNewADD = np.array([k for k in dataSetADD if float(k[15]) > 0 and float(k[12]) > 0])


        draft =np.array([k for k in dtNew if float(k[1]) > 1 and float(k[1])<30 ])[:, 1].astype(float)
        trim = np.array([k for k in dtNew if float(k[10]) < 30])[:, 10].astype(float)
        meanDraft = np.mean(draft)
        minDraft = np.floor(np.min(draft))
        maxDraft = np.ceil(np.max(draft))
        stdDraft = np.std(draft)

        dataModel = KMeans(n_clusters=3)
        draft = draft.reshape(-1, 1)
        dataModel.fit(draft)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        draftsSorted = np.sort(centroids, axis=0)


        workbook = load_workbook(filename=pathToexcel)
        ##########################
        #########
        ###

        ###OUTLIERS DELETE VALUES THAT ARE < MEAN - 3 * STD OR > MEAN + 3 * STD
        '''meanFoc = np.mean(dtNew[:, 8])
        stdFoc = np.std(dtNew[:, 8])
        dtNew = np.array([k for k in dtNew if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''


        ########################################################################


        ########################################################################
        ########################################################################
        ##LOAD EXCEL


        ladenFlag = True
        ballastFlag = True
        thirdDrftFlag = False
        print("Min Draft: " + str(minDraft))
        print("Mean Draft: " +str(meanDraft))
        print("Max Draft: " + str(maxDraft))

        maxBalDraft = (draftsSorted[0][0] + maxDraft) /2

        thirdDrft =  '('+ str(np.floor(draftsSorted[0][0])-1 ) +' - '+ str(np.floor(draftsSorted[1][0])) + ')'
        ballastDrft = '(' + str(np.floor(draftsSorted[1][0])) + ' - ' + str(np.round(draftsSorted[2][0])) + ')'
        ladenDrft = '(' + str(np.floor(draftsSorted[2][0])) + ' - ' + str(maxDraft) + ')'

        categories = ballastDrft +' '+ ladenDrft +' '+ thirdDrft
        categoriesDraft  = [minDraft , np.floor(draftsSorted[1][0]) , np.round(draftsSorted[2][0]) , maxDraft ]
        print("Draft categories: " + categories)
        #ladenDt = np.array([k for k in dtNew if k[1] > np.floor(maxBalDraft) and k[1] <= np.ceil(maxDraft) ]).astype(float)
        

        thridCategoryDraftDt = np.array([k for k in dtNew if k[1] >= np.floor(draftsSorted[0][0])-1 and k[1] <= np.round(draftsSorted[1][0])]).astype(float)
        ballastDt = np.array([k for k in dtNew if k[1] >=  np.round(draftsSorted[1][0]) and k[1] <=  np.round(draftsSorted[2][0])]).astype(float)
        ladenDt = np.array([k for k in dtNew if k[1] > np.round(draftsSorted[2][0]) and k[1] <= maxDraft]).astype(float)

        #thridCategoryDraftDt = np.array([k for k in dtNew if k[1] >= 4 and k[1] <= 10]).astype(float)
        #ballastDt = np.array([k for k in dtNew if k[1] >= 10 and k[1] <= 12]).astype(float)
        ##ladenDt = np.array([k for k in dtNew if k[1] > 12 and k[1] <= 15]).astype(float)

        #ballastDt = np.array([k for k in dtNew if k[1] >= np.floor(draftsSorted[0][0] - 1) and k[1] <= np.floor(maxBalDraft)]).astype(float)
        #ladenDt = np.array([k for k in dtNew if k[1] > np.floor(draftsSorted[0][0]) and k[1] <= np.ceil(maxDraft)]).astype(float)


        #workbook._sheets[1].title = 'Cons profile for Draft (' + str(np.floor(draftsSorted[0][0]-1)) + ' - ' + str(np.ceil(maxDraft)) + ')'
        workbook._sheets[1].title = 'Cons profile for Draft ' + firstDrft
        workbook._sheets[2].title = 'Cons profile for Draft ' + secondDrft
        workbook._sheets[3].title = 'Cons profile for Draft '+ thirdDrft

        
        meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)
        meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
        meanDradftThird = round(float(np.mean(np.array([k for k in thridCategoryDraftDt if k[1] > 0])[:, 1])), 2)

        velocities = (
            np.array((np.array([k for k in dtNew if float(k[5]) > 7 and float(k[5]<=30)])[:, 5])).astype(float))  # and float(k[12]) < 18
        dataModel = KMeans(n_clusters=5)
        velocities = velocities.reshape(-1, 1)
        dataModel.fit(velocities)
        # Extract centroid values
        centroids = dataModel.cluster_centers_
        velocitiesSorted = np.sort(centroids, axis=0)
        ################################################################################################

        #workbook = self.calculateExcelStatistics(workbook, dtNew, velocities, draft, trim, velocities, [],
                                                 #company, vessel, tlgDataset, 'all')

        velMinGen = np.round(velocitiesSorted[0][0])
        velMaxGen = np.round(velocitiesSorted[3][0])

        workbook._sheets[0]['B2'] = vessel
        workbook._sheets[0]['B5'] = round(np.max(velocities), 1)
        workbook._sheets[0]['B7'] = np.min(draft)
        workbook._sheets[0]['B8'] = np.max(draft)

        minAccThres = 1000000000000000
        laddenJSON = '{}'
        json_decoded = json.loads(laddenJSON)
        json_decoded['ConsumptionProfile'] = {"vessel_code": str(imo), 'vessel_name': vessel,
                                              "dateCreated": date.today().strftime("%d/%m/%Y"), "consProfilePORTS":[],"consProfile": []}



        ##FIND CANALS PORTS DATA POINTS 1KM RESOLUTION FORM LAND
        #trDataPorts = self.findCloseToLandDataPoints(dtNew)
        #json_decoded = self.fillPORTSCANALSsheet(workbook,trDataPorts,json_decoded,company,vessel)
        ###################################################################################LADDEN BEST FIT
        maxVel = np.round(np.max(velocities))
        consVelocities = np.arange(np.floor(velocitiesSorted[0][0]), maxVel)
        consVelocitiesJSON = np.arange(np.floor(velocitiesSorted[0][0]), maxVel,0.5)
        stepVelRanges = int(np.round(len(consVelocities)/4))
        consVelocitiesRanges=[]
        for i in range(0,len(consVelocities),stepVelRanges):
         consVelocitiesRanges.append(consVelocities[i])

        r = 10
        rows =[]
        for i in range(0,len(consVelocities)):
            row = np.arange(r,r+64,9)
            rows.append(row)
            r += 79

        ##weather impact ranges

        draftData = dtNew
        #for i in range(0,len(draftData)):
            #draftData[i] = np.mean(draftData[i:i + 15], axis=0)
        minSpeedWF = 11.5
        maxSpeedWF = 12.5
        ##############################################
        minDraftWF = 8
        maxDraftWF = 15

        ##draft weights

        sizesDraft = []
        draft = []
        avgActualFoc = []
        minActualFoc = []
        maxActualFoc = []
        stdActualFoc = []
        maxDraftF = maxDraft
        minDraftF = minDraft
        i = minDraftF
        speedRange2 = np.array([k for k in draftData if k[5] >= minSpeedWF and k[5] <= maxSpeedWF ])
        rawDraft = speedRange2
        #np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 3])
        # rawDraft  = speedRange1
        while i <= maxDraftF:

            speedArray = np.array([k for k in rawDraft if float(k[1]) >= i and float(k[1]) <= i + 2])

            if speedArray.__len__() > 0:
                sizesDraft.append(speedArray.__len__())
                draft.append(i)
                avgActualFoc.append(np.mean(speedArray[:, 8]))
                minActualFoc.append(np.min(speedArray[:, 8]))
                maxActualFoc.append(np.max(speedArray[:, 8]))
                stdActualFoc.append(np.std(speedArray[:, 8]))
            i += 2



        xi = np.array(draft)
        yi = np.array(avgActualFoc)
        zi = np.array(sizesDraft)

        p2 = np.poly1d(np.polyfit(xi, yi, 1))
        draft810 = np.array([random.uniform(9, 11) for p in range(0, 100)])
        draft1012 = np.array([random.uniform(11, 13) for p in range(0, 100)])

        draft1215 = np.array([random.uniform(13, 15) for p in range(0, 100)])

        #draftData = np.array([k for k in rawDraft if float(k[1]) >= minDraft and float(k[1]) <= maxDraft])
        #splineR = SplineRegression.Earth(max_degree=2,)
        #splineR.fit(draftData[:,1].reshape(-1,1),draftData[:,8].reshape(-1,1))

        p2_810 = p2(draft810)
        p2_1012 = p2(draft1012)
        p2_1215 = p2(draft1215)

        #p2_810 = splineR.predict(draft810.reshape(-1,1))
        #p2_1012 = splineR.predict(draft1012.reshape(-1,1))
        #p2_1215 = splineR.predict(draft1215.reshape(-1,1))

        factorDRAFT = [0, abs((np.mean(p2_1012) - np.mean(p2_810)) / np.mean(p2_810)),
                       abs((np.mean(p2_1215) - np.mean(p2_1012)) / np.mean(p2_1012)),
                       ]
        ##################################################

        x=0
        if thirdDrftFlag == True:

            minLaddenSpeedn = consVelocities[0]
            # consVelocities[0]
            maxLaddenSpeedn = np.ceil(np.max(thridCategoryDraftDt[:, 5]))
            consVelocitiesLadden = np.arange(minLaddenSpeedn, np.round(maxLaddenSpeedn) + 1)

            stepVelRanges = int(np.round(len(consVelocitiesLadden) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocitiesLadden), stepVelRanges):
                consVelocitiesRanges.append(consVelocitiesLadden[i])

            consVelocitiesRanges.append(
                consVelocitiesLadden[len(consVelocitiesLadden) - 1]) if consVelocitiesRanges.__len__() < 4 else \
                consVelocitiesRanges
            workbook._sheets[1]['B2'] = meanDradftThird

            print(consVelocitiesRanges)
            #minDraftWF = categoriesDraft[0]
            #maxDraftWF = categoriesDraft[1]

            thirdCategoryDraft = np.array([k for k in dtNew if k[1] > minDraftWF and k[1] <= maxDraftWF]).astype(float)

            speedRange1 = np.array([k for k in thirdCategoryDraft if k[5] >= minSpeedWF and k[5] <= maxSpeedWF])

            try:

                sizesSwell = []
                swell = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxSwell = 8
                minSwell = 0
                i = 0
                rawSwell = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 3])
                while i <= maxSwell:

                    speedArray = np.array([k for k in rawSwell if float(k[13]) >= i and float(k[13]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesSwell.append(speedArray.__len__())
                        swell.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(swell)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesSwell)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                swell01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                swell12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                swell23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                swell34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                swell45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                swell56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                swell67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                swell78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(swell01)
                p2_12 = p2(swell12)
                p2_23 = p2(swell23)
                p2_34 = p2(swell34)
                p2_45 = p2(swell45)
                p2_56 = p2(swell56)
                p2_67 = p2(swell67)
                p2_78 = p2(swell78)

                factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]

                '''factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01))),
                             abs((np.mean(p2_23) - np.mean(p2_12))),
                             abs((np.mean(p2_34) - np.mean(p2_23))),
                             abs((np.mean(p2_45) - np.mean(p2_34))),
                             abs((np.mean(p2_56) - np.mean(p2_45))),
                             abs((np.mean(p2_67) - np.mean(p2_56))),
                             abs((np.mean(p2_78) - np.mean(p2_67)))]'''
                x = 0
            except:
                print('EXCEPTION IN WEIGHTS SWELL THIRD GROUP')
                weightsSWH79 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsSWH911 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsSWH1114 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                ############################################################################################################################
                ############################################################################################################################

        
            wsLen = 1000000000000000000
            try:

                ###############################

                sizesWS = []
                ws = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWS = 8
                minWS = 0
                i = 0
                rawWS = np.array([k for k in speedRange1 if float(k[13]) >= 0 and float(k[13]) <= 2])
                while i <= maxWS:

                    speedArray = np.array([k for k in rawWS if float(k[4]) >= i and float(k[4]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesWS.append(speedArray.__len__())
                        ws.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(ws)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWS)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                ws01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                ws12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                ws23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                ws34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                ws45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                ws56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                ws67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                ws78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(ws01)
                p2_12 = p2(ws12)
                p2_23 = p2(ws23)
                p2_34 = p2(ws34)
                p2_45 = p2(ws45)
                p2_56 = p2(ws56)
                p2_67 = p2(ws67)
                p2_78 = p2(ws78)

                factorSWS = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]

            except:
                print('EXCEPTION IN WEIGHTS WS LADDEN')
                weightsWS79 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsWS911 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsWS1114 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                ##########################################WIND DIRECTION ###############################################
                ##########################################WIND DIRECTION ###############################################

            try:

                ##################################################################################

                sizesWD = []
                wd = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWD = 8
                minWD = 0
                i = 0
                listWD = [0, 22.5, 67.5, 112.5, 157.5, 180]
                rawWD = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 2])
                while i < len(listWD) - 1:

                    speedArray = np.array(
                        [k for k in rawWD if float(k[3]) >= listWD[i] and float(k[3]) <= listWD[i + 1]])

                    if speedArray.__len__() > 0:
                        sizesWD.append(speedArray.__len__())
                        wd.append(np.mean(listWD[i:i + 1]))
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(wd)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWD)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 180)

                wd01 = np.array([random.uniform(0, 22.5) for p in range(0, 100)])
                wd12 = np.array([random.uniform(22.5, 67.5) for p in range(0, 100)])
                wd23 = np.array([random.uniform(67.5, 112.5) for p in range(0, 100)])
                wd34 = np.array([random.uniform(112.5, 157.5) for p in range(0, 100)])
                wd45 = np.array([random.uniform(157.5, 180) for p in range(0, 100)])

                p2_01 = p2(wd01)
                p2_12 = p2(wd12)
                p2_23 = p2(wd23)
                p2_34 = p2(wd34)
                p2_45 = p2(wd45)

                factorSWD = [abs((np.mean(p2_01) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_12) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_23) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             0]

                g = 0

            except Exception as e:
                print('EXCEPTION IN WEIGHTS WD LADDEN')
                print(str(e))
            

            ###############################################################################################
            ###############################################################################################

            speedFoc = np.array(
                [k for k in thridCategoryDraftDt if
                 (k[5] >= consVelocities[0] and k[5] <= maxLaddenSpeedn) and (k[4] >= 0 and k[4] <= 3) and k[8] > 1])#

            '''meanFoc = np.mean(speedFoc[:, 8])
            stdFoc = np.std(speedFoc[:, 8])
            speedFoc = np.array([k for k in speedFoc if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''

            '''for i in range(0, len(speedFoc)):
                  #ballastDt[i,12] = 1 if ballastDt[i,12]==0 else 1
                  speedFoc[i] = np.mean(speedFoc[i:i + 15], axis=0)'''

            foc = np.round(speedFoc[:, 8], 3)  # .reshape(-1,1)
            speed = np.round(speedFoc[:, 5], 3)  # .reshape(-1,1)

            # lrSpeedFoc = LinearRegression()
            # lrSpeedFoc = RandomForestRegressor()
            lrSpeedFoc = SplineRegression.Earth(max_degree=2, )

            lrSpeedFoc.fit(speed.reshape(-1, 1), foc.reshape(-1, 1))
            print(lrSpeedFoc.score(speed.reshape(-1, 1), foc.reshape(-1, 1)))
            # plt.scatter(speed, foc , alpha=0.4, linewidth=4)
            # plt.plot(speed, lrSpeedFoc.predict(np.array(speed).reshape(-1, 1)))
            # plt.show()
            # rfSpeedFoc.fit(trainX.reshape(-1,1), trainY.reshape(-1,1))
            # lrSpeedFoc.fit(trainX.reshape(-1,1), trainY.reshape(-1,1))
            # testPreds = lrSpeedFoc.predict(testX.reshape(-1, 1))
            # print("LR SCORE: "+str(lrSpeedFoc.score(testX.reshape(-1,1),testY.reshape(-1,1))))
            # print("SR SCORE: " + str(lrSpeedFoc.score(testX.reshape(-1,1), testY.reshape(-1,1))))
            # print("SR MAE: " + str(mean_absolute_error(testY.reshape(-1,1), testPreds.reshape(-1,1))))

            minfoc = np.min(foc)
            maxfoc = np.max(foc)

            minspeed = np.min(speed)
            maxspeed = np.max(speed)

            focsApp = []
            meanSpeeds = []
            stdSpeeds = []
            ranges = []
            k = 0
            i = minfoc if minfoc >= 0 else 1
            i = minspeed if minspeed >= 0 else 1

            focsPLot = []
            speedsPlot = []

            while i <= maxspeed:
                # workbook._sheets[sheet].insert_rows(k+27)

                focArray = np.array([k for k in speedFoc if float(k[5]) >= i - 0.25 and float(k[5]) <= i + 0.25])
                # focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
                '''meanFoc = np.mean(focArray[:, 8])
                stdFoc = np.std(focArray[:, 8])
                speedFoc = np.array([k for k in focArray if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''

                if focArray.__len__() > 1:
                    focsPLot.append(focArray.__len__())
                    speedsPlot.append(i)
                    ranges.append(np.mean(focArray[:, 8]))
                    # lrSpeedFoc.fit(focArray[:,5].reshape(-1, 1), focArray[:,8].reshape(-1, 1))
                i += 1
                k += 1

            xi = np.array(speedsPlot)
            yi = np.array(ranges)
            zi = np.array(focsPLot)

            p2 = np.poly1d(np.polyfit(xi, yi, 1, ), )

            # Change color with c and alpha
            plt.clf()
            xp = np.linspace(min(xi), max(xi), 100)

            plt.plot([], [], '.', xp, p2(xp))
            speedList = [8, 9, 10, 11, 12, 13, 14]

            # plt.plot( xi, p2(xi),c='red')

            plt.scatter(xi, yi, s=zi / 10, c="red", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi)), np.ceil(max(xi)) + 1, 1))
            # plt.yticks(np.arange(min(yi), max(yi) + 1, 5))
            plt.xlabel("Speed (knots)")
            plt.ylabel("FOC (MT / day)")
            plt.title("Density plot", loc="center")

            #dataModel = KMeans(n_clusters=3)
            #zi = zi.reshape(-1, 1)
            #dataModel.fit(zi)
            # Extract centroid values
            #centroids = dataModel.cluster_centers_
            #ziSorted = np.sort(centroids, axis=0)

            #for z in ziSorted:
                #plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0] / 10),
                            #label='       ' + str(int(np.floor(z[0]))) + ' obs.')
            #plt.legend(borderpad=4, scatterpoints=1, frameon=True, labelspacing=6, title='# of obs')

            #fig = matplotlib.pyplot.gcf()
            #fig.set_size_inches(17.5, 9.5)
            # fig.savefig('./Figures/' + company + '_' + vessel + '_3.png', dpi=96)
            # plt.clf()
            # img = Image('./Figures/' + company + '_' + vessel + '_3.png')

            # workbook._sheets[1].add_image(img, 'F' + str(490))
            ###END OF BEST FIT
            ###############################################################################

            # [0, 0.25, 0.45, 0.65, 0.75, 1.15, 1.25, 1.35, 1.45]

            velMin = 7.75
            velMax = 8.25
            consVelocities = np.arange(np.round(minspeed), np.ceil(maxspeed), )

            # row = [10, 19, 28, 37, 46, 55, 64, 73]
            windForceWeightsList = [factorSWS, factorSWS, factorSWS]
            windDirWeightsList = [factorSWD, factorSWD, factorSWD]
            swellHeightWeightsList = [factorSWH, factorSWH, factorSWH]
            wind = [0, 22.5, 67.5, 112.5, 157.5, 180]

            windF = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            swellH = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            column = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            step = 1
            arrayFocThirdDraft = []
            arrayFoc = []

            consVelocities = np.arange(np.round(minLaddenSpeedn), np.round(maxLaddenSpeedn), )

            maxLaddenSpeedn = maxLaddenSpeedn if maxLaddenSpeedn <= consVelocities[len(consVelocities) - 1] else \
            consVelocities[len(consVelocities) - 1] - 1

            consVelocitiesJSON = np.arange(np.round(minLaddenSpeedn), np.round(maxLaddenSpeedn), 0.5)
            stepVelRanges = int(np.round(len(consVelocities) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocities), stepVelRanges):
                consVelocitiesRanges.append(consVelocities[i])

            consVelocitiesRanges.append(
                consVelocitiesLadden[len(consVelocities) - 1]) if consVelocitiesRanges.__len__() < 4 else None

            for vel in range(0, len(consVelocitiesJSON)):

                if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                    row = rows[int(vel / 2)]
                    workbook._sheets[1]['B' + str(row[0] - 4)] = consVelocitiesJSON[vel]
                if vel == len(consVelocitiesJSON) - 1:
                    workbook._sheets[1]['A' + str(row[0] - 4)] = 'MAX SPEED GROUP'
                    workbook._sheets[1].delete_rows(row[len(row) - 1] + 6, 1000)

                if consVelocitiesJSON[vel] >= consVelocitiesRanges[0] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[1]:

                    windDirWeights = windDirWeightsList[0]
                    swellHeightWeights = swellHeightWeightsList[0]
                    windForceWeights = windForceWeightsList[0]

                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[1] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[2]:

                    windDirWeights = windDirWeightsList[1]
                    swellHeightWeights = swellHeightWeightsList[1]
                    windForceWeights = windForceWeightsList[1]

                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[2] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[3]:
                    windDirWeights = windDirWeightsList[2]
                    swellHeightWeights = swellHeightWeightsList[2]
                    windForceWeights = windForceWeightsList[2]

                outerItem = {"draft": meanDradftThird, "speed": (consVelocitiesJSON[vel]), "cells": []}

                centralMean = lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1, 1))[0]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                stw = consVelocitiesJSON[vel]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                # lrSpeedFoc.predict(np.array([consVelocities[vel]]).reshape(-1,1))[0]
                # p2(consVelocities[vel])
                thridCategoryDraftDt7801 = []
                for w in range(0, len(windF) - 1):
                    for s in range(0, len(swellH) - 1):
                        thridCategoryDraftDt7_8 = []
                        numberOfApp11_8 = []
                        for i in range(0, len(wind) - 1):
                            ####arrayFoc missing
                            if arrayFoc.__len__() > minAccThres:
                                meanArrayFoc = np.mean(arrayFoc[:, 8])
                                stdArrayFoc = np.std(arrayFoc[:, 8])
                                arrayFoc = np.array([k for k in arrayFoc if
                                                     k[8] >= meanArrayFoc - (2 * stdArrayFoc) and k[
                                                         8] <= meanArrayFoc + (
                                                             2 * stdArrayFoc)])

                                steamTime = arrayFoc[:, 12]

                            numberOfApp11_8.append(arrayFoc.__len__())  # + centralArray.__len__())
                            '''rawFoc = np.array([k for k in thridCategoryDraftDt if (k[5]>= stw - 0.25 and k[5]<=stw+0.25) and
                            (k[4]>=windF[w+1] and k[4]<=windF[w+1]) and (k[13] >=swellH[s] and k[13]<=swellH[s+1] ) and
                            (k[3]>= wind[i] and k[3]< wind[i+1] )])
                            meanRawFoc = np.round(np.mean(rawFoc[:,8]),2) if len(rawFoc) > 0 else 0'''
                            if (s > 0 and w >= 0):
                                cellValue = round(
                                    (thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 5] + (
                                        thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 5] *    swellHeightWeights[s])),
                                    2)
                            elif s == 0 and w == 0:
                                cellValue = round(centralMean,2)  # + (centralMean * windDirWeights[i]), 2)

                            elif s == 0 and w > 0:
                                if w < 3:
                                    cellValue = round(
                                        thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] + (
                                                    thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] * (windForceWeights[w])), 2)
                                elif w == 3:
                                    cellValue = round(
                                        thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] + (
                                                thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] * (windDirWeights[i])), 2)
                                elif w > 3:
                                    cellValue = round(
                                        thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] + (
                                                thridCategoryDraftDt7801[len(thridCategoryDraftDt7801) - 40] * (windForceWeights[w])), 2)

                            # cellValue = meanRawFoc
                            '''lstmPoint =[]
                            pPoint = np.array([meanDraftLadden,(wind[i]), (windF[w]+windF[w+1])/2,consVelocitiesJSON[vel], (swellH[s])])

                            #lstmPoint.append(np.array(
                                    #[meanDraftLadden, (wind[i]), ((windF[w] + windF[w + 1]) / 2) - 1,
                                     #consVelocitiesJSON[vel], (swellH[s]) - 1]))

                            if s==0 and w==0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = w, w + 1, 0.2
                            elif s>0 and w>=0:
                                    startS, endS, stepS = (s+1)-2, s+1, 0.5
                                    startW, endW, stepW = w , w+1, 0.2
                            elif s==0 and w>0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = (w+1)-2, w , 0.5

                            countS = 0
                            countW = 0
                            countWd = 0
                            stepWd=5

                            wd = np.linspace(wind[i] ,wind[i+1],n_steps)
                            wf = np.linspace(windF[w]+0.1, windF[w + 1], n_steps)
                            swh = np.linspace(swellH[s]+0.1, swellH[s + 1], n_steps)
                            stw = np.linspace(consVelocitiesJSON[vel], consVelocitiesJSON[vel+1] if vel < len(consVelocitiesJSON)-1 else consVelocitiesJSON[vel]+1, n_steps)
                            for k in np.arange(0,n_steps):

                                    lstmPoint.append(np.array(
                                        [meanDraftLadden, (wd[k]), wf[k],
                                         stw[k], (swh[k])]))
                                    #lstmPoint.append(np.array(
                                            #[meanDraftLadden, (wd[k]), windF[startW]+ countW,
                                             #stw[k], (swellH[startS])+ countS]))

                                    countS += stepS
                                    countS = countS - stepS if swellH[startS]+ countS > endS else countS
                                    countW += stepW
                                    countW = countW - stepW if windF[startW]+ countW > endW else countW
                            #lstmPoint.append(pPoint)
                            lstmPoint=np.array(lstmPoint).reshape(n_steps,-1)
                            XSplineVectors=[]
                            for j in range(0,len(lstmPoint)):
                                    pPoint = lstmPoint[j]
                                    vector  , interceptsGen = dm.extractFunctionsFromSplines('Gen',pPoint[0], pPoint[1], pPoint[2], pPoint[3],pPoint[4])
                                    #vector = list(np.where(np.array(vector) < 0, 0, np.array(vector)))
                                    #vector = ([abs(k) for k in vector])
                                    XSplineVector = np.append(pPoint, vector)
                                    XSplineVector = np.array(XSplineVector).reshape(1, -1)
                                    XSplineVectors.append(XSplineVector)
                            XSplineVectors = np.array(XSplineVectors).reshape(n_steps,-1)
                            #XSplineVectors = lstmPoint
                            XSplineVector = XSplineVectors.reshape(1,XSplineVectors.shape[0], XSplineVectors.shape[1])
                            cellValue = float(currModeler.predict(XSplineVector)[0][0])
                            cellValue = np.round((cellValue),2)'''
                            item = {"windBFT": w + 1, "windDir": i + 1, "swell": s + 1, "cons": cellValue}
                            outerItem['cells'].append(item)
                            thridCategoryDraftDt7_8.append(cellValue)
                            thridCategoryDraftDt7801.append(cellValue)
                            #if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                                #arrayFocThirdDraft.append(cellValue)

                        if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                            for i in range(row[w], row[w] + 5):
                                try:
                                    workbook._sheets[1][column[s - 1 if s == 8 else s] + str(i)] = str(
                                        thridCategoryDraftDt7_8[i - row[w]])
                                    arrayFocThirdDraft.append(thridCategoryDraftDt7_8[i - row[w]])# + '(' + str(numberOfApp11_8[i - row[w]]) + ')'
                                    workbook._sheets[1][column[s - 1 if s == 8 else s] + str(i)].alignment = Alignment(
                                        horizontal='right')
                                except:
                                    print("Exception")
                    lastLenthridCategoryDraftDt7801 = len(thridCategoryDraftDt7801)
                json_decoded['ConsumptionProfile']['consProfile'].append(outerItem)
            # workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
            # return
            ladenSPEEDMin = thridCategoryDraftDt7_8
            ####################END 8 SPEED ########################################################################
            ####################END 8 SPEED #######################################################################

        if ballastFlag == True:
            print("BALLAST")
            minBallastSpeedn = consVelocities[0]
            maxBallastSpeedn = np.max(ballastDt[:, 5])

            consVelocitiesBallast = np.arange(minBallastSpeedn, np.round(maxBallastSpeedn)+1)

            stepVelRanges = int(np.round(len(consVelocitiesBallast) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocitiesBallast), stepVelRanges):
                consVelocitiesRanges.append(consVelocitiesBallast[i])
            #consVelocitiesRanges.append(17)

            consVelocitiesRanges.append(
                consVelocitiesBallast[len(consVelocitiesBallast) - 1]) if consVelocitiesRanges.__len__() < 4 else \
                consVelocitiesRanges

            workbook._sheets[1]['B2'] = meanDraftBallast


            print(consVelocitiesRanges)

            #minDraftWF = categoriesDraft[1]
            #maxDraftWF = categoriesDraft[2]
            ballast = np.array([k for k in dtNew if k[1] > minDraftWF and k[1] <= maxDraftWF]).astype(float)
            speedRange1 = np.array([k for k in ballast if k[5] >= minSpeedWF and k[5] <= maxSpeedWF])


            try:

                sizesSwell = []
                swell = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxSwell = 8
                minSwell = 0
                i = 0
                rawSwell = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 3])
                while i <= maxSwell:

                    speedArray = np.array([k for k in rawSwell if float(k[13]) >= i and float(k[13]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesSwell.append(speedArray.__len__())
                        swell.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(swell)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesSwell)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                swell01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                swell12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                swell23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                swell34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                swell45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                swell56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                swell67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                swell78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(swell01)
                p2_12 = p2(swell12)
                p2_23 = p2(swell23)
                p2_34 = p2(swell34)
                p2_45 = p2(swell45)
                p2_56 = p2(swell56)
                p2_67 = p2(swell67)
                p2_78 = p2(swell78)

                factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]
                '''factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01)) ),
                             abs((np.mean(p2_23) - np.mean(p2_12)) ),
                             abs((np.mean(p2_34) - np.mean(p2_23)) ),
                             abs((np.mean(p2_45) - np.mean(p2_34)) ),
                             abs((np.mean(p2_56) - np.mean(p2_45)) ),
                             abs((np.mean(p2_67) - np.mean(p2_56)) ),
                             abs((np.mean(p2_78) - np.mean(p2_67)) )]'''
                u=0
            except:
                print('EXCEPTION IN WEIGHTS SWELL BALLAST')
                weightsSWH79 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsSWH911 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsSWH1114 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                ############################################################################################################################
                ############################################################################################################################

            try:

                sizesWS = []
                ws = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWS = 8
                minWS = 0
                i = 0
                rawWS = np.array([k for k in speedRange1 if float(k[13]) >= 0 and float(k[13]) <= 2  and float(k[13])>=0 and float(k[13]) <=2])
                while i <= maxWS:

                    speedArray = np.array([k for k in rawWS if float(k[4]) >= i and float(k[4]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesWS.append(speedArray.__len__())
                        ws.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(ws)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWS)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                ws01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                ws12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                ws23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                ws34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                ws45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                ws56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                ws67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                ws78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(ws01)
                p2_12 = p2(ws12)
                p2_23 = p2(ws23)
                p2_34 = p2(ws34)
                p2_45 = p2(ws45)
                p2_56 = p2(ws56)
                p2_67 = p2(ws67)
                p2_78 = p2(ws78)

                factorSWS = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]

            except:
                print('EXCEPTION IN WEIGHTS WS BALLAST')
                weightsWS79 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsWS911 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                weightsWS1114 = [0, 0.0043, 0.0023, 0.0024, 0.0025, 0.0046, 0.0057, 0.0058, 0.0059]
                ##########################################WIND DIRECTION ###############################################
                ##########################################WIND DIRECTION ###############################################


            try:

                sizesWD = []
                wd = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWD = 8
                minWD = 0
                i = 0
                listWD = [0, 22.5, 67.5, 112.5, 157.5, 180]
                rawWD = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 2  and float(k[13])>=0 and float(k[13]) <=2])
                while i < len(listWD)-1:

                    speedArray = np.array(
                        [k for k in rawWD if float(k[3]) >= listWD[i] and float(k[3]) <= listWD[i + 1]])

                    if speedArray.__len__() > 0:
                        sizesWD.append(speedArray.__len__())
                        wd.append(np.mean(listWD[i:i + 1]))
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(wd)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWD)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 180)

                wd01 = np.array([random.uniform(0, 22.5) for p in range(0, 100)])
                wd12 = np.array([random.uniform(22.5, 67.5) for p in range(0, 100)])
                wd23 = np.array([random.uniform(67.5, 112.5) for p in range(0, 100)])
                wd34 = np.array([random.uniform(112.5, 157.5) for p in range(0, 100)])
                wd45 = np.array([random.uniform(157.5, 180) for p in range(0, 100)])

                p2_01 = p2(wd01)
                p2_12 = p2(wd12)
                p2_23 = p2(wd23)
                p2_34 = p2(wd34)
                p2_45 = p2(wd45)

                factorSWD = [abs((np.mean(p2_01) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_12) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_23) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             0]



            except:
                print('EXCEPTION IN WEIGHTS WD BALLAST')
                weightsWD79 = [0.2, 0.1, 0.09, 0.08, 0.05]
                weightsWD911 = [0.2, 0.1, 0.09, 0.08, 0.05]
                weightsWD1114 = [0.2, 0.1, 0.09, 0.08, 0.05]

            ###############################################################################################
            ###############################################################################################

            speedFoc = np.array(
                [k for k in ballastDt if (k[5] >= consVelocities[0] and k[5] <= maxBallastSpeedn) and (k[4] >= 0 and k[4] <=3) and k[8] > 1])#

            ##ballastDt
            '''meanFoc = np.mean(speedFoc[:, 8])
            stdFoc = np.std(speedFoc[:, 8])
            speedFoc = np.array(
                [k for k in speedFoc if k[8] >= (meanFoc - (2 * stdFoc)) and k[8] <= (meanFoc + (2 * stdFoc))])'''

            '''for i in range(0, len(speedFoc)):
                  #ballastDt[i,12] = 1 if ballastDt[i,12]==0 else 1
                  speedFoc[i] = np.mean(speedFoc[i:i + 15], axis=0)'''

            foc = speedFoc[:, 8]#.reshape(-1,1)
            speed = speedFoc[:, 5]# .reshape(-1,1)

            #lrSpeedFoc = LinearRegression()
            # rfSpeedFoc = RandomForestRegressor()
            lrSpeedFoc = SplineRegression.Earth(max_degree=2,)

            #trainX,testX, trainY,testY = train_test_split(speed,foc, test_size=0.2,random_state=42)
            '''tscv = TimeSeriesSplit()
            for train_index, test_index in tscv.split(speed):
                # print("TRAIN:", train_index, "TEST:", test_index)
                trainX, testX = speed[train_index], speed[test_index]
                trainY, testY = foc[train_index], foc[test_index]'''

            lrSpeedFoc.fit(speed.reshape(-1, 1), foc.reshape(-1, 1))
            print(lrSpeedFoc.score(speed.reshape(-1, 1), foc.reshape(-1, 1)))
                #print("SR SCORE: " + str(lrSpeedFoc.score(testX.reshape(-1, 1), testY.reshape(-1, 1))))

            minfoc = np.min(foc)
            maxfoc = np.max(foc)

            minspeed = np.min(speed)
            maxspeed = np.max(speed)

            focsApp = []
            meanSpeeds = []
            stdSpeeds = []
            ranges = []
            k = 0
            i = minfoc if minfoc >= 0 else 1
            i = minspeed if minspeed >= 0 else 1

            focsPLot = []
            speedsPlot = []

            while i <= maxspeed:
                # workbook._sheets[sheet].insert_rows(k+27)

                focArray = np.array([k for k in speedFoc if float(k[5]) >= i-0.25 and float(k[5]) <= i + 0.25])
                # focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
                '''meanFoc = np.mean(focArray[:, 8])
                stdFoc = np.std(focArray[:, 8])
                focArray = np.array(
                    [k for k in focArray if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''

                if focArray.__len__() > 1:
                    focsPLot.append(focArray.__len__())
                    speedsPlot.append(i)
                    ranges.append(np.mean(focArray[:, 8]))
                    #lrSpeedFoc.fit(focArray[:, 5].reshape(-1, 1), focArray[:, 8].reshape(-1, 1))
                i += 0.5
                k += 1

            xi = np.array(speedsPlot)
            yi = np.array(ranges)
            zi = np.array(focsPLot)

            p2 = np.poly1d(np.polyfit(speed, foc, 1))

            plt.clf()
            xp = np.linspace(min(xi), max(xi), 100)
            plt.plot([], [], '.', xp, p2(xp))

            plt.scatter(xi, yi, s=zi / 10, c="red", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi)), np.ceil(max(xi)) + 1, 1))
            # plt.yticks(np.arange(min(yi), max(yi) + 1, 5))
            plt.xlabel("Speed (knots)")
            plt.ylabel("FOC (MT / day)")
            plt.title("Density plot", loc="center")

            dataModel = KMeans(n_clusters=3)
            zi = zi.reshape(-1, 1)
            dataModel.fit(zi)
            # Extract centroid values
            centroids = dataModel.cluster_centers_
            ziSorted = np.sort(centroids, axis=0)

            for z in ziSorted:
                plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0] / 10),
                            label='       ' + str(int(np.floor(z[0]))) + ' obs.')
            plt.legend(borderpad=4, scatterpoints=1, frameon=True, labelspacing=6, title='# of obs')

            fig = matplotlib.pyplot.gcf()
            fig.set_size_inches(17.5, 9.5)
            fig.savefig('./Figures/' + company + '_' + vessel + '_4.png', dpi=96)
            # plt.clf()
            img = Image('./Figures/' + company + '_' + vessel + '_4.png')
            ###############################################################################
            workbook._sheets[2].add_image(img, 'F' + str(490))


            windForceWeights = [0, 0.25, 0.35, 0.45, 0.55, 0.65, 0.75, 0.85, 1.05]
            # [0, 0.25, 0.45, 0.65, 0.75, 1.15, 1.25, 1.35, 1.45]

            velMin = 7.75
            velMax = 8.25

            # row = [10, 19, 28, 37, 46, 55, 64, 73]
            windDirWeightsList = [factorSWD, factorSWD, factorSWD]
            windForceWeightsList = [factorSWS, factorSWS, factorSWS]
            swellHeightWeightsList = [factorSWH, factorSWH, factorSWH]
            wind = [0, 22.5, 67.5, 112.5, 157.5, 180]
            windF = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            swellH = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            column = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            step = 1
            arrayFocBallast = []
            arrayFoc = []
            consVelocities = np.arange(np.round(minBallastSpeedn),np.round(maxBallastSpeedn), )
            consVelocitiesJSON = np.arange(np.round(minBallastSpeedn), np.round(maxBallastSpeedn),0.5 )
            stepVelRanges = int(np.round(len(consVelocities) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocities), stepVelRanges):
                consVelocitiesRanges.append(consVelocities[i])

            consVelocitiesRanges.append(
                consVelocitiesLadden[len(consVelocities) - 1]) if consVelocitiesRanges.__len__() < 4 else \
                consVelocitiesRanges
            #consVelocitiesRanges.append(17)
            rowCounter = 0
            for vel in range(0, len(consVelocitiesJSON)):
                if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel])==0:
                    row = rows[int(vel/2)]
                    workbook._sheets[2]['B' + str(row[0] - 4)] = consVelocitiesJSON[vel]
                if vel == len(consVelocitiesJSON) - 1:
                    workbook._sheets[2]['A' + str(row[0] - 4)] = 'MAX SPEED GROUP'
                    workbook._sheets[2].delete_rows(row[len(row) - 1] + 6, 1000)
                if consVelocitiesJSON[vel] >= consVelocitiesRanges[0] and consVelocitiesJSON[vel] <= consVelocitiesRanges[1]:
                    windDirWeights = windDirWeightsList[0]
                    swellHeightWeights = swellHeightWeightsList[0]
                    windForceWeights = windForceWeightsList[0]
                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[1] and consVelocitiesJSON[vel] <= consVelocitiesRanges[2]:
                    windDirWeights = windDirWeightsList[1]
                    swellHeightWeights = swellHeightWeightsList[1]
                    windForceWeights = windForceWeightsList[1]
                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[2] and consVelocitiesJSON[vel] <= consVelocitiesRanges[3]:
                    windDirWeights = windDirWeightsList[2]
                    swellHeightWeights = swellHeightWeightsList[2]
                    windForceWeights = windForceWeightsList[2]

                outerItem = {"draft": meanDraftBallast, "speed": (consVelocitiesJSON[vel] ),"cells": []}


                #centralMean = lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                    #lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                    #lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                centralMean  = p2(consVelocitiesJSON[vel])
                #lrSpeedFoc.predict(np.array([consVelocities[vel]]).reshape(-1,1))[0]

                if consVelocitiesJSON[vel]==12:
                    x=0
                ballastDt7801 = []
                for w in range(0, len(windF) - 1):
                    for s in range(0, len(swellH) - 1):
                        ballastDt7_8 = []
                        numberOfApp11_8 = []
                        for i in range(0, len(wind) - 1):
                            ####arrayFoc missing
                            if arrayFoc.__len__() > minAccThres:
                                meanArrayFoc = np.mean(arrayFoc[:, 8])
                                stdArrayFoc = np.std(arrayFoc[:, 8])
                                arrayFoc = np.array([k for k in arrayFoc if
                                                     k[8] >= meanArrayFoc - (2 * stdArrayFoc) and k[
                                                         8] <= meanArrayFoc + (
                                                             2 * stdArrayFoc)])

                                steamTime = arrayFoc[:, 12]
                            # tlgarrayFoc = arrayFoc[:, 9] if arrayFoc.__len__() > minAccThres else []
                            # tlgarrayFoc = np.array([k for k in tlgarrayFoc if k > 5])
                            # tlgarrayFoc = np.array([k for k in ballastDt if k[5] > velMax and k[5] <= velMax and k[8] > 10])
                            tlgarrayFoc = []
                            if tlgarrayFoc.__len__() > lenConditionTlg:
                                tlgarrayFoc = np.array(

                                    [k for k in ballastDt if k[5] > velMax and k[5] <= velMax and k[8] > 10])[:, 9]
                                meanFoc = (np.mean(arrayFoc[:, 8]) + np.mean(
                                    tlgarrayFoc) + centralMean) / 3 if arrayFoc.__len__() > minAccThres else (
                                                                                                                     centralMean + np.mean(
                                                                                                                 tlgarrayFoc)) / 2
                                numberOfApp11_8.append(
                                    arrayFoc.__len__() + tlgarrayFoc.__len__() + centralArray.__len__())
                            else:
                                # np.average(arrayFoc[:, 8],weights=steamTime)
                                # weighted_avgFocArray = np.average(arrayFoc[:, 8],
                                # weights=steamTime) if arrayFoc.__len__() > minAccThres else centralMean
                                # meanFoc = (weighted_avgFocArray + centralMean) / 2 if arrayFoc.__len__() > minAccThres else centralMean
                                numberOfApp11_8.append(arrayFoc.__len__())  # + centralArray.__len__())

                                if (s > 0 and w >= 0):
                                    cellValue = round(
                                        (ballastDt7801[len(ballastDt7801) - 5] + (ballastDt7801[len(ballastDt7801) - 5] * swellHeightWeights[s])),2)
                                elif s == 0 and w == 0:
                                    #if  320 * rowCounter < len(arrayFocThirdDraft):
                                        #cellValue = round(arrayFocThirdDraft[ 320 * rowCounter] + (arrayFocThirdDraft[ 320 * rowCounter] * factorDRAFT[1]), 2)
                                    #else:
                                        #cellValue = round( arrayFocBallast[len(arrayFocBallast)-320] + (arrayFocBallast[len(arrayFocBallast)-320] * factorDRAFT[1]), 2)
                                        #cellValue = round(centralMean + (centralMean * factorDRAFT[1]),2)
                                   #cellValue = round(centralMean + (centralMean * factorDRAFT[1]) , 2)
                                    cellValue = round(centralMean , 2)
                                elif s == 0 and w > 0:
                                    if w < 3:
                                        cellValue = round(
                                            ballastDt7801[len(ballastDt7801) - 40] + (
                                                        ballastDt7801[len(ballastDt7801) - 40] * (windForceWeights[w])), 2)
                                    elif w == 3:
                                        cellValue = round(
                                            ballastDt7801[len(ballastDt7801) - 40] + (
                                                    ballastDt7801[len(ballastDt7801) - 40] * (windDirWeights[i])), 2)
                                    elif w > 3:
                                        cellValue = round(
                                            ballastDt7801[len(ballastDt7801) - 40] + (
                                                    ballastDt7801[len(ballastDt7801) - 40] * (windForceWeights[w])), 2)
                                '''lstmPoint=[]

                                if s==0 and w==0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = w, w + 1, 0.2
                                elif s>0 and w>=0:
                                    startS, endS, stepS = (s+1)-2, s+1, 0.5
                                    startW, endW, stepW = w , w+1, 0.2
                                elif s==0 and w>0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = (w+1)-2, w , 0.5

                                countS = 0
                                countW = 0
                                countWd = 0
                                stepWd=5
                                wd = np.linspace(wind[i] + 0.1, wind[i + 1], n_steps)
                                wf = np.linspace(windF[w] + 0.1, windF[w + 1], n_steps)
                                swh = np.linspace(swellH[s] + 0.1, swellH[s + 1], n_steps)
                                stw = np.linspace(consVelocitiesJSON[vel], consVelocitiesJSON[vel+1] if vel < len(consVelocitiesJSON)-1 else consVelocitiesJSON[vel]+1, n_steps)
                                for k in np.arange(0,n_steps):

                                    lstmPoint.append(np.array(
                                        [meanDraftBallast, (wd[k]), wf[k],
                                        stw[k], (swh[k])]))
                                    #lstmPoint.append(np.array(
                                            #[meanDraftBallast, (wd[k]), windF[startW]+ countW,
                                             #consVelocitiesJSON[vel], (swellH[startS])+ countS]))

                                    countS += stepS
                                    countS = countS - stepS if swellH[startS]+ countS > endS else countS
                                    countW += stepW
                                    countW = countW - stepW if windF[startW]+ countW > endW else countW
                                #lstmPoint.append(pPoint)
                                lstmPoint=np.array(lstmPoint).reshape(n_steps,-1)
                                XSplineVectors=[]
                                for j in range(0,len(lstmPoint)):
                                    pPoint = lstmPoint[j]
                                    vector , interceptsGen = dm.extractFunctionsFromSplines('Gen',pPoint[0], pPoint[1], pPoint[2], pPoint[3],pPoint[4])
                                    #vectorNew = np.array([i + interceptsGen for i in vector])
                                    #vector = ([abs(k) for k in vector])
                                    #vector = list(np.where(np.array(vector) < 0, 0, np.array(vector)))
                                    XSplineVector = np.append(pPoint, vector)
                                    XSplineVector = np.array(XSplineVector).reshape(1, -1)
                                    XSplineVectors.append(XSplineVector)
                                XSplineVectors = np.array(XSplineVectors).reshape(n_steps,-1)
                                #XSplineVectors = lstmPoint
                                XSplineVector = XSplineVectors.reshape(1,XSplineVectors.shape[0], XSplineVectors.shape[1])
                                cellValue = float(currModeler.predict(XSplineVector)[0][0])
                                cellValue = np.round((cellValue), 2)'''
                                item = {"windBFT": w + 1, "windDir": i + 1, "swell": s + 1, "cons": cellValue}
                                outerItem['cells'].append(item)
                            ballastDt7_8.append(cellValue)
                            #arrayFocBallast.append(cellValue)
                            ballastDt7801.append(cellValue)
                        lastLenDt_8 = len(ballastDt7_8)
                        if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel])==0:

                            for i in range(row[w], row[w] + 5):
                                try:
                                    workbook._sheets[2][column[s - 1 if s == 8 else s] + str(i)] = str(
                                        ballastDt7_8[i - row[w]])  # + '(' + str(numberOfApp11_8[i - row[w]]) + ')'
                                    workbook._sheets[2][column[s - 1 if s == 8 else s] + str(i)].alignment = Alignment(
                                        horizontal='right')
                                    arrayFocBallast.append(ballastDt7_8[i - row[w]])
                                except:
                                    print("Exception")
                    lastLenballastDt7801 = len(ballastDt7801)
                if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                    rowCounter += 1

                json_decoded['ConsumptionProfile']['consProfile'].append(outerItem)
            # workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
            # return
            ladenSPEEDMin = ballastDt7_8
            ####################END 8 SPEED ########################################################################
            ####################END 8 SPEED #######################################################################

        if ladenFlag == True:

            minLaddenSpeedn = 9#consVelocities[0]
            # consVelocities[0]
            maxLaddenSpeedn = np.ceil(np.max(ladenDt[:, 5]))
            consVelocitiesLadden = np.arange(minLaddenSpeedn, np.round(maxLaddenSpeedn) + 1)

            stepVelRanges = int(np.round(len(consVelocitiesLadden) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocitiesLadden), stepVelRanges):
                consVelocitiesRanges.append(consVelocitiesLadden[i])

            consVelocitiesRanges.append(
                consVelocitiesLadden[len(consVelocitiesLadden) - 1]) if consVelocitiesRanges.__len__() < 4 else \
                consVelocitiesRanges
            workbook._sheets[3]['B2'] = meanDraftLadden

            print(consVelocitiesRanges)
            #minDraftWF = categoriesDraft[2]
            #maxDraftWF = categoriesDraft[3]

            laden = np.array([k for k in dtNew if k[1] > minDraftWF and k[1] <= maxDraftWF]).astype(float)

            speedRange1 = np.array([k for k in laden if k[5] >= minSpeedWF and k[5] <= maxSpeedWF])

            try:

                sizesSwell = []
                swell = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxSwell = 8
                minSwell = 0
                i = 0
                rawSwell = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 3])
                while i <= maxSwell:

                    speedArray = np.array([k for k in rawSwell if float(k[13]) >= i and float(k[13]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesSwell.append(speedArray.__len__())
                        swell.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(swell)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesSwell)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                swell01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                swell12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                swell23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                swell34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                swell45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                swell56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                swell67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                swell78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(swell01)
                p2_12 = p2(swell12)
                p2_23 = p2(swell23)
                p2_34 = p2(swell34)
                p2_45 = p2(swell45)
                p2_56 = p2(swell56)
                p2_67 = p2(swell67)
                p2_78 = p2(swell78)

                factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]

                '''factorSWH = [0, abs((np.mean(p2_12) - np.mean(p2_01)) ),
                             abs((np.mean(p2_23) - np.mean(p2_12)) ),
                             abs((np.mean(p2_34) - np.mean(p2_23)) ),
                             abs((np.mean(p2_45) - np.mean(p2_34)) ),
                             abs((np.mean(p2_56) - np.mean(p2_45)) ),
                             abs((np.mean(p2_67) - np.mean(p2_56))),
                             abs((np.mean(p2_78) - np.mean(p2_67)))]'''
                x = 0
            except:
                print('EXCEPTION IN WEIGHTS SWELL LADDEN')

                ############################################################################################################################
                ############################################################################################################################

            try:

                ###############################

                sizesWS = []
                ws = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWS = 8
                minWS = 0
                i = 0
                rawWS = np.array([k for k in speedRange1 if float(k[13]) >= 0 and float(k[13]) <= 2])
                while i <= maxWS:

                    speedArray = np.array([k for k in rawWS if float(k[4]) >= i and float(k[4]) <= i + 1])

                    if speedArray.__len__() > 0:
                        sizesWS.append(speedArray.__len__())
                        ws.append(i)
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(ws)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWS)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 8)

                ws01 = np.array([random.uniform(0, 1) for p in range(0, 100)])
                ws12 = np.array([random.uniform(1, 2) for p in range(0, 100)])

                ws23 = np.array([random.uniform(2, 3) for p in range(0, 100)])
                ws34 = np.array([random.uniform(3, 4) for p in range(0, 100)])
                ws45 = np.array([random.uniform(4, 5) for p in range(0, 100)])
                ws56 = np.array([random.uniform(5, 6) for p in range(0, 100)])
                ws67 = np.array([random.uniform(6, 7) for p in range(0, 100)])
                ws78 = np.array([random.uniform(7, 8) for p in range(0, 100)])

                p2_01 = p2(ws01)
                p2_12 = p2(ws12)
                p2_23 = p2(ws23)
                p2_34 = p2(ws34)
                p2_45 = p2(ws45)
                p2_56 = p2(ws56)
                p2_67 = p2(ws67)
                p2_78 = p2(ws78)

                factorSWS = [0, abs((np.mean(p2_12) - np.mean(p2_01)) / np.mean(p2_01)),
                             abs((np.mean(p2_23) - np.mean(p2_12)) / np.mean(p2_12)),
                             abs((np.mean(p2_34) - np.mean(p2_23)) / np.mean(p2_23)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_56) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_67) - np.mean(p2_56)) / np.mean(p2_56)),
                             abs((np.mean(p2_78) - np.mean(p2_67)) / np.mean(p2_67))]

            except:
                print('EXCEPTION IN WEIGHTS WS LADDEN')

                ##########################################WIND DIRECTION ###############################################
                ##########################################WIND DIRECTION ###############################################

            try:

                ##################################################################################

                sizesWD = []
                wd = []
                avgActualFoc = []
                minActualFoc = []
                maxActualFoc = []
                stdActualFoc = []
                maxWD = 8
                minWD = 0
                i = 0
                listWD = [0, 22.5, 67.5, 112.5, 157.5, 180]
                rawWD = np.array([k for k in speedRange1 if float(k[4]) >= 0 and float(k[4]) <= 2])
                while i < len(listWD) - 1:

                    speedArray = np.array(
                        [k for k in rawWD if float(k[3]) >= listWD[i] and float(k[3]) <= listWD[i + 1]])

                    if speedArray.__len__() > 0:
                        sizesWD.append(speedArray.__len__())
                        wd.append(np.mean(listWD[i:i + 1]))
                        avgActualFoc.append(np.mean(speedArray[:, 8]))
                        minActualFoc.append(np.min(speedArray[:, 8]))
                        maxActualFoc.append(np.max(speedArray[:, 8]))
                        stdActualFoc.append(np.std(speedArray[:, 8]))
                    i += 1

                xi = np.array(wd)
                yi = np.array(avgActualFoc)
                zi = np.array(sizesWD)

                p2 = np.poly1d(np.polyfit(xi, yi, 1))

                xp = np.arange(0, 180)

                wd01 = np.array([random.uniform(0, 22.5) for p in range(0, 100)])
                wd12 = np.array([random.uniform(22.5, 67.5) for p in range(0, 100)])
                wd23 = np.array([random.uniform(67.5, 112.5) for p in range(0, 100)])
                wd34 = np.array([random.uniform(112.5, 157.5) for p in range(0, 100)])
                wd45 = np.array([random.uniform(157.5, 180) for p in range(0, 100)])

                p2_01 = p2(wd01)
                p2_12 = p2(wd12)
                p2_23 = p2(wd23)
                p2_34 = p2(wd34)
                p2_45 = p2(wd45)

                factorSWD = [abs((np.mean(p2_01) - np.mean(p2_45)) / np.mean(p2_45)),
                             abs((np.mean(p2_12) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_23) - np.mean(p2_34)) / np.mean(p2_34)),
                             abs((np.mean(p2_45) - np.mean(p2_34)) / np.mean(p2_34)),
                             0]

                g = 0

            except Exception as e:
                print('EXCEPTION IN WEIGHTS WD LADDEN')
                print(str(e))
                weightsWD79 = [0.2, 0.1, 0.09, 0.08, 0.05]
                weightsWD911 = [0.2, 0.1, 0.09, 0.08, 0.05]
                weightsWD1114 = [0.2, 0.1, 0.09, 0.08, 0.05]

            ###############################################################################################
            ###############################################################################################

            speedFoc = np.array(
                [k for k in ladenDt if
                 (k[5] >= consVelocities[0] and k[5] <= maxLaddenSpeedn) and (k[4] >= 0 and k[4] <= 3)  and k[8] > 1]) #
            #ladenDt
            '''meanFoc = np.mean(speedFoc[:, 8])
            stdFoc = np.std(speedFoc[:, 8])
            speedFoc = np.array([k for k in speedFoc if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''

            '''for i in range(0, len(speedFoc)):
                  #ballastDt[i,12] = 1 if ballastDt[i,12]==0 else 1
                  speedFoc[i] = np.mean(speedFoc[i:i + 15], axis=0)'''

            foc = np.round(speedFoc[:, 8], 3)  # .reshape(-1,1)
            speed = np.round(speedFoc[:, 5], 3)  # .reshape(-1,1)

            # lrSpeedFoc = LinearRegression()
            # lrSpeedFoc = RandomForestRegressor()
            lrSpeedFoc = SplineRegression.Earth(max_degree=2, )

            lrSpeedFoc.fit(speed.reshape(-1, 1), foc.reshape(-1, 1))
            print(lrSpeedFoc.score(speed.reshape(-1, 1), foc.reshape(-1, 1)))
            # plt.scatter(speed, foc , alpha=0.4, linewidth=4)
            # plt.plot(speed, lrSpeedFoc.predict(np.array(speed).reshape(-1, 1)))
            # plt.show()
            # rfSpeedFoc.fit(trainX.reshape(-1,1), trainY.reshape(-1,1))
            # lrSpeedFoc.fit(trainX.reshape(-1,1), trainY.reshape(-1,1))
            # testPreds = lrSpeedFoc.predict(testX.reshape(-1, 1))
            # print("LR SCORE: "+str(lrSpeedFoc.score(testX.reshape(-1,1),testY.reshape(-1,1))))
            # print("SR SCORE: " + str(lrSpeedFoc.score(testX.reshape(-1,1), testY.reshape(-1,1))))
            # print("SR MAE: " + str(mean_absolute_error(testY.reshape(-1,1), testPreds.reshape(-1,1))))

            minfoc = np.min(foc)
            maxfoc = np.max(foc)

            minspeed = np.min(speed)
            maxspeed = np.max(speed)

            focsApp = []
            meanSpeeds = []
            stdSpeeds = []
            ranges = []
            k = 0
            i = minfoc if minfoc >= 0 else 1
            i = minspeed if minspeed >= 0 else 1

            focsPLot = []
            speedsPlot = []

            while i <= maxspeed:
                # workbook._sheets[sheet].insert_rows(k+27)

                focArray = np.array([k for k in speedFoc if float(k[5]) >= i - 0.25 and float(k[5]) <= i + 0.25])
                # focsApp.append(str(np.round(focArray.__len__() / focAmount * 100, 2)) + '%')
                '''meanFoc = np.mean(focArray[:, 8])
                stdFoc = np.std(focArray[:, 8])
                speedFoc = np.array([k for k in focArray if k[8] >= (meanFoc - (3 * stdFoc)) and k[8] <= (meanFoc + (3 * stdFoc))])'''

                if focArray.__len__() > 1:
                    focsPLot.append(focArray.__len__())
                    speedsPlot.append(i)
                    ranges.append(np.mean(focArray[:, 8]))
                    # lrSpeedFoc.fit(focArray[:,5].reshape(-1, 1), focArray[:,8].reshape(-1, 1))
                i += 1
                k += 1

            xi = np.array(speedsPlot)
            yi = np.array(ranges)
            zi = np.array(focsPLot)

            p2 = np.poly1d(np.polyfit(xi, yi, 1, ), )

            # Change color with c and alpha
            plt.clf()
            xp = np.linspace(min(xi), max(xi), 100)

            plt.plot([], [], '.', xp, p2(xp))
            speedList = [8, 9, 10, 11, 12, 13, 14]

            # plt.plot( xi, p2(xi),c='red')

            plt.scatter(xi, yi, s=zi / 10, c="red", alpha=0.4, linewidth=4)
            # plt.xticks(np.arange(np.floor(min(xi)), np.ceil(max(xi)) + 1, 1))
            # plt.yticks(np.arange(min(yi), max(yi) + 1, 5))
            plt.xlabel("Speed (knots)")
            plt.ylabel("FOC (MT / day)")
            plt.title("Density plot", loc="center")

            dataModel = KMeans(n_clusters=3)
            zi = zi.reshape(-1, 1)
            dataModel.fit(zi)
            # Extract centroid values
            centroids = dataModel.cluster_centers_
            ziSorted = np.sort(centroids, axis=0)

            for z in ziSorted:
                plt.scatter([], [], c='r', alpha=0.5, s=np.floor(z[0] / 10),
                            label='       ' + str(int(np.floor(z[0]))) + ' obs.')
            plt.legend(borderpad=4, scatterpoints=1, frameon=True, labelspacing=6, title='# of obs')

            fig = matplotlib.pyplot.gcf()
            fig.set_size_inches(17.5, 9.5)
            # fig.savefig('./Figures/' + company + '_' + vessel + '_3.png', dpi=96)
            # plt.clf()
            # img = Image('./Figures/' + company + '_' + vessel + '_3.png')

            # workbook._sheets[3].add_image(img, 'F' + str(490))
            ###END OF BEST FIT
            ###############################################################################

            # [0, 0.25, 0.45, 0.65, 0.75, 1.15, 1.25, 1.35, 1.45]

            velMin = 7.75
            velMax = 8.25
            consVelocities = np.arange(np.round(minspeed), np.ceil(maxspeed), )

            # row = [10, 19, 28, 37, 46, 55, 64, 73]
            windForceWeightsList = [factorSWS, factorSWS, factorSWS]
            windDirWeightsList = [factorSWD, factorSWD, factorSWD]
            swellHeightWeightsList = [factorSWH, factorSWH, factorSWH]
            wind = [0, 22.5, 67.5, 112.5, 157.5, 180]

            windF = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            swellH = [0, 1, 2, 3, 4, 5, 6, 7, 8]
            column = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
            step = 1

            arrayFocLaden = []
            consVelocities = np.arange(np.round(minLaddenSpeedn), np.round(maxLaddenSpeedn), )

            maxLaddenSpeedn = maxLaddenSpeedn if maxLaddenSpeedn <= consVelocities[len(consVelocities) - 1] else \
            consVelocities[len(consVelocities) - 1] - 1

            consVelocitiesJSON = np.arange(np.round(minLaddenSpeedn), np.round(maxLaddenSpeedn), 0.5)
            stepVelRanges = int(np.round(len(consVelocities) / 4))
            consVelocitiesRanges = []
            for i in range(0, len(consVelocities), stepVelRanges):
                consVelocitiesRanges.append(consVelocities[i])

            consVelocitiesRanges.append(
                consVelocitiesLadden[len(consVelocities) - 1]) if consVelocitiesRanges.__len__() < 4 else None
            rowCounter=0
            for vel in range(0, len(consVelocitiesJSON)):

                if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                    row = rows[int(vel / 2)]
                    workbook._sheets[3]['B' + str(row[0] - 4)] = consVelocitiesJSON[vel]
                if vel == len(consVelocitiesJSON) - 1:
                    workbook._sheets[3]['A' + str(row[0] - 4)] = 'MAX SPEED GROUP'
                    workbook._sheets[3].delete_rows(row[len(row) - 1] + 6, 1000)

                if consVelocitiesJSON[vel] >= consVelocitiesRanges[0] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[1]:

                    windDirWeights = windDirWeightsList[0]
                    swellHeightWeights = swellHeightWeightsList[0]
                    windForceWeights = windForceWeightsList[0]

                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[1] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[2]:

                    windDirWeights = windDirWeightsList[1]
                    swellHeightWeights = swellHeightWeightsList[1]
                    windForceWeights = windForceWeightsList[1]

                elif consVelocitiesJSON[vel] >= consVelocitiesRanges[2] and consVelocitiesJSON[vel] <= \
                        consVelocitiesRanges[3]:
                    windDirWeights = windDirWeightsList[2]
                    swellHeightWeights = swellHeightWeightsList[2]
                    windForceWeights = windForceWeightsList[2]

                outerItem = {"draft": meanDraftLadden, "speed": (consVelocitiesJSON[vel]), "cells": []}

                centralMean = lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1, 1))[0]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                stw = consVelocitiesJSON[vel]
                # lrSpeedFoc.predict(np.array([consVelocitiesJSON[vel]]).reshape(-1,1))[0]
                # lrSpeedFoc.predict(np.array([consVelocities[vel]]).reshape(-1,1))[0]
                # p2(consVelocities[vel])
                if stw ==10:
                    x=0
                ladenDt7801 = []
                for w in range(0, len(windF) - 1):
                    for s in range(0, len(swellH) - 1):
                        ladenDt7_8 = []
                        numberOfApp11_8 = []
                        for i in range(0, len(wind) - 1):
                            ####arrayFoc missing
                            if arrayFoc.__len__() > minAccThres:
                                meanArrayFoc = np.mean(arrayFoc[:, 8])
                                stdArrayFoc = np.std(arrayFoc[:, 8])
                                arrayFoc = np.array([k for k in arrayFoc if
                                                     k[8] >= meanArrayFoc - (2 * stdArrayFoc) and k[
                                                         8] <= meanArrayFoc + (
                                                             2 * stdArrayFoc)])

                                steamTime = arrayFoc[:, 12]

                            numberOfApp11_8.append(arrayFoc.__len__())  # + centralArray.__len__())
                            '''rawFoc = np.array([k for k in ladenDt if (k[5]>= stw - 0.25 and k[5]<=stw+0.25) and
                            (k[4]>=windF[w+1] and k[4]<=windF[w+1]) and (k[13] >=swellH[s] and k[13]<=swellH[s+1] ) and
                            (k[3]>= wind[i] and k[3]< wind[i+1] )])
                            meanRawFoc = np.round(np.mean(rawFoc[:,8]),2) if len(rawFoc) > 0 else 0'''
                            if (s > 0 and w >= 0):
                                #cellValue = round(
                                    #(ladenDt7801[len(ladenDt7801) - 5] + (ladenDt7801[len(ladenDt7801) - 5] * swellHeightWeights[s])),2)
                                cellValue = round(
                                (ladenDt7801[len(ladenDt7801) - 5] + (ladenDt7801[len(ladenDt7801) - 5] * swellHeightWeights[s])),2)
                            elif s == 0 and w == 0:
                                #if 320 * rowCounter < len(arrayFocBallast):
                                    #cellValue = round(arrayFocBallast[320 * rowCounter] + (arrayFocBallast[320 * rowCounter] * factorDRAFT[2]), 2)
                                #else:
                                    #cellValue = round(arrayFocLaden[len(arrayFocLaden)-320] + (arrayFocLaden[len(arrayFocLaden)-320] * factorDRAFT[2]), 2)
                                    #cellValue = round(centralMean + (centralMean * factorDRAFT[2]), 2)
                                #cellValue = round(centralMean , 2)
                                #if 320 * rowCounter < len(arrayFocBallast):
                                    #while cellValue<=arrayFocBallast[320 * rowCounter]:
                                #cellValue = round(centralMean + (centralMean * factorDRAFT[2]), 2)
                                cellValue = round(centralMean , 2)

                            elif s == 0 and w > 0:
                                if w < 3:
                                    cellValue = round(
                                        ladenDt7801[len(ladenDt7801) - 40] + (
                                                    ladenDt7801[len(ladenDt7801) - 40] * (windForceWeights[w])), 2)
                                elif w == 3:
                                    cellValue = round(
                                        ladenDt7801[len(ladenDt7801) - 40] + (
                                                ladenDt7801[len(ladenDt7801) - 40] * (windDirWeights[i])), 2)
                                elif w > 3:
                                    cellValue = round(
                                        ladenDt7801[len(ladenDt7801) - 40] + (
                                                ladenDt7801[len(ladenDt7801) - 40] * (windForceWeights[w])), 2)

                            # cellValue = meanRawFoc
                            '''lstmPoint =[]
                            pPoint = np.array([meanDraftLadden,(wind[i]), (windF[w]+windF[w+1])/2,consVelocitiesJSON[vel], (swellH[s])])

                            #lstmPoint.append(np.array(
                                    #[meanDraftLadden, (wind[i]), ((windF[w] + windF[w + 1]) / 2) - 1,
                                     #consVelocitiesJSON[vel], (swellH[s]) - 1]))

                            if s==0 and w==0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = w, w + 1, 0.2
                            elif s>0 and w>=0:
                                    startS, endS, stepS = (s+1)-2, s+1, 0.5
                                    startW, endW, stepW = w , w+1, 0.2
                            elif s==0 and w>0:
                                    startS, endS, stepS = s , s+1, 0.2
                                    startW, endW, stepW = (w+1)-2, w , 0.5

                            countS = 0
                            countW = 0
                            countWd = 0
                            stepWd=5

                            wd = np.linspace(wind[i] ,wind[i+1],n_steps)
                            wf = np.linspace(windF[w]+0.1, windF[w + 1], n_steps)
                            swh = np.linspace(swellH[s]+0.1, swellH[s + 1], n_steps)
                            stw = np.linspace(consVelocitiesJSON[vel], consVelocitiesJSON[vel+1] if vel < len(consVelocitiesJSON)-1 else consVelocitiesJSON[vel]+1, n_steps)
                            for k in np.arange(0,n_steps):

                                    lstmPoint.append(np.array(
                                        [meanDraftLadden, (wd[k]), wf[k],
                                         stw[k], (swh[k])]))
                                    #lstmPoint.append(np.array(
                                            #[meanDraftLadden, (wd[k]), windF[startW]+ countW,
                                             #stw[k], (swellH[startS])+ countS]))

                                    countS += stepS
                                    countS = countS - stepS if swellH[startS]+ countS > endS else countS
                                    countW += stepW
                                    countW = countW - stepW if windF[startW]+ countW > endW else countW
                            #lstmPoint.append(pPoint)
                            lstmPoint=np.array(lstmPoint).reshape(n_steps,-1)
                            XSplineVectors=[]
                            for j in range(0,len(lstmPoint)):
                                    pPoint = lstmPoint[j]
                                    vector  , interceptsGen = dm.extractFunctionsFromSplines('Gen',pPoint[0], pPoint[1], pPoint[2], pPoint[3],pPoint[4])
                                    #vector = list(np.where(np.array(vector) < 0, 0, np.array(vector)))
                                    #vector = ([abs(k) for k in vector])
                                    XSplineVector = np.append(pPoint, vector)
                                    XSplineVector = np.array(XSplineVector).reshape(1, -1)
                                    XSplineVectors.append(XSplineVector)
                            XSplineVectors = np.array(XSplineVectors).reshape(n_steps,-1)
                            #XSplineVectors = lstmPoint
                            XSplineVector = XSplineVectors.reshape(1,XSplineVectors.shape[0], XSplineVectors.shape[1])
                            cellValue = float(currModeler.predict(XSplineVector)[0][0])
                            cellValue = np.round((cellValue),2)'''
                            item = {"windBFT": w + 1, "windDir": i + 1, "swell": s + 1, "cons": cellValue}
                            outerItem['cells'].append(item)
                            ladenDt7_8.append(cellValue)
                            ladenDt7801.append(cellValue)
                            arrayFocLaden.append(cellValue)
                        if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                            for i in range(row[w], row[w] + 5):
                                try:
                                    workbook._sheets[3][column[s - 1 if s == 8 else s] + str(i)] = str(
                                        ladenDt7_8[i - row[w]])  # + '(' + str(numberOfApp11_8[i - row[w]]) + ')'

                                    workbook._sheets[3][column[s - 1 if s == 8 else s] + str(i)].alignment = Alignment(
                                        horizontal='right')
                                except:
                                    print("Exception")
                    lastLenLadenDt7801 = len(ladenDt7801)
                if consVelocitiesJSON[vel] - int(consVelocitiesJSON[vel]) == 0:
                    rowCounter+=1
                json_decoded['ConsumptionProfile']['consProfile'].append(outerItem)
            # workbook.save(filename=pathToexcel.split('.')[0] + '_1.' + pathToexcel.split('.')[1])
            # return
            ladenSPEEDMin = ladenDt7_8
            ####################END 8 SPEED ########################################################################
            ####################END 8 SPEED #######################################################################
        ####################################################LADDEN START ###################################################################
        ####################################################LADDEN START ###################################################################
        ####################################################LADDEN START ###################################################################


        '''wind = [0, 22.5, 67.5, 112.5, 157.5, 180]

          ###SEA STATE
          ballastMaxSeaSate = []
          for i in range(0, len(wind) - 1):
              arrayWF = np.array(
                  [k for k in ballastDt if k[3] >= wind[i] and k[3] <= wind[i + 1] and k[4] <= 9 and k[8] > 0])
              maxSS = np.max(arrayWF[:, 4]) if arrayWF.__len__() > 0 else 0
              ballastMaxSeaSate.append(round(maxSS, 2))

          for i in range(3, 8):
              workbook._sheets[3]['B' + str(i)] = ballastMaxSeaSate[i - 3]
          ####################################################
          laddenMaxSeaSate = []
          for i in range(0, len(wind) - 1):
              arrayWF = np.array(
                  [k for k in ladenDt if k[3] >= wind[i] and k[3] <= wind[i + 1] and k[4] <= 9 and k[8] > 0])
              maxSS = np.max(arrayWF[:, 4]) if arrayWF.__len__() > 0 else 0
              laddenMaxSeaSate.append(round(maxSS, 2))

          for i in range(12, 17):
              workbook._sheets[3]['B' + str(i)] = laddenMaxSeaSate[i - 12]
          ###############################
          ##end of sea state
          ###
          # meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] > 0])[:, 1])), 2)
          workbook._sheets[2]['B2'] = np.floor(meanDraftBallast) + 0.5
          # meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] > 0])[:, 1])), 2)
          workbook._sheets[1]['B2'] = np.ceil(meanDraftLadden) + 0.5

          ##WRITE DRAFTS TO FILE
          drafts = []
          drafts.append(np.floor(meanDraftBallast) + 0.5)
          drafts.append(np.ceil(meanDraftLadden) + 0.5)
          with open('./data/' + company + '/' + vessel + '/ListOfDrafts.csv', mode='w') as data:
              data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
              data_writer.writerow(
                  ['Draft'])
              for i in range(0, 2):
                  data_writer.writerow(
                      [drafts[i]])

          #############################################################################
          ##BALLAST
          ##delete ballast outliers
          # np.delete(ladenDt, [i for (i, v) in enumerate(ladenDt[:, 8]) if v < (
          # np.mean(ladenDt[:, 8]) - np.std(ladenDt[:, 8])) or v > np.mean(
          # ladenDt[:, 8]) + np.std(
          # ladenDt[:, 8])], 0)
          partitionsX = []
          partitionLabels = []
          # For each label
          for curLbl in np.unique(labels):
              # Create a partition for X using records with corresponding label equal to the current
              partitionsX.append(np.asarray(velocitiesB[labels == curLbl]))
              # Create a partition for Y using records with corresponding label equal to the current

              # Keep partition label to ascertain same order of results
              partitionLabels.append(curLbl)

          sorted = []
          initialX = len(partitionsX)
          initialXpart = partitionsX
          while sorted.__len__() < initialX:
              min = 100000000000
              for i in range(0, len(partitionsX)):
                  mean = np.mean(partitionsX[i])
                  if mean < min:
                      min = mean
                      minIndx = i
              sorted.append(partitionsX[minIndx])
              # partitionsX.remove(partitionsX[minIndx])
              partitionsX.pop(minIndx)'''

        ################################################################################################################

        #workbook = self.calculateExcelStatistics(workbook, dtNew, velocities, draft, trim, velocitiesTlg, rawData,
                                                 #company, vessel, tlgDataset, 'all')
        # workbook = self.calculateExcelStatistics(workbook, dtNewBDD, velocities, draft, trim, velocitiesTlgBDD, rawData,
        # company, vessel, tlgDatasetBDD,'bdd')
        # workbook = self.calculateExcelStatistics(workbook, dtNewADD, velocities, draft, trim, velocitiesTlgBDD, rawData,
        # company, vessel, tlgDatasetADD, 'Add')
        ##delete ladden outliers
        # np.delete(ladenDt, [i for (i, v) in enumerate(ladenDt[:, 8]) if v < (
        # np.mean(ladenDt[:, 8]) - np.std(ladenDt[:, 8])) or v > np.mean(
        # ladenDt[:, 8]) + np.std(
        # ladenDt[:, 8])], 0)
        workbook.save(filename=pathToexcel.split('.')[0] + '_' + vessel + '.' + pathToexcel.split('.')[1])
        with open('./consProfileJSON/consProfile_'+vessel+'_.json', 'w') as json_file:
            json.dump(json_decoded, json_file)

        return
        #################################
