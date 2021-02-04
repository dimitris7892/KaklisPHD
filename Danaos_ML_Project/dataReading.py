import numpy as np
import pandas as pd
from scipy.stats import ks_2samp,chisquare,chi2_contingency
import math
#from pyproj import Proj, transform
import pyearth as sp
import matplotlib
from sklearn.cluster import KMeans
from matplotlib.lines import Line2D
from decimal import Decimal
import random
from numpy import inf
#from coordinates.converter import CoordinateConverter, WGS84, L_Est97
import pyodbc
import csv
import locale
#locale.setlocale(locale.LC_NUMERIC, "en_DK.UTF-8")
import datetime
import cx_Oracle
from dateutil.rrule import rrule, DAILY, MINUTELY
from sympy.solvers import solve
from sympy import Symbol
from pathlib import Path
import itertools
from sympy import cos, sin , tan , exp , sqrt , E
from openpyxl import load_workbook
import glob, os
from pathlib import Path
#from openpyxl.styles.colors import YELLOW
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
import shutil
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
#import matplotlib.pyplot as plt
#import seaborn as sns
from openpyxl.drawing.image import Image
from scipy.stats import ttest_ind_from_stats
#import seaborn as sns
#import  generateProfile as genProf
import generateProfile as genProf

generateExcel = genProf.BaseProfileGenerator()

class BaseSeriesReader:

    def decomposeWindVector(self,vector):

        return (vector[0] * np.cos(vector[1])) , (vector[0] * np.sin(vector[1]))

    def calculateSeaElUVcompFromSpeedDir(self, speed, dir):

        if dir < 90:
            Theta = 90 - dir
        else:
            Theta = (360 + (90 - dir)) % 360
        ThetaRad = self.degreeeToRadian(Theta)
        wx = speed * np.cos(ThetaRad)
        wy = speed * np.sin(ThetaRad)

        return np.round(wx, 2), np.round(wy, 2)

    def degreeeToRadian(self,angle):
        return math.pi * angle / 180.0

    def DecomposeUV(self,ucompD,vcompD):
        windSpeed = np.round(math.sqrt(math.pow(ucompD,2)+math.pow(vcompD,2)),2)
        windDir = 0
        if ucompD==0 :
            if vcompD >=0:
                windDir=180
            else:
                windDir = 0
        elif vcompD==0:
            if ucompD>=0:
                windDir=270
            else:
                windDir = 90
        else:
            theta = math.atan2(np.abs(vcompD) , np.abs(ucompD))
            thetaDeg = 57.295779513 * theta
            if ucompD >= 0:
                if vcompD >= 0:
                    windDir = 270 - thetaDeg
                else:
                    windDir = 270 + thetaDeg
            else:
                if vcompD < 0:
                    windDir = 90 - thetaDeg
                else:
                    windDir = 90 + thetaDeg
            windDir = np.round(windDir,2)

        return windSpeed ,  windDir

    def calculateWindUVcompFromSpeedDir(self,windSpeed,windDir):

        if windDir < 270:
            windTheta = 270 - windDir
        else:
            windTheta = (360 + (270 - windDir)) % 360
        windThetaRad = self.degreeeToRadian(windTheta)
        wx = windSpeed * np.cos(windThetaRad)
        wy = windSpeed * np.sin(windThetaRad)

        return np.round(wx,2) , np.round(wy,2)

    def convertApparentToTrueWeather(self,weatherVectorApp,vesselVector):

        weatherU,weatherV = self.calculateWindUVcompFromSpeedDir(weatherVectorApp[0] , weatherVectorApp[1])

        vesselU,vesselV = self.calculateSeaElUVcompFromSpeedDir(vesselVector[0] , vesselVector[1])

        trueWeatherU = weatherU - vesselU
        trueWeatherV = weatherV - vesselV

        return np.array([trueWeatherU,trueWeatherV])

    def getRelativeDirectionWeatherVessel(self , vesDir , weatherDir):
        relativeDir = 0

        if (weatherDir >= vesDir):
            relativeDir = weatherDir - vesDir
        else:
            relativeDir = 360 - (vesDir - weatherDir)

        return relativeDir

    def getActualWindDirFromWindSensor(self,windDir):

        return (360 + windDir) % 360


    def ConvertMSToBeaufort(self , ws):
        wsB = 0
        if ws >= 0 and ws <0.2:
            wsB = 0
        elif ws >= 0.3 and ws <1.5:
            wsB = 1
        elif ws >= 1.6 and ws < 3.3:
            wsB = 2
        elif ws >= 3.4 and ws < 5.4:
            wsB = 3
        elif ws >= 5.5 and ws < 7.9:
            wsB = 4

        elif ws >= 8 and ws < 10.7:
            wsB = 5
        elif ws >= 10.8 and ws < 13.8:
            wsB = 6
        elif ws >= 13.9 and ws < 17.1:
            wsB = 7
        elif ws >= 17.2 and ws < 20.7:
            wsB = 8
        elif ws >= 20.8 and ws < 24.4:
            wsB = 9
        elif ws >= 24.5 and ws < 28.4:
            wsB = 10
        elif ws >= 28.5 and ws < 32.6:
            wsB = 11
        elif ws >= 32.7:
            wsB = 12
        return wsB

    def ConvertKNotsToBeaufort(self , ws):
        wsB = 0
        if ws <= 1:
            wsB = 0
        elif ws >= 1 and ws <3:
            wsB = 1
        elif ws >= 4 and ws < 6:
            wsB = 2
        elif ws >= 7 and ws < 10:
            wsB = 3
        elif ws >= 11 and ws < 16:
            wsB = 4

        elif ws >= 17 and ws < 21:
            wsB = 5
        elif ws >= 22 and ws < 27:
            wsB = 6
        elif ws >= 28 and ws < 33:
            wsB = 7
        elif ws >= 34 and ws < 40:
            wsB = 8
        elif ws >= 41 and ws < 47:
            wsB = 9
        elif ws >= 48 and ws < 55:
            wsB = 10
        elif ws >= 56 and ws < 63:
            wsB = 11
        elif ws >= 64:
            wsB = 12
        return wsB

    def CalculateVesselDirection(self,vlon1,  vlat1,  vlon2,  vlat2):

        angle = 0
        if (vlon1 == vlon2):
            if vlat1 <= vlat2:
                angle = 0
            else:
                angle = 180
        else:
            absDif = abs(vlon1 - vlon2)
            # Checks if it is closer to change directions.
            if (absDif > 360 - absDif):

                vlon1 = (360 + vlon1) % 360
                vlon2 = (360 + vlon2) % 360
            if (vlat1 == vlat2):

                if (vlon1 <= vlon2):
                    angle = 90
                else:
                    angle = 270
            else:
                division = ((vlat2 - vlat1) / (vlon2 - vlon1))
                if (vlon1 < vlon2):
                    angle = 90 - np.rad2deg(math.atan(division))
                else:# if ((vlon1 > vlon2) & & (vlat1 < vlat2))
                    angle = 270 - np.rad2deg(math.atan(division))


        return angle

    def extractRawData(self, newDataSet, telegrams,companyTlgs,company,vessel):

        windDirs = []
        windSpeeds = []
        drafts = []
        blFlags = []
        vCourses = []
        stws = []
        focs = []
        lats = []
        lons = []
        bearings = []
        tlgsFocs = []
        trims=[]
        tlgSpeeds=[]
        steamHours=[]
        combSWHs =[]
        combWDs = []
        swellSWHs = []
        swellDs = []
        waveHs = []
        waveDs=[]
        relCombWD = []
        relSwellWd = []
        relWaveDir = []
        ttimes=[]
        windSpeedsWS=[]
        windSpeedsTlg=[]
        windDirsTlg = []

        if company=='MARMARAS':
            #sFile =  "./neural_data/marmaras_data.csv"
            #data = pd.read_csv(sFile,delimiter=';')
            #data = newDataSet.drop(["wind_speed", "wind_dir"], axis=1)
            data = np.array(newDataSet)
            meanDRaft =np.mean(data[:,1])
            for i in range(0,len(data)):
                if data[i,1] > np.ceil(meanDRaft + 1):
                    blFlags.append('L')
                else:
                    blFlags.append('B')

            stws = np.nan_to_num(data[:,0]).reshape(-1)
            focs = np.nan_to_num(data[:, 9]).reshape(-1)
            for i in range(0,len(data[:,3])):
                windSpeeds.append(self.ConvertKNotsToBeaufort(data[i,3]))
            windSpeeds = np.nan_to_num(windSpeeds).reshape(-1)
            windDirs = np.nan_to_num(data[:,4]).reshape(-1)
            blFlags = np.array(blFlags).reshape(-1)
            drafts = np.array(data[:,1]).reshape(-1)
            vCourses = np.array([0] * len(stws)).reshape(-1)
            tlgSpeeds=np.array([random.randint(3,17)] * len(stws)).reshape(-1)
            firstColumn = np.array([0] * len(stws)).reshape(-1, 1)
            otherColumns = np.array([0] * len(stws)).reshape(-1)
            trims=np.array([0] * len(stws)).reshape(-1)
            tlgsFocs = np.array([0] * len(stws)).reshape(-1)

            newDataSet = np.array(
                np.append(firstColumn, np.asmatrix(
                    [vCourses, blFlags, otherColumns, otherColumns, otherColumns, otherColumns, otherColumns, drafts,
                     otherColumns, windDirs, windSpeeds, stws,
                     otherColumns, otherColumns, focs, tlgsFocs, trims, tlgSpeeds]).T, axis=1))

        if company=='GOLDENPORT':

            ##map weather data from telegrams

            ####
            #len(newDataSet)
            telegrams = telegrams.values
            for i in range(0,len(newDataSet)):

                datetimeV = str(newDataSet[i, 0])
                dateV = datetimeV.split(" ")[0]
                hhMMss = datetimeV.split(" ")[1]
                month = dateV.split("/")[0]
                day = dateV.split("/")[1]
                year = dateV.split("/")[2]
                month = '0' + month if month.__len__() == 1 else month
                day = '0' + day if day.__len__() == 1 else day
                newDate = year + '-' + month + '-' + day
                newDate1 = year + '-' + month + '-' + day + " " + ":".join(hhMMss.split(":")[0:2])
                #rpm = newDataSet[i, 20]
                telegramRow = np.array([row for row in telegrams if str(row[0]).split(" ")[0] == newDate])
                foc = newDataSet[i, 8]


                lat = str(newDataSet[i, 4])
                latDir = str(newDataSet[i, 5])
                lon = str(newDataSet[i, 6])
                lonDir = str(newDataSet[i, 7])
                vCourse = newDataSet[i, 1]
                stw = newDataSet[i, 3]

                power = newDataSet[i, 22]
                #####FIND BEARING
                LAT, LON = self.convertLATLONfromDegMinSec(lat, lon, latDir, lonDir)
                lats.append(LAT)
                lons.append(LON)
                ###prev lat lon
                #prev_lat = str(newDataSet[i - 1, 4])
                #prev_latDir = str(newDataSet[i - 1, 5])
                #prev_lon = str(newDataSet[i - 1, 6])
                #prev_lonDir = str(newDataSet[i - 1, 7])

                #prevLAT, prevLON = self.convertLATLONfromDegMinSec(prev_lat, prev_lon, prev_latDir, prev_lonDir)
                ###########################################

                #bearing = self.CalculateVesselDirection(float(LON), float(LAT), float(prevLON), float(prevLAT))
                bearing = vCourse
                bearings.append(vCourse)



                windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = self.mapWeatherData(bearing, newDate1, np.round(LAT), np.round(LON))
                #windSpeed = self.ConvertMSToBeaufort(windSpeed)
                combSWHs.append(combSWH)
                combWDs.append(relCombWavesdDir)

                swellSWHs.append(swellSWH)
                swellDs.append(relSwelldir)

                waveHs.append(wavesSWH)
                waveDs.append(relWavesdDir)


                windDirs.append(relWindDir)
                windSpeeds.append(windSpeed)
                drafts.append(0 if telegramRow.__len__() == 0 else telegramRow[:, 8][0])

                blFlags.append('nan' if telegramRow.__len__() == 0 else telegramRow[:, 2][0])
                stws.append(stw)

                focMTd = np.round((foc / 1000) * 24, 2)

                tlgFoc = 0 if telegramRow.__len__() == 0 else telegramRow[:, 15][0]
                trim = 0 if telegramRow.__len__() == 0 else telegramRow[:, 16][0]
                tlgSpeed = 0 if telegramRow.__len__()==0 else telegramRow[:,12]
                steamHour =0 if telegramRow.__len__()==0 else np.array(telegramRow[:, 13]).reshape(-1).astype(float) + np.array(
                telegramRow[:, 14]).reshape(-1).astype(float) / 60

                steamHour = steamHour[0] if str(steamHour)[0] == '[' else steamHour
                tlgSpeed = tlgSpeed[0] if str(tlgSpeed)[0] == '[' else tlgSpeed
                #if abs(tlgFoc - focMTd) >= 5:
                    #if stw >= 6:
                        #focs.append(tlgFoc if tlgFoc > focMTd else focMTd)
                    #else:
                        #focs.append(focMTd)
                #else:
                    #focs.append(focMTd)

                focs.append(focMTd)
                tlgsFocs.append(tlgFoc)
                tlgSpeeds.append(tlgSpeed)
                trims.append(trim)
                steamHours.append(steamHour)


                # filteredDTWs = [d for d in dateTimesW if month == str(d).split('/')[1] and day ==
                # str(d).split('/')[0] and year[2:] == str(d).split('/')[2]]

            # stw = np.array([k for k in newDataSet])[:,3].astype(float).reshape(-1)

            # bearings[0].insert(0, 0)
            #bearings = np.asarray(bearings)[0].reshape(-1)


            drafts = np.nan_to_num(drafts).reshape(-1)
            # foc = np.array([k for k in newDataSet])[:,8].astype(float).reshape(-1)

            combSWHs = np.nan_to_num(combSWHs).reshape(-1)
            lats = np.nan_to_num(lats).reshape(-1)
            lons = np.nan_to_num(lons).reshape(-1)
            combWDs = np.nan_to_num(combWDs).reshape(-1)
            swellSWHs = np.nan_to_num(swellSWHs).reshape(-1)
            swellDs = np.nan_to_num(swellDs).reshape(-1)
            waveHs = np.nan_to_num(waveHs).reshape(-1)
            waveDs = np.nan_to_num(waveDs).reshape(-1)

            windSpeeds = np.nan_to_num(windSpeeds).reshape(-1)
            windDirs = np.nan_to_num(windDirs).reshape(-1)
            blFlags = np.array(blFlags).reshape(-1)
            drafts = np.array(drafts).reshape(-1)
            #vCourses = np.array(vCourses).reshape(-1)
            tlgSpeeds = np.array(tlgSpeeds).reshape(-1)
            steamHours = np.array(steamHours).reshape(-1)
            trims = np.array(trims).reshape(-1)
            firstColumn = np.array([0] * len(stws)).reshape(-1, 1)
            otherColumns = np.array([0] * len(stws)).reshape(-1)
            newDataSet = np.array(
                np.append(firstColumn, np.asmatrix(
                    [bearings, blFlags, otherColumns, otherColumns, otherColumns, otherColumns, otherColumns, drafts,
                     otherColumns, windDirs, windSpeeds, stws,
                     otherColumns, otherColumns, focs, tlgsFocs, trims, tlgSpeeds, steamHours,
                     combSWHs, combWDs, swellSWHs, swellDs, waveHs, waveDs,lats,lons]).T, axis=1))


        if company=='MILLENIA':
            telegrams = telegrams.values
            if telegrams!=[] and newDataSet!=[]:
                ########## extract
                d=0
            elif newDataSet==[] and telegrams!=[]:
                for i in range(0,len(telegrams)):
                    for k in range(0,telegrams.shape[1]):
                        value= str(telegrams[i][k])
                        telegrams[i][k] = value.split(',')[0]+'.'+value.split(',')[1] if value.__contains__(',') else value

                for i in range(0, len(telegrams)):

                        newDate1 = telegrams[i, 0]
                        '''latDir = 'N' if float(lat) < 0 else 'W'
                        lon = str(telegrams[i, 3])
                        lonDir = 'N' if float(lon) < 0 else 'W'

                        stw = telegrams[i, 12]'''
                        vCourses.append(telegrams[i, 7])
                        # LAT, LON = self.convertLATLONfromDegMinSec(lat, lon, latDir, lonDir)

                        # bearing = vCourse

                        # bearings.append(vCourse)

                        windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = \
                        telegrams[i, 11], telegrams[i, 10], \
                        telegrams[i, 20], telegrams[i, 21], telegrams[i, 18], telegrams[i, 19], None, None
                        # self.mapWeatherData(
                        # bearing, newDate1, np.round(LAT), np.round(LON))
                        # windSpeed = self.ConvertMSToBeaufort(windSpeed)
                        combSWHs.append(combSWH)
                        combWDs.append(relCombWavesdDir)

                        windSpeeds.append(windSpeed)
                        windDirs.append(relWindDir)

                        swellSWHs.append(swellSWH)
                        swellDs.append(relSwelldir)

                        waveHs.append(wavesSWH)
                        waveDs.append(relWavesdDir)

                        tlgFoc = telegrams[i, 15]
                        focMTd = tlgFoc  # np.round((foc / 1000) * 24, 2)

                        trim = telegrams[i, 16]
                        # tlgSpeed =  telegrams[:, 12]
                        steamHour = np.array(telegrams[i, 13]).reshape(-1).astype(
                            float) + np.array(
                            telegrams[i, 14]).reshape(-1).astype(float) / 60

                        drafts.append(telegrams[i, 8])

                        focs.append(focMTd)
                        tlgsFocs.append(tlgFoc)
                        tlgSpeed = telegrams[i, 12]
                        # tlgSpeeds.append(tlgSpeed[0] if str(tlgSpeed).__len__() > 1 else tlgSpeed)
                        tlgSpeeds.append(tlgSpeed)
                        trims.append(trim)
                        # steamHours.append(steamHour[0] if str(tlgSpeed).__len__() > 1 else steamHour)
                        steamHours.append(steamHour)
                        stws.append(tlgSpeed)
                        blFlags.append(telegrams[i, 2])

                drafts = np.nan_to_num(drafts).reshape(-1)
                # foc = np.array([k for k in newDataSet])[:,8].astype(float).reshape(-1)

                combSWHs = np.nan_to_num(combSWHs).reshape(-1)
                combWDs = np.nan_to_num(combWDs).reshape(-1)
                swellSWHs = np.nan_to_num(swellSWHs).reshape(-1)
                swellDs = np.nan_to_num(swellDs).reshape(-1)
                waveHs = np.nan_to_num(waveHs).reshape(-1)
                waveDs = np.nan_to_num(waveDs).reshape(-1)

                stws = np.nan_to_num(stws).reshape(-1)
                focs = np.nan_to_num(focs).reshape(-1)

                windSpeeds = np.nan_to_num(windSpeeds).reshape(-1)
                windDirs = np.nan_to_num(windDirs).reshape(-1)
                blFlags = np.array(blFlags).reshape(-1)
                drafts = np.nan_to_num(drafts).reshape(-1)
                vCourses = np.nan_to_num(vCourses).reshape(-1)
                tlgSpeeds = np.nan_to_num(tlgSpeeds).reshape(-1)
                tlgsFocs = np.nan_to_num(tlgsFocs).reshape(-1)
                steamHours = np.array(steamHours).reshape(-1)
                trims = np.array(trims).reshape(-1)
                firstColumn = np.array([0] * len(stws)).reshape(-1, 1)
                otherColumns = np.array([0] * len(stws)).reshape(-1)
                newDataSet = np.array(
                    np.append(firstColumn, np.asmatrix(
                        [vCourses, blFlags, otherColumns, otherColumns, otherColumns, otherColumns, otherColumns,
                         drafts,
                         otherColumns, windDirs, windSpeeds, stws,
                         otherColumns, otherColumns, focs, tlgsFocs, trims, tlgSpeeds, steamHours, swellSWHs, swellDs,
                         combSWHs, combWDs, waveHs, waveDs]).T, axis=1))
            return newDataSet
                #np.array(np.append(telegrams[:,telegrams.shape[1]-1].reshape(-1,telegrams.shape[1]-1),np.asmatrix([telegrams[:,16].reshape(-1)]).T,axis=1))

        if company == 'OCEAN_GOLD':

            telegrams = telegrams.values
            for i in range(0, len(newDataSet)):

                if newDataSet[i,4]!='Under way using engines':
                    continue

                datetimeV = str(newDataSet[i, 0])
                dateV = datetimeV.split(" ")[0]
                hhMMss = datetimeV.split(" ")[1]
                month = dateV.split("-")[1]
                day = dateV.split("-")[2]
                year = dateV.split("-")[0]
                month = '0' + month if month.__len__() == 1 else month
                day = '0' + day if day.__len__() == 1 else day
                newDate = year + '-' + month + '-' + day
                newDate1 = year + '-' + month + '-' + day + " " + ":".join(hhMMss.split(":")[0:2])
                # rpm = newDataSet[i, 20]
                telegramRow = np.array([row for row in telegrams if str(row[0]).split(" ")[0] == newDate])
                foc = newDataSet[i, 11]

                lat = str(newDataSet[i, 26])
                #latDir = str(newDataSet[i, 27])
                lon = str(newDataSet[i, 27])
                #lonDir = str(newDataSet[i, 7])
                vCourse = newDataSet[i, 3]
                stw = newDataSet[i, 2]

                #power = newDataSet[i, 22]
                #####FIND BEARING
                LAT, LON = float(lat),float(lon)
                #self.convertLATLONfromDegMinSec(lat, lon, latDir, lonDir)
                # lats.append(LAT)
                # lons.append(LON)
                ###prev lat lon
                #prev_lat = str(newDataSet[i - 1, 4])
                #prev_latDir = str(newDataSet[i - 1, 5])
                #prev_lon = str(newDataSet[i - 1, 6])
                #prev_lonDir = str(newDataSet[i - 1, 7])

                #prevLAT, prevLON = self.convertLATLONfromDegMinSec(prev_lat, prev_lon, prev_latDir, prev_lonDir)
                ###########################################

                #bearing = self.CalculateVesselDirection(float(LON), float(LAT), float(prevLON), float(prevLAT))
                bearing = np.round(vCourse,3)
                #print(bearing)
                bearings.append(vCourse)

                if LAT!=0 and LON!=0:
                    windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = self.mapWeatherData(
                        bearing, newDate1, np.round(LAT), np.round(LON))
                else:
                    print(str(lat) +" "+ str(lon))
                    if telegramRow.__len__() > 0:
                        windSpeedTlg, relWindDirTlg, swellSWHTlg, relSwelldirTlg, wavesSWHTlg, relWavesdDirTlg, combSWHTlg, relCombWavesdDirTlg = \
                            telegrams[i, 11], telegrams[i, 10], \
                            telegrams[i, 20], telegrams[i, 21], telegrams[i, 18], telegrams[i, 19], None, None
                    else:
                        windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = None, None , None, None , None, None , None, None
                # windSpeed = self.ConvertMSToBeaufort(windSpeed)
                combSWHs.append(combSWH)
                combWDs.append(relCombWavesdDir)

                swellSWHs.append(swellSWH)
                swellDs.append(relSwelldir)

                waveHs.append(wavesSWH)
                waveDs.append(relWavesdDir)

                windDirs.append(relWindDir)
                windSpeeds.append(windSpeed)
                drafts.append(0 if telegramRow.__len__() == 0 else telegramRow[:, 8][0])
                vCourses.append(bearing)
                blFlags.append('nan' if telegramRow.__len__() == 0 else telegramRow[:, 2][0])
                stws.append(stw)

                focMTd = np.round((foc / 1000) * 1440, 2)

                tlgFoc = 0 if telegramRow.__len__() == 0 else telegramRow[:, 15][0]
                trim = 0 if telegramRow.__len__() == 0 else telegramRow[:, 16][0]
                tlgSpeed = 0 if telegramRow.__len__() == 0 else telegramRow[:, 12]
                steamHour = 0 if telegramRow.__len__() == 0 else np.array(telegramRow[:, 13]).reshape(-1).astype(float) + \
                                                                 np.array(telegramRow[:, 14]).reshape(-1).astype(float) / 60
                steamHour= steamHour[0] if str(steamHour)[0]=='['  else steamHour
                tlgSpeed = tlgSpeed[0] if str(tlgSpeed)[0]=='[' else tlgSpeed
                # if abs(tlgFoc - focMTd) >= 5:
                # if stw >= 6:
                # focs.append(tlgFoc if tlgFoc > focMTd else focMTd)
                # else:
                # focs.append(focMTd)
                # else:
                # focs.append(focMTd)

                focs.append(focMTd)
                tlgsFocs.append(tlgFoc)
                tlgSpeeds.append(tlgSpeed)
                trims.append(trim)
                steamHours.append(steamHour)
                lats.append(lat)
                lons.append(lon)



            drafts = np.nan_to_num(drafts).reshape(-1)
            # foc = np.array([k for k in newDataSet])[:,8].astype(float).reshape(-1)

            combSWHs = np.nan_to_num(combSWHs).reshape(-1)
            combWDs = np.nan_to_num(combWDs).reshape(-1)
            swellSWHs = np.nan_to_num(swellSWHs).reshape(-1)
            swellDs = np.nan_to_num(swellDs).reshape(-1)
            waveHs = np.nan_to_num(waveHs).reshape(-1)
            waveDs = np.nan_to_num(waveDs).reshape(-1)
            lats = np.nan_to_num(lats).reshape(-1)
            lons = np.nan_to_num(lons).reshape(-1)
            windSpeeds = np.nan_to_num(windSpeeds).reshape(-1)
            windDirs = np.nan_to_num(windDirs).reshape(-1)
            blFlags = np.array(blFlags).reshape(-1)
            drafts = np.array(drafts).reshape(-1)
            vCourses = np.array(vCourses).reshape(-1)
            tlgSpeeds = np.array(tlgSpeeds).reshape(-1)
            steamHours = np.array(steamHours).reshape(-1)
            trims = np.array(trims).reshape(-1)
            firstColumn = np.array([0] * len(stws)).reshape(-1, 1)
            otherColumns = np.array([0] * len(stws)).reshape(-1)
            newDataSet = np.array(
                np.append(firstColumn, np.asmatrix(
                    [vCourses, blFlags, otherColumns, otherColumns, otherColumns, otherColumns, otherColumns, drafts,
                     otherColumns, windDirs, windSpeeds, stws,
                     otherColumns, otherColumns, focs, tlgsFocs, trims, tlgSpeeds, steamHours,
                     combSWHs, combWDs, swellSWHs, swellDs, waveHs, waveDs, lats , lons]).T, axis=1))

        if company=='DANAOS':

            ##map weather data from telegrams

            ####
            #len(newDataSet)

            telegrams = telegrams.values
            if len(telegrams)>0 and len(newDataSet)>0:
                ########## extract
                for i in range(0, len(newDataSet)):
                    datetimeV = str(newDataSet[i, 0])
                    ttime = datetimeV
                    ttimes.append(ttime)
                    dateV = datetimeV.split(" ")[0]
                    hhMMss = datetimeV.split(" ")[1]
                    month = dateV.split("-")[1]
                    day = dateV.split("-")[2]
                    year = dateV.split("-")[0]
                    month = '0' + month if month.__len__() == 1 else month
                    day = '0' + day if day.__len__() == 1 else day
                    newDate = year + '-' + month + '-' + day
                    newDate1 = year + '-' + month + '-' + day + " " + ":".join(hhMMss.split(":")[0:2])
                    #print(str(newDate1))
                    telegramRow = np.array([row for row in telegrams if str(row[0]).split(" ")[0] == newDate])
                    foc = newDataSet[i, 8]

                    lat = str(newDataSet[i, 2])
                    latDir = 'S' if float(lat) < 0 else 'N'
                    lat = lat.split('-')[1] if float(lat) < 0 else lat

                    lon = str(newDataSet[i, 3])
                    lonDir = 'W' if float(lon) < 0 else 'E'
                    lon = lon.split('-')[1] if float(lon) < 0 else lon

                    vCourse = newDataSet[i, 1]
                    stw = newDataSet[i, 6]

                    # power = newDataSet[i, 22]
                    #####FIND BEARING
                    LAT, LON = self.convertLATLONfromDegMinSec(lat, lon, latDir, lonDir)
                    lats.append(LAT)
                    lons.append(LON)
                    ###prev lat lon
                    '''prev_lat = str(newDataSet[i - 1, 2])
                    prev_latDir = str(newDataSet[i - 1, 5])
                    prev_lon = str(newDataSet[i - 1, 3])
                    prev_lonDir = str(newDataSet[i - 1, 7])'''

                    # prevLAT, prevLON = self.convertLATLONfromDegMinSec(prev_lat, prev_lon, prev_latDir, prev_lonDir)
                    ###########################################

                    bearing = vCourse
                    # self.CalculateVesselDirection(float(LON), float(LAT), float(prevLON), float(prevLAT))

                    bearings.append(vCourse)

                    windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = self.mapWeatherData(
                        bearing, newDate1, np.round(LAT), np.round(LON))
                    # windSpeed = self.ConvertMSToBeaufort(windSpeed)
                    combSWHs.append(combSWH)
                    combWDs.append(relCombWavesdDir)

                    swellSWHs.append(swellSWH)
                    swellDs.append(relSwelldir)

                    waveHs.append(wavesSWH)
                    waveDs.append(relWavesdDir)

                    windSpeedSensor = newDataSet[i,5]
                    windDirSensor = newDataSet[i, 4]
                    relWindDirSensor = self.getRelativeDirectionWeatherVessel(vCourse,windDirSensor) if windDirSensor!='nan' else 'nan'

                    windDirs.append(relWindDirSensor if relWindDirSensor!='nan' else relWindDir)
                    windSpeeds.append(windSpeedSensor if windSpeedSensor!='nan' else windSpeed)

                    windSpeedsWS.append(windSpeed)


                    if newDataSet[i,7] !='nan':
                        drafts.append(newDataSet[i, 7])
                    elif newDataSet[i,7] =='nan' and telegramRow.__len__() > 0:
                        drafts.append(telegramRow[:, 8][0])
                    elif newDataSet[i,7] =='nan' and telegramRow.__len__() == 0:
                        drafts.append('nan')
                    #drafts.append(newDataSet[i, 7] if telegramRow.__len__() == 0 else telegramRow[:, 8][0])
                    vCourses.append(bearing)
                    blFlags.append('nan' if telegramRow.__len__() == 0 else telegramRow[:, 2][0])
                    stws.append(stw)

                    focMTd = foc  # np.round((foc / 1000) * 24, 2)

                    tlgFoc = 0 if telegramRow.__len__() == 0 else telegramRow[:, 15][0]
                    trim = 0 if telegramRow.__len__() == 0 else telegramRow[:, 16][0]
                    tlgSpeed = 0 if telegramRow.__len__() == 0 else telegramRow[:, 12]
                    steamHour = 0 if telegramRow.__len__() == 0 else np.array(telegramRow[:, 13]).reshape(-1).astype(
                        float) + np.array(
                        telegramRow[:, 14]).reshape(-1).astype(float) / 60
                    # if abs(tlgFoc - focMTd) >= 5:
                    # if stw >= 6:
                    # focs.append(tlgFoc if tlgFoc > focMTd else focMTd)
                    # else:
                    # focs.append(focMTd)
                    # else:
                    # focs.append(focMTd)

                    focs.append(focMTd)
                    tlgsFocs.append(tlgFoc)
                    tlgSpeeds.append(tlgSpeed[0] if str(tlgSpeed).__len__() > 1 else tlgSpeed)
                    trims.append(trim)
                    steamHours.append(steamHour[0] if str(tlgSpeed).__len__() > 1 else steamHour)

                    # filteredDTWs = [d for d in dateTimesW if month == str(d).split('/')[1] and day ==
                    # str(d).split('/')[0] and year[2:] == str(d).split('/')[2]]
            elif newDataSet == [] and telegrams != []:
                for i in range(0, len(telegrams)):
                    for k in range(0, telegrams.shape[1]):
                        value = str(telegrams[i][k])
                        telegrams[i][k] = value.split(',')[0] + '.' + value.split(',')[1] if value.__contains__(
                            ',') else value
                for i in range(0, len(telegrams)):
                    newDate1 = telegrams[i,0]
                    '''latDir = 'N' if float(lat) < 0 else 'W'
                    lon = str(telegrams[i, 3])
                    lonDir = 'N' if float(lon) < 0 else 'W'
                    
                    stw = telegrams[i, 12]'''
                    vCourses.append(telegrams[i, 7])
                    #LAT, LON = self.convertLATLONfromDegMinSec(lat, lon, latDir, lonDir)

                    #bearing = vCourse

                    #bearings.append(vCourse)

                    windSpeed, relWindDir, swellSWH, relSwelldir, wavesSWH, relWavesdDir, combSWH, relCombWavesdDir = telegrams[i,11] ,telegrams[i,10],\
                    telegrams[i, 20],telegrams[i,21],telegrams[i,18],telegrams[i,19] , None ,None
                        #self.mapWeatherData(
                        #bearing, newDate1, np.round(LAT), np.round(LON))
                    # windSpeed = self.ConvertMSToBeaufort(windSpeed)
                    combSWHs.append(combSWH)
                    combWDs.append(relCombWavesdDir)

                    windSpeeds.append(windSpeed)
                    windDirs.append(relWindDir)

                    swellSWHs.append(swellSWH)
                    swellDs.append(relSwelldir)

                    waveHs.append(wavesSWH)
                    waveDs.append(relWavesdDir)

                    tlgFoc =  telegrams[i, 15]
                    focMTd = tlgFoc  # np.round((foc / 1000) * 24, 2)


                    trim =  telegrams[i, 16]
                    #tlgSpeed =  telegrams[:, 12]
                    steamHour = np.array(telegrams[i, 13]).reshape(-1).astype(
                        float) + np.array(
                        telegrams[i, 14]).reshape(-1).astype(float) / 60


                    drafts.append(telegrams[i,8])

                    focs.append(focMTd)
                    tlgsFocs.append(tlgFoc)
                    tlgSpeed = telegrams[i,12]
                    #tlgSpeeds.append(tlgSpeed[0] if str(tlgSpeed).__len__() > 1 else tlgSpeed)
                    tlgSpeeds.append(tlgSpeed)
                    trims.append(trim)
                    #steamHours.append(steamHour[0] if str(tlgSpeed).__len__() > 1 else steamHour)
                    steamHours.append(steamHour)
                    stws.append(tlgSpeed)
                    blFlags.append(telegrams[i,2])
                '''return np.array(np.append(telegrams[:, :16].reshape(-1, telegrams.shape[1] - 2),
                                          np.asmatrix([telegrams[:, 15].reshape(-1),
                                                       telegrams[:, 16].reshape(-1),
                                                       telegrams[:, 12].reshape(-1),
                                                       steamHours.reshape(-1),
                                                       telegrams[:, 17].reshape(-1)
                                                       ]).T, axis=1))'''


            # stw = np.array([k for k in newDataSet])[:,3].astype(float).reshape(-1)

            # bearings[0].insert(0, 0)
            #bearings = np.asarray(bearings)[0].reshape(-1)


            drafts = np.nan_to_num(drafts).reshape(-1)
            # foc = np.array([k for k in newDataSet])[:,8].astype(float).reshape(-1)

            ttimes = np.nan_to_num(ttimes).reshape(-1, 1)
            combSWHs = np.nan_to_num(combSWHs).reshape(-1)
            combWDs = np.nan_to_num(combWDs).reshape(-1)
            swellSWHs = np.nan_to_num(swellSWHs).reshape(-1)
            swellDs = np.nan_to_num(swellDs).reshape(-1)
            waveHs = np.nan_to_num(waveHs).reshape(-1)
            waveDs = np.nan_to_num(waveDs).reshape(-1)

            stws=np.nan_to_num(stws).reshape(-1)
            focs = np.nan_to_num(focs).reshape(-1)

            windSpeeds = np.nan_to_num(windSpeeds).reshape(-1)
            windSpeedsWS = np.nan_to_num(windSpeedsWS).reshape(-1)
            lats = np.nan_to_num(lats).reshape(-1)
            lons = np.nan_to_num(lons).reshape(-1)
            windDirs = np.nan_to_num(windDirs).reshape(-1)
            blFlags = np.array(blFlags).reshape(-1)
            drafts = np.nan_to_num(drafts).reshape(-1)
            vCourses = np.nan_to_num(vCourses).reshape(-1)
            tlgSpeeds = np.nan_to_num(tlgSpeeds).reshape(-1)
            tlgsFocs = np.nan_to_num(tlgsFocs).reshape(-1)
            steamHours = np.array(steamHours).reshape(-1)
            trims = np.array(trims).reshape(-1)
            firstColumn = np.array([0] * len(stws)).reshape(-1, 1)
            otherColumns = np.array([0] * len(stws)).reshape(-1)
            newDataSet = np.array(
                np.append(ttimes, np.asmatrix(
                    [vCourses, blFlags, otherColumns, otherColumns, otherColumns, otherColumns, otherColumns, drafts,
                     otherColumns, windDirs, windSpeeds, stws,
                     otherColumns, otherColumns, focs, tlgsFocs,trims,tlgSpeeds,steamHours ,swellSWHs , swellDs, combSWHs, combWDs ,  waveHs , waveDs ,
                     lats , lons , windSpeedsWS ]).T, axis=1))


        return newDataSet


    def GenericParserForDataExtraction(self,systemType, company, vessel ,driver=None,server=None,sid=None,usr=None,password=None,rawData=None,telegrams=None,companyTelegrams=None,seperator=None ,fileType=None, granularity=None, fileName=None,
                                       pathOfRawData=None,comparisonTest=None,dataTrain=None):

        boolTlg = telegrams
        pathExcel = '/home/dimitris/Desktop/templateDETAILED.xlsx'
        self.imo = None
        #'/home/dimitris/Desktop/templateDETAILED.xlsx'
        #'/home/dimitris/Desktop/template.xlsx'
        #'C:/Users/dkaklis/Desktop/template.xlsx'
        #'/home/dimitris/Desktop/template.xlsx'
        ##public vars

        if systemType=='LEMAG':


            if comparisonTest: ##run experiments with train/test split
                dataSet = dataTrain
            else:
                if telegrams:
                    my_file = Path('./data/'+company+'/'+vessel+'/TELEGRAMS/'+vessel+'.csv')
                    if my_file.is_file() == False:
                        self.GenericParserForDataExtraction('TELEGRAMS', company, vessel,driver,server,sid,usr,password)
                        tlgs = pd.read_csv('./data/'+company+'/'+vessel+'/TELEGRAMS/'+vessel+'.csv',sep=';')
                        '''tlgsBDD = pd.read_csv('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'B_DD.csv',
                                           sep=';')
                        tlgsADD = pd.read_csv('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'A_DD.csv',
                                              sep=';')'''
                    else:
                        tlgs = pd.read_csv('./data/'+company+'/'+vessel+'/TELEGRAMS/'+vessel+'.csv',sep=';')
                        '''tlgsBDD = pd.read_csv('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'B_DD.csv',
                                           sep=';')
                        tlgsADD = pd.read_csv('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'A_DD.csv',
                                              sep=';')'''

                    telegrams = tlgs#.values

                if Path('./data/' + company + '/' + vessel + '/mappedDataTrain.csv').is_file():
                    data = pd.read_csv('./data/' + company +'/'+vessel +'/mappedDataTrain.csv')
                    newDataSet = data.values


                if os.path.isdir('./data/' + company + '/' + vessel) == False:
                    os.mkdir('./data/' + company + '/' + vessel)
                if os.path.isdir('./data/' + company + '/' + vessel+'/TELEGRAMS') == False:
                    os.mkdir('./data/' + company + '/' + vessel+'/TELEGRAMS')

                dataSet=[]
                if rawData==True:
                    path = './data/' + company + '/' + vessel+'/'
                    #Path('./data/' + company + '/'+vessel).mkdir(parents=True, exist_ok=True)
                    if os.path.isdir('./data/' + company + '/'+vessel)==False:
                        shutil.copytree(pathOfRawData, path)
                    dataSet = []
                    for infile in  sorted(glob.glob(path+'*.csv')):
                        data = pd.read_csv(infile, sep=',', decimal='.',skiprows=1)
                        dataSet.append(data.values)
                        print(str(infile))
                    #if len(dataSet)>1:
                    dataSet = np.concatenate(dataSet)

                ##########################################################


                if companyTelegrams:
                    #my_file = Path('./data/' + company + '/TELEGRAMS/' + companyTlgFIle + '.csv')

                    companyTlgs = pd.read_excel('./data/' + company + '/' + vessel + '/Penelope TrackReport.xls')
                    #CompanyTlgs = pd.read_csv('./data/' + company + '/TELEGRAMS/' + companyTlgFIle + '.csv', sep=';')

                    companyTelegrams = companyTlgs#.values


            ##WRITE NEWDATASET IN A CSV
            #
            #newDataSetBDD = self.extractRawData(dataSet, tlgsBDD, companyTelegrams, company, vessel)
            #newDataSetADD = self.extractRawData(dataSet, tlgsADD, companyTelegrams, company, vessel)
            if Path('./data/' + company + '/' + vessel + '/mappedDataTrain.csv').is_file()==False:
                #dataSet=dataSet[:5,:]
                newDataSet = self.extractRawData(dataSet, telegrams, companyTelegrams, company, vessel)
                with open('./data/' + company +'/'+vessel+'/mappedData.csv', mode='w') as data:
                    data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    for i in range(0, len(newDataSet)):
                        data_writer.writerow(
                            [newDataSet[i][0], newDataSet[i][1], newDataSet[i][2], 0, 0, 0, 0, 0, newDataSet[i][8], 0 ,newDataSet[i][10],
                             newDataSet[i][11], newDataSet[i][12], 0, 0, newDataSet[i][15], newDataSet[i][16],newDataSet[i][17],newDataSet[i][18],newDataSet[i][19],
                             newDataSet[i][20],newDataSet[i][21],newDataSet[i][22],newDataSet[i][23],newDataSet[i][24],newDataSet[i][25],
                             newDataSet[i][26],newDataSet[i][27],])

            #if boolTlg==False and rawData==True:
                #tlgDataset=[]
            #if rawData == False and boolTlg == True:
                #tlgDataset = telegrams
            tlgDataset=telegrams.values
            #genExcel = genProf.BaseProfileGenerator()
            #genExcel.fillExcelProfCons(company,vessel, pathExcel,newDataSet,rawData,tlgDataset,newDataSetBDD,newDataSetADD)
            #self.fillExcelProfCons(company,vessel, pathExcel,newDataSet,rawData,tlgDataset,newDataSetBDD,newDataSetADD)
            #self.fillExcelProfCons(company, vessel, pathExcel, newDataSet, rawData, tlgDataset, [],[])
            dsn = cx_Oracle.makedsn(
                server,
                '1521',
                service_name=sid
            )
            connection = cx_Oracle.connect(
                user=usr,
                password=password,
                dsn=dsn
            )
            cursor_myserver = connection.cursor()
            cursor_myserver.execute('SELECT IMO_NO FROM VESSEL_DATA WHERE VESSEL_CODE =  (SELECT VESSEL_CODE FROM VESSEL_DATA WHERE VESSEL_NAME LIKE '"'%" + vessel + "%'"'  AND ROWNUM=1)  ')
            # if cursor_myserver.fetchall().__len__() > 0:
            for row in cursor_myserver.fetchall():
                imo = row[0]
                print(str(imo))
            #newDataSet[:5000]
            generateExcel.fillDetailedExcelProfCons(company, vessel, pathExcel, newDataSet, rawData, tlgDataset, [], [], imo)
            return

        if systemType=='TELEGRAMS':
            if driver=='ORACLE':
                my_file = Path('./data/'+company+'/'+vessel+'/TELEGRAMS/'+vessel+'.csv')
                if my_file.is_file()==False:
                    if os.path.isdir('./data/'+company+'/'+vessel+'/TELEGRAMS/') == False:
                        Path('./data/'+company+'/'+vessel+'/TELEGRAMS/').mkdir(parents=True, exist_ok=True)

                    dsn = cx_Oracle.makedsn(
                            server,
                            '1521',
                            service_name=sid
                        )
                    connection = cx_Oracle.connect(
                            user=usr,
                            password=password,
                            dsn=dsn
                        )
                    cursor_myserver = connection.cursor()

                    cursor_myserver.execute('SELECT VESSEL_CODE FROM VESSEL_DATA WHERE VESSEL_NAME LIKE  '"'%" + vessel + "%'"' OR VESSEL_SHORT_NAME LIKE '"'%" + vessel + "%'"'   ')
                    #if cursor_myserver.fetchall().__len__() > 0:
                    for row in cursor_myserver.fetchall():
                            vessel_code = row[0]
                            print(str(vessel_code))

                    vessel_code='T003'
                    if usr =='shipping':
                        cursor_myserver.execute(
                            'SELECT  TELEGRAM_DATE , TELEGRAM_TYPE,BALAST_FLAG,LATITUDE_DEGREES , LATITUDE_SECONDS'
                            ' ,LONGITUDE_DEGREES , LONGITUDE_SECONDS ,vessel_course,(DRAFT_AFT + DRAFT_FORE)/2 as DRAFT , ENGINE_RPM , WIND_DIRECTION'
                            ' , WIND_FORCE  ,AVERAGE_SPEED ,hours_slc,minutes_slc, '
                            '(( NVL(ME_FO_CONS,0)))   as ME_CONS_24h ,'
                            ' (DRAFT_AFT-DRAFT_FORE) AS TRIM, '
                            'ROUND( ((( NVL(ME_FO_CONS,0))))  / decode(MILES_SLC,0,1,MILES_SLC),2) as ME_FOC_PER_MILE,'
                            'SEA_FORCE     ,'
                            'SEA_DIRECTION   ,'
                            'SWELL_FORCE      ,'
                            'SWELL_DIRECTION  '
                            'FROM TELEGRAMS where vessel_code = '"'" + vessel_code + "'" 'AND (telegram_type='"'N'"' or telegram_type='"'A'"' )')

                    else:
                        cursor_myserver.execute(
                            'SELECT  TELEGRAM_DATE,'
                            'TELEGRAM_TYPE,'
                            'BALAST_FLAG,'
                            'LATITUDE_DEGREES ,'
                            'LATITUDE_SECONDS,'
                            'LONGITUDE_DEGREES , '
                            'LONGITUDE_SECONDS ,'
                            'vessel_course,'
                            '(DRAFT_AFT + DRAFT_FORE)/2 as DRAFT,'
                            'ENGINE_RPM ,'
                            'WIND_DIRECTION,'
                            'WIND_FORCE  ,'
                            'AVERAGE_SPEED ,'
                            'hours_slc,'
                            'minutes_slc, '
                            '(( NVL(ME_HSFO_CONS,0)+ NVL(ME_LSFO_CONS,0)+ NVL(ME_HSDO_CONS,0 ) + NVL(ME_LSDO_CONS,0)))   as ME_CONS_24h,'
                            ' (DRAFT_AFT-DRAFT_FORE) AS TRIM, '
                            'ROUND( ((( NVL(ME_FO_CONS,0))))  / decode(MILES_SLC,0,1,MILES_SLC),2) as ME_FOC_PER_MILE,'
                            'SEA_FORCE     ,'
                            'SEA_DIRECTION   ,'
                            'SWELL_FORCE      ,'
                            'SWELL_DIRECTION  '
                            'FROM TELEGRAMS where vessel_code = '"'" + vessel_code + "'" 'AND (telegram_type='"'N'"' or telegram_type='"'A'"' )')
                        #and    telegram_date < TO_DATE('"'04/10/2019'"', '"'DD/MM/YY'"') AND telegram_date >= TO_DATE('"'04/06/2019'"', '"'DD/MM/YY'"') ')

                    with open('./data/'+company+'/'+vessel+'/TELEGRAMS/'+vessel+'.csv', mode='w') as data:
                        data_writer = csv.writer(data, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                        data_writer.writerow(['TELEGRAM_DATE','TELEGRAM_TYPE','BALAST_FLAG','LATITUDE_DEGREES','LATITUDE_SECONDS','LONGITUDE_DEGREES','LONGITUDE_SECONDS',
                                              'VESSEL_COURSE','DRAFT','ENGINE_RPM','WIND_DIRECTION','WIND_FORCE','AVERAGE_SPEED','HOURS_SLC','MINUTES_SLC','ME_CONS_24H','TRIM','ME_CONS_PER_MILE','WAVEH','WAVED','SWELLH','SWELLD'])

                        for row in cursor_myserver.fetchall():
                            telegram_date = str(row[0])
                            telegram_type = row[1]
                            ballast_flag = row[2]
                            lat_deg = row[3]
                            lat_sec = row[4]
                            lon_deg = row[5]
                            lon_sec = row[6]
                            vessel_course = row[7]
                            draft = row[8]
                            rpm = row[9]
                            wd = row[10]
                            wf = row[11]
                            speed = row[12]
                            h_slc = np.nan_to_num(row[13]) if row[13]!=None else 0
                            m_slc =np.nan_to_num(row[14]) if row[14]!=None else 0
                            foc = row[15]
                            trim = row[16]
                            focPerMile = row[17]

                            waveH = row[18]
                            waveD = row[19]
                            swellH = row[20]
                            swellD = row[21]
                            #imo  = self.imo

                            try:
                                foc = foc * 24 / ((h_slc + (m_slc / 60))) if ((h_slc + (m_slc / 60))) > 0  else foc
                            except:
                                d=0
                            data_writer.writerow(
                                [telegram_date, telegram_type, ballast_flag, lat_deg, lat_sec, lon_deg, lon_sec, vessel_course, draft,rpm, wd, wf, speed, h_slc,m_slc,foc,trim , focPerMile,waveH,waveD,swellH,swellD])

                    '''cursor_myserver.execute(
                        'SELECT  TELEGRAM_DATE , TELEGRAM_TYPE,BALAST_FLAG,LATITUDE_DEGREES , LATITUDE_SECONDS'
                        ' ,LONGITUDE_DEGREES , LONGITUDE_SECONDS ,vessel_course,(DRAFT_AFT + DRAFT_FORE)/2 as DRAFT , ENGINE_RPM , WIND_DIRECTION'
                        ' , WIND_FORCE  ,AVERAGE_SPEED ,hours_slc,minutes_slc, (( NVL(ME_HSFO_CONS,0)+ NVL(ME_LSFO_CONS,0)+ NVL(ME_HSDO_CONS,0 ) + NVL(ME_LSDO_CONS,0)))   as ME_CONS_24h ,'
                        ' (DRAFT_AFT-DRAFT_FORE) AS TRIM ,'
                        'ROUND( ((( NVL(ME_HSFO_CONS,0)+ NVL(ME_LSFO_CONS,0)+ NVL(ME_HSDO_CONS,0 ) + NVL(ME_LSDO_CONS,0))))  / decode(MILES_SLC,0,1,MILES_SLC),2) as ME_FOC_PER_MILE   '
                        'FROM TELEGRAMS where vessel_code = '"'" + vessel_code + "'" 'AND (telegram_type='"'N'"' or telegram_type='"'A'"' )'
                                                                                 'and (telegram_date >= TO_DATE('"'16/07/2017'"', '"'DD/MM/YY'"') AND telegram_date <= TO_DATE('"'16/07/2019'"', '"'DD/MM/YY'"')  )'
                                                                                 '  ')
                    # and    telegram_date < TO_DATE('"'04/10/2019'"', '"'DD/MM/YY'"') AND telegram_date >= TO_DATE('"'04/06/2019'"', '"'DD/MM/YY'"') ')

                    with open('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'B_DD.csv', mode='w') as data:
                        data_writer = csv.writer(data, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                        data_writer.writerow(
                            ['TELEGRAM_DATE', 'TELEGRAM_TYPE', 'BALAST_FLAG', 'LATITUDE_DEGREES', 'LATITUDE_SECONDS',
                             'LONGITUDE_DEGREES', 'LONGITUDE_SECONDS',
                             'VESSEL_COURSE', 'DRAFT', 'ENGINE_RPM', 'WIND_DIRECTION', 'WIND_FORCE', 'AVERAGE_SPEED',
                             'HOURS_SLC', 'MINUTES_SLC', 'ME_CONS_24H', 'TRIM', 'ME_CONS_PER_MILE'])

                        for row in cursor_myserver.fetchall():
                            telegram_date = str(row[0])
                            telegram_type = row[1]
                            ballast_flag = row[2]
                            lat_deg = row[3]
                            lat_sec = row[4]
                            lon_deg = row[5]
                            lon_sec = row[6]
                            vessel_course = row[7]
                            draft = row[8]
                            rpm = row[9]
                            wd = row[10]
                            wf = row[11]
                            speed = row[12]
                            h_slc = np.nan_to_num(row[13]) if row[13] != None else 0
                            m_slc = np.nan_to_num(row[14]) if row[14] != None else 0
                            foc = row[15]
                            trim = row[16]
                            focPerMile = row[17]

                            foc = foc * 24 / ((h_slc + (m_slc / 60))) if ((h_slc + (m_slc / 60))) > 0 else foc

                            data_writer.writerow(
                                [telegram_date, telegram_type, ballast_flag, lat_deg, lat_sec, lon_deg, lon_sec,
                                 vessel_course, draft, rpm, wd, wf, speed, h_slc, m_slc, foc, trim, focPerMile])

                    x=0

                    cursor_myserver.execute(
                        'SELECT  TELEGRAM_DATE , TELEGRAM_TYPE,BALAST_FLAG,LATITUDE_DEGREES , LATITUDE_SECONDS'
                        ' ,LONGITUDE_DEGREES , LONGITUDE_SECONDS ,vessel_course,(DRAFT_AFT + DRAFT_FORE)/2 as DRAFT , ENGINE_RPM , WIND_DIRECTION'
                        ' , WIND_FORCE  ,AVERAGE_SPEED ,hours_slc,minutes_slc, (( NVL(ME_HSFO_CONS,0)+ NVL(ME_LSFO_CONS,0)+ NVL(ME_HSDO_CONS,0 ) + NVL(ME_LSDO_CONS,0)))   as ME_CONS_24h ,'
                        ' (DRAFT_AFT-DRAFT_FORE) AS TRIM ,'
                        'ROUND( ((( NVL(ME_HSFO_CONS,0)+ NVL(ME_LSFO_CONS,0)+ NVL(ME_HSDO_CONS,0 ) + NVL(ME_LSDO_CONS,0))))  / decode(MILES_SLC,0,1,MILES_SLC),2) as ME_FOC_PER_MILE   '
                        'FROM TELEGRAMS where vessel_code = '"'" + vessel_code + "'" 'AND (telegram_type='"'N'"' or telegram_type='"'A'"' )'
                                                                                 'and (telegram_date > TO_DATE('"'16/07/2019'"', '"'DD/MM/YY'"')  )'
                                                                                 '  ')

                    with open('./data/' + company + '/' + vessel + '/TELEGRAMS/' + vessel + 'A_DD.csv', mode='w') as data:
                        data_writer = csv.writer(data, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                        data_writer.writerow(
                            ['TELEGRAM_DATE', 'TELEGRAM_TYPE', 'BALAST_FLAG', 'LATITUDE_DEGREES', 'LATITUDE_SECONDS',
                             'LONGITUDE_DEGREES', 'LONGITUDE_SECONDS',
                             'VESSEL_COURSE', 'DRAFT', 'ENGINE_RPM', 'WIND_DIRECTION', 'WIND_FORCE', 'AVERAGE_SPEED',
                             'HOURS_SLC', 'MINUTES_SLC', 'ME_CONS_24H', 'TRIM', 'ME_CONS_PER_MILE'])

                        for row in cursor_myserver.fetchall():
                            telegram_date = str(row[0])
                            telegram_type = row[1]
                            ballast_flag = row[2]
                            lat_deg = row[3]
                            lat_sec = row[4]
                            lon_deg = row[5]
                            lon_sec = row[6]
                            vessel_course = row[7]
                            draft = row[8]
                            rpm = row[9]
                            wd = row[10]
                            wf = row[11]
                            speed = row[12]
                            h_slc = row[13]
                            m_slc = row[14]
                            foc = row[15]
                            trim = row[16]
                            focPerMile = row[17]

                            foc = foc * 24 / ((h_slc + (m_slc / 60))) if ((h_slc + (m_slc / 60))) > 0 else foc

                            data_writer.writerow(
                                [telegram_date, telegram_type, ballast_flag, lat_deg, lat_sec, lon_deg, lon_sec,
                                 vessel_course, draft, rpm, wd, wf, speed, h_slc, m_slc, foc, trim, focPerMile])'''

                    x=0
                #self.fillExcelProfCons(vessel, 'C:/Users/dkaklis/Desktop/template.xlsx', newDataSet)
                #self.readExtractNewDataset(company,vessel,'C:/Users/dkaklis/Desktop/template.xlsx',';')
        if systemType == 'LAROS':

            #self.readLarosDAta(0,0)

            #self.ExtractLAROSDataset("MT_DELTA_MARIA")
            data = pd.read_csv('./data/'+company+'/'+vessel+'.csv', sep=',',decimal='.')

            #data = data.drop(['DateTime'], axis=1)
            dtNew =np.nan_to_num(data.values)#.astype(float)

            if telegrams==True:
                tlg_data = pd.read_csv('./data/' + company + '/' +'noon_reports_'+ vessel + '.csv', sep=';', decimal='.')

                # data = data.drop(['DateTime'], axis=1)
                tlg_dtNew = np.nan_to_num(tlg_data.values)  # .astype(float)

            ###CREATE TABLE  OF CURRENT VESSEL FOR COMPANY
            connection = pyodbc.connect('Driver={SQL Server};'
                                        'Server=WEATHERSERVER_DEV;'
                                        'Database=' + str.lower(company) + ';'
                                                                           'UID=sa;'
                                                                           'PWD=sa1!')

            cursor_myserver = connection.cursor()

            if telegrams == True:
                if cursor_myserver.tables(table=vessel+'_TLG', tableType='TABLE').fetchone():
                    print("TABLE " + vessel+'_TLG' + " ON DATABASE " + str.lower(company) + " EXISTS! ")
                else:
                    cursor_myserver.execute(
                        'CREATE TABLE [dbo].[' + vessel+'_TLG' + ']('
                        + '[id] [int] IDENTITY(1,1) NOT NULL,'
                        + '[vdate] [date] NOT NULL,'
                        + '[time_utc] [char](10) NULL,'
                        + '[type] [char](10) NULL,'
                        + '[actl_spd] [decimal](5, 2) NULL,'
                        + '[avg_spd] [decimal](5, 2) NULL,'
                        + '[slip] [decimal](5, 2) NULL,'
                        + '[act_rpm] [decimal](5, 2) NULL,'
                        + '[sail_time] [decimal](5, 2) NULL,'
                        + '[weather_force] [int] NULL,'
                        + '[weather_wind_tv] [char](10) NULL,'
                        + '[sea_condition] [char](10) NULL,'
                        + '[swell_dir] [char](10) NULL,'
                        + '[swell_lvl] [char](10) NULL,'
                        + '[current_dir] [char](10) NULL,'
                        + '[current_speed] [decimal](5, 2) NULL,'
                        + '[draft] [decimal](5, 2) NULL '
                        + 'CONSTRAINT [PK_' + vessel+'_TLG' +'] PRIMARY KEY CLUSTERED '
                        + '('
                        + '     [id] ASC'
                        + ')'
                        + ')'
                        )

                       # + ') WITH (PAD_INDEX = OFF, STATISTICS_NORECOMPUTE = OFF, IGNORE_DUP_KEY = OFF, ALLOW_ROW_LOCKS = ON, ALLOW_PAGE_LOCKS = ON) ON [PRIMARY]'
                        #+ ' ) ON [PRIMARY]')

                    connection.commit()

                if cursor_myserver.tables(table=vessel, tableType='TABLE').fetchone():
                    print("TABLE " + vessel + " ON DATABASE " + str.lower(company) + " EXISTS! ")
                else:
                    cursor_myserver.execute(
                        'CREATE TABLE [dbo].[' + vessel + ']('
                        + '[id] [int] IDENTITY(1,1) NOT NULL,'
                        + '[lat] [decimal](9, 6) NOT NULL,'
                        + '[lon] [decimal](9, 6) NOT NULL,'
                        + '[t_stamp] [datetime] NULL,'
                        + '[course] [decimal](5, 2) NULL,'
                        + '[stw] [decimal](5, 2) NULL,'
                        + '[s_ovg] [decimal](5, 2) NULL,'
                        + '[me_foc] [decimal](5, 2) NULL,'
                        + '[rpm] [decimal](5, 2) NULL,'
                        + '[power] [int] NULL,'
                        + '[draft] [decimal](5, 2) NULL,'
                        + '[trim] [decimal](5, 2) NULL,'
                        + '[at_port] [int] NULL,'
                        + '[daysFromLastPropClean] [int] NULL,'
                        + '[daysFromLastDryDock] [int] NULL,'
                        + '[sensor_wind_speed] [decimal](5, 2) NULL,'
                        + '[sensor_wind_dir] [decimal](5, 2) NULL,'
                        + '[wind_speed] [decimal](5, 2) NULL,'
                        + '[wind_dir] [decimal](5, 2) NULL,'
                        + '[curr_speed] [decimal](5, 2) NULL,'
                        + '[curr_dir] [decimal](5, 2) NULL,'
                        + '[comb_waves_height] [decimal](5, 2) NULL,'
                        + '[comb_waves_dir] [decimal](5, 2) NULL,'
                        + '[swell_height] [decimal](5, 2) NULL,'
                        + '[swell_dir] [decimal](5, 2) NULL,'
                        + '[wind_waves_height] [decimal](5, 2) NULL,'
                        + '[wind_waves_dir] [decimal](5, 2) NULL '
                        + 'CONSTRAINT [PK_' + vessel + '] PRIMARY KEY CLUSTERED '
                        + '('
                        + '     [id] ASC'
                        + ')'
                        + ')'
                        )

                       # + ') WITH (PAD_INDEX = OFF, STATISTICS_NORECOMPUTE = OFF, IGNORE_DUP_KEY = OFF, ALLOW_ROW_LOCKS = ON, ALLOW_PAGE_LOCKS = ON) ON [PRIMARY]'
                        #+ ' ) ON [PRIMARY]')

                    connection.commit()

                ### INSERT TLG DATA ON TLG TABLE
                vdate = 0
                time_utc = 0
                actl_speed = 0
                avg_speed = 0
                slip = 0
                act_rpm = 0
                sail_time = 0
                weather_force = 0
                weather_wind_tv = 0
                sea_condition = 0
                swell_dir = 0
                swell_lvl = 0
                current_lvl = 0
                current_speed = 0
                ##############################################
                for i in range(0, len(tlg_dtNew)):
                    values = np.nan_to_num(np.array(tlg_dtNew[i, :15]))
                    vdate = values[1]
                    time_utc =  values[2]
                    hours = str(time_utc).split(":")[0]
                    minutes = str(time_utc).split(":")[1]
                    actl_speed = values[3]
                    avg_speed = values[4]
                    slip = values[5]
                    act_rpm = values[6]
                    sail_time = values[7]
                    weather_force = values[8]
                    weather_wind_tv = values[9]
                    sea_condition = values[10]
                    swell_dir = values[11]
                    swell_lvl = values[12]
                    current_dir = values[13]
                    current_speed = values[14]

                    swell_lvl= 'n/a' if str(swell_lvl) == 'nan' else str(swell_lvl)
                    try:
                        slip = str(slip).split(',')[0]+"."+str(slip).split(',')[1] if str(slip).split(',').__len__()>0 and str(slip) !='nan' else slip
                    except Exception as e:
                        print(str(e))

                    datetime1 = vdate+" "+time_utc
                    datetime2=datetime.datetime.strptime(datetime1, '%d/%m/%y %H:%M') - datetime.timedelta(hours=float(sail_time), minutes=float(0))
                    datetime2 = str(datetime2)

                    date = vdate
                    month = date.split('/')[1]
                    day = date.split('/')[0]
                    year = '20'+date.split('/')[2]

                    vdate = month+'/'+day+'/'+year

                    hhMMss = time_utc
                    newDate = year + "-" + month + "-" + day + " " + ":".join(hhMMss.split(":")[0:2])
                    datetime1 = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                    datetime1=str(datetime1)
                    try:
                        cursor_myserver.execute(
                            "insert into [dbo].[" + vessel + '_TLG' + "](vdate, time_utc , type ,actl_spd  , avg_spd , slip , act_rpm , sail_time , weather_force , weather_wind_tv , sea_condition , swell_dir,"
                            + "swell_lvl    , current_dir , current_speed , draft ) "
                            + "values ("
                            + "'" + str(vdate) + "', "
                            + "'" + str(time_utc) + "', "
                            + "'" + 'N' + "', "
                            + str(np.nan_to_num(actl_speed)) + ", "
                            + str(np.nan_to_num(avg_speed)) + ", "
                            + str(np.nan_to_num(float(slip))) + ", "
                            + str(np.nan_to_num(act_rpm)) + " , "
                            + str(sail_time) + " , "
                            + str(np.nan_to_num(weather_force)) + " , "
                            + "'" + weather_wind_tv + "', "
                            + "'" + sea_condition + "', "
                            + "'" + swell_dir + "', "
                            + "'" + str(swell_lvl) + "', "
                            + "'" + str(current_dir) + "', "
                            + str(np.nan_to_num(current_speed)) + ", "
                            + str(0) + " )")

                        connection.commit()

                    except Exception as e:
                        print(str(e))

                    cursor_myserver.execute(
                        'SELECT id from ' + vessel + '_TLG' + ' WHERE vdate =' "'" + vdate + "'" + ' AND time_utc =' "'" + time_utc + "'")
                    for row in cursor_myserver.fetchall():
                        val_id = row.id
                    cursor_myserver.execute(
                        'UPDATE ' + vessel + ' SET tlg_id=' + str(val_id) + ' WHERE t_stamp > '"'" + datetime2 + "'" ' AND t_stamp < ' "'" + datetime1 + "'")

                    connection.commit()
                #####END OF INSERTION ON TLG TABLE

                ####START OF INSERTION AT VESSEL TABLE ==> SENSOR DATA PLUS WEATHER HISTORIC DATA INSERTION
                #813602
                cursor_myserver.execute(
                    'UPDATE ' + vessel + ' SET at_port=1 WHERE tlg_id IS NULL')

                cursor_myserver.execute(
                    'UPDATE ' + vessel + ' SET at_port=0 WHERE tlg_id IS NOT NULL')
                connection.commit()
            ####END TELEGRAMS INSERTION IF AVAILABEL

            for i in range(0,len(dtNew)):

                dt = dtNew[i, 13]
                values = np.nan_to_num(np.array(dtNew[i, :13]).astype(float))
                LAT = float(values[11])
                LON = float(values[12])



                ###################################################################
                Vcourse = float(values[3])
                Vstw = float(values[7])
                SensorwindSpeed = float(values[4])
                SensorwindDir = float(values[5])
                SensorwindDir = self.getActualWindDirFromWindSensor(SensorwindDir)
                stw = float(values[7])
                Sovg = float(values[6])
                meFoc = float(values[9])
                rpm = float(values[8])
                draft = float(values[0])
                trim = np.round(float(float(values[2]) - float(values[1])), 4)

                daysFromLastPropClean = 0
                daysFromLastDryDock = 0
                power = values[10]
                atPort = 1 if power == 0 else 0
                ###################################################################
                if stw < 1 and Sovg < 1 and power == 0 and rpm == 0:
                    continue
                if LAT == 0 or LON == 0:
                    continue


                ###CONVERT DECIMAL DEGREE MINUTE LAT , LON INTO DD
                strLATbeforeComma = str(LAT).split('.')[0]
                strLATafterComma  =  str(LAT).split('.')[1]

                pos = 3  if strLATbeforeComma.__len__()==5  else 2
                LATDegrees = float(strLATbeforeComma[0:pos])
                LATmin = float(str(strLATbeforeComma[pos:]+'.'+strLATafterComma))

                strLONbeforeComma = str(LON).split('.')[0]
                strLONafterComma = str(LON).split('.')[1]

                pos = 3 if strLONbeforeComma.__len__() == 5  else 2
                LONDegrees = float(strLONbeforeComma[0:pos])
                LONmin = float(str(strLONbeforeComma[pos:] + '.' + strLONafterComma))

                LAT = LATDegrees + float(LATmin) / 60 #+ float(LATsec) / 3600
                LON = LONDegrees + float(LONmin) / 60 #+ float(LONsec) / 3600
                OrigLAT = LAT
                OrigLON = LON
                #OrigLAT=float('{0:.5f}'.format(OrigLAT))
                #OrigLON=float('{0:.5f}'.format(OrigLON))

                LAT = np.round(LAT)
                LON = np.round(LON)
                ### end lat , lon


                ##FORM WEATHER HISTORY DB NAME TO SEARCH
                date = str(dt).split(" ")[ 0 ]
                day = int(date.split("-")[2])
                month = date.split("-")[1]
                year = date.split("-")[0]
                hhMM = str(dt).split(" ")[ 1 ]
                h = int(hhMM.split(":")[ 0 ])
                M = int(hhMM.split(":")[ 1 ])
                if int(M) > 30:
                    h += 1
                if h % 3 == 2:
                    h += 1
                else:
                    h = h - h % 3

                day = day+1 if h==24 else day
                day = '0'+ str(day) if str(day).__len__()==1 else str(day)
                h = 0 if h==24 else h

                h = '0'+ str(h) if str(h).__len__()==1 else str(h)
                dbDate = year + month + day + h
                #####END WEATHER HISTORY DB NAME

                ##############################
                ##FIND RELATIVES
                ##wind speed
                ##wind dir
                vsl = np.array([Vstw * 0.514, Vcourse])
                wind = np.array([ SensorwindSpeed,SensorwindDir])

                truewindVector = self.convertApparentToTrueWeather(wind,vsl)
                SensorwindSpeed , SensorwindDir = self.DecomposeUV(truewindVector[0],truewindVector[1])#self.decomposeVector(truewindVector)
                relSensorWindDir = self.getRelativeDirectionWeatherVessel( Vcourse,SensorwindDir )

                ##convert cartesian coordinates to polar
                #x = np.array([x[0] * np.cos(x[1]), x[0] * np.sin(x[1])])

                #y = np.array([y[0] * np.cos(y[1]), y[0] * np.sin(y[1])])

                #x_norm = np.sqrt(sum(x ** 2))

                # Apply the formula as mentioned above
                # for projecting a vector onto the orthogonal vector n
                # find dot product using np.dot()
                #proj_of_y_on_x = (np.dot(y, x) / x_norm ** 2) * x

                # vecproj =  np.dot(y, x) / np.dot(y, y) * y
                #relSensorWindSpeed = float(proj_of_y_on_x[0] + SensorwindSpeed)
                #relSensorWindSpeed = float('{0:.3f}'.format(relSensorWindSpeed))
                ##wind dir
                #relSensorWindDir = SensorwindDir - Vcourse
                #relSensorWindDir += 360 if relSensorWindDir < 0 else relSensorWindDir
                ####END RELATIVES

                #SELECT FIELDS FROM WEATHER HISTORY DB
                connWEATHERHISTORY = pyodbc.connect('DRIVER={SQL Server};SERVER=WEATHERSERVER_DEV;'
                                      'DATABASE=WeatherHistoryDB;'
                                      'UID=sa;'
                                      'PWD=sa1!')

                cursor = connWEATHERHISTORY.cursor()

                ##INITIALIZE WEATHER HISTORY DB FIELDS
                #relWindSpeed = 0
                windSpeed = 0
                windDir = 0
                relWindDir = 0
                #relCurrentsSpeed = 0
                relCurrentsDir = 0
                swellSWH = 0
                relSwellDir = 0
                combSWH = 0
                currentsSpeed=0
                relCombDir = 0
                wavesSWH = 0
                relWaveDir = 0
                #####################
                if cursor.tables(table='d' + dbDate , tableType='TABLE').fetchone():
                    #IF WEATHER HISTORY TABLE EXISTS => SELECT FIELDS FROM WEATHER HISTORY DB
                    cursor.execute(
                        'SELECT * FROM d' + dbDate + ' WHERE lat=' + str(LAT) + 'and lon=' + str(
                            LON))

                    for row in cursor.fetchall():
                            windSpeed = float(row.windSpeed)
                            windDir = float(row.windDir)
                            relWindDir = self.getRelativeDirectionWeatherVessel( Vcourse,windDir)
                            currentsSpeed =float( row.currentSpeed)
                            currentsDir = float(row.currentDir)
                            relCurrentsDir = self.getRelativeDirectionWeatherVessel(Vcourse,currentsDir)
                            swellSWH = float(row.swellSWH)
                            swellDir = float(row.swellDir)
                            relSwellDir = self.getRelativeDirectionWeatherVessel(Vcourse,swellDir)
                            combSWH =  float(row.combSWH)
                            combDir =  float(row.combDir)
                            relCombDir = self.getRelativeDirectionWeatherVessel(Vcourse,combDir)
                            wavesSWH =  float(row.wavesSWH)
                            wavesDir =  float(row.wavesDir)
                            relWaveDir = self.getRelativeDirectionWeatherVessel(Vcourse,wavesDir)

                            ##FIND RELATIVES
                            ##wind speed
                            ##wind dir
                            #x = np.array([Vstw * 0.514, Vcourse])
                            #y = np.array([windSpeed, windDir])
                            ##convert cartesian coordinates to polar
                            #x = np.array([x[0] * np.cos(x[1]), x[0] * np.sin(x[1])])

                            #y = np.array([y[0] * np.cos(y[1]), y[0] * np.sin(y[1])])

                            #x_norm = np.sqrt(sum(x ** 2))

                            # Apply the formula as mentioned above
                            # for projecting a vector onto the orthogonal vector n
                            # find dot product using np.dot()
                            #proj_of_y_on_x = (np.dot(y, x) / x_norm ** 2) * x

                            # vecproj =  np.dot(y, x) / np.dot(y, y) * y
                            #relWindSpeed = float(proj_of_y_on_x[0] + windSpeed)
                            #relWindSpeed = float('{0:.3f}'.format(relWindSpeed))

                            ##wind dir
                            #relWindDir = windDir - Vcourse
                            #relWindDir += 360 if relWindDir < 0 else relWindDir

                            ##currents speed
                            ##currents dir
                            #x = np.array([Vstw * 0.514, Vcourse])
                            #y = np.array([currentsSpeed, currentsDir])
                            ##convert cartesian coordinates to polar
                            #x = np.array([x[0] * np.cos(x[1]), x[0] * np.sin(x[1])])

                            #y = np.array([y[0] * np.cos(y[1]), y[0] * np.sin(y[1])])

                            #x_norm = np.sqrt(sum(x ** 2))

                            # Apply the formula as mentioned above
                            # for projecting a vector onto the orthogonal vector n
                            # find dot product using np.dot()
                            #proj_of_y_on_x = (np.dot(y, x) / x_norm ** 2) * x

                            # vecproj =  np.dot(y, x) / np.dot(y, y) * y
                            #relCurrentsSpeed = float(proj_of_y_on_x[0] + currentsSpeed)
                            #relCurrentsSpeed = float('{0:.3f}'.format(relCurrentsSpeed))

                            #relCurrentsDir = currentsDir - Vcourse
                            #relCurrentsDir += 360 if relCurrentsDir < 0 else relCurrentsDir
                            #####################
                            #relWindDir = windDir - Vcourse
                            #relWindDir += 360 if relWindDir < 0 else relWindDir

                            #####################
                            #relWaveDir = wavesDir - Vcourse
                            #relWaveDir += 360 if relWaveDir < 0 else relWaveDir
                            #########
                            #relCombDir = combDir - Vcourse
                            #relCombDir += 360 if relCombDir < 0 else relCombDir
                            #########
                            #relSwellDir = swellDir - Vcourse
                            #relSwellDir += 360 if relSwellDir < 0 else relSwellDir
                            #test = '''ewrwerew
                            #ewrwerwerwe
                            #werwerwe{1}{2}{3}'''
                            #test.format(1,2,3)
                            #dt =  datetime.datetime.strptime(dt, '%Y-%m-%d %H:%M:s')
                            cursor_myserver.execute(
                                "insert into [dbo].[" + vessel + "](lat, lon, t_stamp , course , stw , s_ovg , me_foc , rpm , power , draft , trim , at_port ,daysFromLastPropClean, "
                                + " daysFromLastDryDock , sensor_wind_speed , sensor_wind_dir , wind_speed ,wind_dir ,curr_speed ,curr_dir ,comb_waves_height ,comb_waves_dir , swell_height ,swell_dir,"
                                + "wind_waves_height   , wind_waves_dir ) "

                                + "values ("
                                + str(OrigLAT) + ", "
                                + str(OrigLON) + ", "
                                + "'"+str(dt)+"', "
                                + str(np.nan_to_num(Vcourse)) + ", "
                                + str(np.nan_to_num(stw)) + ", "
                                + str(np.nan_to_num(Sovg)) + " , "
                                + str(np.nan_to_num(meFoc)*1.44) + " , "
                                + str(rpm/10 if str(int(rpm)).__len__()==3 else rpm) + " , "
                                + str(np.nan_to_num(power)) + " , "
                                + str(np.nan_to_num(draft)) + " , "
                                + str(np.nan_to_num(trim)) + " , "
                                + str(atPort) + " , "
                                + str(daysFromLastPropClean) + " , "
                                + str(daysFromLastDryDock) + " , "
                                + str(np.nan_to_num(SensorwindSpeed)) + " , "
                                + str(np.nan_to_num(np.round(relSensorWindDir,2))) + " , "
                                + str(np.nan_to_num(windSpeed)) + " , "
                                + str(np.nan_to_num(relWindDir)) + " , "
                                + str(np.nan_to_num(currentsSpeed)) + ", "
                                + str(np.nan_to_num(relCurrentsDir)) + " , "
                                + str(np.nan_to_num(combSWH)) + " , "
                                + str(np.nan_to_num(relCombDir)) + " , "
                                + str(np.nan_to_num(swellSWH)) + " , "
                                + str(np.nan_to_num(relSwellDir)) + " , "
                                + str(np.nan_to_num(wavesSWH)) + " , "
                                + str(np.nan_to_num(relWaveDir)) + " )")

                            connection.commit()
                ##WEATHER HISTORY TABLE DOES NOT EXISTS
                else:
                    try:
                        cursor_myserver.execute(
                            "insert into [dbo].[" + vessel + "](lat, lon, t_stamp , course , stw , s_ovg , me_foc , rpm , power , draft , trim , at_port ,daysFromLastPropClean, "
                            + " daysFromLastDryDock , sensor_wind_speed , sensor_wind_dir , wind_speed ,wind_dir ,curr_speed ,curr_dir ,comb_waves_height ,comb_waves_dir , swell_height ,swell_dir,"
                            + "wind_waves_height   , wind_waves_dir ) "

                            + "values ("
                            + str(OrigLAT) + ", "
                            + str(OrigLON) + ", "
                            + "'" + str(dt) + "', "
                            + str(Vcourse) + ", "
                            + str(stw) + ", "
                            + str(Sovg) + " , "
                            + str(meFoc * 1.44) + " , "
                            + str( rpm/10 if str(int(rpm)).__len__()==3 else rpm) + " , "
                            + str(power) + " , "
                            + str(draft) + " , "
                            + str(trim) + " , "
                            + str(atPort) + " , "
                            + str(daysFromLastPropClean) + " , "
                            + str(daysFromLastDryDock) + " , "
                            + str(np.nan_to_num(SensorwindSpeed)) + " , "
                            + str(np.nan_to_num(relSensorWindDir)) + " , "
                            + str(windSpeed) + " , "
                            + str(relWindDir) + " , "
                            + str(currentsSpeed) + ", "
                            + str(relCurrentsDir) + " , "
                            + str(combSWH) + " , "
                            + str(relCombDir) + " , "
                            + str(swellSWH) + " , "
                            + str(relSwellDir) + " , "
                            + str(wavesSWH) + " , "
                            + str(relWaveDir) + " )")

                        connection.commit()
                    except Exception as e:
                        print(str(e))
                        #cursor.execute('SELECT * FROM ' + company + 'dbo.data')


                        print("TABLE "'WeatherHistoryDB.dbo.d' + dbDate + " ON DATABASE WeatherHistoryDB DOES NOT EXISTS...INSERTING FIELDS FROM SENSORS.. ")

    def convertLATLONfromDegMinSec(self,lat, lon ,latDir , lonDir ):

        LAT = lat[0:10]
        LON = lon[0:10]
        ###CONVERT DECIMAL DEGREE MINUTE LAT , LON INTO DD
        strLATbeforeComma = str(LAT).split('.')[0]
        strLATafterComma = str(LAT).split('.')[1]

        len = 4 - strLATbeforeComma.__len__()
        strLATbeforeComma = ''.join(['0' for _ in range(len)]) + strLATbeforeComma if strLATbeforeComma.__len__() < 4 else strLATbeforeComma
        pos = 3 if strLATbeforeComma.__len__() == 5   else 2
        LATDegrees = strLATbeforeComma[0:pos]
        LATmin = (str(strLATbeforeComma[pos:]  + strLATafterComma))


        strLONbeforeComma = str(LON).split('.')[0]
        strLONafterComma = str(LON).split('.')[1]

        len = 4 - strLONbeforeComma.__len__()
        strLONbeforeComma = ''.join(['0' for _ in range(len)])+strLONbeforeComma  if strLONbeforeComma.__len__() < 4 else strLONbeforeComma
        #strLONbeforeComma = '0'+strLONbeforeComma if strLONbeforeComma.__len__() < 4 else strLONbeforeComma
        pos = 3 if strLONbeforeComma.__len__() == 5   else 2
        LONDegrees = strLONbeforeComma[0:pos]
        LONmin = (str(strLONbeforeComma[pos:]  + strLONafterComma))


        LAT = float(LATDegrees+'.' + LATmin) #/ 60  # + float(LATsec) / 3600
        LON = float(LONDegrees +'.' +LONmin) #/ 60  # + float(LONsec) / 3600
        OrigLAT = LAT
        OrigLON = LON
        # OrigLAT=float('{0:.5f}'.format(OrigLAT))
        # OrigLON=float('{0:.5f}'.format(OrigLON))

        #LAT = np.round(LAT)
        #LON = np.round(LON)
        LAT = -LAT if latDir == 'S'  else LAT
        LON = -LON if lonDir == 'W'  else LON
        ### end lat , lon
        return LAT , LON

    def mapWeatherData(self,Vcourse,dt ,LAT , LON  ):

        #LAT , LON  = self.convertLATLONfromDegMinSec(lat, lon ,latDir , lonDir)

        ##FORM WEATHER HISTORY DB NAME TO SEARCH
        date = str(dt).split(" ")[0]
        day = int(date.split("-")[2])
        month = date.split("-")[1]
        year = date.split("-")[0]
        hhMM = str(dt).split(" ")[1]
        h = int(hhMM.split(":")[0])
        M = int(hhMM.split(":")[1])
        if int(M) > 30:
            h += 1
        if h % 3 == 2:
            h += 1
        else:
            h = h - h % 3

        day = day + 1 if h == 24 else day
        day = '0' + str(day) if str(day).__len__() == 1 else str(day)
        h = 0 if h == 24 else h

        h = '0' + str(h) if str(h).__len__() == 1 else str(h)
        dbDate = year + month + day + h
        #####END WEATHER HISTORY DB NAME

        connWEATHERHISTORY = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER= 10.2.4.84;'
                                            'DATABASE=WeatherProdData;'
                                            'UID=sa;'
                                            'PWD=sa1!')

        cursor = connWEATHERHISTORY.cursor()


        ##INITIALIZE WEATHER HISTORY DB FIELDS
        # relWindSpeed = 0
        windSpeed = -1
        windDir = -1
        relWindDir = -1
        combSWH = -1
        combDir = -1
        relCombWavesdDir=-1
        swellSWH = -1
        swellDir = -1
        relSwelldir=-1
        wavesSWH =-1
        wavesDir =-1
        relWavesdDir=-1


        #####################
        if cursor.tables(table='d' + dbDate, tableType='TABLE').fetchone():

            # IF WEATHER HISTORY TABLE EXISTS => SELECT FIELDS FROM WEATHER HISTORY DB
            cursor.execute(
                'SELECT * FROM d' + dbDate + ' WHERE lat=' + str(LAT) + 'and lon=' + str(
                    LON))

            for row in cursor.fetchall():
                windSpeed = float(row.windSpeed)
                windDir = float(row.windDir)

                combSWH = float(row.combSWH)
                combDir = float(row.combDir)
                swellSWH = float(row.swellSWH)
                swellDir = float(row.swellDir)
                wavesSWH = float(row.wavesSWH)
                wavesDir = float(row.wavesDir)

                relWindDir = self.getRelativeDirectionWeatherVessel(Vcourse,windDir)
                relSwelldir = self.getRelativeDirectionWeatherVessel(Vcourse, swellDir)
                relWavesdDir = self.getRelativeDirectionWeatherVessel(Vcourse, wavesDir)
                relCombWavesdDir = self.getRelativeDirectionWeatherVessel(Vcourse, combDir)


        return windSpeed ,relWindDir ,swellSWH, relSwelldir , wavesSWH , relWavesdDir ,combSWH , relCombWavesdDir

    def readExtractNewDataset(self, company,vessel,pathToexcel,separator=None):
        ####################
        ####################
        ####################
        if company == 'MILLENIA' or company=='GOLDENPORT':
            ####DOMINIA
            data = pd.read_csv('./data/' + company + '/TELEGRAMS/' + vessel + '.csv', sep=separator, decimal=',')

            # data = data.drop(['DateTime'], axis=1)
            dtNew = np.nan_to_num(data.values)  # .astype(float)
            ballastDt = np.array([k for k in dtNew if k[2] == 'B'])[:, 7:].astype(float)
            ladenDt = np.array([k for k in dtNew if k[2] == 'L'])[:, 7:].astype(float)

            meanDraftBallast = round(float(np.mean(np.array([k for k in ballastDt if k[1] >= 0])[:, 1])), 2)
            meanDraftLadden = round(float(np.mean(np.array([k for k in ladenDt if k[1] >= 0])[:, 1])), 2)

            #draft = (np.array([k for k in dtNew  ])[:,5] + np.array([k for k in dtNew ])[:,6])/2
            #dtNew = np.array(np.append(dtNew.reshape(-1,dtNew.shape[1]) ,  np.asmatrix([draft]).T,axis=1))
            foc24h=[]
            for i in range(0,len(dtNew)):
                dtNew[i,10] = self.getRelativeDirectionWeatherVessel(float(dtNew[i,7]),float(dtNew[i,10]))
                if str(dtNew[i,2])=='nan':
                    if float(dtNew[i,8])>math.floor(meanDraftLadden):
                        dtNew[i,2]='L'
                    else:
                        dtNew[i,2]='B'

                if float(dtNew[i,15])>0:
                    if float(dtNew[i,13]) + float(dtNew[i,14]) > 0:
                        dtNew[i, 15]=( (float(dtNew[i,15])  / (float(dtNew[i,13]) + float(dtNew[i,14]) /60 )) *24)

            #self.fillExcelProfCons(vessel,pathToexcel,dtNew)
            #return

            draft = {"data": [ ]}
            course = {"data": [ ]}
            trim = {"data": [ ]}
            windSpeed = {"data": [ ]}
            windAngle = {"data": [ ]}
            blFlag = {"data": [ ]}
            STW = {"data": [ ]}
            rpm = {"data": [ ]}

            foc = {"data": [ ]}

            dateD = {"data": [ ]}

            data = pd.read_csv('./data/MARIA_M_TLG.csv', sep='\t')
            # data = data.values[0:, :]

            rows = data.values[ :, : ]
            for srow in rows:
                row = srow[ 0 ].split(';')
                for id in range(0, len(row)):

                    if id == 0:
                        dt = {}
                        dt[ "value" ] = row[ id ]
                        dateD[ "data" ].append(dt)

                    if id == 13:
                        drft = {}
                        drft[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        draft[ "data" ].append(drft)

                    if id == 8:
                        stw = {}
                        stw[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        STW[ "data" ].append(stw)

                    if id == 12:
                        wa = {}
                        wa[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        windAngle[ "data" ].append(wa)

                    if id == 11:
                        ws = {}
                        ws[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        windSpeed[ "data" ].append(ws)

                    if id == 10:
                        vcourse = {}
                        vcourse[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        course[ "data" ].append(vcourse)

                    if id == 8:
                        rpms = {}
                        rpms[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        rpm[ "data" ].append(rpms)

                    if id == 15:
                        fo = {}
                        fo[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        foc[ "data" ].append(fo)

                    if id == 1:
                        bl = {}
                        bl[ "value" ] = row[ id ]
                        blFlag[ "data" ].append(bl)

                    if id == 14:
                        vtrim = {}
                        vtrim[ "value" ] = row[ id ].replace(',', '.') if ',' in row[ id ] else row[ id ]
                        trim[ "data" ].append(vtrim)

            with open('./data/MARIA_M.csv', mode='w') as data:
                data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ 'DateTime', 'BL Flag', 'Draft', 'Course', 'WindSpeed', 'WindAngle', 'STW', 'Rpm', 'Trim',
                      'M/E FOC (MT/day)',
                      ])

                for i in range(0, len(dateD[ 'data' ])):
                    dt = dateD[ 'data' ][ i ][ 'value' ]

                    bl = blFlag[ 'data' ][ i ][ 'value' ]
                    vcourse = course[ 'data' ][ i ][ 'value' ]
                    drft = draft[ 'data' ][ i ][ 'value' ]

                    stw = STW[ 'data' ][ i ][ 'value' ]

                    wa = windAngle[ 'data' ][ i ][ 'value' ]

                    ws = windSpeed[ 'data' ][ i ][ 'value' ]

                    vfoc = foc[ 'data' ][ i ][ 'value' ]
                    vtrim = trim[ 'data' ][ i ][ 'value' ]
                    vrpm = rpm[ 'data' ][ i ][ 'value' ]

                    data_writer.writerow(
                        [ dt, bl, drft, vcourse, ws, wa, stw, vrpm, vtrim, vfoc ])

        dataV = pd.read_csv('./data/MARIA_M.csv')
        dataV = dataV.drop([ 'DateTime' ], axis=1)
        dtNew = dataV.values
        ballastDt = np.array([ k for k in dtNew if k[ 0 ] == 'B' ])[ :, 1: ].astype(float)
        ladenDt = np.array([ k for k in dtNew if k[ 0 ] == 'L' ])[ :, 1: ].astype(float)
        x = 0

        if company == 'OCEAN_GOLD':
            dataV = pd.read_csv('./data/PERSEFONE/PERSEFONE_1-30_06_19.csv')
            # dataV=dataV.drop([ 't_stamp', 'vessel status' ], axis=1)
            dtNew = dataV.values  # .astype(float)

            data = pd.read_csv('./export telegram.csv', sep='\t')
            DTLGtNew = data.values[ 0:, : ]
            dateTimesW = [ ]
            ws = [ ]
            arrivalsDate = [ ]
            departuresDate = [ ]

            for i in range(0, len(DTLGtNew[ :, 1 ])):
                if DTLGtNew[ i, 0 ] == 'T004':
                    tlgType = DTLGtNew[ i, 2 ]

                    date = DTLGtNew[ i, 1 ].split(" ")[ 0 ]
                    month = date.split('/')[ 1 ]
                    day = date.split('/')[ 0 ]
                    year = date.split('/')[ 2 ]

                    hhMMss = DTLGtNew[ i, 1 ].split(" ")[ 1 ]
                    newDate = year + "-" + month + "-" + day + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                    dtObj = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                    if tlgType == 'A':
                        arrivalsDate.append(dtObj)
                        for k in range(i, len(DTLGtNew[ :, 1 ])):
                            if DTLGtNew[ k, 0 ] == 'T004':
                                if DTLGtNew[ k, 2 ] == 'D':
                                    date = DTLGtNew[ k, 1 ].split(" ")[ 0 ]
                                    month = date.split('/')[ 1 ]
                                    day = date.split('/')[ 0 ]
                                    year = date.split('/')[ 2 ]

                                    hhMMss = DTLGtNew[ k, 1 ].split(" ")[ 1 ]
                                    newDate = year + "-" + month + "-" + day + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                                    dtObj = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                                    departuresDate.append(dtObj)
                                    break
                    # elif tlgType == 'D':
                    # departuresDate.append(dtObj)

                    dt = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                    # if DdtNew[ i, 7 ] != 'No data' and DdtNew[ i, 8 ] != 'No data':
                    dateTimesW.append(dt)

            dateTimesV = [ ]
            dateTimesVstr = [ ]
            for row in dtNew[ :, 0 ]:
                date = row.split(" ")[ 0 ]
                hhMMss = row.split(" ")[ 1 ]
                newDate = date + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                dtV = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                #################################
                date = str(dtV).split(" ")[ 0 ]
                month = date.split('-')[ 1 ]

                day = date.split('-')[ 2 ]
                # hhMMssV = str(dateTimesV[i]).split(" ")[1]
                year = date.split('-')[ 0 ]

                ##find dateTimesV at sea
                filteredDTWs = [ d for d in dateTimesW if month == str(d).split(' ')[ 0 ].split('-')[ 1 ] and day ==
                                 str(d).split(' ')[ 0 ].split('-')[ 2 ] and year == str(d).split(' ')[ 0 ].split('-')[ 0 ] ]
                for i in range(0, len(arrivalsDate)):
                    if dtV > arrivalsDate[ i ] and dtV < departuresDate[ i ]:
                        dateTimesV.append(dtV)
                        dateTimesVstr.append(str(dtV))
                # for filteredDate in filteredDTWs:
                # date = str(filteredDate).split(" ")[0]
                # month = date.split('-')[1]
                # month = month[1] if month[0] == '0' else month
                # day = date.split('-')[2]
                # day = day[1] if day[0] == '0' else day
                # year = date.split('-')[0]
                # hhMMss = str(filteredDate).split(" ")[ 1 ]

                # newDate = day + "/" + month + "/" + year + " " + ":".join(hhMMss.split(":"))
                # tlgType=np.array(data.loc[ data[ 'telegram_date' ] == newDate ].values)[ 0 ][ 2 ]
                # if tlgType == 'A' or tlgType=='N':
                # x=0

                ############################

            draftsVmapped = [ ]
            ballastLadenFlagsVmapped = [ ]
            for i in range(0, len(dateTimesV)):

                datesWs = [ ]

                date = str(dateTimesV[ i ]).split(" ")[ 0 ]
                month = date.split('-')[ 1 ]

                day = date.split('-')[ 2 ]

                hhMMssV = str(dateTimesV[ i ]).split(" ")[ 1 ]

                year = date.split('-')[ 0 ]
                dt = dateTimesV[ i ]
                day = '0' + day if day.__len__() == 1 else day
                month = '0' + month if month.__len__() == 1 else month
                filteredDTWs = [ d for d in dateTimesW if month == str(d).split(' ')[ 0 ].split('-')[ 1 ] and day ==
                                 str(d).split(' ')[ 0 ].split('-')[ 2 ] and year == str(d).split(' ')[ 0 ].split('-')[ 0 ] ]

                date2numFiltered = {}
                sorted = [ ]

                for k in range(0, len(filteredDTWs)):
                    # dtDiff[filteredDTWs[k]]=abs((filteredDTWs[k]-dt).total_seconds() / 60.0)
                    date = str(filteredDTWs[ k ]).split(" ")[ 0 ]
                    month = date.split('-')[ 1 ]
                    month = month[ 1 ] if month[ 0 ] == '0' else month
                    day = date.split('-')[ 2 ]
                    day = day[ 1 ] if day[ 0 ] == '0' else day
                    year = date.split('-')[ 0 ]
                    hhMMss = str(filteredDTWs[ k ]).split(" ")[ 1 ]
                    # newDate = month + "/" + day + "/" + year + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                    newDate = day + "/" + month + "/" + year + " " + ":".join(hhMMss.split(":"))
                    datesWs.append(newDate)

                    date2numFilteredList = [ ]

                    sorted.append(newDate)
                    # date2numFiltered.append(matplotlib.dates.date2num(k))

                    # dtFiltered[ filteredDTWs[ k ] ] = matplotlib.dates.date2num(k)
                    # dt2num = matplotlib.dates.date2num(dt)
                    # date2numFiltered.append(dt2num)
                date = str(dt).split(" ")[ 0 ]
                month = date.split('-')[ 1 ]
                month = month[ 1 ] if month[ 0 ] == '0' else month
                day = date.split('-')[ 2 ]
                day = day[ 1 ] if day[ 0 ] == '0' else day
                year = date.split('-')[ 0 ]

                hhMMss = hhMMssV

                # newDate = month + "/" + day + "/" + year + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                newDate = day + "/" + month + "/" + year + " " + ":".join(hhMMss.split(":"))
                sorted.append(newDate)
                sorted.sort()
                for i in range(0, len(sorted)):
                    if sorted[ i ] == newDate:
                        try:
                            if i > 0:
                                if i == len(sorted) - 1:
                                    previous = sorted[ i - 1 ]
                                    blFlag = np.array(data.loc[ data[ 'telegram_date' ] == previous ].values)[ 0 ][ 49 ]
                                    if blFlag == 'B':
                                        ballastLadenFlagsVmapped.append('B')
                                    else:
                                        ballastLadenFlagsVmapped.append('L')
                                else:
                                    try:
                                        previous = sorted[ i - 1 ]
                                        next = sorted[ i + 1 ]

                                        # date1 = {k: v for k, v in dtFiltered.items() if v == previous}
                                        # date2 = {k: v for k, v in dtFiltered.items() if v == next}
                                        blFlag1 = np.array(data.loc[ data[ 'telegram_date' ] == previous ].values)[ 0 ][
                                            49 ]
                                        blFlag2 = np.array(data.loc[ data[ 'telegram_date' ] == next ].values)[ 0 ][ 49 ]
                                    except:
                                        x = 0
                                    if blFlag1 == 'L' and blFlag2 == 'L':
                                        ballastLadenFlagsVmapped.append('L')
                                    elif blFlag1 == 'L' and blFlag2 == 'B':
                                        ballastLadenFlagsVmapped.append('L')
                                    elif blFlag1 == 'B' and blFlag2 == 'L':
                                        ballastLadenFlagsVmapped.append('B')
                                    else:
                                        ballastLadenFlagsVmapped.append('B')

                            elif i == 0:
                                # day =
                                # day = float(day)-1
                                # filteredDTWs=[]
                                # for i in range(0,len(dateTimesW)):
                                # yearW =  str(dateTimesW[i]).split(' ')[ 0 ].split('-')[ 0 ]
                                # monthW =  str(dateTimesW[i]).split(' ')[ 0 ].split('-')[ 1 ]
                                # dayW = str(dateTimesW[i]).split(' ')[ 0 ].split('-')[ 2 ]
                                # while filteredDTWs.__len__()==0:
                                # if month==monthW and day ==dayW and year==yearW:
                                # filteredDTWs.append(dateTimesW[i])
                                newDate = dateTimesV[ i ] - datetime.timedelta(1)

                                date = str(newDate).split(" ")[ 0 ]
                                month = date.split('-')[ 1 ]

                                day = date.split('-')[ 2 ]

                                hhMMssV = str(newDate).split(" ")[ 1 ]

                                year = date.split('-')[ 0 ]
                                dt = dateTimesV[ i ]
                                day = '0' + day if day.__len__() == 1 else day
                                month = '0' + month if month.__len__() == 1 else month

                                filteredDTWsBL = [ d for d in dateTimesW if
                                                   month == str(d).split(' ')[ 0 ].split('-')[ 1 ] and day ==
                                                   str(d).split(' ')[ 0 ].split('-')[ 2 ] and year ==
                                                   str(d).split(' ')[ 0 ].split('-')[ 0 ] ]

                                # next = sorted[ i + 1 ]
                                previous = str(filteredDTWsBL[ len(filteredDTWsBL) - 1 ])
                                dtBL = previous.split(' ')[ 0 ]
                                month = dtBL.split('-')[ 1 ]
                                month = month[ 1 ] if month[ 0 ] == '0' else month
                                day = dtBL.split('-')[ 2 ]
                                day = day[ 1 ] if day[ 0 ] == '0' else day
                                year = dtBL.split('-')[ 0 ]
                                hhMMss = str(previous).split(" ")[ 1 ]
                                previous = day + "/" + month + "/" + year + " " + ":".join(hhMMss.split(":"))

                                blFlag = np.array(data.loc[ data[ 'telegram_date' ] == previous ].values)[ 0 ][ 49 ]
                                if blFlag == 'B':
                                    ballastLadenFlagsVmapped.append('B')
                                else:
                                    ballastLadenFlagsVmapped.append('L')
                            elif i == len(sorted) - 1:
                                previous = sorted[ i - 1 ]
                                blFlag = np.array(data.loc[ data[ 'telegram_date' ] == previous ].values)[ 0 ][ 49 ]
                                if blFlag == 'B':
                                    ballastLadenFlagsVmapped.append('B')
                                else:
                                    ballastLadenFlagsVmapped.append('L')
                        except Exception as e:

                            d = 0
                            print(e)
                            ballastLadenFlagsVmapped.append('N')

                            # firstMin = min(dtDiff.items(), key=lambda x: x[1])
                # del dtDiff[ firstMin[ 0 ] ]
                # secMin = min(dtDiff.items(), key=lambda x: x[ 1 ] )

                drafts = [ ]
                for date in datesWs:
                    drftAFT = str(np.nan_to_num(np.array(data.loc[ data[ 'telegram_date' ] == date ].values)[ 0 ][ 27 ]))
                    drftFORE = str(np.nan_to_num(np.array(data.loc[ data[ 'telegram_date' ] == date ].values)[ 0 ][ 28 ]))

                    drftAFT = float(drftAFT.split(',')[ 0 ] + "." + drftAFT.split(',')[ 1 ]) if drftAFT.__contains__(
                        ',') else float(drftAFT)
                    drftFORE = float(
                        drftFORE.split(',')[ 0 ] + "." + drftFORE.split(',')[ 1 ]) if drftFORE.__contains__(
                        ',') else float(drftFORE)

                    drft = (drftAFT + drftFORE) / 2
                    drafts.append(float(drft))
                if len(datesWs) > 1:
                    dx = 0
                try:
                    y = np.array(drafts)
                    x = matplotlib.dates.date2num(filteredDTWs)
                    f = sp.Earth()
                    f.fit(x.reshape(-1, 1), y.reshape(-1, 1))
                    pred = f.predict(np.array(matplotlib.dates.date2num(dt)).reshape(-1, 1))
                    draftsVmapped.append(pred)
                except Exception as e:
                    dx = 0
                    # if np.nan_to_num(np.array(data.loc[ data[ 'telegram_date' ] == date ].values)[ 0 ][ 27 ]).__len__()==0 or \
                    # np.nan_to_num(np.array(data.loc[ data[ 'telegram_date' ] == date ].values)[ 0 ][ 28 ]).__len__() == 0:
                    print(e)
                    if drafts.__len__() == 0:
                        draftsVmapped.append([ 0 ])

            ##########

            windSpeedVmapped = [ ]
            windDirsVmapped = [ ]

            # dateTimesV = [ ]
            # for row in dtNew[ :, 0 ]:
            # date = row.split(" ")[ 0 ]
            # hhMMss = row.split(" ")[ 1 ]
            # newDate = date + " " + ":".join(hhMMss.split(":")[ 0:2 ])
            # dt = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
            # dateTimesV.append(dt)

            data = pd.read_excel('./data/Persefone/Persefone TrackReport.xls')

            WdtNew = data.values[ 0:, : ]
            dateTimesW = [ ]
            ws = [ ]
            newDataset = [ ]
            for i in range(0, len(WdtNew[ :, 1 ])):
                date = WdtNew[ i, 1 ].split(" ")[ 0 ]
                month = date.split('.')[ 1 ]
                day = date.split('.')[ 0 ]
                year = date.split('.')[ 2 ]

                hhMMss = WdtNew[ i, 1 ].split(" ")[ 1 ]
                newDate = year + "-" + month + "-" + day + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                dt = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')
                if WdtNew[ i, 7 ] != 'No data' and WdtNew[ i, 8 ] != 'No data':
                    dateTimesW.append(dt)

            for i in range(0, len(dtNew)):

                # dtFiltered = {}
                #########################
                if str(dtNew[ i ][ 0 ].split(".")[ 0 ]) in dateTimesVstr:
                    newDataset.append(dtNew[ i, : ].reshape(1, -1))

            for i in range(0, len(dateTimesV)):

                datesWs = [ ]

                date = str(dateTimesV[ i ]).split(" ")[ 0 ]
                month = date.split('-')[ 1 ]
                day = date.split('-')[ 2 ]
                year = date.split('-')[ 0 ]
                dt = dateTimesV[ i ]

                filteredDTWs = [ d for d in dateTimesW if month == str(d).split(' ')[ 0 ].split('-')[ 1 ] and day ==
                                 str(d).split(' ')[ 0 ].split('-')[ 2 ] and year == str(d).split(' ')[ 0 ].split('-')[ 0 ] ]
                date2numFiltered = {}

                for k in range(0, len(filteredDTWs)):
                    date = str(filteredDTWs[ k ]).split(" ")[ 0 ]
                    month = date.split('-')[ 1 ]
                    day = date.split('-')[ 2 ]
                    year = date.split('-')[ 0 ]
                    hhMMss = str(filteredDTWs[ k ]).split(" ")[ 1 ]
                    newDate = day + "." + month + "." + year + " " + ":".join(hhMMss.split(":")[ 0:2 ])

                    datesWs.append(newDate)

                windSpeeds = [ ]
                windDirs = [ ]
                for date in datesWs:
                    ws = np.array(data.loc[ data[ 'Date' ] == date + " Z" ].values)[ 0 ][ 7 ]
                    wd = np.array(data.loc[ data[ 'Date' ] == date + " Z" ].values)[ 0 ][ 8 ]
                    if ws != 'No data':
                        windSpeeds.append(float(ws.split('m')[ 0 ]))
                    if wd != 'No data':
                        windDirs.append(float(wd[ :len(wd) - 1 ]))
                try:
                    y = np.array(windSpeeds)
                    x = matplotlib.dates.date2num(filteredDTWs)
                    f = sp.Earth(max_degree=2)
                    f.fit(x.reshape(-1, 1), y.reshape(-1, 1))
                    pred = f.predict(np.array(matplotlib.dates.date2num(dt)).reshape(-1, 1))
                    windSpeedVmapped.append(pred)
                except:

                    if windSpeeds.__len__() == 0:
                        windSpeedVmapped.append(None)

                try:
                    y = np.array(windDirs)
                    x = matplotlib.dates.date2num(filteredDTWs)
                    f = sp.Earth(max_degree=2)

                    f.fit(x.reshape(-1, 1), y.reshape(-1, 1))
                    pred = f.predict(np.array(matplotlib.dates.date2num(dt)).reshape(-1, 1))
                    # f=interp1d(list(x), list(y), kind='cubic')
                    windDirsVmapped.append(pred)
                except:

                    if windDirs.__len__() == 0:
                        windDirsVmapped.append(None)

            # pred = f(matplotlib.dates.date2num(dt))
            # data = pd.read_excel('./Penelope TrackReport.xls')
            # WdtNew = data.values[ 0:, : ]
            # print(pred)

            dataV = pd.read_csv('./data/Persefone/PERSEFONE_1-30_06_19.csv')
            newDataset = np.array(newDataset).reshape(-1, dtNew.shape[ 1 ])

            dataV = dataV.drop([ 't_stamp', 'vessel status' ], axis=1)
            dtNew = dataV.values.astype(float)

            newMappedData = np.array(

                np.append(newDataset, np.asmatrix(
                    [ np.array(windSpeedVmapped).reshape(-1), np.array(windDirsVmapped).reshape(-1),
                      np.array(ballastLadenFlagsVmapped)[ 0:len(draftsVmapped) ] ]).T,
                          axis=1))

            Vcourse = np.array(newMappedData[ :, 2 ]).astype(float)
            Vstw = np.array(newMappedData[ :, 1 ]).astype(float)
            windDir = np.array(newMappedData[ :, 26 ]).astype(float)
            windSpeed = np.array(newMappedData[ :, 25 ]).astype(float)
            ###projection of vector wind speed direction vector into vessel speed direction orthogonal system of coordinates
            ##convert vessel speed (knots) to m/s
            relWindSpeeds = [ ]
            relWindDirs = [ ]
            for i in range(0, len(Vcourse)):
                #x = np.array([ Vstw[ i ] * 0.514, Vcourse[ i ] ])
                #y = np.array([ windSpeed[ i ], windDir[ i ] ])
                ##convert cartesian coordinates to polar
                #x = np.array([ x[ 0 ] * np.cos(x[ 1 ]), x[ 0 ] * np.sin(x[ 1 ]) ])

                #y = np.array([ y[ 0 ] * np.cos(y[ 1 ]), y[ 0 ] * np.sin(y[ 1 ]) ])

                #x_norm = np.sqrt(sum(x ** 2))

                # Apply the formula as mentioned above
                # for projecting a vector onto the orthogonal vector n
                # find dot product using np.dot()
                #proj_of_y_on_x = (np.dot(y, x) / x_norm ** 2) * x

                # vecproj =  np.dot(y, x) / np.dot(y, y) * y
                #relWindSpeed = proj_of_y_on_x[ 0 ] + windSpeed[ i ]
                relDir = self.getRelativeDirectionWeatherVessel(Vcourse[i],windDir[i])
                relWindDirs.append(relDir)
                relWindSpeeds.append(windSpeed[i])
            newMappedData = np.array(
                np.append(newDataset,

                          np.asmatrix([ np.array(relWindSpeeds), np.array(relWindDirs), np.array(draftsVmapped).reshape(-1),
                                        np.array(ballastLadenFlagsVmapped)[ 0:len(draftsVmapped)] ]).T , axis=1))  # .astype(float)

            #####################
            ############################MAP TELEGRAM DATES FOR DRAFT

            ####STATSTICAL ANALYSIS PENELOPE

            with open('./PERSEFONE_mapped_new_3.csv', mode='w') as dataw:

                data_writer = csv.writer(dataw, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow([ 'STW', 'WindSpeed', 'WindAngle', 'Draft'
                                         , 'M/E FOC (MT/days)', 'B/L Flag' ])
                for k in range(0, len(newMappedData)):
                    data_writer.writerow([ newMappedData[ k ][ 1 ], newMappedData[ k ][ 25 ], newMappedData[ k ][ 26 ],
                                           newMappedData[ k ][ 27 ], newMappedData[ k ][ 8 ], newMappedData[ k ][ 28 ] ])

            x = 0

    def readLarosDAta(self, dtFrom, dtTo):

        UNK = {"data": [ ]}
        draft = {"data": [ ]}
        draftAFT = {"data": [ ]}
        draftFORE = {"data": [ ]}
        windSpeed = {"data": [ ]}
        windAngle = {"data": [ ]}
        SpeedOvg = {"data": [ ]}
        STW = {"data": [ ]}
        rpm = {"data": [ ]}
        power = {"data": [ ]}
        foc = {"data": [ ]}
        foct = {"data": [ ]}
        dateD = {"data": [ ]}
        vCourse = {"data": []}
        lat = {"data": []}
        lon = {"data": []}
        ListTimeStamps = [ ]

        drftFList = [ ]
        drftAList = [ ]
        drftList = [ ]
        STWList = [ ]
        POWList = [ ]
        RPMList = [ ]
        WAList = [ ]
        WSList = [ ]
        SOVGList = [ ]
        FOCList = [ ]
        FOCTList = [ ]
        UNKList = [ ]
        vCourseList=[]
        latList=[]
        lonList = []

        timeStampsCounter = 0
        drftFCounter = 0
        drftACounter = 0
        drftCounter = 0
        STWCounter = 0
        POWCounter = 0
        RPMCounter = 0
        WACounter = 0
        WSCounter = 0
        SOVGCounter = 0
        FOCCounter = 0
        FOCTCounter = 0
        UNKCounter = 0
        vCourseCounter=0
        latCounter=0
        lonCounter = 0

        endDate=""
        startDate=""
        # print(sum(1 for line in open('./MT_DELTA_MARIA.txt')))
        with open('./MT_DELTA_MARIA.txt') as csv_file:

            csv_reader = csv.reader(csv_file, delimiter=';')
            line_count = 0

            for row in csv_reader:
                try:

                    if row[ 1 ] == "STATUS_PARAMETER_ID":
                        continue
                    if row[ 2 ] == '122':
                        x = 0
                    id = row[ 1 ]
                    date = row[ 4 ].split(" ")[ 0 ]
                    hhMMss = row[ 4 ].split(" ")[ 1 ]
                    newDate = date + " " + ":".join(hhMMss.split(":")[ 0:2 ])
                    dt = datetime.datetime.strptime(newDate, '%Y-%m-%d %H:%M')

                    value = row[ 3 ]
                    siteId = row[ 2 ]
                    # if dt >  datetime.datetime.strptime('2018-05-25', '%Y-%m-%d'):

                    if id == "7":
                        if ListTimeStamps != [ ]:
                            if dt != ListTimeStamps[ timeStampsCounter - 1 ]:
                                datt = {}
                                datt[ "value" ] = 0
                                datt[ "dt" ] = dt
                                dateD[ "data" ].append(datt)
                                print(dt)
                                ListTimeStamps.append(dt)
                                timeStampsCounter = timeStampsCounter + 1
                        else:
                            datt = {}
                            datt[ "value" ] = 0
                            datt[ "dt" ] = dt
                            dateD[ "data" ].append(datt)
                            print(dt)
                            ListTimeStamps.append(dt)
                            timeStampsCounter = timeStampsCounter + 1

                    if id == '7':

                            unknown = {}
                            unknown[ "value" ] = value
                            unknown[ "dt" ] = dt
                            UNK[ "data" ].append(unknown)
                            #UNKList.append(dt)
                            UNKCounter = UNKCounter + 1

                    if id == '65':
                        #if vCourseList != []:
                            #if dt != vCourseList[vCourseCounter - 1]:
                                course = {}
                                course["value"] = value
                                course["dt"] = dt
                                vCourse["data"].append(course)
                                vCourseList.append(dt)
                                vCourseCounter = vCourseCounter + 1
                                # else:
                    #course = {}
                            #course["value"] = value
                            #course["dt"] = dt
                            #vCourse["data"].append(course)
                            #vCourseList.append(dt)
                            #vCourseCounter = vCourseCounter + 1


                    if id == '21':
                        #if drftAList != [ ]:
                            #if dt != drftAList[ drftACounter - 1 ]:
                                drftA = {}
                                drftA[ "value" ] = value
                                drftA[ "dt" ] = dt
                                draftAFT[ "data" ].append(drftA)
                                drftAList.append(dt)
                                drftACounter = drftACounter + 1
                                #else:
                    #drftA = {}
                            #drftA[ "value" ] = value
                            #drftA[ "dt" ] = dt
                            #draftAFT[ "data" ].append(drftA)
                            #drftAList.append(dt)
                            #drftACounter = drftACounter + 1

                    if id == '35':
                        #if latList != []:
                        #if dt != latList[latCounter - 1]:
                                latitude = {}
                                latitude["value"] = value
                                latitude["dt"] = dt
                                lat["data"].append(latitude)
                                latList.append(dt)
                                latCounter = latCounter + 1
                                #else:
                    #latitude = {}
                            #latitude["value"] = value
                            #latitude["dt"] = dt
                            #lat["data"].append(latitude)
                            #latList.append(dt)
                            #latCounter = latCounter + 1

                    if id == '38':
                        #if  lonList!= []:
                            #if dt != lonList[lonCounter - 1]:
                                longitude = {}
                                longitude["value"] = value
                                longitude["dt"] = dt
                                lon["data"].append(longitude)
                                lonList.append(dt)
                                lonCounter = lonCounter + 1
                                #else:
                    #longitude = {}
                            #longitude["value"] = value
                            #longitude["dt"] = dt
                            #lon["data"].append(longitude)
                            #lonList.append(dt)
                            #lonCounter = lonCounter + 1

                    if id == '67':
                        #if STWList != [ ]:
                            #if dt != STWList[ STWCounter - 1 ]:
                                stw = {}
                                stw[ "value" ] = value
                                stw[ "dt" ] = dt
                                STW[ "data" ].append(stw)
                                STWList.append(dt)
                                STWCounter = STWCounter + 1
                                #else:
                    #stw = {}
                            #stw[ "value" ] = value
                            #stw[ "dt" ] = dt
                            #STW[ "data" ].append(stw)
                            #STWList.append(dt)
                            #STWCounter = STWCounter + 1

                    if id == '22':
                        #if drftFList != [ ]:
                        #if dt != drftFList[ drftFCounter - 1 ]:
                                drftF = {}
                                drftF[ "value" ] = value
                                drftF[ "dt" ] = dt
                                draftFORE[ "data" ].append(drftF)
                                drftFList.append(dt)
                                drftFCounter = drftFCounter + 1
                                #else:
                    #drftF = {}
                            #drftF[ "value" ] = value
                            #drftF[ "dt" ] = dt
                            #draftFORE[ "data" ].append(drftF)
                            #drftFList.append(dt)
                            #drftFCounter = drftFCounter + 1

                    if id == '61':
                        #if drftList != [ ]:
                            #if dt != drftList[ drftCounter - 1 ]:
                                drft = {}
                                drft[ "value" ] = value
                                drft[ "dt" ] = dt
                                draft[ "data" ].append(drft)
                                drftList.append(dt)
                                drftCounter = drftCounter + 1
                                #else:
                    #drft = {}
                            #drft[ "value" ] = value
                            #drft[ "dt" ] = dt
                            #draft[ "data" ].append(drft)
                            #drftList.append(dt)
                            #drftCounter = drftCounter + 1

                    if id == '81':
                        #if WAList != [ ]:
                        #if dt != WAList[ WACounter - 1 ]:
                                wa = {}
                                wa[ "value" ] = value
                                wa[ "dt" ] = dt
                                windAngle[ "data" ].append(wa)
                                WAList.append(dt)
                                WACounter = WACounter + 1
                                #else:
                    #wa = {}
                            #wa[ "value" ] = value
                            #wa[ "dt" ] = dt
                            #windAngle[ "data" ].append(wa)
                            #WAList.append(dt)
                            #WACounter = WACounter + 1

                    if id == '82':
                        #if WSList != [ ]:
                            #if dt != WSList[ WSCounter - 1 ]:
                                ws = {}
                                ws[ "value" ] = value
                                ws[ "dt" ] = dt
                                windSpeed[ "data" ].append(ws)
                                WSList.append(dt)
                                WSCounter = WSCounter + 1
                                #else:
                    #ws = {}
                            #ws[ "value" ] = value
                            #ws[ "dt" ] = dt
                            #windSpeed[ "data" ].append(ws)
                            #WSList.append(dt)
                            #WSCounter = WSCounter + 1

                    if id == '84':
                        #if SOVGList != [ ]:
                        #if dt != SOVGList[ SOVGCounter - 1 ]:
                                so = {}
                                so[ "value" ] = value
                                so[ "dt" ] = dt
                                SpeedOvg[ "data" ].append(so)
                                SOVGList.append(dt)
                                SOVGCounter = SOVGCounter + 1
                                #else:
                    #so = {}
                            #so[ "value" ] = value
                            #so[ "dt" ] = dt
                            #SpeedOvg[ "data" ].append(so)
                            #SOVGList.append(dt)
                            #SOVGCounter = SOVGCounter + 1

                    if id == '89':
                        #if RPMList != [ ]:
                            #if dt != RPMList[ RPMCounter - 1 ]:
                                rpms = {}
                                rpms[ "value" ] = value
                                rpms[ "dt" ] = dt
                                rpm[ "data" ].append(rpms)
                                RPMList.append(dt)
                                RPMCounter = RPMCounter + 1
                                #else:
                    #rpms = {}
                            #rpms[ "value" ] = value
                            #rpms[ "dt" ] = dt
                            #rpm[ "data" ].append(rpms)
                            #RPMList.append(dt)
                            #RPMCounter = RPMCounter + 1

                    if id == '137':
                        #if FOCList != [ ]:
                        #if dt != FOCList[ FOCCounter - 1 ]:
                                fo = {}
                                fo[ "value" ] = value
                                fo[ "dt" ] = dt
                                foc[ "data" ].append(fo)
                                FOCList.append(dt)
                                FOCCounter = FOCCounter + 1
                                #else:
                    #fo = {}
                            #fo[ "value" ] = value
                            #fo[ "dt" ] = dt
                            #foc[ "data" ].append(fo)
                            #FOCList.append(dt)
                            #FOCCounter = FOCCounter + 1



                    if id == '353':
                        #if POWList != [ ]:
                            #if dt != POWList[ POWCounter - 1 ]:
                                pow = {}
                                pow[ "value" ] = value
                                pow[ "dt" ] = dt
                                power[ "data" ].append(pow)
                                POWList.append(dt)
                                POWCounter = POWCounter + 1
                        #else:
                            #pow = {}
                            #pow[ "value" ] = value
                        #pow[ "dt" ] = dt
                        #power[ "data" ].append(pow)
                        #POWList.append(dt)
                        #POWCounter = POWCounter + 1

                    line_count += 1

                    if int(id) >353:
                        break

                    # print('Processed '+str(line_count)+ ' lines.')
                except Exception as e:
                    p = 0
                    print(str(e))
        try:
            # with open('./WSraw.csv', mode='w') as data:
            # ws_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            # ws_writer.writerow(['Field', 'DateTime'])
            # print(len(windSpeed["data"]))
            # for i in range(0, len(windSpeed["data"])):
            # ws_writer.writerow([windSpeed["data"][i]['value'], windSpeed["data"][i]['dt']])
            # return
            endDate = str(dateD[ "data" ][ len(dateD[ "data" ]) - 1 ][ "dt" ])
            startDate = str(dateD[ "data" ][ 0 ][ "dt" ])

            continuousDTS = (pd.DataFrame(columns=[ 'NULL' ],
                                          index=pd.date_range(startDate, endDate,
                                                              freq='1T'))
                # .between_time('00:00', '15:11')
                .index.strftime('%Y-%m-%d %H:%M:%S')
                .tolist()
                )

            try:
                with open('./data/LAROS/UNK.csv', mode='w') as data:
                    course_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    course_writer.writerow(['Field', 'DateTime'])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(UNK["data"][k]['dt']) == str(UNK["data"][k - 1]['dt']) else k
                        if str(UNK["data"][k]['dt']) == continuousDTS[i]:
                            course_writer.writerow([UNK["data"][k]['value'], UNK["data"][k]['dt']])
                            k = k + 1
                            UNKList.append(UNK["data"][k]['dt'])
                        else:
                            course_writer.writerow(["", continuousDTS[i]])
                            UNKList.append(continuousDTS[i])

            except:
                d = 0


            with open('./data/LAROS/DateTime.csv', mode='w') as data:
                dt_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                dt_writer.writerow([ 'DateTimes' ])
                for i in range(0, len(continuousDTS)):
                    dt_writer.writerow([ continuousDTS[ i ] ])

            try:
                with open('./data/LAROS/Course.csv', mode='w') as data:
                    course_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    course_writer.writerow(['Field', 'DateTime'])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(vCourse["data"][k]['dt']) == str(vCourse["data"][k - 1]['dt']) else k
                        if str(vCourse["data"][k]['dt']) == continuousDTS[i]:
                            course_writer.writerow([vCourse["data"][k]['value'], vCourse["data"][k]['dt']])
                            k = k + 1
                        else:
                            course_writer.writerow(["", continuousDTS[i]])

            except:
                d = 0

            try:
                with open('./data/LAROS/LAT.csv', mode='w') as data:
                    lat_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    lat_writer.writerow(['Field', 'DateTime'])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(lat["data"][k]['dt']) == str(lat["data"][k - 1]['dt']) else k
                        if str(lat["data"][k]['dt']) == continuousDTS[i]:
                            lat_writer.writerow([lat["data"][k]['value'], lat["data"][k]['dt']])
                            k = k + 1
                        else:
                            lat_writer.writerow(["", continuousDTS[i]])

            except:
                d = 0

            try:
                with open('./data/LAROS/LON.csv', mode='w') as data:
                    lon_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    lon_writer.writerow(['Field', 'DateTime'])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(lon["data"][k]['dt']) == str(lon["data"][k - 1]['dt']) else k
                        if str(lon["data"][k]['dt']) == continuousDTS[i]:
                            lon_writer.writerow([lon["data"][k]['value'], lon["data"][k]['dt']])
                            k = k + 1
                        else:
                            lon_writer.writerow(["", continuousDTS[i]])

            except:
                d = 0

            try:
                with open('./data/LAROS/Draft.csv', mode='w') as data:
                    draft_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    draft_writer.writerow([ 'Field', 'DateTime' ])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(draft[ "data" ][ k ][ 'dt' ]) == str(draft[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(draft[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            draft_writer.writerow([ draft[ "data" ][ k ][ 'value' ], draft[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1

                        else:
                            draft_writer.writerow([ "", continuousDTS[ i ] ])

            except:
                d = 0

            try:
                with open('./data/LAROS/WS.csv', mode='w') as data:
                    ws_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    ws_writer.writerow([ 'Field', 'DateTime' ])

                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(windSpeed[ "data" ][ k ][ 'dt' ]) == str(
                                windSpeed[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(windSpeed[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            ws_writer.writerow([ windSpeed[ "data" ][ k ][ 'value' ], windSpeed[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD["data"][i]['dt'] > windSpeed["data"][k]['dt']:
                            # k = k + 1
                            # ws_writer.writerow([windSpeed["data"][k]['value'], dateD["data"][i]['dt']])
                            # k = k + 1
                            # else:
                            ws_writer.writerow([ "", continuousDTS[ i ] ])

                            # for i in range(0, len(windSpeed[ "data" ])):
                            # ws_writer.writerow([ windSpeed[ "data" ][ i ][ 'value' ], windSpeed[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/DraftAFT.csv', mode='w') as data:
                    draftA_writer = csv.writer(data, delimiter=',', quotechar='"',
                                               quoting=csv.QUOTE_MINIMAL)
                    draftA_writer.writerow([ 'Field', 'DateTime' ])

                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if str(dateD[ "data" ][ i ][ 'dt' ]) == str(
                                datetime.datetime.strptime('2018-10-09 13:01', '%Y-%m-%d %H:%M')):
                            print(dt)

                        if k > 0:
                            k = k + 1 if str(draftAFT[ "data" ][ k ][ 'dt' ]) == str(
                                draftAFT[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(draftAFT[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            draftA_writer.writerow([ draftAFT[ "data" ][ k ][ 'value' ], draftAFT[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > draftAFT[ "data" ][ k ][ 'dt' ]:
                            # k=k+1
                            # draftA_writer.writerow([ draftAFT[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k=k+1
                            # else:

                            draftA_writer.writerow([ "", continuousDTS[ i ] ])
            except:
                ex = 0
                # for i in range(0, len(draftAFT[ "data" ])):
                # draftA_writer.writerow([ draftAFT[ "data" ][ i ][ 'value' ], draftAFT[ "data" ][ i ][ 'dt' ] ])

            try:
                with open('./data/LAROS/DraftFORE.csv', mode='w') as data:
                    draftF_writer = csv.writer(data, delimiter=',', quotechar='"',
                                               quoting=csv.QUOTE_MINIMAL)
                    draftF_writer.writerow([ 'Field', 'DateTime' ])

                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(draftFORE[ "data" ][ k ][ 'dt' ]) == str(draftFORE[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(draftFORE[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            draftF_writer.writerow(
                                [ draftFORE[ "data" ][ k ][ 'value' ], draftFORE[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k-1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > draftFORE[ "data" ][ k ][ 'dt' ]:
                            # k=k+1
                            # draftF_writer.writerow([ draftFORE[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k=k+1
                            # else:
                            draftF_writer.writerow([ "", continuousDTS[ i ] ])
                        # k = +1

                # for i in range(0, len(draftFORE[ "data" ])):
                # draftF_writer.writerow([ draftFORE[ "data" ][ i ][ 'value' ], draftFORE[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/WA.csv', mode='w') as data:
                    wa_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    wa_writer.writerow([ 'Field', 'DateTime' ])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(windAngle[ "data" ][ k ][ 'dt' ]) == str(
                                windAngle[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(windAngle[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            wa_writer.writerow([ windAngle[ "data" ][ k ][ 'value' ], windAngle[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD["data"][i]['dt'] > windAngle["data"][k]['dt']:
                            # k = k + 1
                            # wa_writer.writerow([windAngle["data"][k]['value'], dateD["data"][i]['dt']])
                            # k = k + 1
                            # else:
                            wa_writer.writerow([ "", continuousDTS[ i ] ])
                            # for i in range(0, len(windAngle[ "data" ])):
                            # wa_writer.writerow([ windAngle[ "data" ][ i ][ 'value' ], windAngle[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/STW.csv', mode='w') as data:
                    stw_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    stw_writer.writerow([ 'Field', 'DateTime' ])

                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(STW[ "data" ][ k ][ 'dt' ]) == str(
                                STW[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(STW[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            stw_writer.writerow([ STW[ "data" ][ k ][ 'value' ], STW[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > STW[ "data" ][ k ][ 'dt' ]:
                            # k=k+1
                            # stw_writer.writerow([ STW[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k=k+1
                            # else:
                            stw_writer.writerow([ "", continuousDTS[ i ] ])

                # for i in range(0, len(STW[ "data" ])):
                # stw_writer.writerow([ STW[ "data" ][ i ][ 'value' ], STW[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/RPM.csv', mode='w') as data:
                    rpm_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    rpm_writer.writerow([ 'Field', 'DateTime' ])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(rpm[ "data" ][ k ][ 'dt' ]) == str(
                                rpm[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(rpm[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            rpm_writer.writerow([ rpm[ "data" ][ k ][ 'value' ], rpm[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > rpm[ "data" ][ k ][ 'dt' ]:
                            # k=k+1
                            # rpm_writer.writerow([ rpm[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k=k+1
                            # else:
                            rpm_writer.writerow([ "", continuousDTS[ i ] ])

                # for i in range(0, len(rpm[ "data" ])):
                # rpm_writer.writerow([ rpm[ "data" ][ i ][ 'value' ], rpm[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/FOC.csv', mode='w') as data:
                    foc_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    foc_writer.writerow([ 'Field', 'DateTime' ])

                    k = 0
                    for i in range(0, len(continuousDTS[ i ])):
                        if k > 0:
                            k = k + 1 if str(foc[ "data" ][ k ][ 'dt' ]) == str(
                                foc[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(foc[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            foc_writer.writerow([ foc[ "data" ][ k ][ 'value' ], foc[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > foc[ "data" ][ k ][ 'dt' ]:
                            # k = k + 1
                            # foc_writer.writerow(
                            # [ foc[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k = k + 1
                            # else:
                            foc_writer.writerow([ "", continuousDTS[ i ] ])

                # for i in range(0, len(foc[ "data" ])):
                # foc_writer.writerow([ foc[ "data" ][ i ][ 'value' ], foc[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0



            try:
                with open('./data/LAROS/SOVG.csv', mode='w') as data:
                    sovg_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    sovg_writer.writerow([ 'Field', 'DateTime' ])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(SpeedOvg[ "data" ][ k ][ 'dt' ]) == str(
                                SpeedOvg[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(SpeedOvg[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            sovg_writer.writerow([ SpeedOvg[ "data" ][ k ][ 'value' ], SpeedOvg[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k = k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > SpeedOvg[ "data" ][ k ][ 'dt' ]:
                            # k = k + 1
                            # sovg_writer.writerow(
                            # [ SpeedOvg[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k = k + 1
                            # else:
                            sovg_writer.writerow([ "", continuousDTS[ i ] ])
                    # sovg_writer.writerow([ SpeedOvg[ "data" ][ i ][ 'value' ], SpeedOvg[ "data" ][ i ][ 'dt' ] ])
            except:
                d = 0

            try:
                with open('./data/LAROS/POW.csv', mode='w') as data:
                    pow_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                    pow_writer.writerow([ 'Field', 'DateTime' ])
                    k = 0
                    for i in range(0, len(continuousDTS)):
                        if k > 0:
                            k = k + 1 if str(power[ "data" ][ k ][ 'dt' ]) == str(
                                power[ "data" ][ k - 1 ][ 'dt' ]) else k
                        if str(power[ "data" ][ k ][ 'dt' ]) == continuousDTS[ i ]:
                            pow_writer.writerow([ power[ "data" ][ k ][ 'value' ], power[ "data" ][ k ][ 'dt' ] ])
                            k = k + 1
                        else:
                            # k=k - 1 if k > 0 else 0
                            # if dateD[ "data" ][ i ][ 'dt' ] > power[ "data" ][ k ][ 'dt' ]:
                            # k = k + 1
                            # pow_writer.writerow(
                            # [ power[ "data" ][ k ][ 'value' ], dateD[ "data" ][ i ][ 'dt' ] ])
                            # k = k + 1
                            # else:
                            pow_writer.writerow([ "", continuousDTS[ i ] ])
            except:
                d = 0


        except Exception as e:
            print(str(e))
        self.ExtractLAROSDataset('',startDate,endDate)

    def ExtractLAROSDataset(self,vessel,startDate,endDate):

        drftFList = [ ]
        drftAList = [ ]
        drftList = [ ]
        STWList = [ ]
        POWList = [ ]
        RPMList = [ ]
        WAList = [ ]
        WSList = [ ]
        SOVGList = [ ]
        FOCList = [ ]
        FOCTList = [ ]
        LatList = []
        LonList = []
        CourseList=[]
        unkList=[]

        continuousDTS = (pd.DataFrame(columns=['NULL'],
                                      index=pd.date_range(startDate, endDate,
                                                          freq='1T'))
                         # .between_time('00:00', '15:11')
                         .index.strftime('%Y-%m-%d %H:%M:%S')
                         .tolist()
                         )

        with open('./MT_DELTA_MARIA.csv', mode='w') as data:
            data_writer = csv.writer(data, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

            data_writer.writerow(
                [ 'Draft', 'DrfAft', 'DrftFore','Course', 'WindSpeed', 'WindAngle', 'SpeedOvg', 'STW', 'Rpm', 'M/E FOC (kg/min)',
                  'Power','Lat','Lon', 'DateTime' ])
            # data_writer.writerow(
            # ['Draft','DateTime' ])

            unk = {}

            with open('./data/LAROS/UNK.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            unk[row[1]] = row[0]
                            unkList.append(row[0])



            drft = {}
            drftCount = 0
            with open('./data/LAROS/Draft.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != [ ]:
                        if row[ 0 ] != 'Field':
                            drft[ row[ 1 ] ] = row[ 0 ]
                            drftList.append(row[ 0 ])
                            # drft[ "dt" ] = row[1]
                            # draft[ "data" ].append(drft)
                            drftCount = drftCount + 1
                        # if drftCount > 1000 : break

            dateD = [ ]
            dateCount = 0
            with open('./data/LAROS/DateTime.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != [ ]:
                        if row[ 0 ] != 'DateTimes':
                            # datt[ "value" ] = row[0]
                            dateD.append(row[ 0 ])
                            dateCount = dateCount + 1
                            # if dateCount>1000: break
                            # dateD[ "data" ].append(datt)

            from itertools import islice
            #endDate = str(dateD["data"][len(dateD["data"]) - 1]["dt"])
            #startDate = str(dateD["data"][0]["dt"])
            lat = {}
            with open('./data/LAROS/LAT.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            lat[row[1]] = row[0]
                            LatList.append(row[0])

            lon = {}
            with open('./data/LAROS/LON.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            lon[row[1]] = row[0]
                            LonList.append(row[0])

            vcourse = {}
            with open('./data/LAROS/Course.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            vcourse[row[1]] = row[0]
                            CourseList.append(row[0])

            drftA = {}
            with open('./data/LAROS/DraftAFT.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            drftA[row[1]] = row[0]
                            drftAList.append(row[0])

            drftF = {}
            with open('./data/LAROS/DraftFORE.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            drftF[row[1]] = row[0]
                            drftFList.append(row[0])



                            # drftF[ "dt" ] = row[1]
                            # draftFORE[ "data" ].append(drftF)

            stw = {}
            with open('./data/LAROS/STW.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            stw[row[1]] = row[0]
                            STWList.append(row[0])
                            # stw[ "dt" ] = row[1]
                            # STW[ "data" ].append(stw)
            wa = {}
            with open('./data/LAROS/WA.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            wa[row[1]] = row[0]
                            WAList.append(row[0])
                            # wa[ "dt" ] = row[1]
                            # windAngle[ "data" ].append(wa)
            ws = {}
            with open('./data/LAROS/WS.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            ws[row[1]] = row[0]
                            WSList.append(row[0])
                            # ws[ "dt" ] =
                            # windSpeed[ "data" ].append(ws)
            so = {}
            with open('./data/LAROS/SOVG.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            so[row[1]] = row[0]
                            SOVGList.append(row[0])
                            # so[ "dt" ] = row[1]
                            # SpeedOvg[ "data" ].append(so)
            rpms = {}
            with open('./data/LAROS/RPM.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            rpms[row[1]] = row[0]
                            RPMList.append(row[0])
                            # rpms[ "dt" ] = row[1]
                            # rpm[ "data" ].append(rpms)
            fo = {}
            with open('./data/LAROS/FOC.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            fo[row[1]] = row[0]
                            FOCList.append(row[0])
                            # fo[ "dt" ] = row[1]
                            # foc[ "data" ].append(fo)
            fot = {}
            with open('./data/LAROS/FOCt.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            fot[row[1]] = row[0]
                            FOCTList.append(row[0])
                            # fot[ "dt" ] = row[1]
                            # foct[ "data" ].append(fot)
            pow = {}
            with open('./data/LAROS/POW.csv') as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')
                for row in csv_reader:
                    if row != []:
                        if row[0] != 'Field':
                            pow[row[1]] = row[0]
                            POWList.append(row[0])
                            # pow[ "dt" ] = row[1]
                            # power[ "data" ].append(pow)

            for i in range(0, len(continuousDTS)):
                dt = continuousDTS[ i ]
                #print(str(dt))
                if dt=='2018-10-09 09:02:00':
                    print(drftList[i]+" "+ drftAList[i]+" "+drftFList[i])
                value = ""
                value7 = ""
                value8 = ""
                value9 = ""
                value10 = ""
                value2 = ""
                value1 = ""
                value6 = ""
                value5 = ""
                value4 = ""
                value3 = ""

                lat=""
                lon=""
                course=""

                try:

                    # if i > 5:
                    # drftSliced = dict(list(drft.items())[i-5:i + 5]).items()
                    # else:
                    # drftSliced = dict(list(drft.items())[i:i + 5]).items()

                    # drftList=list(dict(filter(lambda elem: elem[ 0 ] == dt, dict(islice(drftSliced.items(), i+10)).items())).values())
                    # drftList = list(dict(filter(lambda elem: elem[0] == dt,
                    # drftSliced)).values())
                    # if drftList.__len__()>0: value = drftList[0]
                    try:
                        value = drftList[ i ]
                    except:
                        r = 0

                    try:
                        lat = LatList[i]
                    except:
                        r = 0

                    try:
                        lon = LonList[i]
                    except:
                        r = 0

                    try:
                        course = CourseList[i]
                    except:
                        r = 0

                    # if draft["data"][i]['dt']==dt:
                    # value=draft['data'][ i ][ 'value' ]
                    # value = [k for k in draft[ "data" ] if k['dt']==dt][0]['value']
                    # drftAList = list(dict(filter(lambda elem: elem[ 0 ] == dt, drftA.items())).values())
                    # if drftAList.__len__() > 0:
                    try:
                        value7 = drftAList[ i ]
                    except:
                        r = 0
                    # if draftAFT["data"][i]['dt']==dt:
                    # value7=draftAFT['data'][ i ][ 'value' ]
                    # value7 = [k for k in draftAFT[ "data" ] if k['dt']==dt][0]['value']

                    # drftFList = list(dict(filter(lambda elem: elem[ 0 ] == dt, drftF.items())).values())
                    # if drftFList.__len__() > 0: value7  = drftFList[0]
                    try:
                        value8 = drftFList[ i ]
                    except:
                        r = 0

                    # if draftFORE[ "data" ][ i ][ 'dt' ] == dt:
                    # value8 = draftFORE[ 'data' ][ i ][ 'value' ]
                    # value8 =[k for k in draftFORE[ "data" ] if k['dt']==dt][0]['value']

                    # STWList = list(dict(filter(lambda elem: elem[ 0 ] == dt, stw.items())).values())
                    # if STWList.__len__() > 0:
                    try:
                        value9 = STWList[ i ]
                    except:
                        r = 0

                    # if STW[ "data" ][ i ][ 'dt' ] == dt:
                    # value9 = STW[ 'data' ][ i ][ 'value' ]
                    # value9 =[k for k in STW[ "data" ] if k['dt']==dt][0]['value']
                    # FOCTList = list(dict(filter(lambda elem: elem[ 0 ] == dt, fot.items())).values())
                    # if FOCTList.__len__() > 0:
                    try:
                        value10 = FOCTList[ i ]
                    except:
                        r = 0

                    # if foct[ "data" ][ i ][ 'dt' ] == dt:
                    # value10 = foct[ 'data' ][ i ][ 'value' ]
                    # value10 = [k for k in foct[ "data" ] if k['dt']==dt][0]['value']

                    # WAList = list(dict(filter(lambda elem: elem[ 0 ] == dt, wa.items())).values())
                    # if WAList.__len__() > 0:
                    try:
                        value1 = WAList[ i ]
                    except:
                        r = 0
                    # if windAngle[ "data" ][ i ][ 'dt' ] == dt:
                    # value1 = windAngle[ 'data' ][ i ][ 'value' ]
                    # value1 =[k for k in windAngle[ "data" ] if k['dt']==dt][0]['value']

                    # WSList = list(dict(filter(lambda elem: elem[ 0 ] == dt, ws.items())).values())
                    # if WSList.__len__() > 0:
                    try:
                        value2 = WSList[ i ]
                    except:
                        r = 0
                    # if windSpeed[ "data" ][ i ][ 'dt' ] == dt:
                    # value2 = windSpeed[ 'data' ][ i ][ 'value' ]
                    # value2 = [k for k in windSpeed[ "data" ] if k['dt']==dt][0]['value']
                    # POWList = list(dict(filter(lambda elem: elem[ 0 ] == dt, pow.items())).values())
                    # if POWList.__len__() > 0:
                    try:
                        value3 = POWList[ i ]
                    except:
                        r = 0

                    # if power[ "data" ][ i ][ 'dt' ] == dt:
                    # value3 = power[ 'data' ][ i ][ 'value' ]
                    # value3 = [k for k in power[ "data" ] if k['dt']==dt][0]['value']
                    # POWList = list(dict(filter(lambda elem: elem[ 0 ] == dt, pow.items())).values())
                    # if POWList.__len__() > 0:
                    try:
                        value3 = POWList[ i ]
                    except:
                        r = 0

                    # FOCList = list(dict(filter(lambda elem: elem[ 0 ] == dt, foc.items())).values())
                    # if FOCList.__len__() > 0:
                    try:
                        value4 = FOCList[ i ]
                    except:
                        r = 0
                    # if foc[ "data" ][ i ][ 'dt' ] == dt:
                    # value4 = foc[ 'data' ][ i ][ 'value' ]
                    # value4 =[k for k in foc[ "data" ] if k['dt']==dt][0]['value']

                    # RPMList = list(dict(filter(lambda elem: elem[ 0 ] == dt, rpm.items())).values())
                    # if RPMList.__len__() > 0:
                    try:
                        value5 = RPMList[ i ]
                    except:
                        r = 0

                    # if rpm[ "data" ][ i ][ 'dt' ] == dt:
                    # value5 = rpm[ 'data' ][ i ][ 'value' ]
                    # value5 = [k for k in rpm[ "data" ] if k['dt']==dt][0]['value']

                    # SOVGList = list(dict(filter(lambda elem: elem[ 0 ] == dt, so.items())).values())
                    # if SOVGList.__len__() > 0:
                    try:
                        value6 = SOVGList[ i ]
                    except:
                        r = 0
                    # if SpeedOvg[ "data" ][ i ][ 'dt' ] == dt:
                    # value6 = SpeedOvg[ 'data' ][ i ][ 'value' ]
                    # value6 =[k for k in SpeedOvg[ "data" ] if k['dt']==dt][0]['value']

                #####
                except Exception as e:
                    print(str(e))
                data_writer.writerow(
                    [ value, value7, value8,course, value2 ,value1, value6, value9, value5, value4 , value3,lat, lon ,str(dt) ])


    def readLarosDataFromCsvNewExtractExcels(self, data):
        # Load file
        dataNew = data.drop([ 'SpeedOvg' ], axis=1)

        dtNew = dataNew.values[ 484362:569386, : ]

        daysSincelastPropClean = [ ]
        daysSinceLastDryDock = [ ]

        with open('./daysSincelastPropClean.csv') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if row.__len__() == 1:
                    daysSincelastPropClean.append(float(row[ 0 ]))
        with open('./daysSinceLastDryDock.csv') as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if row.__len__() == 1:
                    daysSinceLastDryDock.append(float(row[ 0 ]))

        daysSincelastPropClean = np.array(daysSincelastPropClean).reshape(-1, ).astype(float)
        daysSinceLastDryDock = np.array(daysSinceLastDryDock).reshape(-1, ).astype(float)

        dataNew = data.drop([ 'SpeedOvg', 'DateTime' ], axis=1)

        dtNew = dataNew.values[ 484362:569386, : ].astype(float)

        draftMean = (dtNew[ :, 1 ] + dtNew[ :, 2 ]) / 2
        dtNew[ :, 0 ] = np.round(draftMean, 3)

        # dtNew = dtNew[~np.isnan(dtNew).any(axis=1)]

        trim = (dtNew[ :, 1 ] - dtNew[ :, 2 ])

        dtNew1 = np.array(np.append(dtNew[ :, 0 ].reshape(-1, 1), np.asmatrix(
            [ np.round(trim, 3), dtNew[ :, 3 ], dtNew[ :, 4 ], dtNew[ :, 5 ], dtNew[ :, 6 ], daysSincelastPropClean,
              daysSinceLastDryDock, dtNew[ :, 7 ] ]).T, axis=1))

        # dtNew1 = np.delete(dtNew1, [i for (i, v) in enumerate(dtNew1[0:, 0]) if v <= 6], 0)
        dtNew1 = np.delete(dtNew1, [ i for (i, v) in enumerate(dtNew1[ 0:, 4 ]) if v <= 0 ], 0)

        rpmB = np.array([ i for i in dtNew1 if i[ 5 ] > 90 and i[ 7 ] < 100 ])

        # dtNew1 = np.delete(dtNew1, [i for (i, v) in enumerate(dtNew1[0:, 5]) if v <= 0 or v > 90], 0)  # or (v>0 and v <=10)

        # dtNew = np.delete(dtNew, [ i for (i, v) in enumerate(dtNew[ 0:, 7 ]) ], 0) # if v <= 7

        ##statistics STD
        # draftSmaller6 = np.array([drft for drft in dtNew1 if drft[0]< 6])
        # draftBigger6 = np.array([ drft for drft in dtNew1 if drft[0] > 6 ])
        #########DRAFT < 9
        drftStw8 = [ ]

        # stw8DrftS4 = np.array([i for i in dtNew1 if i[4] >= 8 and i[4] < 8.5 and i[0] <= 9])

        with open('./MT_DELTA_MARIA_draftS_9_analysis.csv', mode='w') as rows:
            for dr in range(6, 9):
                stw8DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 8 and i[ 4 ] < 8.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw8DrftS4 = np.delete(stw8DrftS4, [ i for (i, v) in enumerate(stw8DrftS4[ :, 8 ]) if v < (
                        np.mean(stw8DrftS4[ :, 8 ]) - np.std(stw8DrftS4[ :, 8 ])) or v > np.mean(
                    stw8DrftS4[ :, 8 ]) + np.std(
                    stw8DrftS4[ :, 8 ]) ], 0)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 8 < STW < 8.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw8DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            # drftStw8.append(stw8DrftS4)

            drftStw9 = [ ]
            stw9DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] <= 9 ])

            for dr in range(6, 9):
                stw9DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw9DrftS4 = np.delete(stw9DrftS4, [ i for (i, v) in enumerate(stw9DrftS4[ :, 8 ]) if v < (
                        np.mean(stw9DrftS4[ :, 8 ]) - np.std(stw9DrftS4[ :, 8 ])) or v > np.mean(
                    stw9DrftS4[ :, 8 ]) + np.std(
                    stw9DrftS4[ :, 8 ]) ], 0)
                drftStw9.append(stw9DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 8.5 < STW < 9                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw9DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw10 = [ ]
            stw10DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] <= 9 ])

            for dr in range(6, 9):
                stw10DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw10DrftS4 = np.delete(stw10DrftS4, [ i for (i, v) in enumerate(stw10DrftS4[ :, 8 ]) if v < (
                        np.mean(stw10DrftS4[ :, 8 ]) - np.std(stw10DrftS4[ :, 8 ])) or v > np.mean(
                    stw10DrftS4[ :, 8 ]) + np.std(
                    stw10DrftS4[ :, 8 ]) ], 0)
                drftStw10.append(stw10DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 9 < STW < 9.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw10DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw11 = [ ]
            stw11DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] <= 9 ])

            for dr in range(6, 9):
                stw11DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw11DrftS4 = np.delete(stw10DrftS4, [ i for (i, v) in enumerate(stw11DrftS4[ :, 8 ]) if v < (
                        np.mean(stw11DrftS4[ :, 8 ]) - np.std(stw11DrftS4[ :, 8 ])) or v > np.mean(
                    stw11DrftS4[ :, 8 ]) + np.std(
                    stw11DrftS4[ :, 8 ]) ], 0)
                drftStw11.append(stw11DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 9.5 < STW < 10                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw11DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw12 = [ ]
            stw12DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] <= 9 ])

            for dr in range(6, 9):
                stw12DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw12DrftS4 = np.delete(stw12DrftS4, [ i for (i, v) in enumerate(stw12DrftS4[ :, 8 ]) if v < (
                        np.mean(stw12DrftS4[ :, 8 ]) - np.std(stw12DrftS4[ :, 8 ])) or v > np.mean(
                    stw12DrftS4[ :, 8 ]) + np.std(
                    stw12DrftS4[ :, 8 ]) ], 0)
                drftStw12.append(stw11DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 10 < STW < 10.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw12DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw13 = [ ]
            # stw13DrftS4 = np.array([i for i in dtNew1 if i[4] >= 10.5 and i[4] < 11 and i[0] <= 9])

            for dr in range(6, 9):
                stw13DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw13DrftS4 = np.delete(stw13DrftS4, [ i for (i, v) in enumerate(stw13DrftS4[ :, 8 ]) if v < (
                        np.mean(stw13DrftS4[ :, 8 ]) - np.std(stw13DrftS4[ :, 8 ])) or v > np.mean(
                    stw13DrftS4[ :, 8 ]) + np.std(
                    stw13DrftS4[ :, 8 ]) ], 0)
                drftStw13.append(drftStw13)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 10.5 < STW < 11                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw13DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw14 = [ ]
            # stw14DrftS4 = np.array([i for i in dtNew1 if i[4] >= 11 and i[4] < 11.5 and i[0] <= 9])

            for dr in range(6, 9):
                stw14DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw14DrftS4 = np.delete(stw14DrftS4, [ i for (i, v) in enumerate(stw14DrftS4[ :, 8 ]) if v < (
                        np.mean(stw14DrftS4[ :, 8 ]) - np.std(stw14DrftS4[ :, 8 ])) or v > np.mean(
                    stw14DrftS4[ :, 8 ]) + np.std(
                    stw14DrftS4[ :, 8 ]) ], 0)
                drftStw14.append(stw14DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 11 < STW < 11.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw14DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw15 = [ ]
            # stw15DrftS4 = np.array([i for i in dtNew1 if i[4] >= 11.5 and i[4] < 12 and i[0] <= 9])

            for dr in range(6, 9):
                stw15DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw15DrftS4 = np.delete(stw15DrftS4, [ i for (i, v) in enumerate(stw15DrftS4[ :, 8 ]) if v < (
                        np.mean(stw15DrftS4[ :, 8 ]) - np.std(stw15DrftS4[ :, 8 ])) or v > np.mean(
                    stw15DrftS4[ :, 8 ]) + np.std(
                    stw15DrftS4[ :, 8 ]) ], 0)
                drftStw15.append(stw15DrftS4)
                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 11.5 < STW < 12                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw15DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw16 = [ ]
            # stw16DrftS4 = np.array([i for i in dtNew1 if i[4] >= 12 and i[4] < 12.5 and i[0] <= 9])

            for dr in range(6, 9):
                stw16DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw16DrftS4 = np.delete(stw16DrftS4, [ i for (i, v) in enumerate(stw16DrftS4[ :, 8 ]) if v < (
                        np.mean(stw16DrftS4[ :, 8 ]) - np.std(stw16DrftS4[ :, 8 ])) or v > np.mean(
                    stw16DrftS4[ :, 8 ]) + np.std(
                    stw16DrftS4[ :, 8 ]) ], 0)
                drftStw16.append(stw16DrftS4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 12 < STW < 12.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw16DrftS4:
                    data_writer.writerow([ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw17 = [ ]
            # stw17DrftS4 = np.array([i for i in dtNew1 if i[4] >= 12.5 and i[4] < 13 and i[0] <= 9])

            for dr in range(6, 9):
                stw17DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw17DrftS4 = np.delete(stw17DrftS4, [ i for (i, v) in enumerate(stw17DrftS4[ :, 8 ]) if v < (
                        np.mean(stw17DrftS4[ :, 8 ]) - np.std(stw17DrftS4[ :, 8 ])) or v > np.mean(
                    stw17DrftS4[ :, 8 ]) + np.std(
                    stw17DrftS4[ :, 8 ]) ], 0)
                drftStw17.append(drftStw17)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 12 < STW < 12.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw17DrftS4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            drftStw18 = [ ]
            # stw18DrftS4 = np.array([i for i in dtNew1 if i[4] >= 13 and i[4] < 13.5 and i[0] <= 9])

            for dr in range(6, 9):
                stw18DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(

                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 13 < STW < 13.5                   ' ], )
                if stw18DrftS4.shape[ 0 ] != 0:
                    stw18DrftS4 = np.delete(stw18DrftS4, [ i for (i, v) in enumerate(stw18DrftS4[ :, 8 ]) if v < (
                            np.mean(stw18DrftS4[ :, 8 ]) - np.std(stw18DrftS4[ :, 8 ])) or v > np.mean(
                        stw18DrftS4[ :, 8 ]) + np.std(
                        stw18DrftS4[ :, 8 ]) ], 0)
                    drftStw18.append(stw18DrftS4)

                    data_writer.writerow([ ], )
                    data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW',
                                           'M/E FOC (kg/min)', 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock' ])
                    for data in stw18DrftS4:
                        data_writer.writerow(
                            [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                              data[ 8 ] ])
                    data_writer.writerow([ ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ ], )

            drftStw19 = [ ]
            # stw19DrftS4 = np.array([i for i in dtNew1 if i[4] >= 13.5 and i[4] < 14 and i[0] <= 9])

            for dr in range(6, 9):
                stw19DrftS4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 13.5 < STW < 14                   ' ], )

                if stw19DrftS4.shape[ 0 ] != 0:
                    stw19DrftS4 = np.delete(stw19DrftS4, [ i for (i, v) in enumerate(stw19DrftS4[ :, 8 ]) if v < (
                            np.mean(stw19DrftS4[ :, 8 ]) - np.std(stw19DrftS4[ :, 8 ])) or v > np.mean(
                        stw19DrftS4[ :, 8 ]) + np.std(
                        stw19DrftS4[ :, 8 ]) ], 0)
                    drftStw19.append(stw19DrftS4)

                    data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                    data_writer.writerow([ ], )
                    data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                             , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                    for data in stw19DrftS4:
                        data_writer.writerow(
                            [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                              data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

        ########################################################################################################################################
        ########################################################################################################################################
        ########################################################################################################################################
        ########################################################################################################################################

        drftStw8 = [ ]
        # stw8DrftB4 = np.array([i for i in dtNew1 if i[4] >= 8 and i[4] < 8.5 and i[0] > 9])
        with open('./MT_DELTA_MARIA_draftB_9_analysis.csv', mode='w') as rows:
            for dr in range(9, 14):
                stw8DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 8 and i[ 4 ] < 8.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw8DrftB4 = np.delete(stw8DrftB4, [ i for (i, v) in enumerate(stw8DrftB4[ :, 8 ]) if v < (
                        np.mean(stw8DrftB4[ :, 8 ]) - np.std(stw8DrftB4[ :, 8 ])) or v > np.mean(
                    stw8DrftB4[ :, 8 ]) + np.std(
                    stw8DrftB4[ :, 8 ]) ], 0)
                drftStw8.append(stw8DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 8 < STW < 8.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw8DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw8DrftB4 = np.concatenate(stw8DrftB4)

            drftStw9 = [ ]
            # stw9DrftB4 = np.array([i for i in dtNew1 if i[4] >= 8.5 and i[4] < 9 and i[0] > 9])
            for dr in range(9, 14):
                stw9DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw9DrftB4 = np.delete(stw9DrftB4, [ i for (i, v) in enumerate(stw9DrftB4[ :, 8 ]) if v < (
                        np.mean(stw9DrftB4[ :, 8 ]) - np.std(stw9DrftB4[ :, 8 ])) or v > np.mean(
                    stw9DrftB4[ :, 8 ]) + np.std(
                    stw9DrftB4[ :, 8 ]) ], 0)
                drftStw9.append(stw9DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 8.5 < STW < 9                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw9DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw9DrftB4 = np.concatenate(drftStw9)

            # drftStw10 = []
            # stw10DrftB4 = np.array([i for i in dtNew1 if i[4] >= 9 and i[4] < 9.5 and i[0] > 9])
            for dr in range(9, 14):
                stw10DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw10DrftB4 = np.delete(stw10DrftB4, [ i for (i, v) in enumerate(stw10DrftB4[ :, 8 ]) if v < (
                        np.mean(stw10DrftB4[ :, 8 ]) - np.std(stw10DrftB4[ :, 8 ])) or v > np.mean(
                    stw10DrftB4[ :, 8 ]) + np.std(
                    stw10DrftB4[ :, 8 ]) ], 0)
                drftStw10.append(stw10DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 9 < STW < 9.5                  ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw10DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            # stw10DrftB4 = np.concatenate(drftStw10)

            # drftStw11 = []
            # stw11DrftB4 = np.array([i for i in dtNew1 if i[4] >= 9.5 and i[4] < 10 and i[0] > 9])
            for dr in range(9, 14):
                stw11DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw11DrftB4 = np.delete(stw11DrftB4, [ i for (i, v) in enumerate(stw11DrftB4[ :, 8 ]) if v < (
                        np.mean(stw11DrftB4[ :, 8 ]) - np.std(stw11DrftB4[ :, 8 ])) or v > np.mean(
                    stw11DrftB4[ :, 8 ]) + np.std(
                    stw11DrftB4[ :, 8 ]) ], 0)
                drftStw11.append(stw11DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 9.5 < STW < 10                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw11DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw11DrftB4 = np.concatenate(drftStw11)

            # drftStw12 = []
            # stw12DrftB4 = np.array([i for i in dtNew1 if i[4] >= 10 and i[4] < 10.5 and i[0] > 9])
            for dr in range(9, 14):
                stw12DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw12DrftB4 = np.delete(stw12DrftB4, [ i for (i, v) in enumerate(stw12DrftB4[ :, 8 ]) if v < (
                        np.mean(stw12DrftB4[ :, 8 ]) - np.std(stw12DrftB4[ :, 8 ])) or v > np.mean(
                    stw12DrftB4[ :, 8 ]) + np.std(
                    stw12DrftB4[ :, 8 ]) ], 0)
                drftStw12.append(drftStw12)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 10 < STW < 10.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw12DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            # stw12DrftB4 = np.concatenate(drftStw12)

            # drftStw13 = []
            # stw13DrftB4 = np.array([i for i in dtNew1 if i[4] >= 10.5 and i[4] < 11 and i[0] > 9])
            for dr in range(9, 14):
                stw13DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw13DrftB4 = np.delete(stw13DrftB4, [ i for (i, v) in enumerate(stw13DrftB4[ :, 8 ]) if v < (
                        np.mean(stw13DrftB4[ :, 8 ]) - np.std(stw13DrftB4[ :, 8 ])) or v > np.mean(
                    stw13DrftB4[ :, 8 ]) + np.std(
                    stw13DrftB4[ :, 8 ]) ], 0)
                drftStw13.append(stw13DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 10.5 < STW < 11                 ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw13DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            # stw13DrftB4 = np.concatenate(drftStw13)

            # drftStw14 = []
            # stw14DrftB4 = np.array([i for i in dtNew1 if i[4] >= 11 and i[4] < 11.5 and i[0] > 9])
            for dr in range(9, 14):
                stw14DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw14DrftB4 = np.delete(stw14DrftB4, [ i for (i, v) in enumerate(stw14DrftB4[ :, 8 ]) if v < (
                        np.mean(stw14DrftB4[ :, 8 ]) - np.std(stw14DrftB4[ :, 8 ])) or v > np.mean(
                    stw14DrftB4[ :, 8 ]) + np.std(
                    stw14DrftB4[ :, 8 ]) ], 0)
                drftStw14.append(stw14DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 11 < STW < 11.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw14DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            # stw14DrftB4 = np.concatenate(drftStw14)

            # drftStw15 = []
            # stw15DrftB4 = np.array([i for i in dtNew1 if i[4] >= 11.5 and i[4] < 12 and i[0] > 9])
            for dr in range(9, 14):
                stw15DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw15DrftB4 = np.delete(stw15DrftB4, [ i for (i, v) in enumerate(stw15DrftB4[ :, 8 ]) if v < (
                        np.mean(stw15DrftB4[ :, 8 ]) - np.std(stw15DrftB4[ :, 8 ])) or v > np.mean(
                    stw15DrftB4[ :, 8 ]) + np.std(
                    stw15DrftB4[ :, 8 ]) ], 0)
                drftStw15.append(stw15DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 11.5 < STW < 12                  ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw15DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw15DrftB4 = np.concatenate(drftStw15)

            # drftStw16 = []
            # stw16DrftB4 = np.array([i for i in dtNew1 if i[4] >= 12 and i[4] < 12.5 and i[0] > 9])
            for dr in range(9, 14):
                stw16DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                stw16DrftB4 = np.delete(stw16DrftB4, [ i for (i, v) in enumerate(stw16DrftB4[ :, 8 ]) if v < (
                        np.mean(stw16DrftB4[ :, 8 ]) - np.std(stw16DrftB4[ :, 8 ])) or v > np.mean(
                    stw16DrftB4[ :, 8 ]) + np.std(
                    stw16DrftB4[ :, 8 ]) ], 0)
                drftStw16.append(stw16DrftB4)

                data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 12 < STW < 12.5                   ' ], )
                data_writer.writerow([ ], )
                data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                         , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                for data in stw16DrftB4:
                    data_writer.writerow(
                        [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                          data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw16DrftB4 = np.concatenate(drftStw16)

            # drftStw17 = []
            # stw17DrftB4 = np.array([i for i in dtNew1 if i[4] >= 12.5 and i[4] < 13 and i[0] > 9])
            for dr in range(9, 14):
                stw17DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 12.5 and i[ 4 ] < 13 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 12.5 < STW < 13                 ' ], )

                if stw17DrftB4.shape[ 0 ] != 0:
                    stw17DrftB4 = np.delete(stw17DrftB4, [ i for (i, v) in enumerate(stw17DrftB4[ :, 8 ]) if v < (
                            np.mean(stw17DrftB4[ :, 8 ]) - np.std(stw17DrftB4[ :, 8 ])) or v > np.mean(
                        stw17DrftB4[ :, 8 ]) + np.std(
                        stw17DrftB4[ :, 8 ]) ], 0)
                    drftStw17.append(stw17DrftB4)

                    data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                    data_writer.writerow([ ], )
                    data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW',
                                           'M/E FOC (kg/min)', 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock' ])
                    for data in stw17DrftB4:
                        data_writer.writerow(
                            [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                              data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw17DrftB4 = np.concatenate(drftStw17)

            # drftStw18 = []
            # stw18DrftB4 = np.array([i for i in dtNew1 if i[4] >= 13 and i[4] < 13.5 and i[0] > 9])
            for dr in range(9, 14):
                stw18DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 13 < STW < 13.5                 ' ], )

                if stw18DrftB4.shape[ 0 ] != 0:
                    stw18DrftB4 = np.delete(stw18DrftB4, [ i for (i, v) in enumerate(stw18DrftB4[ :, 8 ]) if v < (
                            np.mean(stw18DrftB4[ :, 8 ]) - np.std(stw18DrftB4[ :, 8 ])) or v > np.mean(
                        stw18DrftB4[ :, 8 ]) + np.std(
                        stw18DrftB4[ :, 8 ]) ], 0)
                    drftStw18.append(stw18DrftB4)

                    data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                    data_writer.writerow([ ], )
                    data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW',
                                           'M/E FOC (kg/min)', 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock' ])
                    for data in stw18DrftB4:
                        data_writer.writerow(
                            [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                              data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

            # stw18DrftB4 = np.concatenate(drftStw18)

            # drftStw19 = []
            # stw19DrftB4 = np.array([i for i in dtNew1 if i[4] >= 13.5 and i[4] < 14 and i[0] > 9])
            for dr in range(9, 14):
                stw19DrftB4 = np.array(
                    [ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
                data_writer.writerow(
                    [ str(dr) + ' < DRAFT < ' + str(dr + 1) + '   ' + ' 13.5 < STW < 14                 ' ], )
                if stw19DrftB4.shape[ 0 ] != 0:
                    stw19DrftB4 = np.delete(stw19DrftB4, [ i for (i, v) in enumerate(stw19DrftB4[ :, 8 ]) if v < (
                            np.mean(stw19DrftB4[ :, 8 ]) - np.std(stw19DrftB4[ :, 8 ])) or v > np.mean(
                        stw19DrftB4[ :, 8 ]) + np.std(
                        stw19DrftB4[ :, 8 ]) ], 0)
                    drftStw19.append(stw19DrftB4)

                    data_writer = csv.writer(rows, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

                    data_writer.writerow([ ], )
                    data_writer.writerow([ 'Draft', 'Trim', 'WindSpeed', 'WindAngle', 'STW', 'RPM'
                                             , 'DaysSinceLastPropCleaning', 'DaysSinceLastDryDock', 'M/E FOC (kg/min)' ])
                    for data in stw19DrftB4:
                        data_writer.writerow(
                            [ data[ 0 ], data[ 1 ], data[ 2 ], data[ 3 ], data[ 4 ], data[ 5 ], data[ 6 ], data[ 7 ],
                              data[ 8 ] ])
                data_writer.writerow([ ], )
            data_writer.writerow([ ], )
            data_writer.writerow([ ], )

        # tw19DrftB4 = np.concatenate(drftStw19)

        x = 0


    def readLarosDataFromCsvNew(self, data):
        # Load file
        from sklearn import preprocessing
        from sklearn.decomposition import PCA
        # if self.__class__.__name__ == 'UnseenSeriesReader':
        # dt = data.sample(n=2880).values[ :90000:, 3:23 ]
        # dt = data.values[ 0:, 2:23 ]
        # else:
        # dt = data.values[0:,2:23]
        # dataNew = data.drop(['M/E FOC (kg/min)', 'DateTime'], axis=1)
        # names = dataNew.columns

        dataNew = data.drop([ 'DateTime', 'SpeedOvg' ], axis=1)
        # 'M/E FOC (kg/min)'

        # dataNew=dataNew[['Draft', 'SpeedOvg']]

        dtNew = dataNew.values[ 484362:569386, : ].astype(float)

        draftMean = (dtNew[ :, 1 ] + dtNew[ :, 2 ]) / 2
        dtNew[ :, 0 ] = np.round(draftMean, 3)

        dtNew = np.delete(dtNew, [ i for (i, v) in enumerate(dtNew[ 0:, 0 ]) if v <= 6 ], 0)
        dtNew = np.delete(dtNew, [ i for (i, v) in enumerate(dtNew[ 0:, 5 ]) if v <= 8 or v > 14 ], 0)
        dtNew = np.delete(dtNew, [ i for (i, v) in enumerate(dtNew[ 0:, 6 ]) if v <= 0 or v > 90 ],
                          0)  # or (v>0 and v <=10)
        # dtNew = np.delete(dtNew, [ i for (i, v) in enumerate(dtNew[ 0:, 7 ]) ], 0) # if v <= 7

        dtNew = dtNew[ ~np.isnan(dtNew).any(axis=1) ]

        trim = (dtNew[ :, 1 ] - dtNew[ :, 2 ])

        dtNew1 = np.array(np.append(dtNew[ :, 0 ].reshape(-1, 1), np.asmatrix(
            [ np.round(trim, 3), dtNew[ :, 3 ], dtNew[ :, 4 ], dtNew[ :, 5 ], dtNew[ :, 7 ] ]).T, axis=1))

        ##statistics STD
        # draftSmaller6 = np.array([drft for drft in dtNew1 if drft[0]< 6])
        # draftBigger6 = np.array([ drft for drft in dtNew1 if drft[0] > 6 ])
        #########DRAFT < 9
        drftStw8 = [ ]
        # stw8DrftS4 = np.array([i for i in dtNew1 if i[4]>=8 and i[4]<8.5 and i[0]<=9])

        for dr in range(6, 9):
            stw8DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 8 and i[ 4 ] < 8.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw8DrftS4.shape[ 0 ] != 0:
                stw8DrftS4 = np.delete(stw8DrftS4, [ i for (i, v) in enumerate(stw8DrftS4[ :, 5 ]) if v < (
                        np.mean(stw8DrftS4[ :, 5 ]) - np.std(stw8DrftS4[ :, 5 ])) or v > np.mean(
                    stw8DrftS4[ :, 5 ]) + np.std(stw8DrftS4[ :, 5 ]) ], 0)
                drftStw8.append(stw8DrftS4)
        stw8DrftS4 = np.concatenate(drftStw8)

        drftStw9 = [ ]
        stw9DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw9DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw9DrftS4 = np.delete(stw9DrftS4, [ i for (i, v) in enumerate(stw9DrftS4[ :, 5 ]) if v < (
                    np.mean(stw9DrftS4[ :, 5 ]) - np.std(stw9DrftS4[ :, 5 ])) or v > np.mean(stw9DrftS4[ :, 5 ]) + np.std(
                stw9DrftS4[ :, 5 ]) ], 0)
            drftStw9.append(stw9DrftS4)

        stw9DrftS4 = np.concatenate(drftStw9)

        drftStw10 = [ ]
        # stw10DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw10DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw10DrftS4 = np.delete(stw10DrftS4, [ i for (i, v) in enumerate(stw10DrftS4[ :, 5 ]) if v < (
                    np.mean(stw10DrftS4[ :, 5 ]) - np.std(stw10DrftS4[ :, 5 ])) or v > np.mean(
                stw10DrftS4[ :, 5 ]) + np.std(stw10DrftS4[ :, 5 ]) ], 0)
            drftStw10.append(stw10DrftS4)

        stw10DrftS4 = np.concatenate(drftStw10)

        drftStw11 = [ ]
        # stw11DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw11DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw11DrftS4 = np.delete(stw10DrftS4, [ i for (i, v) in enumerate(stw11DrftS4[ :, 5 ]) if v < (
                    np.mean(stw11DrftS4[ :, 5 ]) - np.std(stw11DrftS4[ :, 5 ])) or v > np.mean(
                stw11DrftS4[ :, 5 ]) + np.std(stw11DrftS4[ :, 5 ]) ], 0)
            drftStw11.append(stw11DrftS4)

        stw11DrftS4 = np.concatenate(drftStw11)

        drftStw12 = [ ]
        # stw12DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw12DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw12DrftS4 = np.delete(stw12DrftS4, [ i for (i, v) in enumerate(stw12DrftS4[ :, 5 ]) if v < (
                    np.mean(stw12DrftS4[ :, 5 ]) - np.std(stw12DrftS4[ :, 5 ])) or v > np.mean(
                stw12DrftS4[ :, 5 ]) + np.std(stw12DrftS4[ :, 5 ]) ], 0)
            drftStw12.append(stw12DrftS4)

        stw12DrftS4 = np.concatenate(drftStw12)

        drftStw13 = [ ]
        # stw13DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw13DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw13DrftS4 = np.delete(stw13DrftS4, [ i for (i, v) in enumerate(stw13DrftS4[ :, 5 ]) if v < (
                    np.mean(stw13DrftS4[ :, 5 ]) - np.std(stw13DrftS4[ :, 5 ])) or v > np.mean(
                stw13DrftS4[ :, 5 ]) + np.std(stw13DrftS4[ :, 5 ]) ], 0)
            drftStw13.append(stw13DrftS4)

        stw13DrftS4 = np.concatenate(drftStw13)

        drftStw14 = [ ]
        # stw14DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw14DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw14DrftS4 = np.delete(stw14DrftS4, [ i for (i, v) in enumerate(stw14DrftS4[ :, 5 ]) if v < (
                    np.mean(stw14DrftS4[ :, 5 ]) - np.std(stw14DrftS4[ :, 5 ])) or v > np.mean(
                stw14DrftS4[ :, 5 ]) + np.std(stw14DrftS4[ :, 5 ]) ], 0)
            drftStw14.append(stw14DrftS4)

        stw14DrftS4 = np.concatenate(drftStw14)

        drftStw15 = [ ]
        stw15DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] <= 9 ])

        for dr in range(6, 9):
            stw15DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw15DrftS4 = np.delete(stw15DrftS4, [ i for (i, v) in enumerate(stw15DrftS4[ :, 5 ]) if v < (
                    np.mean(stw15DrftS4[ :, 5 ]) - np.std(stw15DrftS4[ :, 5 ])) or v > np.mean(
                stw15DrftS4[ :, 5 ]) + np.std(stw15DrftS4[ :, 5 ]) ], 0)
            drftStw15.append(stw15DrftS4)

        stw15DrftS4 = np.concatenate(drftStw15)

        drftStw16 = [ ]
        # stw16DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] <= 9 ])
        for dr in range(6, 9):
            stw16DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw16DrftS4 = np.delete(stw16DrftS4, [ i for (i, v) in enumerate(stw16DrftS4[ :, 5 ]) if v < (
                    np.mean(stw16DrftS4[ :, 5 ]) - np.std(stw16DrftS4[ :, 5 ])) or v > np.mean(
                stw16DrftS4[ :, 5 ]) + np.std(stw16DrftS4[ :, 5 ]) ], 0)
            drftStw16.append(stw16DrftS4)

        stw16DrftS4 = np.concatenate(drftStw16)

        drftStw17 = [ ]
        # stw17DrftS4 = np.array([i for i in dtNew1 if i[4] >= 12.5 and i[4] < 13 and i[0] <= 9])

        for dr in range(6, 9):
            stw17DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw17DrftS4.shape[ 0 ] != 0:
                stw17DrftS4 = np.delete(stw17DrftS4, [ i for (i, v) in enumerate(stw17DrftS4[ :, 5 ]) if v < (
                        np.mean(stw17DrftS4[ :, 5 ]) - np.std(stw17DrftS4[ :, 5 ])) or v > np.mean(
                    stw17DrftS4[ :, 5 ]) + np.std(stw17DrftS4[ :, 5 ]) ], 0)
                drftStw17.append(stw17DrftS4)

        stw17DrftS4 = np.concatenate(drftStw17)

        drftStw18 = [ ]
        # stw18DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] <= 9 ])
        for dr in range(6, 9):
            stw18DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw18DrftS4.shape[ 0 ] != 0:
                stw18DrftS4 = np.delete(stw18DrftS4, [ i for (i, v) in enumerate(stw18DrftS4[ :, 5 ]) if v < (
                        np.mean(stw18DrftS4[ :, 5 ]) - np.std(stw18DrftS4[ :, 5 ])) or v > np.mean(
                    stw18DrftS4[ :, 5 ]) + np.std(stw18DrftS4[ :, 5 ]) ], 0)
                drftStw18.append(stw18DrftS4)

        stw18DrftS4 = np.concatenate(drftStw18)

        drftStw19 = [ ]
        # stw19DrftS4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] <= 9 ])
        for dr in range(6, 9):
            stw19DrftS4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw19DrftS4.shape[ 0 ] != 0:
                stw19DrftS4 = np.delete(stw19DrftS4, [ i for (i, v) in enumerate(stw19DrftS4[ :, 5 ]) if v < (
                        np.mean(stw19DrftS4[ :, 5 ]) - np.std(stw19DrftS4[ :, 5 ])) or v > np.mean(
                    stw19DrftS4[ :, 5 ]) + np.std(stw19DrftS4[ :, 5 ]) ], 0)
                drftStw19.append(stw19DrftS4)

        stw19DrftS4 = np.concatenate(drftStw19)

        ##################################################################################
        ##################################################################################
        ##################################################################################
        #########DRAFT > 9
        drftStw8 = [ ]
        stw8DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 8 and i[ 4 ] < 8.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw8DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 8 and i[ 4 ] < 8.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw8DrftB4 = np.delete(stw8DrftB4, [ i for (i, v) in enumerate(stw8DrftB4[ :, 5 ]) if v < (
                    np.mean(stw8DrftB4[ :, 5 ]) - np.std(stw8DrftB4[ :, 5 ])) or v > np.mean(stw8DrftB4[ :, 5 ]) + np.std(
                stw8DrftB4[ :, 5 ]) ], 0)
            drftStw8.append(stw8DrftB4)
        stw8DrftB4 = np.concatenate(drftStw8)

        drftStw9 = [ ]
        stw9DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw9DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 8.5 and i[ 4 ] < 9 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw9DrftB4 = np.delete(stw9DrftB4, [ i for (i, v) in enumerate(stw9DrftB4[ :, 5 ]) if v < (
                    np.mean(stw9DrftB4[ :, 5 ]) - np.std(stw9DrftB4[ :, 5 ])) or v > np.mean(stw9DrftB4[ :, 5 ]) + np.std(
                stw9DrftB4[ :, 5 ]) ], 0)
            drftStw9.append(stw9DrftB4)

        stw9DrftB4 = np.concatenate(drftStw9)

        drftStw10 = [ ]
        stw10DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw10DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 9 and i[ 4 ] < 9.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw10DrftB4 = np.delete(stw10DrftB4, [ i for (i, v) in enumerate(stw10DrftB4[ :, 5 ]) if v < (
                    np.mean(stw10DrftB4[ :, 5 ]) - np.std(stw10DrftB4[ :, 5 ])) or v > np.mean(
                stw10DrftB4[ :, 5 ]) + np.std(stw10DrftB4[ :, 5 ]) ], 0)
            drftStw10.append(stw10DrftB4)
        stw10DrftB4 = np.concatenate(drftStw10)

        drftStw11 = [ ]
        stw11DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw11DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 9.5 and i[ 4 ] < 10 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw11DrftB4 = np.delete(stw11DrftB4, [ i for (i, v) in enumerate(stw11DrftB4[ :, 5 ]) if v < (
                    np.mean(stw11DrftB4[ :, 5 ]) - np.std(stw11DrftB4[ :, 5 ])) or v > np.mean(
                stw11DrftB4[ :, 5 ]) + np.std(stw11DrftB4[ :, 5 ]) ], 0)
            drftStw11.append(stw11DrftB4)
        stw11DrftB4 = np.concatenate(drftStw11)

        drftStw12 = [ ]
        stw12DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw12DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 10 and i[ 4 ] < 10.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw12DrftB4 = np.delete(stw12DrftB4, [ i for (i, v) in enumerate(stw12DrftB4[ :, 5 ]) if v < (
                    np.mean(stw12DrftB4[ :, 5 ]) - np.std(stw12DrftB4[ :, 5 ])) or v > np.mean(
                stw12DrftB4[ :, 5 ]) + np.std(stw12DrftB4[ :, 5 ]) ], 0)
            drftStw12.append(stw12DrftB4)

        stw12DrftB4 = np.concatenate(drftStw12)

        drftStw13 = [ ]
        stw13DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw13DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 10.5 and i[ 4 ] < 11 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw13DrftB4 = np.delete(stw13DrftB4, [ i for (i, v) in enumerate(stw13DrftB4[ :, 5 ]) if v < (
                    np.mean(stw13DrftB4[ :, 5 ]) - np.std(stw13DrftB4[ :, 5 ])) or v > np.mean(
                stw13DrftB4[ :, 5 ]) + np.std(stw13DrftB4[ :, 5 ]) ], 0)
            drftStw13.append(stw13DrftB4)
        stw13DrftB4 = np.concatenate(drftStw13)

        drftStw14 = [ ]
        stw14DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw14DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 11 and i[ 4 ] < 11.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw14DrftB4 = np.delete(stw14DrftB4, [ i for (i, v) in enumerate(stw14DrftB4[ :, 5 ]) if v < (
                    np.mean(stw14DrftB4[ :, 5 ]) - np.std(stw14DrftB4[ :, 5 ])) or v > np.mean(
                stw14DrftB4[ :, 5 ]) + np.std(stw14DrftB4[ :, 5 ]) ], 0)
            drftStw14.append(stw14DrftB4)
        stw14DrftB4 = np.concatenate(drftStw14)

        drftStw15 = [ ]
        stw15DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw15DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 11.5 and i[ 4 ] < 12 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw15DrftB4 = np.delete(stw15DrftB4, [ i for (i, v) in enumerate(stw15DrftB4[ :, 5 ]) if v < (
                    np.mean(stw15DrftB4[ :, 5 ]) - np.std(stw15DrftB4[ :, 5 ])) or v > np.mean(
                stw15DrftB4[ :, 5 ]) + np.std(stw15DrftB4[ :, 5 ]) ], 0)
            drftStw15.append(stw15DrftB4)
        stw15DrftB4 = np.concatenate(drftStw15)

        drftStw16 = [ ]
        stw16DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw16DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            stw16DrftB4 = np.delete(stw16DrftB4, [ i for (i, v) in enumerate(stw16DrftB4[ :, 5 ]) if v < (
                    np.mean(stw16DrftB4[ :, 5 ]) - np.std(stw16DrftB4[ :, 5 ])) or v > np.mean(
                stw16DrftB4[ :, 5 ]) + np.std(stw16DrftB4[ :, 5 ]) ], 0)
            drftStw16.append(stw16DrftB4)
        stw16DrftB4 = np.concatenate(drftStw16)

        drftStw17 = [ ]
        stw17DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 12.5 and i[ 4 ] < 13 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw17DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 12.5 and i[ 4 ] < 13 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw17DrftB4.shape[ 0 ] != 0:
                stw17DrftB4 = np.delete(stw17DrftB4, [ i for (i, v) in enumerate(stw17DrftB4[ :, 5 ]) if v < (
                        np.mean(stw17DrftB4[ :, 5 ]) - np.std(stw17DrftB4[ :, 5 ])) or v > np.mean(
                    stw17DrftB4[ :, 5 ]) + np.std(stw17DrftB4[ :, 5 ]) ], 0)
                drftStw17.append(stw17DrftB4)
        stw17DrftB4 = np.concatenate(drftStw17)

        drftStw18 = [ ]
        stw18DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw18DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 13 and i[ 4 ] < 13.5 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw18DrftB4.shape[ 0 ] != 0:
                stw18DrftB4 = np.delete(stw18DrftB4, [ i for (i, v) in enumerate(stw18DrftB4[ :, 5 ]) if v < (
                        np.mean(stw18DrftB4[ :, 5 ]) - np.std(stw18DrftB4[ :, 5 ])) or v > np.mean(
                    stw18DrftB4[ :, 5 ]) + np.std(stw18DrftB4[ :, 5 ]) ], 0)
                drftStw18.append(stw18DrftB4)
        stw18DrftB4 = np.concatenate(drftStw18)

        drftStw19 = [ ]
        stw19DrftB4 = np.array([ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] > 9 ])
        for dr in range(9, 14):
            stw19DrftB4 = np.array(
                [ i for i in dtNew1 if i[ 4 ] >= 13.5 and i[ 4 ] < 14 and i[ 0 ] > dr and i[ 0 ] <= dr + 1 ])
            if stw19DrftB4.shape[ 0 ] != 0:
                stw19DrftB4 = np.delete(stw19DrftB4, [ i for (i, v) in enumerate(stw19DrftB4[ :, 5 ]) if v < (
                        np.mean(stw19DrftB4[ :, 5 ]) - np.std(stw19DrftB4[ :, 5 ])) or v > np.mean(
                    stw19DrftB4[ :, 5 ]) + np.std(stw19DrftB4[ :, 5 ]) ], 0)
                drftStw19.append(stw19DrftB4)
        stw19DrftB4 = np.concatenate(drftStw19)

        ################################################
        concatS = np.concatenate(
            [ stw8DrftS4, stw9DrftS4, stw10DrftS4, stw11DrftS4, stw12DrftS4, stw13DrftS4, stw14DrftS4, stw15DrftS4,
              stw16DrftS4, stw17DrftS4, stw18DrftS4, stw19DrftS4 ])
        concatB = np.concatenate(
            [ stw8DrftB4, stw9DrftB4, stw10DrftB4, stw11DrftB4, stw12DrftB4, stw13DrftB4, stw14DrftB4, stw15DrftB4,
              stw16DrftB4, stw17DrftB4, stw18DrftB4, stw19DrftB4 ])
        dtNew1 = np.concatenate([ concatB, concatS ])
        # partitions = np.append([concatB,concatS])
        # stdFOC4_89 = np.std(stw8DrftS4[:,5])

        concatS = [ stw8DrftS4, stw9DrftS4, stw10DrftS4, stw11DrftS4, stw12DrftS4, stw13DrftS4, stw14DrftS4, stw15DrftS4,
                    stw16DrftS4, stw17DrftS4, stw18DrftS4, stw19DrftS4 ]
        concatB = [ stw8DrftB4, stw9DrftB4, stw10DrftB4, stw11DrftB4, stw12DrftB4, stw13DrftB4, stw14DrftB4, stw15DrftB4,
                    stw16DrftB4, stw17DrftB4, stw18DrftB4, stw19DrftB4 ]
        # concat =[ stw8DrftS4, stw9DrftS4, stw10DrftS4, stw11DrftS4, stw12DrftS4, stw13DrftS4, stw14DrftS4, stw15DrftS4,
        # stw16DrftS4, stw17DrftS4, stw18DrftS4, stw19DrftS4,
        # stw8DrftB4, stw9DrftB4, stw10DrftB4, stw11DrftB4, stw12DrftB4, stw13DrftB4, stw14DrftB4, stw15DrftB4,
        # stw16DrftB4, stw17DrftB4, stw18DrftB4, stw19DrftB4 ]

        # partitionsX=[k[:,0:5] for k in concat]
        # partitionsY = [ k[ :, 5 ] for k in concat ]
        # draftBigger9 = np.array([ drft for drft in dtNew1 if drft[ 0 ] > 9 ])

        # stw8Drft9 = np.array([ i for i in draftBigger9 if i[ 4 ] >= 12 and i[ 4 ] < 12.5 ])

        # stdFOC9_12 = np.std(stw8Drft9[ :, 5 ])

        seriesX = dtNew1[ :, 0:5 ]
        # seriesX=dtNew[:,0:5]
        # pca = PCA()
        # pca.fit(seriesX)
        # seriesX = pca.transform(seriesX)
        # scaler = preprocessing.StandardScaler()
        # Fit your data on the scaler object
        # scaled_df = scaler.fit_transform(seriesX)
        # seriesX = scaled_df
        # scaled_df = pd.DataFrame(scaled_df, columns=names)
        # dataNew = scaled_df

        # seriesX=seriesX[ 0:, 0:3 ]
        # seriesX=preprocessing.normalize([seriesX])
        newSeriesX = [ ]
        # for x in seriesX:
        # dmWS = x[0] * x[1]
        # waSO = x[2] * x[3]
        # newSeriesX.append([x[0],x[1],x[2],x[3],np.round(dmWS,3),np.round(waSO,3)])
        # newSeriesX = np.array(newSeriesX)
        ##seriesX = newSeriesX

        UnseenSeriesX = dtNew1[ 10000:11000, 0:5 ]
        newUnseenSeriesX = [ ]
        # for x in UnseenSeriesX:
        # dmWS = x[0] * x[1]
        # waSO = x[2] * x[3]
        # newUnseenSeriesX.append([x[0], x[1], x[2], x[3], np.round(dmWS, 3), np.round(waSO, 3)])
        # newUnseenSeriesX = np.array(newUnseenSeriesX)
        # UnseenSeriesX = newUnseenSeriesX
        ##RPM
        # Rpm = dtNew[0:150000,4]
        # unseenRpm = dtNew[160000:161000,4]

        ##FOC
        FOC = dtNew1[ :, 5 ]
        # scaled_df = scaler.fit_transform(FOC)
        # FOC = scaled_df
        # normalized_Y = preprocessing.normalize([FOC])

        unseenFOC = dtNew1[ 10000:11000, 5 ]
        # WS= np.asarray([x for x in dt[ :, 1 ] ])
        # WA = np.asarray([ x for x in dt[ :, 2 ] ])
        # SO = np.asarray([ y for y in dt[ :, 3 ] ])
        # RPM = np.asarray([ y for y in dt[ :, 4 ] ])
        # FOC = np.asarray([ y for y in dt[ :, 5 ] ])
        # POW = np.asarray([ y for y in dt[ :, 6 ] ])
        # DA = np.asarray([ np.nan_to_num(np.float(y)) for y in dt[ :, 13 ] ])
        # DF = np.asarray([ y for y in dt[ :, 0 ] ])
        WaFoc = [ ]
        # newArray = np.append([WS,WA,SO,RPM,POW,DF])
        partionLabels = np.arange(0, 24)
        return seriesX, FOC, UnseenSeriesX, unseenFOC, None, None, None, None, None, None, partionLabels  # draftBigger6[:,0:5],draftSmaller6[:,0:5] , draftBigger6[:,5] , draftSmaller6[:,5]


    def readLarosDataFromCsv(self, data):
        # Load file
        # if self.__class__.__name__ == 'UnseenSeriesReader':
        # dt = data.sample(n=2880).values[ :90000:, 3:23 ]
        # dt = data.values[ 0:, 2:23 ]
        # else:
        # dt = data.values[0:,2:23]
        dt = data.values[ 0:, 2:8 ].astype(float)
        dt = dt[ ~np.isnan(dt).any(axis=1) ]
        WS = np.asarray([ x for x in dt[ :, 0 ] ])
        WA = np.asarray([ x for x in dt[ :, 1 ] ])
        SO = np.asarray([ y for y in dt[ :, 3 ] ])
        RPM = np.asarray([ y for y in dt[ :, 4 ] ])
        FOC = np.asarray([ y for y in dt[ :, 5 ] ])
        POW = np.asarray([ y for y in dt[ :, 5 ] ])
        # DA = np.asarray([ np.nan_to_num(np.float(y)) for y in dt[ :, 13 ] ])
        DF = np.asarray([ np.nan_to_num(np.float(y)) for y in dt[ :, 6 ] ])
        WaFoc = [ ]
        for i in range(0, len(WA)):
            prev = np.append(WA[ i ], FOC[ i ])
            WaFoc.append(prev)

        WaFoc = np.array(WaFoc).reshape(-1, 2)
        # WaFoc = np.array(np.stack([WA,FOC],axis=0))
        firstFOC = [ ]
        secondFOC = [ ]
        thirdFOC = [ ]
        fourthFOC = [ ]
        fifthFOC = [ ]
        for val in WaFoc:
            if val[ 0 ] <= 22.5 or (val[ 0 ] <= 360 and val[ 0 ] > 337.5):
                firstFOC.append(val[ 1 ])
            elif (val[ 0 ] > 22.5 and val[ 0 ] < 67.5) or (val[ 0 ] > 292.5 and val[ 0 ] < 247.5):
                secondFOC.append(val[ 1 ])
            elif (val[ 0 ] > 67.5 and val[ 0 ] < 112.5) or (val[ 0 ] > 247.5 and val[ 0 ] < 292.5):
                thirdFOC.append(val[ 1 ])
            elif (val[ 0 ] > 112.5 and val[ 0 ] < 157.5) or (val[ 0 ] > 202.5 and val[ 0 ] < 247.5):
                fourthFOC.append(val[ 1 ])
            elif val[ 0 ] > 157.5 and val[ 0 ] < 202.5:
                fifthFOC.append(val[ 1 ])

        ##extract FOC depending on wind angle
        import matplotlib.pyplot as plt
        # arraysToplot = np.array(np.stack([firstFOC,secondFOC],axis=0))
        df = pd.DataFrame(np.array(firstFOC),
                          columns=[ '' ])
        boxplot = df.boxplot(column=[ '' ])
        plt.ylabel('M/E FOC (kg/min)')
        plt.title("STD and outliers of FOC with 0< WA <= 22.5")
        plt.show()
        df = pd.DataFrame(np.array(secondFOC),
                          columns=[ '' ])
        boxplot1 = df.boxplot(
            column=[ '' ])
        plt.ylabel('M/E FOC (kg/min)')
        plt.title("STD and outliers of FOC with WA > 22.5")
        plt.show()
        x = 1

        # x = [ math.sin(Lon[ i ] - Lon[ i - 1 ]) * math.cos(Lat[ i ]) for i in range(1, len(Lat)) ]
        # y = [ math.cos(Lat[ i ]) * math.sin(Lat[ i - 1 ])
        # - math.sin(Lat[ i - 1 ]) * math.sin(Lat[ i ]) * math.cos(Lon[ i ] - Lon[ i - 1 ]) for i in
        # range(1, len(Lat)) ]
        # bearing = [ ]

        # bearing.append([ math.atan2(x[ i ], y[ i ]) for i in range(0, len(x)) ])
        # bearing[ 0 ].insert(0, 0)
        # bearing = np.asarray(bearing)[ 0 ]
        # WA = np.asarray([ np.nan_to_num(np.float(y)) for y in dt[ :, 5 ] ])
        return WS, WA
